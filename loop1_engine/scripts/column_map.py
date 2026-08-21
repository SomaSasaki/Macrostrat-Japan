# -*- coding: utf-8 -*-
"""Create a review map and Google Earth KML for Macrostrat columns.

The GSJ ``geo_A.shp`` file stores one polygon record per ``MAJOR_CODE``.  A
review workbook stores the column assignment for each geological unit.  This
module joins the two without a GIS dependency, builds *candidate* column
regions, chooses an interior representative point where possible, and writes:

* a compact PNG intended for embedding in the review workbook; and
* a KML file that can be opened directly in Google Earth.

Column boundaries are not an official GSJ product.  In particular, polygons
for a unit assigned to more than one column must be divided spatially.  We use
exclusive-unit polygons as seeds and assign each shared polygon component to
the nearest seed.  Both output files explicitly describe this as an inferred,
human-reviewable result.

Only geographic (longitude/latitude) Shapefiles can be exported to KML.  The
currently published GSJ 1:50,000 vector packages use JGD2011 geographic
coordinates.  JGD2011 coordinates are used directly for Google Earth review;
the sub-metre datum difference is immaterial for choosing a column point.  A
projected CRS is rejected rather than silently writing invalid KML.
"""

from __future__ import annotations

import argparse
import html
import json
import math
import re
import struct
import textwrap
import unicodedata
import zlib
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence

try:
    from shape_source import read_dbf
except ImportError:  # pragma: no cover - package-style import
    from .shape_source import read_dbf


Point = tuple[float, float]
Color = tuple[int, int, int]


PALETTE: tuple[Color, ...] = (
    (52, 120, 246),
    (234, 88, 73),
    (31, 162, 106),
    (155, 89, 182),
    (242, 153, 32),
    (24, 157, 178),
    (214, 68, 126),
    (105, 113, 125),
)


@dataclass
class ShapeRecord:
    """One Polygon/PolygonZ/PolygonM record from a Shapefile."""

    record_index: int
    rings: list[list[Point]]
    bbox: tuple[float, float, float, float]
    attributes: dict[str, Any] = field(default_factory=dict)


@dataclass
class PolygonComponent:
    """An outer ring and its holes, derived from one Shapefile record."""

    record_index: int
    outer: list[Point]
    holes: list[list[Point]]
    area: float
    major_code: str = ""
    unit_name: str = ""
    eligible_columns: tuple[str, ...] = ()
    assigned_column: str = ""
    assignment_method: str = "unassigned"

    @property
    def bbox(self) -> tuple[float, float, float, float]:
        xs = [p[0] for p in self.outer]
        ys = [p[1] for p in self.outer]
        return min(xs), min(ys), max(xs), max(ys)


@dataclass
class ColumnDefinition:
    col_id: str
    name: str
    lat: float | None = None
    lng: float | None = None


@dataclass
class UnitAssignment:
    unit_id: str
    columns: tuple[str, ...]
    unit_name: str = ""
    reference_name: str = ""
    major_code: str = ""
    row_number: int = 0
    place_names: str = ""


@dataclass
class ColumnRegion:
    definition: ColumnDefinition
    color: Color
    components: list[PolygonComponent]
    candidate_lng: float
    candidate_lat: float
    candidate_method: str
    candidate_inside_region: bool
    shared_component_count: int = 0


@dataclass
class ColumnMapResult:
    png_path: Path
    kml_path: Path
    columns: list[ColumnRegion]
    warnings: list[str]
    source_crs: str
    matched_units: int
    unmatched_units: int
    unassigned_records: int

    def as_dict(self) -> dict[str, Any]:
        return {
            "png_path": str(self.png_path),
            "kml_path": str(self.kml_path),
            "source_crs": self.source_crs,
            "matched_units": self.matched_units,
            "unmatched_units": self.unmatched_units,
            "unassigned_records": self.unassigned_records,
            "warnings": list(self.warnings),
            "columns": [
                {
                    "col_id": region.definition.col_id,
                    "col_name": region.definition.name,
                    "lng": round(region.candidate_lng, 8),
                    "lat": round(region.candidate_lat, 8),
                    "method": region.candidate_method,
                    "inside_region": region.candidate_inside_region,
                    "components": len(region.components),
                    "shared_components": region.shared_component_count,
                    "units": sorted({
                        component.unit_name
                        for component in region.components
                        if component.unit_name
                    }),
                }
                for region in self.columns
            ],
        }


class UnsupportedCrsError(ValueError):
    """Raised when source coordinates cannot safely be written to KML."""


def _clean_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if re.fullmatch(r"[+-]?\d+\.0+", text):
        return text.split(".", 1)[0]
    return text


def _split_ids(value: Any) -> tuple[str, ...]:
    """Split a review ``column_id`` while preserving stable order."""

    text = _clean_id(value)
    if not text:
        return ()
    output: list[str] = []
    for part in re.split(r"\s*[,;]\s*", text):
        part = _clean_id(part)
        if part and part not in output:
            output.append(part)
    return tuple(output)


def _normalise_name(value: Any) -> str:
    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    text = re.sub(r"\([^)]*\)\s*$", "", text)
    return "".join(ch for ch in text if ch.isalnum())


def _field_value(row: dict[str, Any], *names: str) -> Any:
    lowered = {str(key).casefold(): value for key, value in row.items()}
    for name in names:
        if name.casefold() in lowered:
            return lowered[name.casefold()]
    return None


def read_review_workbook(path: str | Path) -> tuple[list[ColumnDefinition], list[UnitAssignment]]:
    """Read the minimal column and unit assignment data from a review workbook."""

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - environment error
        raise RuntimeError("Reading review workbooks requires openpyxl") from exc

    path = Path(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_lookup = {name.casefold(): name for name in workbook.sheetnames}
    column_sheet_name = sheet_lookup.get("columns_review") or sheet_lookup.get("columns")
    unit_sheet_name = sheet_lookup.get("units_review") or sheet_lookup.get("review") or sheet_lookup.get("units")
    if not column_sheet_name or not unit_sheet_name:
        raise ValueError(f"Review workbook lacks units/columns sheets: {path}")

    def rows_for(sheet_name: str) -> Iterable[tuple[int, dict[str, Any]]]:
        sheet = workbook[sheet_name]
        values = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(values)]
        except StopIteration:
            return
        for row_number, values_row in enumerate(values, start=2):
            yield row_number, dict(zip(headers, values_row))

    columns: list[ColumnDefinition] = []
    for _, row in rows_for(column_sheet_name):
        col_id = _clean_id(_field_value(row, "col_id", "column_id"))
        if not col_id:
            continue
        name = str(_field_value(row, "col_name", "column_name") or col_id).strip()
        lat = _as_float(_field_value(row, "lat", "latitude"))
        lng = _as_float(_field_value(row, "lng", "lon", "longitude"))
        columns.append(ColumnDefinition(col_id, name, lat=lat, lng=lng))

    assignments: list[UnitAssignment] = []
    for row_number, row in rows_for(unit_sheet_name):
        col_ids = _split_ids(_field_value(row, "column_id", "col_id"))
        if not col_ids:
            continue
        major_value = _field_value(
            row,
            "REF_shape_major_code",
            "shape_major_code",
            "major_code",
            "REF_major_code",
        )
        major_code = _clean_id(major_value)
        if not major_code:
            source = str(_field_value(row, "REF_shape_source") or "")
            match = re.search(r"MAJOR_CODE\s*[:=]\s*([\w.-]+)", source, re.I)
            major_code = _clean_id(match.group(1)) if match else ""
        assignments.append(
            UnitAssignment(
                unit_id=_clean_id(_field_value(row, "unit_id")),
                columns=col_ids,
                unit_name=str(_field_value(row, "unit_name") or "").strip(),
                reference_name=str(
                    _field_value(
                        row,
                        "REF_shape_unit_name",
                        "REF_unit_name_en",
                        "shape_unit_name",
                    )
                    or ""
                ).strip(),
                major_code=major_code,
                row_number=row_number,
            )
        )
    workbook.close()
    if not columns:
        raise ValueError(f"No column definitions found in {path}")
    if not assignments:
        raise ValueError(f"No unit-to-column assignments found in {path}")
    return columns, assignments


def _as_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        result = float(value)
    except (TypeError, ValueError):
        return None
    return result if math.isfinite(result) else None


def read_polygon_shapefile(path: str | Path) -> tuple[list[ShapeRecord], tuple[float, float, float, float]]:
    """Read Polygon, PolygonZ, or PolygonM records using the public SHP format."""

    path = Path(path)
    records: list[ShapeRecord] = []
    with path.open("rb") as stream:
        header = stream.read(100)
        if len(header) != 100 or struct.unpack(">I", header[:4])[0] != 9994:
            raise ValueError(f"Invalid Shapefile header: {path}")
        shape_type = struct.unpack("<i", header[32:36])[0]
        if shape_type not in (5, 15, 25):
            raise ValueError(f"geo_A must be a polygon Shapefile, got type {shape_type}: {path}")
        dataset_bbox = struct.unpack("<4d", header[36:68])

        while True:
            record_header = stream.read(8)
            if not record_header:
                break
            if len(record_header) != 8:
                raise ValueError(f"Incomplete Shapefile record header: {path}")
            record_number, words = struct.unpack(">2i", record_header)
            content = stream.read(words * 2)
            if len(content) != words * 2:
                raise ValueError(f"Incomplete Shapefile record {record_number}: {path}")
            record_type = struct.unpack("<i", content[:4])[0]
            if record_type == 0:
                continue
            if record_type not in (5, 15, 25) or len(content) < 44:
                continue
            bbox = struct.unpack("<4d", content[4:36])
            n_parts, n_points = struct.unpack("<2i", content[36:44])
            if n_parts < 1 or n_points < 4:
                continue
            part_end = 44 + 4 * n_parts
            point_end = part_end + 16 * n_points
            if point_end > len(content):
                raise ValueError(f"Invalid part/point counts in record {record_number}: {path}")
            starts = list(struct.unpack(f"<{n_parts}i", content[44:part_end]))
            points = [
                struct.unpack("<2d", content[part_end + i * 16:part_end + (i + 1) * 16])
                for i in range(n_points)
            ]
            rings: list[list[Point]] = []
            for part_index, start in enumerate(starts):
                end = starts[part_index + 1] if part_index + 1 < n_parts else n_points
                ring = points[start:end]
                if len(ring) >= 4:
                    rings.append(_close_ring(ring))
            if rings:
                records.append(ShapeRecord(record_number, rings, bbox))
    return records, dataset_bbox


def _close_ring(ring: Sequence[Point]) -> list[Point]:
    output = [(float(x), float(y)) for x, y in ring]
    if output and output[0] != output[-1]:
        output.append(output[0])
    return output


def _signed_area(ring: Sequence[Point]) -> float:
    return 0.5 * sum(
        x1 * y2 - x2 * y1
        for (x1, y1), (x2, y2) in zip(ring, ring[1:])
    )


def _ring_bbox(ring: Sequence[Point]) -> tuple[float, float, float, float]:
    xs = [point[0] for point in ring]
    ys = [point[1] for point in ring]
    return min(xs), min(ys), max(xs), max(ys)


def _bbox_contains(bbox: tuple[float, float, float, float], point: Point) -> bool:
    return bbox[0] <= point[0] <= bbox[2] and bbox[1] <= point[1] <= bbox[3]


def _point_in_ring(point: Point, ring: Sequence[Point]) -> bool:
    x, y = point
    inside = False
    for (x1, y1), (x2, y2) in zip(ring, ring[1:]):
        if (y1 > y) != (y2 > y):
            cross_x = x1 + (y - y1) * (x2 - x1) / (y2 - y1)
            if cross_x > x:
                inside = not inside
    return inside


def point_in_component(point: Point, component: PolygonComponent) -> bool:
    return _point_in_ring(point, component.outer) and not any(
        _point_in_ring(point, hole) for hole in component.holes
    )


def _rings_to_components(record: ShapeRecord) -> list[PolygonComponent]:
    """Group clockwise SHP outer rings with their counter-clockwise holes."""

    ring_data = [(ring, _signed_area(ring), _ring_bbox(ring)) for ring in record.rings]
    outers = [(ring, area, bbox) for ring, area, bbox in ring_data if area < 0]
    holes = [(ring, area, bbox) for ring, area, bbox in ring_data if area >= 0]
    if not outers:
        # Some non-conforming producers reverse all rings.  The largest ring is
        # still a defensible outer boundary; remaining rings are tested below.
        largest = max(ring_data, key=lambda item: abs(item[1]))
        outers = [largest]
        holes = [item for item in ring_data if item is not largest]

    components = [
        PolygonComponent(record.record_index, ring, [], abs(area))
        for ring, area, _ in outers
    ]
    outer_boxes = [_ring_bbox(component.outer) for component in components]
    for hole, _, _ in holes:
        probe = hole[0]
        containers = [
            (index, components[index].area)
            for index, bbox in enumerate(outer_boxes)
            if _bbox_contains(bbox, probe) and _point_in_ring(probe, components[index].outer)
        ]
        if containers:
            index = min(containers, key=lambda item: item[1])[0]
            components[index].holes.append(hole)
            components[index].area = max(0.0, components[index].area - abs(_signed_area(hole)))
        else:
            # A positive ring outside all negative rings is treated as another
            # outer boundary.  This is preferable to silently losing geometry.
            components.append(
                PolygonComponent(record.record_index, list(reversed(hole)), [], abs(_signed_area(hole)))
            )
            outer_boxes.append(_ring_bbox(hole))
    return components


def _component_point(component: PolygonComponent) -> Point:
    """Return an interior point, preferring an area centroid when valid."""

    ring = component.outer
    signed = _signed_area(ring)
    if signed:
        cx = sum(
            (x1 + x2) * (x1 * y2 - x2 * y1)
            for (x1, y1), (x2, y2) in zip(ring, ring[1:])
        ) / (6 * signed)
        cy = sum(
            (y1 + y2) * (x1 * y2 - x2 * y1)
            for (x1, y1), (x2, y2) in zip(ring, ring[1:])
        ) / (6 * signed)
        if point_in_component((cx, cy), component):
            return cx, cy

    xmin, ymin, xmax, ymax = component.bbox
    centre = ((xmin + xmax) / 2, (ymin + ymax) / 2)
    if point_in_component(centre, component):
        return centre

    # Equivalent in spirit to a GIS representative_point(): find the longest
    # interior segment on several horizontal scan lines and use its midpoint.
    best: tuple[float, Point] | None = None
    for fraction in (0.5, 0.375, 0.625, 0.25, 0.75, 0.125, 0.875):
        y = ymin + (ymax - ymin) * fraction
        intersections: list[float] = []
        for candidate_ring in [component.outer, *component.holes]:
            for (x1, y1), (x2, y2) in zip(candidate_ring, candidate_ring[1:]):
                if (y1 > y) != (y2 > y):
                    intersections.append(x1 + (y - y1) * (x2 - x1) / (y2 - y1))
        intersections.sort()
        for left, right in zip(intersections[0::2], intersections[1::2]):
            point = ((left + right) / 2, y)
            width = right - left
            if width > 0 and point_in_component(point, component) and (best is None or width > best[0]):
                best = (width, point)
    if best:
        return best[1]

    # Degenerate last resort: an outer-ring point nudged towards the bbox centre.
    x, y = ring[0]
    for factor in (0.001, 0.01, 0.05, 0.1, 0.25):
        candidate = (x + (centre[0] - x) * factor, y + (centre[1] - y) * factor)
        if point_in_component(candidate, component):
            return candidate
    return centre


def _read_crs(shp_path: Path, bbox: tuple[float, float, float, float]) -> tuple[str, list[str]]:
    prj_path = shp_path.with_suffix(".prj")
    warnings: list[str] = []
    if prj_path.exists():
        text = prj_path.read_text(encoding="utf-8", errors="replace").strip()
        upper = text.upper()
        if "GEOGCS" not in upper or "DEGREE" not in upper:
            raise UnsupportedCrsError(
                f"KML requires longitude/latitude coordinates; projected CRS found in {prj_path}"
            )
        if "JGD_2011" in upper or "JGD2011" in upper:
            return "JGD2011 geographic (used directly as WGS84 for review)", warnings
        if "JGD_2000" in upper or "JGD2000" in upper:
            warnings.append("JGD2000 coordinates are used directly for review; verify if sub-metre precision matters.")
            return "JGD2000 geographic (used directly as WGS84 for review)", warnings
        if "WGS_1984" in upper or "WGS 84" in upper or "WGS84" in upper:
            return "WGS84 geographic", warnings
        warnings.append("Unrecognised geographic datum; coordinates are used directly in KML for visual review.")
        return "Unrecognised geographic longitude/latitude", warnings

    if -180 <= bbox[0] <= bbox[2] <= 180 and -90 <= bbox[1] <= bbox[3] <= 90:
        warnings.append("No .prj file was found; coordinate ranges look geographic and are assumed WGS84.")
        return "Assumed WGS84 geographic (missing .prj)", warnings
    raise UnsupportedCrsError(f"No .prj file and coordinates are not longitude/latitude: {shp_path}")


def _attach_attributes(records: list[ShapeRecord], dbf_path: Path) -> None:
    by_index = {int(row.get("_record_index", 0)): row for row in read_dbf(dbf_path)}
    for record in records:
        record.attributes = by_index.get(record.record_index, {})


def _record_code(record: ShapeRecord) -> str:
    return _clean_id(record.attributes.get("MAJOR_CODE"))


def _record_name(record: ShapeRecord) -> str:
    return str(record.attributes.get("LEGEND03E") or record.attributes.get("LEGEND03") or "").strip()


def _match_assignments(
    records: list[ShapeRecord], assignments: list[UnitAssignment]
) -> tuple[dict[int, tuple[str, ...]], int, list[UnitAssignment], list[str]]:
    """Join review rows to SHP records by code, name, then safe ordered fallback."""

    mapping: dict[int, tuple[str, ...]] = {}
    matched_rows: set[int] = set()
    warnings: list[str] = []
    by_code: dict[str, list[ShapeRecord]] = {}
    by_name: dict[str, list[ShapeRecord]] = {}
    for record in records:
        code = _record_code(record)
        name = _normalise_name(_record_name(record))
        if code:
            by_code.setdefault(code, []).append(record)
        if name:
            by_name.setdefault(name, []).append(record)

    used_records: set[int] = set()
    for assignment_index, assignment in enumerate(assignments):
        candidates: list[ShapeRecord] = []
        if assignment.major_code:
            candidates = by_code.get(assignment.major_code, [])
        if not candidates:
            for raw_name in (assignment.unit_name, assignment.reference_name):
                name = _normalise_name(raw_name)
                if name and name in by_name:
                    candidates = by_name[name]
                    break
        candidate = next((item for item in candidates if item.record_index not in used_records), None)
        if candidate is not None:
            mapping[candidate.record_index] = assignment.columns
            used_records.add(candidate.record_index)
            matched_rows.add(assignment_index)

    unmatched = [a for index, a in enumerate(assignments) if index not in matched_rows]

    # Existing review v1 files do not expose MAJOR_CODE.  GSJ geo_A and the
    # generated review rows retain the same MAJOR_CODE order, so ordered fallback
    # is permitted only when the remaining cardinalities match exactly.
    remaining_records = [
        record
        for record in records
        if record.record_index not in used_records
        and _record_code(record)
        and _record_name(record)
    ]
    if unmatched and len(unmatched) == len(remaining_records):
        for assignment, record in zip(unmatched, remaining_records):
            mapping[record.record_index] = assignment.columns
            used_records.add(record.record_index)
        warnings.append(
            f"Matched {len(unmatched)} unit(s) by GSJ record order because the review file has no MAJOR_CODE."
        )
        matched_rows.update(index for index, _ in enumerate(assignments))
        unmatched = []

    if unmatched:
        labels = [assignment.unit_id or assignment.unit_name or f"row {assignment.row_number}" for assignment in unmatched]
        warnings.append("Unmatched review units: " + ", ".join(labels))
    return mapping, len(matched_rows), unmatched, warnings


def _directional_anchor(name: str, bbox: tuple[float, float, float, float]) -> Point | None:
    text = unicodedata.normalize("NFKC", name).casefold()
    xmin, ymin, xmax, ymax = bbox
    x, y = (xmin + xmax) / 2, (ymin + ymax) / 2
    found = False
    if any(token in text for token in ("west", "western", "西部", "西側")):
        x, found = xmin + (xmax - xmin) * 0.22, True
    elif any(token in text for token in ("east", "eastern", "東部", "東側")):
        x, found = xmin + (xmax - xmin) * 0.78, True
    if any(token in text for token in ("north", "northern", "北部", "北側")):
        y, found = ymin + (ymax - ymin) * 0.78, True
    elif any(token in text for token in ("south", "southern", "南部", "南側")):
        y, found = ymin + (ymax - ymin) * 0.22, True
    elif any(token in text for token in ("central", "center", "centre", "中央")):
        found = True
    return (x, y) if found else None


def _weighted_anchor(components: Sequence[PolygonComponent]) -> Point | None:
    if not components:
        return None
    weighted = [(_component_point(component), max(component.area, 1e-20)) for component in components]
    total = sum(weight for _, weight in weighted)
    return (
        sum(point[0] * weight for point, weight in weighted) / total,
        sum(point[1] * weight for point, weight in weighted) / total,
    )


def _build_components(
    records: list[ShapeRecord], record_columns: dict[int, tuple[str, ...]]
) -> list[PolygonComponent]:
    components: list[PolygonComponent] = []
    for record in records:
        code = _record_code(record)
        name = _record_name(record)
        assignments = record_columns.get(record.record_index, ())
        for component in _rings_to_components(record):
            component.major_code = code
            component.unit_name = name
            component.eligible_columns = assignments
            components.append(component)
    return components


def _assign_components(
    components: list[PolygonComponent],
    columns: list[ColumnDefinition],
    bbox: tuple[float, float, float, float],
) -> dict[str, Point]:
    col_ids = {column.col_id for column in columns}
    exclusive: dict[str, list[PolygonComponent]] = {column.col_id: [] for column in columns}
    for component in components:
        eligible = tuple(col_id for col_id in component.eligible_columns if col_id in col_ids)
        component.eligible_columns = eligible
        if len(eligible) == 1:
            component.assigned_column = eligible[0]
            component.assignment_method = "exclusive unit"
            exclusive[eligible[0]].append(component)

    anchors: dict[str, Point] = {}
    coordinate_counts: dict[Point, int] = {}
    for column in columns:
        if column.lat is not None and column.lng is not None:
            key = (column.lng, column.lat)
            coordinate_counts[key] = coordinate_counts.get(key, 0) + 1
    for index, column in enumerate(columns):
        seed = _weighted_anchor(exclusive[column.col_id])
        if seed is not None:
            anchors[column.col_id] = seed
            continue
        explicit = (column.lng, column.lat) if column.lng is not None and column.lat is not None else None
        if explicit is not None and coordinate_counts.get(explicit, 0) == 1:
            anchors[column.col_id] = explicit
            continue
        directional = _directional_anchor(column.name, bbox)
        if directional is not None:
            anchors[column.col_id] = directional
            continue
        angle = 2 * math.pi * index / max(1, len(columns))
        anchors[column.col_id] = (
            (bbox[0] + bbox[2]) / 2 + math.cos(angle) * (bbox[2] - bbox[0]) * 0.28,
            (bbox[1] + bbox[3]) / 2 + math.sin(angle) * (bbox[3] - bbox[1]) * 0.28,
        )

    latitude = (bbox[1] + bbox[3]) / 2
    x_scale = max(0.1, math.cos(math.radians(latitude)))
    for component in components:
        if len(component.eligible_columns) <= 1:
            continue
        point = _component_point(component)
        assigned = min(
            component.eligible_columns,
            key=lambda col_id: (
                ((point[0] - anchors[col_id][0]) * x_scale) ** 2
                + (point[1] - anchors[col_id][1]) ** 2
            ),
        )
        component.assigned_column = assigned
        component.assignment_method = "shared unit; nearest exclusive-unit seed"
    return anchors


def _column_regions(
    columns: list[ColumnDefinition],
    components: list[PolygonComponent],
    bbox: tuple[float, float, float, float],
) -> list[ColumnRegion]:
    regions: list[ColumnRegion] = []
    for index, column in enumerate(columns):
        assigned = [component for component in components if component.assigned_column == column.col_id]
        explicit = (
            (column.lng, column.lat)
            if column.lat is not None and column.lng is not None
            else None
        )
        containing = (
            next((component for component in assigned if point_in_component(explicit, component)), None)
            if explicit is not None
            else None
        )
        if containing is not None:
            point = explicit
            inside = True
            method = "PDF-constrained candidate inside assigned polygon"
        elif assigned:
            largest = max(assigned, key=lambda component: component.area)
            point = _component_point(largest)
            inside = point_in_component(point, largest)
            method = "largest assigned polygon interior point"
        elif explicit is not None:
            point = explicit
            inside = False
            method = "FALLBACK: canonical coordinate; no assigned polygon"
        else:
            point = ((bbox[0] + bbox[2]) / 2, (bbox[1] + bbox[3]) / 2)
            inside = False
            method = "FALLBACK: Shapefile bbox centre; no assigned polygon"
        regions.append(
            ColumnRegion(
                definition=column,
                color=PALETTE[index % len(PALETTE)],
                components=assigned,
                candidate_lng=point[0],
                candidate_lat=point[1],
                candidate_method=method,
                candidate_inside_region=inside,
                shared_component_count=sum(
                    component.assignment_method.startswith("shared unit") for component in assigned
                ),
            )
        )
    return regions


def generate_column_map(
    review_path: str | Path,
    shp_path: str | Path,
    png_path: str | Path,
    kml_path: str | Path,
    *,
    width: int = 1200,
    height: int = 900,
    title: str | None = None,
) -> ColumnMapResult:
    """Generate PNG/KML review artifacts and return their candidate metadata."""

    review_path = Path(review_path)
    shp_path = Path(shp_path)
    png_path = Path(png_path)
    kml_path = Path(kml_path)
    if shp_path.suffix.casefold() != ".shp":
        raise ValueError(f"Expected a .shp file: {shp_path}")
    dbf_path = shp_path.with_suffix(".dbf")
    if not dbf_path.exists():
        raise FileNotFoundError(f"Matching DBF not found: {dbf_path}")

    columns, assignments = read_review_workbook(review_path)
    records, bbox = read_polygon_shapefile(shp_path)
    source_crs, warnings = _read_crs(shp_path, bbox)
    _attach_attributes(records, dbf_path)
    record_columns, matched, unmatched, match_warnings = _match_assignments(records, assignments)
    warnings.extend(match_warnings)
    components = _build_components(records, record_columns)
    _assign_components(components, columns, bbox)
    regions = _column_regions(columns, components, bbox)

    unknown_column_ids = sorted(
        {
            col_id
            for assignment in assignments
            for col_id in assignment.columns
            if col_id not in {column.col_id for column in columns}
        }
    )
    if unknown_column_ids:
        warnings.append("Unit rows refer to undefined column IDs: " + ", ".join(unknown_column_ids))
    for region in regions:
        if not region.candidate_inside_region:
            warnings.append(
                f"{region.definition.col_id}: representative point is a visible fallback and needs review."
            )

    png_path.parent.mkdir(parents=True, exist_ok=True)
    kml_path.parent.mkdir(parents=True, exist_ok=True)
    map_title = title or f"Column review: {review_path.stem}"
    write_map_png(png_path, components, regions, bbox, width=width, height=height, title=map_title)
    write_kml(kml_path, regions, source_crs, warnings, document_name=map_title)
    return ColumnMapResult(
        png_path=png_path,
        kml_path=kml_path,
        columns=regions,
        warnings=warnings,
        source_crs=source_crs,
        matched_units=matched,
        unmatched_units=len(unmatched),
        unassigned_records=sum(1 for record in records if record.record_index not in record_columns),
    )


# ---------------------------------------------------------------------------
# Dependency-free PNG renderer
# ---------------------------------------------------------------------------


class _Canvas:
    def __init__(self, width: int, height: int, background: Color = (255, 255, 255)):
        self.width = width
        self.height = height
        self.data = bytearray(background * (width * height))

    def pixel(self, x: int, y: int, color: Color) -> None:
        if 0 <= x < self.width and 0 <= y < self.height:
            offset = (y * self.width + x) * 3
            self.data[offset:offset + 3] = bytes(color)

    def rectangle(self, x0: int, y0: int, x1: int, y1: int, color: Color) -> None:
        x0, x1 = sorted((max(0, x0), min(self.width - 1, x1)))
        y0, y1 = sorted((max(0, y0), min(self.height - 1, y1)))
        row = bytes(color) * max(0, x1 - x0 + 1)
        for y in range(y0, y1 + 1):
            offset = (y * self.width + x0) * 3
            self.data[offset:offset + len(row)] = row

    def line(self, x0: int, y0: int, x1: int, y1: int, color: Color, width: int = 1) -> None:
        dx, sx = abs(x1 - x0), 1 if x0 < x1 else -1
        dy, sy = -abs(y1 - y0), 1 if y0 < y1 else -1
        error = dx + dy
        while True:
            radius = max(0, width // 2)
            self.rectangle(x0 - radius, y0 - radius, x0 + radius, y0 + radius, color)
            if x0 == x1 and y0 == y1:
                break
            twice = 2 * error
            if twice >= dy:
                error += dy
                x0 += sx
            if twice <= dx:
                error += dx
                y0 += sy

    def circle(self, x: int, y: int, radius: int, color: Color) -> None:
        for py in range(y - radius, y + radius + 1):
            extent = int(math.sqrt(max(0, radius * radius - (py - y) ** 2)))
            self.rectangle(x - extent, py, x + extent, py, color)

    def fill_even_odd(self, rings: Sequence[Sequence[tuple[int, int]]], color: Color) -> None:
        intersections_by_y: dict[int, list[float]] = {}
        for ring in rings:
            for (x1, y1), (x2, y2) in zip(ring, ring[1:]):
                if y1 == y2:
                    continue
                low, high = sorted((y1, y2))
                start = max(0, int(math.ceil(low - 0.5)))
                stop = min(self.height - 1, int(math.ceil(high - 0.5)) - 1)
                for y in range(start, stop + 1):
                    scan = y + 0.5
                    x = x1 + (scan - y1) * (x2 - x1) / (y2 - y1)
                    intersections_by_y.setdefault(y, []).append(x)
        for y, crossings in intersections_by_y.items():
            crossings.sort()
            for left, right in zip(crossings[0::2], crossings[1::2]):
                self.rectangle(int(math.ceil(left)), y, int(math.floor(right)), y, color)

    def save_png(self, path: Path) -> None:
        raw = b"".join(
            b"\x00" + bytes(self.data[y * self.width * 3:(y + 1) * self.width * 3])
            for y in range(self.height)
        )

        def chunk(kind: bytes, payload: bytes) -> bytes:
            return (
                struct.pack(">I", len(payload))
                + kind
                + payload
                + struct.pack(">I", zlib.crc32(kind + payload) & 0xFFFFFFFF)
            )

        content = (
            b"\x89PNG\r\n\x1a\n"
            + chunk(b"IHDR", struct.pack(">IIBBBBB", self.width, self.height, 8, 2, 0, 0, 0))
            + chunk(b"IDAT", zlib.compress(raw, 9))
            + chunk(b"IEND", b"")
        )
        path.write_bytes(content)


_FONT: dict[str, tuple[str, ...]] = {
    "A": ("01110", "10001", "10001", "11111", "10001", "10001", "10001"),
    "B": ("11110", "10001", "10001", "11110", "10001", "10001", "11110"),
    "C": ("01111", "10000", "10000", "10000", "10000", "10000", "01111"),
    "D": ("11110", "10001", "10001", "10001", "10001", "10001", "11110"),
    "E": ("11111", "10000", "10000", "11110", "10000", "10000", "11111"),
    "F": ("11111", "10000", "10000", "11110", "10000", "10000", "10000"),
    "G": ("01111", "10000", "10000", "10111", "10001", "10001", "01111"),
    "H": ("10001", "10001", "10001", "11111", "10001", "10001", "10001"),
    "I": ("11111", "00100", "00100", "00100", "00100", "00100", "11111"),
    "J": ("00111", "00010", "00010", "00010", "10010", "10010", "01100"),
    "K": ("10001", "10010", "10100", "11000", "10100", "10010", "10001"),
    "L": ("10000", "10000", "10000", "10000", "10000", "10000", "11111"),
    "M": ("10001", "11011", "10101", "10101", "10001", "10001", "10001"),
    "N": ("10001", "11001", "10101", "10011", "10001", "10001", "10001"),
    "O": ("01110", "10001", "10001", "10001", "10001", "10001", "01110"),
    "P": ("11110", "10001", "10001", "11110", "10000", "10000", "10000"),
    "Q": ("01110", "10001", "10001", "10001", "10101", "10010", "01101"),
    "R": ("11110", "10001", "10001", "11110", "10100", "10010", "10001"),
    "S": ("01111", "10000", "10000", "01110", "00001", "00001", "11110"),
    "T": ("11111", "00100", "00100", "00100", "00100", "00100", "00100"),
    "U": ("10001", "10001", "10001", "10001", "10001", "10001", "01110"),
    "V": ("10001", "10001", "10001", "10001", "10001", "01010", "00100"),
    "W": ("10001", "10001", "10001", "10101", "10101", "10101", "01010"),
    "X": ("10001", "10001", "01010", "00100", "01010", "10001", "10001"),
    "Y": ("10001", "10001", "01010", "00100", "00100", "00100", "00100"),
    "Z": ("11111", "00001", "00010", "00100", "01000", "10000", "11111"),
    "0": ("01110", "10001", "10011", "10101", "11001", "10001", "01110"),
    "1": ("00100", "01100", "00100", "00100", "00100", "00100", "01110"),
    "2": ("01110", "10001", "00001", "00010", "00100", "01000", "11111"),
    "3": ("11110", "00001", "00001", "01110", "00001", "00001", "11110"),
    "4": ("00010", "00110", "01010", "10010", "11111", "00010", "00010"),
    "5": ("11111", "10000", "10000", "11110", "00001", "00001", "11110"),
    "6": ("01110", "10000", "10000", "11110", "10001", "10001", "01110"),
    "7": ("11111", "00001", "00010", "00100", "01000", "01000", "01000"),
    "8": ("01110", "10001", "10001", "01110", "10001", "10001", "01110"),
    "9": ("01110", "10001", "10001", "01111", "00001", "00001", "01110"),
    "-": ("00000", "00000", "00000", "11111", "00000", "00000", "00000"),
    ".": ("00000", "00000", "00000", "00000", "00000", "00110", "00110"),
    ":": ("00000", "00110", "00110", "00000", "00110", "00110", "00000"),
    "/": ("00001", "00010", "00100", "01000", "10000", "00000", "00000"),
    "[": ("01110", "01000", "01000", "01000", "01000", "01000", "01110"),
    "]": ("01110", "00010", "00010", "00010", "00010", "00010", "01110"),
    "*": ("00000", "10101", "01110", "11111", "01110", "10101", "00000"),
    " ": ("00000",) * 7,
}


def _draw_text(canvas: _Canvas, x: int, y: int, text: str, color: Color, scale: int = 2) -> None:
    cursor = x
    for char in unicodedata.normalize("NFKC", text).upper():
        glyph = _FONT.get(char, ("11111", "10001", "00110", "00110", "00110", "10001", "11111"))
        for gy, row in enumerate(glyph):
            for gx, value in enumerate(row):
                if value == "1":
                    canvas.rectangle(
                        cursor + gx * scale,
                        y + gy * scale,
                        cursor + (gx + 1) * scale - 1,
                        y + (gy + 1) * scale - 1,
                        color,
                    )
        cursor += 6 * scale
def _pastel(color: Color) -> Color:
    return tuple(int(channel * 0.58 + 255 * 0.42) for channel in color)  # type: ignore[return-value]


def _unit_label_key(value: object) -> str:
    """地図ラベル用の比較キー。大小文字と空白の違いは同じユニットとみなす。

    2026-08-13: bootstrapが "Oritsumedake Fan Deposits"、レビュー側が
    "Oritsumedake fan deposits" を持つと、同じ地層に別コードが振られ、
    凡例にも地図にも二重に出ていた。
    """

    return " ".join(str(value or "").split()).casefold()


def _region_unit_names(region: ColumnRegion) -> list[str]:
    """表示名を1つに畳んで返す（大小文字違いは最初に出た綴りへ寄せる）。"""

    by_key: dict[str, str] = {}
    for component in region.components:
        name = str(component.unit_name or "").strip()
        key = _unit_label_key(name)
        if key and key not in by_key:
            by_key[key] = name
    return sorted(by_key.values(), key=_unit_label_key)


def _unit_legend_lines(code: str, name: str, max_chars: int) -> list[str]:
    chunks = textwrap.wrap(
        name,
        width=max(8, max_chars - 4),
        break_long_words=False,
        break_on_hyphens=False,
    ) or [name]
    return [f"{code} {chunks[0]}", *(f"    {chunk}" for chunk in chunks[1:])]


def write_map_png(
    path: str | Path,
    components: Sequence[PolygonComponent],
    regions: Sequence[ColumnRegion],
    bbox: tuple[float, float, float, float],
    *,
    width: int = 1200,
    height: int = 900,
    title: str = "Candidate column map",
    place_name_pins: Sequence[dict] | None = None,
    centroid_pins: Sequence[dict] | None = None,
    underground_units: Sequence[str] | None = None,
) -> None:
    """Render a candidate map with unit codes and a complete unit legend."""

    if width < 700 or height < 500:
        raise ValueError("Map PNG must be at least 700 x 500 pixels for readable labels")
    # 大小文字違いを同一ユニットとして1コードに畳む。
    display_by_key: dict[str, str] = {}
    for region in regions:
        for name in _region_unit_names(region):
            display_by_key.setdefault(_unit_label_key(name), name)
    unit_names = [display_by_key[key] for key in sorted(display_by_key)]
    unit_codes = {
        key: f"U{index:02d}" for index, key in enumerate(sorted(display_by_key), start=1)
    }
    legend_width = min(400, max(300, width // 3))
    max_legend_chars = max(12, (legend_width - 62) // 6)
    legend_lines = sum(
        3 + sum(
            len(_unit_legend_lines(unit_codes[_unit_label_key(name)], name, max_legend_chars))
            for name in _region_unit_names(region)
        )
        for region in regions
    )
    height = max(height, 220 + legend_lines * 11)
    canvas = _Canvas(width, height, (250, 251, 253))
    left, right, top, bottom = 64, width - legend_width - 32, 74, height - 68
    xmin, ymin, xmax, ymax = bbox
    mean_lat = (ymin + ymax) / 2
    longitude_factor = max(0.1, math.cos(math.radians(mean_lat)))
    display_width = max((xmax - xmin) * longitude_factor, 1e-12)
    display_height = max(ymax - ymin, 1e-12)
    scale = min((right - left) / display_width, (bottom - top) / display_height)
    drawn_width = display_width * scale
    drawn_height = display_height * scale
    x_offset = left + ((right - left) - drawn_width) / 2
    y_offset = top + ((bottom - top) - drawn_height) / 2

    def project(point: Point) -> tuple[int, int]:
        x = x_offset + (point[0] - xmin) * longitude_factor * scale
        y = y_offset + drawn_height - (point[1] - ymin) * scale
        return int(round(x)), int(round(y))

    canvas.rectangle(left - 2, top - 2, right + 2, bottom + 2, (72, 79, 91))
    canvas.rectangle(left, top, right, bottom, (239, 243, 247))
    color_by_column = {region.definition.col_id: _pastel(region.color) for region in regions}
    color_map = {region.definition.col_id: region.color for region in regions}

    # Largest polygons first makes narrow or small units remain visible.
    for component in sorted(components, key=lambda item: item.area, reverse=True):
        fill = color_by_column.get(component.assigned_column, (220, 224, 229))
        pixel_rings = [[project(point) for point in ring] for ring in [component.outer, *component.holes]]
        canvas.fill_even_odd(pixel_rings, fill)
        for ring in pixel_rings:
            for start, end in zip(ring, ring[1:]):
                canvas.line(*start, *end, (115, 121, 130), 1)

    # Label each geological unit once per Column, using a compact code so the
    # map remains readable.  The full English name is in the side legend and
    # every KML polygon remains individually clickable.
    occupied: list[tuple[int, int]] = []
    for region in regions:
        by_name: dict[str, list[PolygonComponent]] = {}
        for component in region.components:
            if component.unit_name:
                by_name.setdefault(_unit_label_key(component.unit_name), []).append(component)
        for name in sorted(by_name):
            component = max(by_name[name], key=lambda item: item.area)
            x, y = project(_component_point(component))
            if any((x - px) ** 2 + (y - py) ** 2 < 18 ** 2 for px, py in occupied):
                continue
            code = unit_codes[name]
            label_width = len(code) * 6 + 4
            canvas.rectangle(x - 2, y - 2, x + label_width, y + 10, (255, 255, 255))
            canvas.rectangle(x - 2, y - 2, x + label_width, y - 1, region.color)
            _draw_text(canvas, x, y, code, (31, 36, 45), 1)
            occupied.append((x, y))

    _draw_text(canvas, 30, 25, title[:72], (31, 36, 45), 2)
    _draw_text(canvas, width - legend_width + 5, 76, "COLOR = COLUMN / UXX = UNIT", (31, 36, 45), 1)
    legend_y = 112
    for region in regions:
        x, y = project((region.candidate_lng, region.candidate_lat))
        canvas.circle(x, y, 9, (255, 255, 255))
        canvas.circle(x, y, 6, region.color)
        canvas.line(x - 11, y, x + 11, y, (25, 25, 25), 1)
        canvas.line(x, y - 11, x, y + 11, (25, 25, 25), 1)

        canvas.rectangle(width - legend_width + 8, legend_y, width - legend_width + 34, legend_y + 18, _pastel(region.color))
        canvas.rectangle(width - legend_width + 8, legend_y, width - legend_width + 34, legend_y + 1, region.color)
        label = f"{region.definition.col_id} {region.definition.name}"
        if not region.candidate_inside_region:
            label += " [FALLBACK]"
        _draw_text(canvas, width - legend_width + 44, legend_y + 2, label[:34], (31, 36, 45), 1)
        _draw_text(
            canvas,
            width - legend_width + 44,
            legend_y + 18,
            f"{region.candidate_lng:.5f}, {region.candidate_lat:.5f}",
            (79, 86, 98),
            1,
        )
        legend_y += 42
        names = _region_unit_names(region)
        _draw_text(
            canvas,
            width - legend_width + 12,
            legend_y,
            f"GEOLOGICAL UNITS ({len(names)})",
            region.color,
            1,
        )
        legend_y += 12
        for name in names:
            code = unit_codes[_unit_label_key(name)]
            for line in _unit_legend_lines(code, name, max_legend_chars):
                _draw_text(
                    canvas,
                    width - legend_width + 18,
                    legend_y,
                    line,
                    (31, 36, 45),
                    1,
                )
                legend_y += 11
        legend_y += 10

    # Optional place-name pins.
    if place_name_pins:
        for pin in place_name_pins:
            plat = pin.get("lat")
            plng = pin.get("lng")
            cid = pin.get("col_id", "")
            pcolor = color_map.get(cid, (200, 50, 50))
            if plat is not None and plng is not None:
                px, py = project((plng, plat))
                if left <= px <= right and top <= py <= bottom:
                    canvas.circle(px, py, 6, (255, 255, 255))
                    canvas.circle(px, py, 4, pcolor)
                    _draw_text(canvas, px + 6, py - 4, f"PIN:{pin.get('name','')[:10]}", (40, 40, 40), 2)

    # Optional centroid pins.
    if centroid_pins:
        for cpin in centroid_pins:
            clat = cpin.get("lat")
            clng = cpin.get("lng")
            cid = cpin.get("col_id", "")
            scolor = color_map.get(cid, (220, 120, 0))
            if clat is not None and clng is not None:
                cx, cy = project((clng, clat))
                if left <= cx <= right and top <= cy <= bottom:
                    canvas.circle(cx, cy, 10, (255, 255, 255))
                    canvas.circle(cx, cy, 7, scolor)
                    canvas.line(cx - 10, cy - 10, cx + 10, cy + 10, (255, 255, 255), 2)
                    canvas.line(cx - 10, cy + 10, cx + 10, cy - 10, (255, 255, 255), 2)
                    _draw_text(canvas, cx + 15, cy - 12, f"*Centroid({cid})", (200, 80, 0), 2)

    note_y = min(height - 128, legend_y + 8)
    _draw_text(canvas, width - legend_width + 8, note_y, "CANDIDATE MAP - REVIEW REQUIRED", (168, 50, 50), 1)
    _draw_text(canvas, width - legend_width + 8, note_y + 16, "SHARED UNITS: NEAREST SEED", (79, 86, 98), 1)
    if underground_units:
        _draw_text(canvas, width - legend_width + 8, note_y + 32, f"UNDERGROUND UNITS: {len(underground_units)}", (100, 100, 120), 1)
    _draw_text(canvas, 64, height - 42, f"BBOX {xmin:.4f}, {ymin:.4f} / {xmax:.4f}, {ymax:.4f}", (79, 86, 98), 1)
    canvas.save_png(Path(path))


def _kml_color(color: Color, alpha: int) -> str:
    return f"{alpha:02x}{color[2]:02x}{color[1]:02x}{color[0]:02x}"



def _xml_id(value: str) -> str:
    output = re.sub(r"[^A-Za-z0-9_.-]+", "-", value).strip("-") or "column"
    return output if output[0].isalpha() or output[0] == "_" else "c-" + output


def _coordinates(ring: Sequence[Point]) -> str:
    return " ".join(f"{x:.8f},{y:.8f},0" for x, y in ring)


def write_kml(
    path: str | Path,
    regions: Sequence[ColumnRegion],
    source_crs: str,
    warnings: Sequence[str],
    *,
    document_name: str = "Macrostrat column candidates",
    place_name_pins: Sequence[dict] | None = None,
    centroid_pins: Sequence[dict] | None = None,
    underground_units: Sequence[str] | None = None,
) -> None:
    """Write candidate regions and representative points as KML 2.2."""

    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        "<Document>",
        f"<name>{html.escape(document_name)}</name>",
        "<description>Candidate column regions for human review. "
        f"Source CRS: {html.escape(source_crs)}. "
        "Shared-unit polygon components are assigned to the nearest exclusive-unit seed. "
        f"Warnings: {html.escape(' | '.join(warnings) or 'none')}.</description>",
    ]
    for region in regions:
        style_id = _xml_id("column-" + region.definition.col_id)
        pin_style_id = _xml_id("pin-" + region.definition.col_id)
        star_style_id = _xml_id("star-" + region.definition.col_id)
        lines.extend(
            [
                f'<Style id="{style_id}">',
                f"<LineStyle><color>{_kml_color(region.color, 220)}</color><width>2</width></LineStyle>",
                f"<PolyStyle><color>{_kml_color(region.color, 105)}</color></PolyStyle>",
                f"<IconStyle><color>{_kml_color(region.color, 255)}</color><scale>1.15</scale></IconStyle>",
                "</Style>",
                f'<Style id="{pin_style_id}">',
                f"<IconStyle><color>{_kml_color(region.color, 255)}</color><scale>1.2</scale><Icon><href>http://maps.google.com/mapfiles/kml/pushpin/blue-pushpin.png</href></Icon></IconStyle>",
                "</Style>",
                f'<Style id="{star_style_id}">',
                f"<IconStyle><color>{_kml_color(region.color, 255)}</color><scale>1.4</scale><Icon><href>http://maps.google.com/mapfiles/kml/shapes/star.png</href></Icon></IconStyle>",
                "</Style>",
            ]
        )

    lines.append("<Folder><name>Candidate column regions</name>")
    for region in regions:
        style_id = _xml_id("column-" + region.definition.col_id)
        method_counts: dict[str, int] = {}
        for component in region.components:
            method_counts[component.assignment_method] = method_counts.get(component.assignment_method, 0) + 1
        details = "; ".join(f"{key}: {value}" for key, value in method_counts.items()) or "no polygon"
        unit_names = _region_unit_names(region)
        unit_detail = "; ".join(unit_names) or "none"
        lines.extend(
            [
                "<Folder>",
                f"<name>{html.escape(region.definition.name)}</name>",
                "<description>INFERRED candidate region; review required. "
                f"Component assignment: {html.escape(details)}. "
                f"Assigned geological units: {html.escape(unit_detail)}.</description>",
            ]
        )
        for index, component in enumerate(region.components, start=1):
            unit_name = component.unit_name or "Unnamed geological unit"
            lines.extend(
                [
                    "<Placemark>",
                    f"<name>{html.escape(unit_name)} — {html.escape(region.definition.name)}</name>",
                    f"<styleUrl>#{style_id}</styleUrl>",
                    "<description>CANDIDATE - REVIEW REQUIRED. "
                    f"Geological unit: {html.escape(unit_name)}. "
                    f"Column: {html.escape(region.definition.name)}. "
                    f"Assignment: {html.escape(component.assignment_method)}.</description>",
                    "<ExtendedData>",
                    f'<Data name="column_id"><value>{html.escape(region.definition.col_id)}</value></Data>',
                    f'<Data name="geological_unit"><value>{html.escape(unit_name)}</value></Data>',
                    f'<Data name="major_code"><value>{html.escape(component.major_code)}</value></Data>',
                    f'<Data name="component_index"><value>{index}</value></Data>',
                    f'<Data name="assignment_method"><value>{html.escape(component.assignment_method)}</value></Data>',
                    '<Data name="status"><value>CANDIDATE - REVIEW REQUIRED</value></Data>',
                    "</ExtendedData>",
                    "<Polygon><tessellate>1</tessellate>",
                    "<outerBoundaryIs><LinearRing><coordinates>",
                    _coordinates(component.outer),
                    "</coordinates></LinearRing></outerBoundaryIs>",
                ]
            )
            for hole in component.holes:
                lines.extend(
                    [
                        "<innerBoundaryIs><LinearRing><coordinates>",
                        _coordinates(hole),
                        "</coordinates></LinearRing></innerBoundaryIs>",
                    ]
                )
            lines.extend(["</Polygon>", "</Placemark>"])
        lines.append("</Folder>")
    lines.append("</Folder>")

    lines.append("<Folder><name>Representative point candidates</name>")
    for region in regions:
        style_id = _xml_id("column-" + region.definition.col_id)
        status = "INSIDE CANDIDATE REGION" if region.candidate_inside_region else "FALLBACK - REVIEW REQUIRED"
        lines.extend(
            [
                "<Placemark>",
                f"<name>{html.escape(region.definition.name)} candidate point</name>",
                f"<styleUrl>#{style_id}</styleUrl>",
                f"<description>{html.escape(status)}. Method: {html.escape(region.candidate_method)}.</description>",
                "<ExtendedData>",
                f'<Data name="column_id"><value>{html.escape(region.definition.col_id)}</value></Data>',
                f'<Data name="candidate_method"><value>{html.escape(region.candidate_method)}</value></Data>',
                f'<Data name="inside_region"><value>{str(region.candidate_inside_region).lower()}</value></Data>',
                "</ExtendedData>",
                f"<Point><coordinates>{region.candidate_lng:.8f},{region.candidate_lat:.8f},0</coordinates></Point>",
                "</Placemark>",
            ]
        )
    lines.append("</Folder>")

    # Optional place-name candidates.
    if place_name_pins:
        lines.append("<Folder><name>Place Name Candidates (Text Extracted 📌)</name>")
        for pin in place_name_pins:
            col_id = pin.get("col_id", "default")
            style_id = _xml_id("pin-" + col_id)
            name = pin.get("name", "Place")
            lat = pin.get("lat")
            lng = pin.get("lng")
            if lat is not None and lng is not None:
                lines.extend(
                    [
                        "<Placemark>",
                        f"<name>📌 {html.escape(name)} ({html.escape(col_id)})</name>",
                        f"<styleUrl>#{style_id}</styleUrl>",
                        f"<description>Extracted from report text for column: {html.escape(col_id)}</description>",
                        f"<Point><coordinates>{lng:.8f},{lat:.8f},0</coordinates></Point>",
                        "</Placemark>",
                    ]
                )
        lines.append("</Folder>")

    # Optional centroid candidates.
    if centroid_pins:
        lines.append("<Folder><name>Centroid Candidates (Place Average ★)</name>")
        for cpin in centroid_pins:
            col_id = cpin.get("col_id", "default")
            style_id = _xml_id("star-" + col_id)
            lat = cpin.get("lat")
            lng = cpin.get("lng")
            if lat is not None and lng is not None:
                lines.extend(
                    [
                        "<Placemark>",
                        f"<name>★ Centroid Candidate: {html.escape(col_id)}</name>",
                        f"<styleUrl>#{style_id}</styleUrl>",
                        f"<description>Calculated centroid from place names for column: {html.escape(col_id)}</description>",
                        f"<Point><coordinates>{lng:.8f},{lat:.8f},0</coordinates></Point>",
                        "</Placemark>",
                    ]
                )
        lines.append("</Folder>")

    # Units that are present in stratigraphy but absent from surface polygons.
    if underground_units:
        lines.append("<Folder><name>Underground-Only Units (No Surface Polygon)</name>")
        for u_name in underground_units:
            lines.extend(
                [
                    "<Placemark>",
                    f"<name>📄 {html.escape(u_name)} (Underground / Subsurface)</name>",
                    "<description>This unit is present in column stratigraphy but has no surface polygon in Shapefile.</description>",
                    "</Placemark>",
                ]
            )
        lines.append("</Folder>")

    lines.extend(["</Document>", "</kml>"])
    Path(path).write_text("\n".join(lines) + "\n", encoding="utf-8")


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--review", required=True, type=Path, help="Review .xlsx file")
    parser.add_argument("--shape", required=True, type=Path, help="GSJ geo_A.shp file")
    parser.add_argument("--png", required=True, type=Path, help="Output PNG path")
    parser.add_argument("--kml", required=True, type=Path, help="Output KML path")
    parser.add_argument(
        "--json",
        type=Path,
        default=None,
        help="Optional machine-readable candidate metadata (including representative points)",
    )
    parser.add_argument("--title", default=None, help="Map/KML title")
    parser.add_argument("--width", type=int, default=1200)
    parser.add_argument("--height", type=int, default=900)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    result = generate_column_map(
        args.review,
        args.shape,
        args.png,
        args.kml,
        width=args.width,
        height=args.height,
        title=args.title,
    )
    payload = json.dumps(result.as_dict(), ensure_ascii=False, indent=2)
    if args.json is not None:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(payload + "\n", encoding="utf-8")
    print(payload)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
