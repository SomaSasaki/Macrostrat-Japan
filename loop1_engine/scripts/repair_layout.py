# -*- coding: utf-8 -*-
"""
repair_layout.py — 既存のレビュー用Excelの「見た目」だけを修復する

セルの値は一切変更しない。次の3点だけを直す:
  1. フリーズペインが広すぎて右の列にスクロールできない状態を解消
  2. 非表示になっている列を再表示
  3. 列幅を現行の推奨値に揃える

使い方:
  python run.py repair                     # data/02_review 以下を全部修復
  python scripts/repair_layout.py <path>   # 個別に指定
"""

import argparse
import glob
import os
import shutil
import sys
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import (  # noqa: E402
    DEFAULT_COLUMN_WIDTH, NO_WRAP_COLS, REVIEW_UNIT_COLS, column_width,
    auto_t_pos, intervals_for_excel, is_blank,
    truncate_for_cell)
from make_review_sheet import write_prop_formulas  # noqa: E402

# 既定は見出し行だけの固定（A2）。列は固定しない。
# 列を固定すると幅の広い参照列が画面を占有して右へスクロールできなくなるうえ、
# 「A列だけ動かない」という分かりにくい挙動になる。
DEFAULT_FREEZE = "A2"


def freezes_columns(freeze_panes):
    """列を固定しているか。'A2' は行だけなので False、'B2'/'G2' は True。"""
    if not freeze_panes:
        return False
    s = str(freeze_panes)
    col = "".join(c for c in s if c.isalpha())
    return bool(col) and col.upper() != "A"


def needs_freeze_fix(ws):
    """
    直すのは「列を固定していて横スクロールできない」場合だけ。

    行だけの固定（A2 / A24 など）はユーザーが意図して設定していることがあるので
    触らない。ここで一律 A2 に戻すと、利用者が手で直した設定を奪ってしまう。
    """
    if ws.freeze_panes is None:
        return True                      # 未設定なら見出し行を固定してあげる
    return freezes_columns(ws.freeze_panes)

# 列幅は common.COLUMN_WIDTHS が唯一の定義（二重管理をやめた）
WIDTHS = None
DEFAULT_WIDTH = DEFAULT_COLUMN_WIDTH


def repair(path, verbose=True):
    if not os.path.exists(path):
        print(f"[skip] 見つかりません: {path}")
        return False
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        print(f"[skip] 読み込めません: {path} ({e})")
        return False

    changes = []
    for name in wb.sheetnames:
        ws = wb[name]

        # 1. フリーズペイン（列を固定している場合だけ直す）
        if needs_freeze_fix(ws):
            reason = "列固定を解除" if ws.freeze_panes else "見出し行を固定"
            changes.append(f"{name}: フリーズ {ws.freeze_panes} -> {DEFAULT_FREEZE}（{reason}）")
            ws.freeze_panes = DEFAULT_FREEZE

        # 2. 非表示列の解除 + 3. 列幅
        headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        for i, h in enumerate(headers, 1):
            letter = get_column_letter(i)
            dim = ws.column_dimensions[letter]
            if dim.hidden:
                changes.append(f"{name}: {letter}列({h}) の非表示を解除")
                dim.hidden = False
            w = column_width(h)
            if dim.width is None or abs((dim.width or 0) - w) > 0.5:
                dim.width = w
        # 見出しより右に幅指定だけ残っている列も再表示しておく
        for letter, dim in list(ws.column_dimensions.items()):
            if dim.hidden:
                changes.append(f"{name}: {letter}列 の非表示を解除")
                dim.hidden = False

    if not changes:
        if verbose:
            print(f"[ok] 修復不要: {path}")
        return True

    try:
        wb.save(path)
    except PermissionError:
        print(f"[ERROR] 保存できません（Excelで開いていませんか？）: {path}")
        return False

    print(f"[fixed] {path}")
    for c in changes[:12]:
        print(f"        - {c}")
    if len(changes) > 12:
        print(f"        ... 他 {len(changes) - 12} 件")
    return True


def needs_migration(path):
    """現行スキーマに足りないもの（列・シート）を列挙する。"""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path)
    except Exception:
        return None
    if "units_review" not in wb.sheetnames:
        return None
    ws = wb["units_review"]
    have = [c.value for c in ws[1]]

    # 長すぎるセル（Excelで折り返しきれず読めなくなる）も移行対象にする
    too_long = []
    for i, h in enumerate(have, 1):
        if not h:
            continue
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=i).value
            if isinstance(v, str) and not v.startswith("=") \
                    and len(v) > len(str(truncate_for_cell(v, h))):
                too_long.append(h)
                break

    # 退役した列。値が入っていなければ落とす（入っていれば末尾に温存する）
    from common import RETIRED_REF_COLS
    retired = []
    for c in RETIRED_REF_COLS:
        if c not in have:
            continue
        i = have.index(c) + 1
        if all(ws.cell(row=r, column=i).value in (None, "")
               for r in range(2, ws.max_row + 1)):
            retired.append(c)

    return {
        "missing_cols": [c for c in REVIEW_UNIT_COLS if c not in have],
        "retired_cols": retired,
        "extra_cols": [c for c in have
                       if c and c not in REVIEW_UNIT_COLS and c not in retired],
        "missing_sheets": [s for s in ("abstract", "intervals", "descriptions",
                                       "thickness_notes")
                           if s not in wb.sheetnames],
        "wrong_order": [c for c in have if c in REVIEW_UNIT_COLS] !=
                       [c for c in REVIEW_UNIT_COLS if c in have],
        "too_long_cols": too_long,
    }


def migrate(path, dry=False):
    """
    古いレビューファイルを現行スキーマへ引き上げる。**入力済みの値は保持する。**

    やること:
      1. 足りない列を正しい位置に追加し、列順を現行に揃える
      2. 足りないシート（intervals / abstract）を追加
      3. t_prop / b_prop の数式を作り直す
      4. 書式（色分け・列幅・折り返し）を付け直す

    なぜ列を「挿入」しないか:
      openpyxl の insert_cols は数式の参照を書き換えない。
      途中に列を挿すと t_prop/b_prop の VLOOKUP が別の列を指してしまう。
      そこで値を読み出して並べ直し、数式は作り直す方式にしている。
    """
    import openpyxl
    import pandas as pd
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    info = needs_migration(path)
    if info is None:
        print(f"[skip] units_review が読めません: {path}")
        return False
    if not (info["missing_cols"] or info["missing_sheets"] or info["wrong_order"]
            or info["too_long_cols"] or info["retired_cols"]):
        print(f"[ok] 既に現行スキーマです: {os.path.basename(path)}")
        return True

    print(f"[migrate] {path}")
    if info["missing_cols"]:
        print(f"          追加する列: {', '.join(info['missing_cols'])}")
    if info["missing_sheets"]:
        print(f"          追加するシート: {', '.join(info['missing_sheets'])}")
    if info["extra_cols"]:
        print(f"          現行にない列（末尾に温存）: {', '.join(info['extra_cols'])}")
    if info["retired_cols"]:
        print(f"          削除する列（全行が空）: {', '.join(info['retired_cols'])}")
    if info["too_long_cols"]:
        print(f"          長すぎるセルを切り詰める: {', '.join(info['too_long_cols'])}")
    if dry:
        print("          (--dry のため変更していません)")
        return True

    # --- バックアップ ---
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.replace(".xlsx", f".bak_{stamp}.xlsx")
    shutil.copy2(path, backup)
    print(f"          バックアップ: {os.path.basename(backup)}")

    # --- 値を読み出す（数式ではなく値。数式は後で作り直す） ---
    wb = openpyxl.load_workbook(path)
    ws = wb["units_review"]
    have = [c.value for c in ws[1]]
    rows = []
    for r in range(2, ws.max_row + 1):
        rec = {}
        for i, h in enumerate(have, 1):
            if not h:
                continue
            v = ws.cell(row=r, column=i).value
            # 数式は捨てる（作り直すため）
            rec[h] = None if isinstance(v, str) and v.startswith("=") else v
        if any(v not in (None, "") for v in rec.values()):
            rows.append(rec)

    order = REVIEW_UNIT_COLS + info["extra_cols"]      # 現行にない列は末尾に温存
    df = pd.DataFrame(rows).reindex(columns=order).astype(object)

    # --- 本文由来の値を ZFK キャッシュから入れ直す ---
    #   REF_ 列は「自動取得の参照専用」なので上書きしてよい。
    #   編集列は空欄のときだけ埋める（人が入れた値は壊さない）。
    print("          ZFK本文から REF_ 列を作り直します（PDFのページ照合を含む）...")
    zfk = _zfk_ref_fields(path)
    n_ref = n_fill = 0
    if zfk and "unit_id" in df.columns:
        for i in df.index:
            info_u = zfk.get(str(df.at[i, "unit_id"] or "").strip())
            if not info_u:
                continue
            for c, v in info_u["_ref"].items():
                if c in df.columns and v not in (None, ""):
                    df.at[i, c] = v
                    n_ref += 1
            for c, v in info_u["_fill"].items():
                if c in df.columns and v not in (None, "") and is_blank(df.at[i, c]):
                    df.at[i, c] = v
                    n_fill += 1
        print(f"          REF_ 列を更新 {n_ref} セル / 空欄を補完 {n_fill} セル")

    # --- t_pos: 各Columnの最上位に入れる（無いと取り込み時に落ちる） ---
    if {"sort_order", "column_id", "t_pos"} <= set(df.columns):
        n_tpos = 0
        tps = auto_t_pos(list(df["column_id"]), list(df["sort_order"]))
        for i, tp in zip(df.index, tps):
            if tp != "" and is_blank(df.at[i, "t_pos"]):
                df.at[i, "t_pos"] = tp
                n_tpos += 1
        if n_tpos:
            print(f"          t_pos を自動入力（{n_tpos} 行）")

    # --- units_review を並べ直して書き戻す ---
    idx = wb.sheetnames.index("units_review")
    del wb["units_review"]
    new = wb.create_sheet("units_review", idx)
    new.append(order)
    for _, rec in df.iterrows():
        new.append(["" if pd.isna(v) else truncate_for_cell(v, c)
                    for c, v in zip(order, rec.tolist())])

    # --- 足りないシートを足す ---
    if "intervals" not in wb.sheetnames:
        iv = intervals_for_excel()
        s = wb.create_sheet("intervals")
        s.append(["interval", "b_age_ma", "t_age_ma", "int_type"])
        for k, v in sorted(iv.items(), key=lambda x: (-float(x[1]["b_age"]), x[0])):
            s.append([k, v["b_age"], v["t_age"], v.get("int_type", "")])
        print(f"          intervals シートを追加（{len(iv)} 件）")

    if "abstract" not in wb.sheetnames:
        s = wb.create_sheet("abstract")
        s.append(["no", "text"])
        txt = _find_abstract_text(path)
        if txt:
            for i, para in enumerate(txt, 1):
                s.append([i, para[:32000]])
            print(f"          abstract シートを追加（{len(txt)} 段落）")
        else:
            s.append([1, "英文Abstractが未取得です。"
                         "`python run.py abstract <図幅名>` を実行してください。"])

    # --- descriptions / thickness_notes を ZFK キャッシュから作る ---
    if "descriptions" not in wb.sheetnames or "thickness_notes" not in wb.sheetnames:
        docs, notes = _zfk_texts(path)
        if "descriptions" not in wb.sheetnames:
            sh = wb.create_sheet("descriptions")
            sh.append(["unit_id", "unit_name", "unit_name_ja", "text"])
            for d in docs:
                sh.append([d["unit_id"], d["unit_name"], d["unit_name_ja"], d["text"][:32000]])
            print(f"          descriptions シートを追加（{len(docs)} 件）")
        if "thickness_notes" not in wb.sheetnames:
            sh = wb.create_sheet("thickness_notes")
            sh.append(["unit_id", "unit_name", "thickness_note"])
            for n_ in notes:
                sh.append([n_["unit_id"], n_["unit_name"], n_["thickness_note"]])
            print(f"          thickness_notes シートを追加（{len(notes)} 件）")

        # REF_thickness が空なら本文から埋める
        if "REF_thickness" in order and docs:
            by_id = {}
            for n_ in notes:
                by_id.setdefault(n_["unit_id"], []).append(n_["thickness_note"])
            uid_col = order.index("unit_id") + 1 if "unit_id" in order else None
            th_col = order.index("REF_thickness") + 1
            filled = 0
            for r in range(2, len(df) + 2):
                if new.cell(row=r, column=th_col).value:
                    continue
                uid = str(new.cell(row=r, column=uid_col).value or "") if uid_col else ""
                if by_id.get(uid):
                    new.cell(row=r, column=th_col).value = truncate_for_cell(
                        " / ".join(by_id[uid]), "REF_thickness")
                    filled += 1
            if filled:
                print(f"          REF_thickness を本文から補完（{filled} 行）")

    # --- 数式を作り直す ---
    n = write_prop_formulas(new, order, len(df))
    if n:
        print(f"          t_prop / b_prop の数式を再生成（{n} セル）")

    # --- 書式 ---
    fill_ref = PatternFill("solid", start_color="F2F2F2")
    fill_input = PatternFill("solid", start_color="FFF2CC")
    fill_calc = PatternFill("solid", start_color="DEEAF6")
    fill_out = PatternFill("solid", start_color="D9EAD3")
    required = {"unit_name", "sort_order", "column_id", "lithology", "t_int", "b_int",
                "t_age_ma", "b_age_ma"}
    new.freeze_panes = "A2"
    new.auto_filter.ref = new.dimensions
    for i, name in enumerate(order, 1):
        c = new.cell(row=1, column=i)
        c.font = Font(bold=True)
        c.fill = (fill_ref if str(name).startswith("REF_") or name == "unit_id"
                  else fill_calc if name in ("t_prop", "b_prop")
                  else fill_input if name in required else fill_out)
        new.column_dimensions[get_column_letter(i)].width = column_width(name)
        # 長文列は折り返さない（全文を入れているので折り返すと1行が数百行になる）
        al = Alignment(wrapText=name not in NO_WRAP_COLS, vertical="top")
        for r in range(2, len(df) + 2):
            new.cell(row=r, column=i).alignment = al

    try:
        wb.save(path)
    except PermissionError:
        print("[ERROR] Excelで開かれているため保存できません。閉じてから再実行してください。")
        return False
    print(f"          完了: {len(df)} 行 / {len(order)} 列")
    return True


def _zfk_ref_fields(review_path, quiet=False):
    """
    ZFKキャッシュから、本文由来の REF_ 列と編集列の初期値を組み立てる。

    migrate から呼ぶ。REF_ 列は「自動取得の参照専用」なので上書きしてよい。
    編集列（strat_name / basal_surface / min_thickness / max_thickness）は
    **空欄のときだけ** 埋める。人が入れた値は壊さない。

    戻り値: {unit_id: {列名: 値}}
    """
    import glob as _glob

    from common import load_json

    import gsj_derived as G
    from pdf_locate import index_for

    mid = os.path.basename(review_path).split("_")[0].lstrip("m")
    unit_dir = os.path.join("data", "raw", "zfk", f"m{mid}", "units")
    files = sorted(_glob.glob(os.path.join(unit_dir, "*.json")))
    if not files:
        return {}

    ref_dir = os.path.join(os.path.dirname(review_path), "references")
    pdf_index = index_for(mid, ref_dir, quiet=quiet)

    out = {}
    for f in files:
        u = load_json(f)
        if not u:
            continue
        uid = u.get("id") or os.path.basename(f)[:-5]
        desc = (u.get("target") or {}).get("text") or ""
        g_mn, g_mx = G.best_thickness(u)
        g_li = G.lithologies(u, pdf_index)
        g_bs, g_sn = G.basal_surface(u, pdf_index), G.strat_name(u)

        out[uid] = {
            "_ref": {                                  # 上書きしてよい（参照専用）
                "REF_source": G.describe_source(u, pdf_index),
                "REF_desc": desc,
                "REF_thickness": G.thickness_block(u, pdf_index),
                "REF_lithology_gsj": (g_li or {}).get("major", ""),
                "REF_minor_lith_gsj": (g_li or {}).get("minor", ""),
                "REF_strat_name": g_sn or "",
                "REF_basal_surface": (g_bs or {}).get("text", ""),
            },
            "_fill": {                                 # 空欄のときだけ埋める
                "strat_name": g_sn or "",
                "basal_surface": (g_bs or {}).get("value", ""),
                "min_thickness": "" if g_mn is None else g_mn,
                "max_thickness": "" if g_mx is None else g_mx,
            },
        }
    return out


def _zfk_texts(review_path):
    """ZFKキャッシュから地層ごとの全文と層厚記述を集める。"""
    import glob as _glob
    from common import extract_thickness_notes, load_json, strip_trailing_paren

    mid = os.path.basename(review_path).split("_")[0].lstrip("m")
    unit_dir = os.path.join("data", "raw", "zfk", f"m{mid}", "units")
    docs, notes = [], []
    for f in sorted(_glob.glob(os.path.join(unit_dir, "*.json"))):
        u = load_json(f)
        if not u:
            continue
        legend = u.get("legend") or {}
        name_en = (legend.get("parent_facies") or {}).get("label_en") or ""
        name_ja = (legend.get("parent_facies") or {}).get("label_ja") or ""
        lith_en = (legend.get("focus") or {}).get("label_en") or ""
        name = name_en or strip_trailing_paren(lith_en)
        text = (u.get("target") or {}).get("text") or ""
        uid = u.get("id") or os.path.basename(f)[:-5]
        docs.append({"unit_id": uid, "unit_name": name,
                     "unit_name_ja": name_ja, "text": text})
        for note in extract_thickness_notes(text, max_items=12):
            notes.append({"unit_id": uid, "unit_name": name, "thickness_note": note})
    return docs, notes


def _find_abstract_text(review_path):
    """references/m{id}_abstract.txt があれば段落に割って返す。"""
    base = os.path.basename(review_path)
    mid = base.split("_")[0].lstrip("m")
    p = os.path.join(os.path.dirname(review_path), "references", f"m{mid}_abstract.txt")
    if not os.path.exists(p):
        return []
    with open(p, encoding="utf-8") as f:
        text = f.read()
    paras, buf = [], []
    for line in text.splitlines():
        if line.strip():
            buf.append(line.strip())
        elif buf:
            paras.append(" ".join(buf))
            buf = []
    if buf:
        paras.append(" ".join(buf))
    return [p for p in paras if len(p) > 2]


def migrate_all(root="data", dry=False):
    files = sorted(p for p in glob.glob(os.path.join(root, "**", "*_review.xlsx"),
                                        recursive=True)
                   if not os.path.basename(p).startswith("~$")
                   and ".bak_" not in os.path.basename(p))
    if not files:
        print(f"{root} 以下にレビューファイルがありません。")
        return
    print(f"{len(files)} 件を確認します。\n")
    for f in files:
        migrate(f, dry=dry)


def repair_all(root="data"):
    files = sorted(
        p for p in glob.glob(os.path.join(root, "**", "*_review.xlsx"), recursive=True)
        if not os.path.basename(p).startswith("~$")
    )
    if not files:
        print(f"{root} 以下にレビューファイルがありません。")
        return
    print(f"{len(files)} 件を確認します。\n")
    for f in files:
        repair(f)


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="レビューExcelの修復・スキーマ移行")
    p.add_argument("path", nargs="?", help="対象ファイル。省略時は data 以下すべて")
    p.add_argument("--migrate", action="store_true",
                   help="現行スキーマへ引き上げる（列・シートの追加。値は保持）")
    p.add_argument("--dry", action="store_true", help="変更せず内容だけ表示")
    a = p.parse_args()
    if a.migrate:
        if a.path:
            sys.exit(0 if migrate(a.path, dry=a.dry) else 1)
        migrate_all(dry=a.dry)
    elif a.path:
        sys.exit(0 if repair(a.path) else 1)
    else:
        repair_all()
