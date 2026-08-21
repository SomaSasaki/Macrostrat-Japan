from pathlib import Path
# -*- coding: utf-8 -*-
"""
make_review_200k.py — GSJ 20万分の1図幅からレビュー用 Excel を生成する（50k完全準拠・厳密単調性・Python直接prop計算・LLM抜き）

Macrostrat column-ingestion format v0.1.1 準拠のレビューシートを作成。
- 年代単調増加ソート（Bottom: 最古 → Top: 最新）の徹底
- Python側での厳密な prop 直接計算（数式キャッシュ非依存）
- 根拠のない接触関係（10 Maルール）の完全廃止（深成岩以外は unknown/空欄）
- 50k完全準拠のカラーシステム、Meiryo UIフォント、オートフィルター、列幅最適化
"""

import io
import json
import os
import re
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, BASE_DIR)

import scripts.common as common

CONFIG_PATH = os.path.join(BASE_DIR, 'config', 'map_index_200k.json')
CACHE_DIR = os.path.join(BASE_DIR, 'data', '200k', 'raw', 'seamless_200k')
REVIEW_BASE_DIR = os.path.join(BASE_DIR, 'data', '200k', '02_review')
VOCAB_PATH = os.path.join(Path(__file__).resolve().parents[2], 'loop2_governance', 'config', 'vocab.json')

intervals_dict = common.load_intervals()
int_names = set(intervals_dict.keys())

with open(VOCAB_PATH, 'r', encoding='utf-8') as f:
    vocab_data = json.load(f)
lith_vocab = {v.lower(): v for v in vocab_data.get('lithology', [])}

def load_200k_index() -> list[dict]:
    """200k 図幅インデックス (112図幅) をロードして返す"""
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        return json.load(f)

def load_sheet_legends(sheet_code: str, name_en: str = "") -> list[dict]:
    """図幅の凡例リストをキャッシュからロードして返す"""
    cache_file = os.path.join(CACHE_DIR, f"{sheet_code}.json")
    aid_cache = os.path.join(CACHE_DIR, f"{name_en}.json")
    if os.path.exists(cache_file):
        with open(cache_file, 'r', encoding='utf-8') as cf:
            return json.load(cf)
    elif name_en and os.path.exists(aid_cache):
        with open(aid_cache, 'r', encoding='utf-8') as af:
            return json.load(af)
    return []

# 200k レビュー用列定義（50k順序に準拠）
REVIEW_200K_EDIT_COLS = [
    "unit_id",
    "col_id",
    "sort_order",
    "unit_name",
    "strat_name",
    "t_int",
    "t_age_ma",
    "t_prop",
    "b_int",
    "b_age_ma",
    "b_prop",
    "lithology",
    "minor_lith",
    "environment",
    "unit_description",
    "min_thickness",
    "max_thickness",
    "basal_surface",
    "lateral_relationship",
    "comments",
    "t_pos",
]

REVIEW_200K_REF_COLS = [
    "REF_symbol",
    "REF_age_ja",
    "REF_age_en",
    "REF_lith_ja",
    "REF_lith_en",
    "REF_group_ja",
    "REF_group_en",
    "REF_domain",
]

REVIEW_200K_UNIT_COLS = REVIEW_200K_EDIT_COLS + REVIEW_200K_REF_COLS

def parse_seamless_age(age_str):
    """
    シームレス地質図の年代文字列から、b_int, t_int, b_age(Ma), t_age(Ma), b_prop, t_prop を厳密に計算
    """
    if not age_str:
        return 'Phanerozoic', 'Phanerozoic', 0.0, 1.0, 541.0, 0.0
    parts = re.split(r'\s*[-–—〜~]\s*', age_str)

    def find_best_interval(part):
        tokens = [t.strip(' ,;') for t in part.split() if t.strip(' ,;')]
        for i in range(len(tokens) - 1, -1, -1):
            t = tokens[i]
            for candidate in [t, t.capitalize()]:
                if candidate in int_names:
                    return candidate
            if i > 0:
                two = f"{tokens[i-1].capitalize()} {tokens[i].capitalize()}"
                if two in int_names:
                    return two
        return None

    b_cand = find_best_interval(parts[0])
    t_cand = find_best_interval(parts[-1]) if len(parts) > 1 else b_cand

    if not b_cand and not t_cand:
        b_cand, t_cand = 'Phanerozoic', 'Phanerozoic'
    elif not b_cand:
        b_cand = t_cand
    elif not t_cand:
        t_cand = b_cand

    b_bounds = common.interval_bounds(b_cand)
    t_bounds = common.interval_bounds(t_cand)

    b_age = b_bounds[0] if b_bounds else 541.0
    t_age = t_bounds[1] if t_bounds else 0.0

    # prop 直接計算（b_int の下端 = 0.0, t_int の上端 = 1.0）
    bp = 0.0
    tp = 1.0

    return b_cand, t_cand, bp, tp, b_age, t_age

def extract_macrostrat_lithologies(text):
    if not text:
        return 'sedimentary', None
    text_clean = text.lower().replace('&', ' and ').replace('/', ' ')
    tokens = [t.strip(' ,;') for t in text_clean.split()]

    found = []
    for i in range(len(tokens) - 1):
        two = f"{tokens[i]} {tokens[i+1]}"
        if two in lith_vocab and two not in found:
            found.append(lith_vocab[two])
    for t in tokens:
        if t in lith_vocab and lith_vocab[t] not in found:
            found.append(lith_vocab[t])

    if not found:
        if 'gravel' in text_clean or 'terrace' in text_clean:
            found.append('gravel')
        elif 'mud' in text_clean or 'clay' in text_clean:
            found.append('mudstone')
        elif 'sand' in text_clean:
            found.append('sandstone')
        elif 'volcanic' in text_clean or 'lava' in text_clean:
            found.append('volcanic')
        elif 'plutonic' in text_clean or 'granit' in text_clean:
            found.append('granite')
        elif 'schist' in text_clean or 'metamorph' in text_clean:
            found.append('schist')
        else:
            found.append('sedimentary')

    main_lith = found[0]
    minor_lith = found[1] if len(found) > 1 else None
    return main_lith, minor_lith

def classify_legend_domain(leg):
    grp = leg.get('group_en', '')
    age = leg.get('formationAge_en', '')
    lith = leg.get('lithology_en', '')

    if 'Accretionary' in grp or 'accretionary' in lith.lower():
        return 'Accretionary Complex'
    if 'Metamorphic' in grp or 'schist' in lith.lower() or 'gneiss' in lith.lower():
        return 'Metamorphic Belt'
    if 'Plutonic' in grp or 'granit' in lith.lower() or 'gabbro' in lith.lower() or 'diorite' in lith.lower():
        return 'Plutonic Complex'
    if 'Volcanic' in grp or 'lava' in lith.lower() or 'pyroclastic' in lith.lower() or 'basalt' in lith.lower() or 'andesite' in lith.lower() or 'rhyolite' in lith.lower() or 'dacite' in lith.lower():
        if any(q in age for q in ['Quaternary', 'Holocene', 'Pleistocene']):
            return 'Quaternary Volcanic Arc'
        return 'Neogene-Paleogene Volcanic Complex'
    if any(q in age for q in ['Quaternary', 'Holocene', 'Pleistocene']):
        return 'Quaternary Basin & Terrace Deposits'
    if any(n in age for n in ['Neogene', 'Miocene', 'Pliocene']):
        return 'Neogene Sedimentary Basin'
    if any(m in age for m in ['Mesozoic', 'Cretaceous', 'Jurassic', 'Triassic']):
        return 'Mesozoic Sedimentary Basin'
    if any(p in age for p in ['Paleozoic', 'Permian', 'Carboniferous', 'Devonian', 'Silurian', 'Ordovician', 'Cambrian']):
        return 'Paleozoic Sedimentary Succession'
    return f'{grp} Succession' if grp else 'Regional Sedimentary Succession'

def infer_environment(domain_name, lith_text):
    text = (domain_name + " " + lith_text).lower()
    if 'accretionary' in text or 'turbidite' in text or 'marine' in text or 'pelagic' in text:
        return 'marine'
    if 'river' in text or 'fluvial' in text or 'floodplain' in text:
        return 'fluvial'
    if 'coastal' in text or 'beach' in text or 'tidal' in text:
        return 'coastal'
    if 'alluvial' in text or 'fan' in text:
        return 'alluvial fan'
    if 'volcanic' in text or 'pyroclastic' in text or 'lava' in text:
        return 'volcanic'
    if 'plutonic' in text or 'granit' in text:
        return 'plutonic'
    if 'metamorphic' in text or 'schist' in text:
        return 'metamorphic'
    return 'unknown'

def infer_basal_surface(domain_name):
    """
    接触関係の推測: 証拠のない10Maルールは廃止。深成岩のみintrusiveとし、他はunknownとする。
    """
    if 'Plutonic' in domain_name:
        return 'intrusive'
    return 'unknown'

def make_review_for_sheet(sheet_meta, legends, out_excel_path):
    sheet_code = sheet_meta['sheet_code']
    name_ja = sheet_meta.get('name_ja', sheet_code)
    name_en = sheet_meta.get('name_en', sheet_code) or sheet_code
    region = sheet_meta.get('region', '00_Other')
    pub_year = sheet_meta.get('pub_year', 2010)

    # 1. 地質帯クラスタリング
    if not legends:
        domains = {'Marine Geological Succession': [{
            'symbol': 'Marine_01',
            'formationAge_en': 'Cenozoic Quaternary Holocene',
            'formationAge_ja': '新生代 第四紀 完新世',
            'lithology_en': 'marine mudstone & seafloor sediments',
            'lithology_ja': '海洋泥・海底堆積物',
            'group_en': 'Marine',
            'group_ja': '海洋堆積物',
        }]}
    else:
        domains = {}
        for leg in legends:
            dom = classify_legend_domain(leg)
            domains.setdefault(dom, []).append(leg)

    cols_review_rows = []
    units_review_rows = []
    desc_rows = []
    global_u_idx = 1

    for c_idx, (dom_name, dom_legs) in enumerate(domains.items(), start=1):
        col_id_str = f"col_{c_idx}"
        dom_clean = dom_name.replace(' ', '_').replace('&', 'and')
        col_name = f"{name_en} ({dom_name}) Column"
        col_group = f"GSJ_200K_{sheet_code.replace('-', '_')}_{dom_clean}"

        dom_units = []
        seen = set()
        for leg in dom_legs:
            age_en = leg.get('formationAge_en', '')
            lith_en = leg.get('lithology_en', '')
            grp_en = leg.get('group_en', '')
            age_ja = leg.get('formationAge_ja', '')
            lith_ja = leg.get('lithology_ja', '')
            grp_ja = leg.get('group_ja', '')
            sym = leg.get('symbol', '')

            k = (age_en, lith_en, grp_en)
            if k in seen:
                continue
            seen.add(k)

            b_int, t_int, bp, tp, b_age, t_age = parse_seamless_age(age_en)
            main_lith, minor_lith = extract_macrostrat_lithologies(lith_en)
            env = infer_environment(dom_name, lith_en)

            unit_name = f"{t_int} {main_lith.capitalize()} ({name_en} {dom_name.split()[0]})"
            strat_name = f"{name_en} {dom_name}"
            unit_id = f"m200k_{sheet_code}_u{str(global_u_idx).zfill(3)}"
            global_u_idx += 1

            basal_surf = infer_basal_surface(dom_name)

            dom_units.append({
                'unit_id': unit_id,
                'col_id': col_id_str,
                'unit_name': unit_name,
                'strat_name': strat_name,
                't_int': t_int,
                't_age_ma': t_age,
                't_prop': tp,
                'b_int': b_int,
                'b_age_ma': b_age,
                'b_prop': bp,
                'lithology': main_lith,
                'minor_lith': minor_lith or '',
                'environment': env,
                'unit_description': f"{age_en}; {lith_en} [{dom_name}]",
                'min_thickness': '',
                'max_thickness': '',
                'basal_surface': basal_surf,
                'lateral_relationship': '',
                'comments': f"GSJ 200k Seamless {sym}",
                't_pos': '',
                'REF_symbol': sym,
                'REF_age_ja': age_ja,
                'REF_age_en': age_en,
                'REF_lith_ja': lith_ja,
                'REF_lith_en': lith_en,
                'REF_group_ja': grp_ja,
                'REF_group_en': grp_en,
                'REF_domain': dom_name,
                'b_age': b_age,
                't_age': t_age,
            })

            desc_rows.append({
                'unit_id': unit_id,
                'unit_name': unit_name,
                'unit_name_ja': f"{age_ja} {lith_ja}",
                'text': f"{age_ja}（{age_en}）に形成された{lith_ja}（{lith_en}）。産総研シームレス地質図V2凡例記号: {sym}",
            })

        # ★ 重要: 層序順序の厳密な単調性（Bottom: 最古 → Top: 最新）
        # 中心年代 (b_age + t_age)/2 および b_age, t_age の降順（古い順）でソート
        dom_units.sort(key=lambda u: ((u['b_age'] + u['t_age']) / 2.0, u['b_age'], u['t_age']), reverse=True)

        for pos, u in enumerate(dom_units, start=1):
            u['sort_order'] = pos
            units_review_rows.append(u)

        # Column の年代範囲（配下の全Unitを完全に包含する最古 b_int 〜 最新 t_int）
        if dom_units:
            oldest_u = max(dom_units, key=lambda u: u['b_age'])
            youngest_u = min(dom_units, key=lambda u: u['t_age'])
            col_b_int = oldest_u['b_int']
            col_t_int = youngest_u['t_int']
        else:
            col_b_int = 'Phanerozoic'
            col_t_int = 'Holocene'

        cols_review_rows.append({
            'col_id': col_id_str,
            'col_name': col_name,
            'col_group': col_group,
            'domain_name': dom_name,
            'b_int': col_b_int,
            't_int': col_t_int,
            'b_prop': 0.0,
            't_prop': 1.0,
            'geom': sheet_meta.get('wkt_geom', ''),
            'rgeom': f"POINT({sheet_meta['center'][1]} {sheet_meta['center'][0]})" if 'center' in sheet_meta else '',
            'comments': f"GSJ 200k Tectonic Skeleton: {dom_name} in {name_ja} ({sheet_code})",
        })

    # DataFrames 構築
    df_inst = pd.DataFrame([
        {"項目": "0. このファイルの役割",
         "説明": "産総研20万分の1シームレス地質図V2から自動抽出した広域地質凡例インベントリおよび人手レビュー用スケルトンです。ここを編集して `python run.py export-200k <コード/地名>` を実行すると、Macrostrat公式形式(v0.1.1)の提出ファイルが作られます。"},
        {"項目": "1. 色の意味（50k準拠）",
         "説明": "黄色=必ず確認・入力する必須列。 / 水色=年代(Ma)から自動計算される数式列。 / 緑=そのまま提出物に転記される列。 / グレー(REF_*)=自動取得の参照専用データ。"},
        {"項目": "2. 地質帯Columnの構成",
         "説明": "同一図幅内の異なるテクトニック単元（付加体、変成帯、火山岩、堆積盆等）を独立したColumn（col_1, col_2...）に分離しています。columns_review で一覧を確認できます。"},
        {"項目": "3. sort_order / position（層序順序）",
         "説明": "最下部（最古の地層）を 1 とし、上位（新しい地層）へ向かって番号が進みます（下から上への単調増加）。必要に応じて順序を書き換えてください。"},
        {"項目": "4. t_prop / b_prop（自動計算）",
         "説明": "水色の t_prop・b_prop はPythonで直接初期計算済みかつ数式連動です。t_age_ma・b_age_ma を書き換えると、intervals シートの時代境界から自動で 0.000〜1.000 の比率が再計算されます。"},
        {"項目": "5. 岩相（lithology / minor_lith）",
         "説明": "Macrostrat公式の214岩相オントロジーに自動正規化してあります。"},
        {"項目": "6. 事前チェックと提出",
         "説明": "`python run.py check-200k <地名>` で年代逆転・prop範囲・語彙矛盾を厳格に検証し、`python run.py export-200k <地名>` でエクスポートします。"},
    ])

    df_units = pd.DataFrame(units_review_rows)[REVIEW_200K_UNIT_COLS]
    df_cols = pd.DataFrame(cols_review_rows)[["col_id", "col_name", "col_group", "domain_name", "b_int", "t_int", "b_prop", "t_prop", "geom", "rgeom", "comments"]]
    
    df_refs = pd.DataFrame([{
        "ref_id": 1,
        "title": f"1:200,000 Geological Map of Japan: {name_ja} ({name_en})",
        "authors": "Geological Survey of Japan, AIST",
        "publication": "Geological Map of Japan 1:200,000",
        "compilation": "GSJ Seamless Geological Map of Japan V2",
        "organization": "Geological Survey of Japan, AIST",
        "date": str(pub_year or 2010),
        "doi": "",
        "url": "https://www.gsj.jp/Map/JP/geology2.html",
        "comments": f"IMW Sheet {sheet_code}",
    }], columns=common.SUBMISSION_REF_COLS)

    df_images = pd.DataFrame([{c: "" for c in common.SUBMISSION_IMAGE_COLS}], columns=common.SUBMISSION_IMAGE_COLS)

    df_meta = pd.DataFrame([
        {"key": "project_name", "value": "GSJ Japan 200k Regional Skeleton"},
        {"key": "organization", "value": "Macrostrat Japan Project / GSJ"},
        {"key": "url", "value": "https://macrostrat.org"},
        {"key": "project_id", "value": f"GSJ_200K_{sheet_code.replace('-', '_')}"},
        {"key": "compile_date", "value": "2026-08-14"},
        {"key": "compiler_name", "value": "MacroStrat Automated Pipeline"},
        {"key": "col_type", "value": "column"},
        {"key": "axis_type", "value": "age"},
        {"key": "timescale", "value": "ICS 2020"},
        {"key": "srid", "value": "4326"},
    ])

    df_gsj = pd.DataFrame([
        {"key": "sheet_code", "value": sheet_code},
        {"key": "name_ja", "value": name_ja},
        {"key": "name_en", "value": name_en},
        {"key": "region", "value": region},
        {"key": "pub_year", "value": str(pub_year)},
        {"key": "bbox", "value": str(sheet_meta.get('bbox', []))},
    ])

    df_intervals = pd.DataFrame(common.intervals_for_excel(), columns=["interval_name", "b_age", "t_age"])
    df_desc = pd.DataFrame(desc_rows, columns=["unit_id", "unit_name", "unit_name_ja", "text"])

    # Excel 書き出し
    os.makedirs(os.path.dirname(out_excel_path), exist_ok=True)
    with pd.ExcelWriter(out_excel_path, engine="openpyxl") as writer:
        df_inst.to_excel(writer, sheet_name="Instructions", index=False)
        df_units.to_excel(writer, sheet_name="units_review", index=False)
        df_cols.to_excel(writer, sheet_name="columns_review", index=False)
        df_refs.to_excel(writer, sheet_name="refs_review", index=False)
        df_images.to_excel(writer, sheet_name="images_review", index=False)
        df_meta.to_excel(writer, sheet_name="project_meta", index=False)
        df_gsj.to_excel(writer, sheet_name="gsj_meta", index=False)
        df_intervals.to_excel(writer, sheet_name="intervals", index=False)
        df_desc.to_excel(writer, sheet_name="descriptions", index=False)

        # スタイリング適用
        _format_200k_workbook(writer, df_units, df_cols, df_refs, df_images, df_meta, df_gsj, df_inst, df_intervals, df_desc)

    return len(cols_review_rows), len(units_review_rows)

def _format_200k_workbook(writer, df_units, df_cols, df_refs, df_images, df_meta, df_gsj, df_inst, df_intervals, df_desc):
    fill_ref = PatternFill("solid", start_color="F2F2F2")        # グレー: 参照専用
    fill_input = PatternFill("solid", start_color="FFF2CC")      # 黄色: 要確認・必須
    fill_calc = PatternFill("solid", start_color="DEEAF6")       # 水色: 自動計算
    fill_out = PatternFill("solid", start_color="D9EAD3")        # 緑: 提出物へ転記
    
    font_main = Font(name="Meiryo UI", size=9)
    font_bold = Font(name="Meiryo UI", size=10, bold=True)
    font_header = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
    header_fill_main = PatternFill("solid", start_color="1F4E79")
    header_fill_ref = PatternFill("solid", start_color="595959")
    
    border_thin = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    wrap = Alignment(wrapText=True, vertical="center")
    nowrap = Alignment(wrapText=False, vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # 1. Instructions
    ws = writer.sheets["Instructions"]
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 100
    ws.freeze_panes = "A2"
    for r in range(1, len(df_inst) + 2):
        ws.cell(row=r, column=1).font = font_bold if r == 1 else font_main
        ws.cell(row=r, column=2).font = font_bold if r == 1 else font_main
        ws.cell(row=r, column=1).alignment = wrap
        ws.cell(row=r, column=2).alignment = wrap
        ws.cell(row=r, column=1).border = border_thin
        ws.cell(row=r, column=2).border = border_thin

    # 2. units_review
    ws = writer.sheets["units_review"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    required_cols = {"unit_name", "sort_order", "col_id", "lithology", "t_int", "b_int", "t_age_ma", "b_age_ma"}
    
    col_widths = {
        "unit_id": 14, "col_id": 10, "sort_order": 10, "unit_name": 32, "strat_name": 28,
        "t_int": 16, "t_age_ma": 12, "t_prop": 10, "b_int": 16, "b_age_ma": 12, "b_prop": 10,
        "lithology": 18, "minor_lith": 16, "environment": 16, "unit_description": 40,
        "min_thickness": 12, "max_thickness": 12, "basal_surface": 16, "lateral_relationship": 18,
        "comments": 25, "t_pos": 10,
        "REF_symbol": 14, "REF_age_ja": 20, "REF_age_en": 25, "REF_lith_ja": 22,
        "REF_lith_en": 25, "REF_group_ja": 18, "REF_group_en": 20, "REF_domain": 24
    }

    cols_map = {name: idx for idx, name in enumerate(df_units.columns, 1)}

    for i, name in enumerate(df_units.columns, 1):
        letter = get_column_letter(i)
        header_cell = ws.cell(row=1, column=i)
        header_cell.font = font_header
        header_cell.alignment = center
        header_cell.border = border_thin

        if name.startswith("REF_"):
            header_cell.fill = header_fill_ref
        else:
            header_cell.fill = header_fill_main

        ws.column_dimensions[letter].width = col_widths.get(name, 16)

        for r in range(2, len(df_units) + 2):
            cell = ws.cell(row=r, column=i)
            cell.font = font_main
            cell.border = border_thin

            if name.startswith("REF_"):
                cell.fill = fill_ref
                cell.alignment = nowrap
            elif name in ("t_prop", "b_prop"):
                cell.fill = fill_calc
                cell.alignment = right
                cell.number_format = "0.000"
            elif name in required_cols:
                cell.fill = fill_input
                cell.alignment = center if name in ("sort_order", "col_id", "t_int", "b_int") else (right if "age" in name else nowrap)
            else:
                cell.fill = fill_out
                cell.alignment = nowrap

    # t_prop / b_prop の VLOOKUP 数式を注入
    t_int_c = get_column_letter(cols_map["t_int"])
    t_age_c = get_column_letter(cols_map["t_age_ma"])
    b_int_c = get_column_letter(cols_map["b_int"])
    b_age_c = get_column_letter(cols_map["b_age_ma"])
    t_prop_col = cols_map["t_prop"]
    b_prop_col = cols_map["b_prop"]

    for r in range(2, len(df_units) + 2):
        iv_t, ag_t = f"${t_int_c}{r}", f"${t_age_c}{r}"
        b_t = f"VLOOKUP({iv_t},intervals!$A:$C,2,FALSE)"
        t_t = f"VLOOKUP({iv_t},intervals!$A:$C,3,FALSE)"
        ws.cell(row=r, column=t_prop_col).value = f'=IF(OR({iv_t}="",{ag_t}=""),"",IFERROR(({b_t}-{ag_t})/({b_t}-{t_t}),""))'
        
        iv_b, ag_b = f"${b_int_c}{r}", f"${b_age_c}{r}"
        b_b = f"VLOOKUP({iv_b},intervals!$A:$C,2,FALSE)"
        t_b = f"VLOOKUP({iv_b},intervals!$A:$C,3,FALSE)"
        ws.cell(row=r, column=b_prop_col).value = f'=IF(OR({iv_b}="",{ag_b}=""),"",IFERROR(({b_b}-{ag_b})/({b_b}-{t_b}),""))'

    # 3. columns_review
    ws = writer.sheets["columns_review"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for i, name in enumerate(df_cols.columns, 1):
        letter = get_column_letter(i)
        ws.cell(row=1, column=i).font = font_header
        ws.cell(row=1, column=i).fill = header_fill_main
        ws.cell(row=1, column=i).alignment = center
        ws.cell(row=1, column=i).border = border_thin
        ws.column_dimensions[letter].width = 20 if "name" in name or "geom" in name else 15
        for r in range(2, len(df_cols) + 2):
            c = ws.cell(row=r, column=i)
            c.font = font_main
            c.border = border_thin
            c.alignment = nowrap

    # 4. 残りのシート
    for sheet_name, df_obj in [
        ("refs_review", df_refs),
        ("images_review", df_images),
        ("project_meta", df_meta),
        ("gsj_meta", df_gsj),
        ("intervals", df_intervals),
        ("descriptions", df_desc),
    ]:
        ws = writer.sheets[sheet_name]
        ws.freeze_panes = "A2"
        for i, name in enumerate(df_obj.columns, 1):
            letter = get_column_letter(i)
            ws.cell(row=1, column=i).font = font_header
            ws.cell(row=1, column=i).fill = header_fill_main
            ws.cell(row=1, column=i).alignment = center
            ws.cell(row=1, column=i).border = border_thin
            ws.column_dimensions[letter].width = 25
            for r in range(2, len(df_obj) + 2):
                c = ws.cell(row=r, column=i)
                c.font = font_main
                c.border = border_thin
                c.alignment = nowrap

def make_all_reviews():
    if not os.path.exists(CONFIG_PATH):
        print(f"[ERROR] {CONFIG_PATH} not found.")
        return

    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        sheets = json.load(f)

    total_cols = 0
    total_units = 0

    print(f"Generating Review Sheets for {len(sheets)} 200k sheets (50k compliant format)...")
    for idx, sheet in enumerate(sheets, start=1):
        sheet_code = sheet['sheet_code']
        name_en = sheet.get('name_en', sheet_code) or sheet_code
        name_ja = sheet.get('name_ja', sheet_code)
        region = sheet.get('region', '00_Other')

        cache_file = os.path.join(CACHE_DIR, f"{sheet_code}.json")
        aid_cache = os.path.join(CACHE_DIR, f"{name_en}.json")
        legends = []
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as cf:
                legends = json.load(cf)
        elif os.path.exists(aid_cache):
            with open(aid_cache, 'r', encoding='utf-8') as af:
                legends = json.load(af)

        out_path = os.path.join(REVIEW_BASE_DIR, region, f"m200k_{sheet_code}_{name_en}", f"m200k_{sheet_code}_review.xlsx")
        c_count, u_count = make_review_for_sheet(sheet, legends, out_path)
        total_cols += c_count
        total_units += u_count
        print(f"[{idx}/{len(sheets)}] {sheet_code} ({name_ja}): {c_count} Columns, {u_count} Units -> {os.path.basename(out_path)}")

    print(f"\n=======================================================")
    print(f"SUCCESS: Generated {len(sheets)} Review Workbooks (Total {total_cols} Columns, {total_units} Units)")
    print(f"Output Directory: {REVIEW_BASE_DIR}")
    print(f"=======================================================")

if __name__ == '__main__':
    make_all_reviews()
