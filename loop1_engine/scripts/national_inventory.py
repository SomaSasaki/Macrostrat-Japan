# -*- coding: utf-8 -*-
"""全国 1:50,000 図幅のデータ経路と作業状態を一つの表にする。

通常実行はローカルキャッシュとファイルだけを読むため、Codex/LLM/API usage を
消費しない。``--refresh`` のときだけ GSJ 出版物 API の未取得図幅を照会し、
成功した JSON を図幅ごとに保存するので中断後も再開できる。
"""

from __future__ import annotations

import concurrent.futures
import csv
import json
import os
import re
import sys
import urllib.request
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import get_region_folder, normalize_sheet_code


ROOT = Path(__file__).resolve().parents[2]
MAP_INDEX = ROOT / "config" / "map_index.json"
ZFK_INDEX = ROOT / "config" / "zfk_index.json"
PUB_CACHE = ROOT / "data" / "50k" / "raw" / "publication" / "g050"
OUT_DIR = ROOT / "data" / "50k" / "00_management"
OUT_JSON = OUT_DIR / "gsj_50k_inventory.json"
OUT_CSV = OUT_DIR / "gsj_50k_inventory.csv"
API = "https://gbank.gsj.jp/ld/resource/publication/map/g050/map{}.json"
UA = {"User-Agent": "MacroStrat-GSJ-inventory/1.0"}


INVENTORY_COLUMNS = [
    "map_id", "title_ja", "title_en", "sheet_code", "region_code", "region_folder",
    "pub_year", "publication_status", "zfk_available", "zfk_units",
    "shape_available", "pdf_available", "attribute_available", "xml_available",
    "kml_available", "geotiff_available", "viewer_available", "legend_image_available",
    "map_image_available", "local_shape", "local_pdf", "local_zfk",
    "source_combination", "recommended_route", "confidence_ceiling", "requires_llm",
    "work_status", "review_units", "unit_name_filled", "lithology_filled",
    "sort_order_filled", "conflict_count", "next_action", "review_path",
    "submission_path", "shape_url", "pdf_url", "attribute_url", "viewer_url",
    "legend_image_url", "source_scale",
]


def _load_json(path: Path, default: Any = None) -> Any:
    try:
        with path.open(encoding="utf-8") as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return default


def _save_json_atomic(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(value, f, ensure_ascii=False, indent=2)
    os.replace(tmp, path)


def _fetch_publication(map_id: str, timeout: int = 45) -> tuple[str, dict[str, Any] | None, str]:
    try:
        req = urllib.request.Request(API.format(map_id), headers=UA)
        with urllib.request.urlopen(req, timeout=timeout) as response:
            data = json.loads(response.read().decode("utf-8"))
        return map_id, data, ""
    except Exception as exc:  # 通信エラーを値で返し、既存キャッシュは壊さない
        return map_id, None, str(exc)


def refresh_publication_cache(map_ids: list[str], workers: int = 8,
                              force: bool = False, limit: int | None = None) -> dict[str, str]:
    PUB_CACHE.mkdir(parents=True, exist_ok=True)
    targets = [mid for mid in map_ids if force or not (PUB_CACHE / f"m{mid}.json").exists()]
    if limit is not None:
        targets = targets[:max(0, limit)]
    if not targets:
        print("出版物APIキャッシュは取得済みです。")
        return {}
    workers = max(1, min(int(workers), 12))
    print(f"GSJ 出版物API: {len(targets)} 図幅を照会（並列 {workers}）")
    errors: dict[str, str] = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as executor:
        for done, (mid, data, error) in enumerate(executor.map(_fetch_publication, targets), 1):
            if data:
                _save_json_atomic(PUB_CACHE / f"m{mid}.json", data)
            else:
                errors[mid] = error
            if done % 50 == 0 or done == len(targets):
                print(f"  {done}/{len(targets)}（失敗 {len(errors)}）")
    return errors


def _download_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    rows = data.get("downloadData") or []
    return [row for row in rows if isinstance(row, dict)]


def _asset_flags(data: dict[str, Any] | None) -> dict[str, Any]:
    flags: dict[str, Any] = {
        "shape_available": False, "pdf_available": False, "attribute_available": False,
        "xml_available": False, "kml_available": False, "geotiff_available": False,
        "viewer_available": False, "legend_image_available": False, "map_image_available": False,
        "shape_url": "", "pdf_url": "", "attribute_url": "", "viewer_url": "",
        "legend_image_url": "",
    }
    if not data:
        return flags
    for item in _download_rows(data):
        dtype = str(item.get("data_type") or "").casefold()
        title = str(item.get("title") or "").casefold()
        url = str(item.get("@id") or "")
        lower_url = url.casefold()
        is_shape = ("shapefile" in dtype or "shapefile" in title
                    or "シェープ" in title or "shape" in title)
        is_pdf = dtype == "pdf" or lower_url.endswith(".pdf")
        is_attr = ("attribute" in title or "属性" in title
                   or lower_url.endswith((".xlsx", ".xls")))
        is_xml = "xml" in dtype or lower_url.endswith(".xml")
        is_kml = "kml" in dtype or lower_url.endswith(".kml")
        is_tiff = "geotiff" in dtype or "geotiff" in title or lower_url.endswith((".tif", ".tiff"))
        if is_shape:
            flags["shape_available"] = True
            flags["shape_url"] = flags["shape_url"] or url
        if is_pdf:
            flags["pdf_available"] = True
            flags["pdf_url"] = flags["pdf_url"] or url
        if is_attr:
            flags["attribute_available"] = True
            flags["attribute_url"] = flags["attribute_url"] or url
        flags["xml_available"] = flags["xml_available"] or is_xml
        flags["kml_available"] = flags["kml_available"] or is_kml
        flags["geotiff_available"] = flags["geotiff_available"] or is_tiff
    flags["viewer_url"] = str(data.get("viewer_url") or "")
    pages = data.get("page") or []
    flags["viewer_available"] = bool(flags["viewer_url"] or any(
        page.get("tile_xyz_template") or page.get("tilejson") for page in pages if isinstance(page, dict)))
    for page in pages:
        if not isinstance(page, dict):
            continue
        for legend in page.get("legend") or []:
            for resource in legend.get("resource") or []:
                url = str(resource.get("@id") or "")
                if url:
                    flags["legend_image_available"] = True
                    flags["legend_image_url"] = flags["legend_image_url"] or url
        flags["map_image_available"] = flags["map_image_available"] or bool(page.get("org_data"))
    return flags


def _publication_meta(data: dict[str, Any] | None, fallback: dict[str, Any]) -> dict[str, Any]:
    data = data or {}
    title_ja = str(data.get("title_j") or data.get("title_ja") or "").strip()
    label = str(data.get("label") or data.get("title_e") or "").strip()
    title_en = str(data.get("title_en") or "").strip()
    if not title_en and label:
        match = re.search(r"GeoMap[:']\s*([A-Za-z0-9' -]+)", label)
        title_en = match.group(1).strip(" '") if match else ""
    title_en = title_en or str(fallback.get("name_en") or "")
    title_ja = title_ja or str(fallback.get("name_ja") or "")
    sheet_code = normalize_sheet_code(data.get("sheet_code") or "")
    if not sheet_code:
        for item in _download_rows(data):
            url = str(item.get("@id") or "")
            match = re.search(r"G0?50[_-]?(\d{5})", url, re.I)
            if match:
                sheet_code = normalize_sheet_code(match.group(1))
                break
    if not sheet_code:
        for page in data.get("page") or []:
            tile_name = str(page.get("tms_dir") or page.get("tilejson") or "")
            match = re.search(r"G50[_-]?(\d{2})[_-]?(\d{3})", tile_name, re.I)
            if match:
                sheet_code = normalize_sheet_code(match.group(1) + match.group(2))
                break
    region_code = sheet_code[:2] if len(sheet_code) == 5 else ""
    return {
        "title_ja": title_ja,
        "title_en": title_en,
        "sheet_code": sheet_code,
        "region_code": region_code,
        "region_folder": get_region_folder(sheet_code) if sheet_code else str(fallback.get("region_folder") or ""),
        "pub_year": data.get("pub_year") or data.get("date") or "",
    }


def _map_id_from_path(path: Path) -> str:
    for part in reversed(path.parts):
        match = re.match(r"m(\d+)(?:_|$)", part, re.I)
        if match:
            return match.group(1)
    match = re.search(r"m(\d+)", path.name, re.I)
    return match.group(1) if match else ""


def _local_state() -> dict[str, dict[str, Any]]:
    state: dict[str, dict[str, Any]] = {}
    review_root = ROOT / "data" / "50k" / "02_review"
    submission_root = ROOT / "data" / "50k" / "03_submission"
    for references in review_root.rglob("references") if review_root.exists() else []:
        mid = _map_id_from_path(references)
        if not mid:
            continue
        entry = state.setdefault(mid, {})
        entry["local_shape"] = any(p.name.casefold() == "geo_a.dbf" for p in references.rglob("*.dbf"))
        entry["local_pdf"] = any(p.suffix.casefold() == ".pdf" for p in references.rglob("*.pdf"))

    for folder in (ROOT / "data" / "50k" / "raw" / "zfk").glob("m*"):
        mid = _map_id_from_path(folder)
        if mid and folder.is_dir():
            state.setdefault(mid, {})["local_zfk"] = any((folder / "units").glob("*.json"))

    for path in review_root.rglob("*_review.xlsx") if review_root.exists() else []:
        if path.name.startswith("~$") or ".bak_" in path.name:
            continue
        mid = _map_id_from_path(path)
        if not mid:
            continue
        entry = state.setdefault(mid, {})
        entry["review_path"] = str(path.relative_to(ROOT))
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
            ws = wb["units_review"]
            headers = {str(cell.value): i + 1 for i, cell in enumerate(ws[1]) if cell.value is not None}
            data_rows = [row for row in ws.iter_rows(min_row=2, values_only=True)
                         if any(value not in (None, "") for value in row)]
            entry["review_units"] = len(data_rows)
            for field in ("unit_name", "lithology", "sort_order"):
                index = headers.get(field)
                entry[f"{field}_filled"] = (sum(1 for row in data_rows
                                                  if index and row[index - 1] not in (None, "")))
            conflict_index = headers.get("REF_conflict")
            entry["conflict_count"] = (sum(1 for row in data_rows
                                             if conflict_index and row[conflict_index - 1] not in (None, "")))
            wb.close()
        except Exception as exc:
            entry["review_error"] = str(exc)

    for path in submission_root.rglob("*.xlsx") if submission_root.exists() else []:
        if path.name.startswith("~$"):
            continue
        mid = _map_id_from_path(path)
        if mid:
            state.setdefault(mid, {})["submission_path"] = str(path.relative_to(ROOT))
    return state


def _source_combination(zfk: bool, shape: bool, pdf: bool, viewer_image: bool) -> str:
    names = [name for name, present in (("ZFK", zfk), ("Shape", shape), ("PDF", pdf)) if present]
    return "+".join(names) if names else ("ViewerImage" if viewer_image else "なし/未確認")


def _decision(zfk: bool, shape: bool, pdf: bool, viewer_image: bool, publication_status: str,
              work_status: str, conflicts: int) -> tuple[str, str, bool, str]:
    if zfk:
        route, confidence, llm = "ZFK→Shape検算→PDF補完", "A", False
    elif shape:
        route, confidence, llm = "Shape→PDF補完", "A", False
    elif pdf:
        route, confidence, llm = "PDF→LLM→引用検証", "C", True
    elif viewer_image:
        route, confidence, llm = "図幅画像→OCR/Vision→人確認", "C", True
    elif publication_status != "cached":
        route, confidence, llm = "出版物APIを再照会", "未判定", False
    else:
        route, confidence, llm = "50k資料を追加探索", "D", False

    if work_status == "提出済み":
        action = "提出済み（定期QAのみ）"
    elif conflicts:
        action = f"shape/ZFK競合 {conflicts} 件を確認"
    elif work_status == "レビュー中":
        action = "不足セルと出典を確認"
    elif zfk or shape:
        action = "レビューExcelを機械生成"
    elif pdf:
        action = "PDFをLLM抽出し引用を機械照合"
    elif viewer_image:
        action = "50k凡例画像をOCR/Vision抽出し目視確認"
    elif publication_status != "cached":
        action = "出版物APIキャッシュを更新"
    else:
        action = "50kの冊子・画像・機関リポジトリを追加探索"
    return route, confidence, llm, action


def build_inventory(refresh: bool = False, force: bool = False, workers: int = 8,
                    limit: int | None = None) -> dict[str, Any]:
    maps = _load_json(MAP_INDEX, []) or []
    if not maps:
        raise FileNotFoundError(f"図幅索引がありません: {MAP_INDEX}")
    map_ids = [str(row.get("map_id")) for row in maps if row.get("map_id")]
    errors = refresh_publication_cache(map_ids, workers, force, limit) if refresh else {}

    zfk_data = _load_json(ZFK_INDEX, {}) or {}
    zfk_rows = zfk_data.get("maps", []) if isinstance(zfk_data, dict) else zfk_data
    zfk_by_id = {str(row.get("map_id")): row for row in zfk_rows if row.get("map_id")}
    local = _local_state()
    rows: list[dict[str, Any]] = []

    for base in maps:
        mid = str(base.get("map_id") or "")
        pub_path = PUB_CACHE / f"m{mid}.json"
        pub = _load_json(pub_path)
        zfk = zfk_by_id.get(mid, {})
        loc = local.get(mid, {})
        pub_status = "cached" if pub else ("error" if mid in errors else "unknown")
        meta = _publication_meta(pub, {**base, **zfk})
        assets = _asset_flags(pub)
        zfk_available = bool(zfk)
        shape_available = bool(assets["shape_available"] or loc.get("local_shape"))
        pdf_available = bool(assets["pdf_available"] or loc.get("local_pdf"))
        viewer_image = bool(assets["viewer_available"] or assets["legend_image_available"]
                            or assets["map_image_available"])
        if loc.get("submission_path"):
            work_status = "提出済み"
        elif loc.get("review_path"):
            work_status = "レビュー中"
        else:
            work_status = "未着手"
        conflicts = int(loc.get("conflict_count") or 0)
        route, confidence, requires_llm, next_action = _decision(
            zfk_available, shape_available, pdf_available, viewer_image,
            pub_status, work_status, conflicts)

        row = {
            "map_id": mid,
            **meta,
            "publication_status": pub_status,
            "zfk_available": zfk_available,
            "zfk_units": zfk.get("n_units", ""),
            **assets,
            "shape_available": shape_available,
            "pdf_available": pdf_available,
            "local_shape": bool(loc.get("local_shape")),
            "local_pdf": bool(loc.get("local_pdf")),
            "local_zfk": bool(loc.get("local_zfk")),
            "source_combination": _source_combination(
                zfk_available, shape_available, pdf_available, viewer_image),
            "recommended_route": route,
            "confidence_ceiling": confidence,
            "requires_llm": requires_llm,
            "work_status": work_status,
            "review_units": loc.get("review_units", 0),
            "unit_name_filled": loc.get("unit_name_filled", 0),
            "lithology_filled": loc.get("lithology_filled", 0),
            "sort_order_filled": loc.get("sort_order_filled", 0),
            "conflict_count": conflicts,
            "next_action": next_action,
            "review_path": loc.get("review_path", ""),
            "submission_path": loc.get("submission_path", ""),
            "source_scale": "1:50,000",
        }
        rows.append({column: row.get(column, "") for column in INVENTORY_COLUMNS})

    def sort_key(row: dict[str, Any]) -> tuple[str, str, int]:
        mid = str(row["map_id"])
        return (str(row["region_code"] or "99"), str(row["sheet_code"] or "99999"),
                int(mid) if mid.isdigit() else 999999)

    rows.sort(key=sort_key)
    output = {
        "generated_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "total_maps": len(rows),
        "publication_api": "https://gbank.gsj.jp/ld/resource/publication/map/g050/",
        "zfk_source": zfk_data.get("source", "") if isinstance(zfk_data, dict) else "",
        "refresh_errors": errors,
        "columns": INVENTORY_COLUMNS,
        "maps": rows,
    }
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    _save_json_atomic(OUT_JSON, output)
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=INVENTORY_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)
    return output


def print_summary(output: dict[str, Any]) -> None:
    rows = output["maps"]
    count = lambda key: sum(bool(row.get(key)) for row in rows)
    combos: dict[str, int] = {}
    for row in rows:
        combos[row["source_combination"]] = combos.get(row["source_combination"], 0) + 1
    print(f"\n全国50k管理データ: {len(rows)} 図幅")
    print(f"  出版物APIキャッシュ {sum(r['publication_status']=='cached' for r in rows)}")
    print(f"  ZFK {count('zfk_available')} / Shape {count('shape_available')} / PDF {count('pdf_available')}")
    print(f"  レビュー中 {sum(r['work_status']=='レビュー中' for r in rows)} / "
          f"提出済み {sum(r['work_status']=='提出済み' for r in rows)}")
    for name, n in sorted(combos.items(), key=lambda item: (-item[1], item[0])):
        print(f"    {name}: {n}")
    print(f"  JSON: {OUT_JSON.relative_to(ROOT)}")
    print(f"  CSV : {OUT_CSV.relative_to(ROOT)}")


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="全国50k図幅の管理データを生成")
    parser.add_argument("--refresh", action="store_true", help="未取得の出版物APIを取得")
    parser.add_argument("--force", action="store_true", help="既存の出版物APIキャッシュも更新")
    parser.add_argument("--workers", type=int, default=8, help="API並列数（最大12）")
    parser.add_argument("--limit", type=int, default=None, help="今回照会する図幅数（試験用）")
    args = parser.parse_args()
    try:
        result = build_inventory(args.refresh, args.force, args.workers, args.limit)
        print_summary(result)
    except Exception as exc:
        print(f"[ERROR] {exc}")
        sys.exit(1)
