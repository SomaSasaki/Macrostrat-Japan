# -*- coding: utf-8 -*-
"""ダッシュボードが読む 1 枚の索引 JSON を、既存のローカル資産だけから組み立てる。

ネットワークにも LLM にも触れない。読むのは次の 5 つと 02_review の実体だけ。

    config/gsj_50k_grid.json                     図幅の正規グリッド（scripts/sheet_geometry.py が生成）
    data/50k/gsj_50k_catalog.json                GSJ カタログ（区画名・ベクタ有無）
    data/50k/00_management/gsj_50k_full_census.json  PDF 実査結果（頁数・章立て・分類）
    config/zfk_index.json                        ZFK 保有図幅
    data/50k/00_management/gsj_50k_inventory.json 全国管理表（あれば shape 有無を補う）

出力:
    dashboard/data/index.json          全図幅の軽量索引（地図描画に必要な最小限）
    dashboard/data/detail/m<id>.json   ワークスペースを持つ図幅の詳細カード

    python scripts/dashboard_data.py            生成
    python scripts/dashboard_data.py --summary  生成せず内訳だけ表示
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
GRID = (ROOT / "loop2_governance" / "config" / "gsj_50k_grid.json") if (ROOT / "loop2_governance" / "config" / "gsj_50k_grid.json").is_file() else (ROOT / "config" / "gsj_50k_grid.json")
CATALOG = (ROOT / "loop2_governance" / "data" / "50k" / "gsj_50k_catalog.json") if (ROOT / "loop2_governance" / "data" / "50k" / "gsj_50k_catalog.json").is_file() else (ROOT / "data" / "50k" / "gsj_50k_catalog.json")
CENSUS = (ROOT / "loop2_governance" / "data" / "50k" / "00_management" / "gsj_50k_full_census.json") if (ROOT / "loop2_governance" / "data" / "50k" / "00_management" / "gsj_50k_full_census.json").is_file() else (ROOT / "data" / "50k" / "00_management" / "gsj_50k_full_census.json")
INVENTORY = (ROOT / "loop2_governance" / "data" / "50k" / "00_management" / "gsj_50k_inventory.json") if (ROOT / "loop2_governance" / "data" / "50k" / "00_management" / "gsj_50k_inventory.json").is_file() else (ROOT / "data" / "50k" / "00_management" / "gsj_50k_inventory.json")
ZFK_INDEX = (ROOT / "loop2_governance" / "config" / "zfk_index.json") if (ROOT / "loop2_governance" / "config" / "zfk_index.json").is_file() else (ROOT / "config" / "zfk_index.json")
REVIEW_ROOT = (ROOT / "loop2_governance" / "data" / "50k" / "02_review") if (ROOT / "loop2_governance" / "data" / "50k" / "02_review").is_dir() else (ROOT / "data" / "50k" / "02_review")
SUBMISSION_ROOT = (ROOT / "loop2_governance" / "data" / "50k" / "03_submission") if (ROOT / "loop2_governance" / "data" / "50k" / "03_submission").is_dir() else (ROOT / "data" / "50k" / "03_submission")
OUT_DIR = (ROOT / "loop1_engine" / "dashboard" / "data") if (ROOT / "loop1_engine" / "dashboard").is_dir() else (ROOT / "dashboard" / "data")

# 進捗の段階。左ほど進んでいる。UI の凡例もこの順に並べる。
STAGES = [
    ("submitted", "提出済み"),
    ("review", "レビュー中"),
    ("vector_ready", "ベクタ入手済み"),
    ("pdf_only", "PDFのみ"),
    ("no_source", "資料未確認"),
    ("unpublished", "未刊行"),
]

# データソース種別バッジ。TASK.md の 4 区分に対応する。
SOURCE_BADGES = {
    "zfk_pdf": "ZFK+PDF",
    "shape_pdf": "Shapefile+PDF",
    "pdf_only": "PDF-Only",
    "vector_only": "ベクタのみ",
    "none": "200k-Only",
}

REVIEW_SHEET_CANDIDATES = ("Review", "units_review", "units")
UNIT_FIELDS = ("unit_name", "lithology", "environment", "t_int", "b_int",
               "t_age_ma", "b_age_ma", "b_prop", "t_prop", "sort_order")


def _load_json(path: Path, default: Any = None) -> Any:
    try:
        with path.open(encoding="utf-8") as handle:
            return json.load(handle)
    except (OSError, json.JSONDecodeError):
        return default


def _rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT)).replace(os.sep, "/")
    except ValueError:
        return str(path).replace(os.sep, "/")


# ---------------------------------------------------------------------------
# ワークスペース（02_review / 03_submission）の実体を見る
# ---------------------------------------------------------------------------

def _map_id_from_name(name: str) -> str:
    match = re.match(r"m(\d+)(?:[_\s]|$)", name)
    return match.group(1) if match else ""


def _pick_workbook(folder: Path) -> Path | None:
    """人が編集した本体を最優先し、無ければ最新の candidate を返す。

    本体が雛形のまま（中身が候補より明らかに小さい）場合は candidate を優先する。
    これを見ないと「本体は 19KB の空箱、実体は candidate 側」という現状を見誤る。
    """
    def usable(path: Path) -> bool:
        return path.is_file() and not path.name.startswith("~$") and ".bak_" not in path.name

    primary = [p for p in folder.glob("m*_review.xlsx") if usable(p)]
    candidates = [p for p in folder.glob("m*_review.candidate-*.xlsx") if usable(p)]
    newest_candidate = max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None
    if primary:
        main = max(primary, key=lambda p: p.stat().st_mtime)
        if newest_candidate and newest_candidate.stat().st_size > main.stat().st_size * 2:
            return newest_candidate
        return main
    return newest_candidate


def read_workbook_stats(path: Path) -> dict[str, Any]:
    """レビュー簿から層数と各フィールドの充足数を数える。openpyxl が無ければ空で返す。"""
    stats: dict[str, Any] = {"workbook": _rel(path), "units": 0, "filled": {},
                             "columns": [], "unresolved": [], "error": ""}
    try:
        from openpyxl import load_workbook
    except ImportError:
        stats["error"] = "openpyxl が無いため層数を数えられません"
        return stats
    try:
        book = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:                       # 壊れた簿を握りつぶさず記録する
        stats["error"] = f"読込エラー: {exc}"
        return stats

    try:
        name = next((n for n in REVIEW_SHEET_CANDIDATES if n in book.sheetnames), "")
        if not name:
            stats["error"] = f"層シートが見つかりません: {book.sheetnames}"
            return stats
        sheet = book[name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return stats
        header = {str(cell): index for index, cell in enumerate(rows[0]) if cell is not None}
        data = [row for row in rows[1:] if any(v not in (None, "") for v in row)]
        stats["units"] = len(data)
        stats["sheet_name"] = name

        def value(row: tuple, field: str) -> Any:
            index = header.get(field)
            return None if index is None or index >= len(row) else row[index]

        for field in UNIT_FIELDS:
            if field in header:
                stats["filled"][field] = sum(1 for row in data
                                             if value(row, field) not in (None, ""))
        # 未解決ユニット: 年代・環境・基底関係のいずれかが空のもの
        for row in data:
            missing = [f for f in ("t_int", "b_int", "environment", "b_prop")
                       if f in header and value(row, f) in (None, "")]
            if missing:
                stats["unresolved"].append({
                    "unit_id": str(value(row, "unit_id") or ""),
                    "unit_name": str(value(row, "unit_name") or ""),
                    "missing": missing,
                })

        if "Columns" in book.sheetnames:
            columns = list(book["Columns"].iter_rows(values_only=True))
            if columns:
                head = {str(c): i for i, c in enumerate(columns[0]) if c is not None}
                for row in columns[1:]:
                    if not any(v not in (None, "") for v in row):
                        continue
                    stats["columns"].append({
                        "col_id": str(row[head.get("col_id", 0)] or ""),
                        "col_name": str(row[head.get("col_name", 1)] or ""),
                        "status": str(row[head["status"]] or "") if "status" in head else "",
                    })
    finally:
        book.close()
    return stats


def scan_workspaces() -> dict[str, dict[str, Any]]:
    """02_review 以下を 1 回だけ走査し、図幅 ID ごとの実体を集める。"""
    found: dict[str, dict[str, Any]] = {}
    if not REVIEW_ROOT.is_dir():
        return found
    for region_dir in sorted(REVIEW_ROOT.iterdir()):
        if not region_dir.is_dir():
            continue
        for folder in sorted(region_dir.iterdir()):
            map_id = _map_id_from_name(folder.name) if folder.is_dir() else ""
            if not map_id:
                continue
            entry: dict[str, Any] = {
                "map_id": map_id,
                "folder": _rel(folder),
                "region_folder": region_dir.name,
                "artifacts": {},
                "notes": [],
            }
            column_map = folder / "column_map.png"
            if column_map.is_file():
                entry["artifacts"]["column_map"] = _rel(column_map)
            kml = folder / "column_map.kml"
            if kml.is_file():
                entry["artifacts"]["column_kml"] = _rel(kml)
            pdfs = sorted((folder / "references").rglob("*.pdf")) if (folder / "references").is_dir() else []
            if pdfs:
                entry["artifacts"]["pdf"] = _rel(pdfs[0])
            if (folder / "references").is_dir() and any(
                    p.name.casefold() == "geo_a.dbf" for p in (folder / "references").rglob("*.dbf")):
                entry["local_shape"] = True
            workbook = _pick_workbook(folder)
            if workbook is not None:
                entry["artifacts"]["workbook"] = _rel(workbook)
                entry["workbook_stats"] = read_workbook_stats(workbook)
                primary = folder / f"m{map_id}_review.xlsx"
                if primary.is_file() and primary != workbook:
                    entry["notes"].append(
                        f"本体 {primary.name} は candidate より小さい（雛形のまま）ため、"
                        f"{workbook.name} を集計に使用")
            log = folder / "build_log.jsonl"
            if log.is_file():
                entry["artifacts"]["build_log"] = _rel(log)
            submission = SUBMISSION_ROOT / region_dir.name / folder.name
            exported = sorted(submission.glob("*.xlsx")) if submission.is_dir() else []
            if exported:
                entry["artifacts"]["submission"] = _rel(exported[0])
            found[map_id] = entry
    return found


# ---------------------------------------------------------------------------
# 索引の組み立て
# ---------------------------------------------------------------------------

def _source_badge(zfk: bool, shape: bool, pdf: bool) -> str:
    if zfk and pdf:
        return "zfk_pdf"
    if shape and pdf:
        return "shape_pdf"
    if zfk or shape:
        return "vector_only"
    if pdf:
        return "pdf_only"
    return "none"


def _stage(has_submission: bool, has_workspace: bool, published: bool,
           vector: bool, pdf: bool) -> str:
    if has_submission:
        return "submitted"
    if has_workspace:
        return "review"
    if not published:
        return "unpublished"
    if vector:
        return "vector_ready"
    if pdf:
        return "pdf_only"
    return "no_source"


def _completion(stats: dict[str, Any] | None) -> float:
    """達成率。層数に対する必須 4 フィールドの充足率の平均。"""
    if not stats or not stats.get("units"):
        return 0.0
    units = stats["units"]
    filled = stats.get("filled") or {}
    keys = [k for k in ("unit_name", "lithology", "t_int", "b_int") if k in filled]
    if not keys:
        return 0.0
    return round(sum(min(filled[k], units) for k in keys) / (units * len(keys)), 4)


def build_index() -> dict[str, Any]:
    grid = _load_json(GRID)
    if not grid or not grid.get("sheets"):
        raise FileNotFoundError(
            f"{_rel(GRID)} がありません。先に python scripts/sheet_geometry.py を実行してください。")

    catalog = _load_json(CATALOG, {}) or {}
    region_names = {str(k).zfill(2): str((v or {}).get("name") or "")
                    for k, v in (catalog.get("region_summary") or {}).items()}

    census_rows = (_load_json(CENSUS, {}) or {}).get("results") or []
    census = {str(row.get("map_id")): row for row in census_rows if row.get("map_id")}

    inventory_rows = (_load_json(INVENTORY, {}) or {}).get("maps") or []
    inventory = {str(row.get("map_id")): row for row in inventory_rows if row.get("map_id")}

    zfk_raw = _load_json(ZFK_INDEX, {}) or {}
    zfk_rows = zfk_raw.get("maps", []) if isinstance(zfk_raw, dict) else zfk_raw
    zfk_by_code = {str(r.get("sheet_code")): r for r in zfk_rows or [] if r.get("sheet_code")}
    zfk_by_map = {str(r.get("map_id")): r for r in zfk_rows or [] if r.get("map_id")}

    workspaces = scan_workspaces()

    sheets: list[dict[str, Any]] = []
    details: dict[str, dict[str, Any]] = {}
    for entry in grid["sheets"]:
        code = entry["sheet_code"]
        map_ids = entry.get("map_ids") or []
        map_id = str(entry.get("latest_map_id") or (map_ids[0] if map_ids else ""))
        cen = census.get(map_id, {})
        inv = inventory.get(map_id, {})
        zfk = zfk_by_code.get(code) or zfk_by_map.get(map_id)
        workspace = workspaces.get(map_id)

        published = bool(map_ids)
        has_pdf = bool(cen.get("pdf_url")) or bool(inv.get("pdf_available")) or bool(
            (workspace or {}).get("artifacts", {}).get("pdf"))
        has_shape = bool(inv.get("shape_available")) or bool((workspace or {}).get("local_shape"))
        has_zfk = zfk is not None
        has_submission = bool((workspace or {}).get("artifacts", {}).get("submission"))

        stats = (workspace or {}).get("workbook_stats")
        row = {
            "sheet_code": code,
            "map_id": map_id,
            "region_code": entry["region_code"],
            "region_name": entry.get("region_name") or region_names.get(entry["region_code"], ""),
            "sheet_number": entry["sheet_number"],
            "name_ja": _clean_title(entry.get("title_ja") or cen.get("title_ja") or ""),
            "name_en": _clean_title_en(entry.get("title_en") or cen.get("title_en") or ""),
            "year": entry.get("pub_year") or cen.get("pub_year") or None,
            "bbox": entry["bbox_wgs84"],
            "row": entry["grid_row"],
            "col": entry["grid_col"],
            "p_row": entry["parent_200k_row"],
            "p_col": entry["parent_200k_col"],
            "geometry_source": entry["geometry_source"],
            "published": published,
            "badge": _source_badge(has_zfk, has_shape, has_pdf),
            "stage": _stage(has_submission, workspace is not None, published, has_zfk or has_shape, has_pdf),
            "units": (stats or {}).get("units", 0),
            "completion": _completion(stats),
            "unresolved": len((stats or {}).get("unresolved") or []),
            "columns": len((stats or {}).get("columns") or []),
            "pages": cen.get("total_pages") or 0,
            "column_structure": cen.get("column_structure") or "",
            "pdf_class": cen.get("classified_type") or "",
            "has_workspace": workspace is not None,
        }
        sheets.append(row)

        if workspace is not None:
            details[map_id] = {
                **row,
                "folder": workspace["folder"],
                "artifacts": workspace["artifacts"],
                "notes": workspace["notes"],
                "workbook_stats": stats or {},
                "zfk_units": (zfk or {}).get("n_units", ""),
                "authors": (zfk or {}).get("authors", ""),
                "viewer_url": f"https://gbank.gsj.jp/geonavi/geonavi.php?layers={map_id}" if map_id else "",
                "pdf_url": cen.get("pdf_url") or inv.get("pdf_url") or "",
                "next_action": inv.get("next_action") or "",
            }

    regions = _aggregate_regions(sheets, region_names)
    parents = _aggregate_parents(sheets)
    return {
        "schema": "macrostrat_dashboard/1",
        "generated_at": datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds"),
        "stages": [{"key": k, "label": v} for k, v in STAGES],
        "badges": SOURCE_BADGES,
        "totals": _totals(sheets),
        "grid_validation": grid.get("validation", {}),
        "regions": regions,
        "parents": parents,
        "sheets": sheets,
        "details": sorted(details),
        "_details_payload": details,
    }


def _clean_title(value: str) -> str:
    """「5万分の1地質図幅「一戸」 (2018)」から図幅名だけを取り出す。"""
    match = re.search(r"「([^」]+)」", value or "")
    name = match.group(1) if match else (value or "")
    return re.sub(r"\s*\(\d{4}\)\s*$", "", name).strip()


def _clean_title_en(value: str) -> str:
    match = re.search(r"'([^']+)'", value or "")
    name = match.group(1) if match else (value or "")
    return re.sub(r"\s*\(\d{4}\)\s*$", "", name).strip()


def _totals(sheets: list[dict[str, Any]]) -> dict[str, Any]:
    stages = {key: 0 for key, _ in STAGES}
    badges = {key: 0 for key in SOURCE_BADGES}
    for sheet in sheets:
        stages[sheet["stage"]] = stages.get(sheet["stage"], 0) + 1
        badges[sheet["badge"]] = badges.get(sheet["badge"], 0) + 1
    return {
        "sheets": len(sheets),
        "published": sum(1 for s in sheets if s["published"]),
        "with_workspace": sum(1 for s in sheets if s["has_workspace"]),
        "units": sum(s["units"] for s in sheets),
        "unresolved": sum(s["unresolved"] for s in sheets),
        "by_stage": stages,
        "by_badge": badges,
    }


def _aggregate_regions(sheets: list[dict[str, Any]],
                       region_names: dict[str, str]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for sheet in sheets:
        grouped.setdefault(sheet["region_code"], []).append(sheet)
    out = []
    for code in sorted(grouped):
        members = grouped[code]
        out.append({
            "region_code": code,
            "region_name": members[0].get("region_name") or region_names.get(code, ""),
            "bbox": _union_bbox(members),
            "sheets": len(members),
            **_totals(members),
        })
    return out


def _aggregate_parents(sheets: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[int, int], list[dict[str, Any]]] = {}
    for sheet in sheets:
        grouped.setdefault((sheet["p_row"], sheet["p_col"]), []).append(sheet)
    out = []
    for (prow, pcol) in sorted(grouped):
        members = grouped[(prow, pcol)]
        out.append({
            "p_row": prow,
            "p_col": pcol,
            "bbox": _union_bbox(members),
            "region_codes": sorted({m["region_code"] for m in members}),
            **_totals(members),
        })
    return out


def _union_bbox(sheets: list[dict[str, Any]]) -> list[float]:
    south = min(s["bbox"][0] for s in sheets)
    west = min(s["bbox"][1] for s in sheets)
    north = max(s["bbox"][2] for s in sheets)
    east = max(s["bbox"][3] for s in sheets)
    return [round(south, 6), round(west, 6), round(north, 6), round(east, 6)]


def write_outputs(payload: dict[str, Any], out_dir: Path = OUT_DIR) -> Path:
    details = payload.pop("_details_payload", {})
    out_dir.mkdir(parents=True, exist_ok=True)
    detail_dir = out_dir / "detail"
    detail_dir.mkdir(parents=True, exist_ok=True)
    for map_id, body in details.items():
        _write_json(detail_dir / f"m{map_id}.json", body)
    index_path = out_dir / "index.json"
    _write_json(index_path, payload)
    return index_path


def _write_json(path: Path, body: Any) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as handle:
        json.dump(body, handle, ensure_ascii=False, separators=(",", ":"))
    os.replace(tmp, path)


def print_summary(payload: dict[str, Any]) -> None:
    totals = payload["totals"]
    print(f"図幅 {totals['sheets']} 件（刊行済み {totals['published']} / "
          f"ワークスペース {totals['with_workspace']}）")
    labels = dict(STAGES)
    for key, count in totals["by_stage"].items():
        print(f"  進捗 {labels.get(key, key)}: {count}")
    for key, count in totals["by_badge"].items():
        print(f"  ソース {SOURCE_BADGES.get(key, key)}: {count}")
    print(f"  取り込み済み層数 {totals['units']} / 未解決 {totals['unresolved']}")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="ダッシュボード索引 JSON を生成する")
    parser.add_argument("--summary", action="store_true", help="生成せず内訳だけ表示")
    args = parser.parse_args(argv)
    payload = build_index()
    if args.summary:
        payload.pop("_details_payload", None)
    else:
        path = write_outputs(payload)
        print(f"書き出し: {_rel(path)}")
    print_summary(payload)
    return 0


if __name__ == "__main__":
    for _stream in (sys.stdout, sys.stderr):
        if hasattr(_stream, "reconfigure"):
            _stream.reconfigure(encoding="utf-8", errors="replace")
    raise SystemExit(main())
