# -*- coding: utf-8 -*-
"""
build_200k_skeleton.py — Phase 1: 200k ＋ シームレス地質図V2 全国スケルトン自動生成スクリプト（案B：地質帯分割版）

日本全国の20万分の1地質図幅（全区画・陸域及び海域）について、
1つの図幅内に同居する異なる地質帯・テクトニック単元（付加体、変成帯、深成岩体、中生代盆地、新第三紀盆地、火山弧、第四紀被覆層、海洋地質）ごとに
完全に独立した複合柱状図（Composite Column）を分割生成し、Macrostrat公式フォーマット（v0.1.1）で出力する。
"""

import io
import json
import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, BASE_DIR)

import scripts.common as common

CONFIG_PATH = os.path.join(BASE_DIR, 'config', 'map_index_200k.json')
CACHE_DIR = os.path.join(BASE_DIR, 'data', '200k', 'raw', 'seamless_200k')
OUTPUT_BASE_DIR = os.path.join(BASE_DIR, 'data', '200k', '03_submission', '00_SKELETON_200K')
VOCAB_PATH = os.path.join(BASE_DIR, 'config', 'vocab.json')

# 語彙のロード
intervals_dict = common.load_intervals()
int_names = set(intervals_dict.keys())

with open(VOCAB_PATH, 'r', encoding='utf-8') as f:
    vocab_data = json.load(f)
lith_vocab = {v.lower(): v for v in vocab_data.get('lithology', [])}

def parse_seamless_age(age_str):
    """
    シームレス地質図の formationAge_en から b_int, t_int, b_prop, t_prop, b_age, t_age を解決する
    """
    if not age_str:
        return 'Phanerozoic', 'Phanerozoic', 0.0, 1.0, 541.0, 0.0

    parts = re.split(r'\s*[-–—〜~]\s*', age_str)

    def find_best_interval(part):
        tokens = [t.strip(' ,;') for t in part.split() if t.strip(' ,;')]
        for i in range(len(tokens) - 1, -1, -1):
            t = tokens[i]
            for candidate in [t, t.capitalize()]:
                if candidate in int_names:
                    return candidate
            if i > 0:
                two = f"{tokens[i-1].capitalize()} {tokens[i].capitalize()}"
                if two in int_names:
                    return two
        return None

    b_cand = find_best_interval(parts[0])
    t_cand = find_best_interval(parts[-1]) if len(parts) > 1 else b_cand

    if not b_cand and not t_cand:
        b_cand, t_cand = 'Phanerozoic', 'Phanerozoic'
    elif not b_cand:
        b_cand = t_cand
    elif not t_cand:
        t_cand = b_cand

    b_bounds = common.interval_bounds(b_cand)
    t_bounds = common.interval_bounds(t_cand)

    b_age = b_bounds[0] if b_bounds else 0.0
    t_age = t_bounds[1] if t_bounds else 0.0

    return b_cand, t_cand, 0.0, 1.0, b_age, t_age

def extract_macrostrat_lithologies(text):
    """
    シームレス地質図の lithology_en から Macrostrat公式の main lithology と minor lithology を抽出
    """
    if not text:
        return 'sedimentary rock', None
    text_clean = text.lower().replace('&', ' and ').replace('/', ' ')
    tokens = [t.strip(' ,;') for t in text_clean.split()]

    found = []
    for i in range(len(tokens) - 1):
        two = f"{tokens[i]} {tokens[i+1]}"
        if two in lith_vocab and two not in found:
            found.append(two)
    for t in tokens:
        if t in lith_vocab and t not in found:
            found.append(t)

    if not found:
        if 'volcanic' in text_clean or 'lava' in text_clean:
            return 'volcanic rock', None
        if 'plutonic' in text_clean or 'granit' in text_clean:
            return 'granite', None
        if 'schist' in text_clean or 'metamorph' in text_clean:
            return 'schist', None
        return 'sedimentary rock', None

    main_lith = found[0]
    minor_lith = found[1] if len(found) > 1 else None
    return main_lith, minor_lith

def classify_legend_domain(leg):
    """
    凡例を地質帯・テクトニック単元（Tectonic / Stratigraphic Domain）に分類
    """
    grp = leg.get('group_en', '')
    age = leg.get('formationAge_en', '')
    lith = leg.get('lithology_en', '')

    if 'Accretionary' in grp or 'accretionary' in lith.lower():
        return 'Accretionary Complex'
    if 'Metamorphic' in grp or 'schist' in lith.lower() or 'gneiss' in lith.lower():
        return 'Metamorphic Belt'
    if 'Plutonic' in grp or 'granit' in lith.lower() or 'gabbro' in lith.lower() or 'diorite' in lith.lower():
        return 'Plutonic Complex'
    if 'Volcanic' in grp or 'lava' in lith.lower() or 'pyroclastic' in lith.lower() or 'basalt' in lith.lower() or 'andesite' in lith.lower() or 'rhyolite' in lith.lower() or 'dacite' in lith.lower():
        if any(q in age for q in ['Quaternary', 'Holocene', 'Pleistocene']):
            return 'Quaternary Volcanic Arc'
        return 'Neogene-Paleogene Volcanic Complex'
    if any(q in age for q in ['Quaternary', 'Holocene', 'Pleistocene']):
        return 'Quaternary Basin & Terrace Deposits'
    if any(n in age for n in ['Neogene', 'Miocene', 'Pliocene']):
        return 'Neogene Sedimentary Basin'
    if any(m in age for m in ['Mesozoic', 'Cretaceous', 'Jurassic', 'Triassic']):
        return 'Mesozoic Sedimentary Basin'
    if any(p in age for p in ['Paleozoic', 'Permian', 'Carboniferous', 'Devonian', 'Silurian', 'Ordovician', 'Cambrian']):
        return 'Paleozoic Sedimentary Succession'
    return f'{grp} Succession' if grp else 'Regional Sedimentary Succession'

def infer_environment(domain_name, lith_text):
    text = (domain_name + " " + lith_text).lower()
    if 'accretionary' in text or 'turbidite' in text or 'marine' in text or 'pelagic' in text:
        return 'marine'
    if 'river' in text or 'fluvial' in text or 'floodplain' in text:
        return 'fluvial'
    if 'coastal' in text or 'beach' in text or 'tidal' in text:
        return 'coastal'
    if 'alluvial' in text or 'fan' in text:
        return 'alluvial fan'
    if 'volcanic' in text or 'pyroclastic' in text or 'lava' in text:
        return 'volcanic'
    if 'plutonic' in text or 'granit' in text:
        return 'plutonic'
    if 'metamorphic' in text or 'schist' in text:
        return 'metamorphic'
    return 'unknown'

def infer_basal_surface(domain_name):
    """
    接触関係の推測: 証拠のない10Maルールは廃止。深成岩のみintrusiveとし、他はunknownとする。
    """
    if 'Plutonic' in domain_name:
        return 'intrusive'
    return 'unknown'

def build_columns_for_sheet(sheet_meta, legends, start_col_id):
    """
    1つの図幅から地質帯ごとに複数のColumnオブジェクトを生成する
    """
    sheet_code = sheet_meta['sheet_code']
    name_ja = sheet_meta.get('name_ja', sheet_code)
    name_en = sheet_meta.get('name_en', sheet_code) or sheet_code
    region = sheet_meta.get('region', '00_Other')
    pub_year = sheet_meta.get('pub_year', 2010)

    # 海域図幅の場合（凡例が空）
    if not legends:
        col_id = start_col_id
        col_name = f"{name_en} Marine Succession Column"
        units = [{
            'unit_id': 1,
            'col_id': col_id,
            'section_id': 1,
            'position': 1,
            'b_int': 'Holocene',
            'b_prop': 0.0,
            't_int': 'Holocene',
            't_prop': 1.0,
            'unit_name': f"Marine Sediments ({name_en})",
            'strat_name': f"{name_en} Marine Group",
            'environment': 'marine',
            'unit_description': f"Quaternary Marine Sediments and Sea Floor Deposits ({name_en})",
            'lithology': 'sedimentary rock',
            'minor_lith': 'mudstone',
            'min_thickness': '',
            'max_thickness': '',
            'basal_surface': 'conformable',
            'lateral_relationship': '',
            'comments': f"GSJ 200k Marine Sheet {sheet_code}",
            't_pos': '',
            'source_unit_id': 'Marine_01',
            'source_unit_name_ja': '海洋堆積物・海底地質',
        }]
        col_data = {
            'col_id': col_id,
            'col_name': col_name,
            'col_group': f"GSJ_200K_{sheet_code.replace('-', '_')}_Marine",
            'ref_ids': 1,
            'date_collected': f"{pub_year}-01-01" if pub_year else "2010-01-01",
            'col_type': 'column',
            'axis_type': 'age',
            'b_int': 'Holocene',
            't_int': 'Holocene',
            'b_prop': 0.0,
            't_prop': 1.0,
            'geom': sheet_meta.get('wkt_geom', ''),
            'rgeom': f"POINT({sheet_meta['center'][1]} {sheet_meta['center'][0]})" if 'center' in sheet_meta else '',
            'comments': f"GSJ 200k Marine Succession for {name_ja} ({sheet_code})",
        }
        ref_data = {
            'ref_id': 1,
            'title': f"1:200,000 Geological Map of Japan: {name_ja} ({name_en})",
            'authors': "Geological Survey of Japan, AIST",
            'publication': "Geological Map of Japan 1:200,000",
            'compilation': "GSJ Seamless Geological Map of Japan V2",
            'organization': "Geological Survey of Japan, AIST",
            'date': str(pub_year or 2010),
            'doi': "",
            'url': f"https://www.gsj.jp/Map/JP/geology2.html",
            'comments': f"IMW Sheet {sheet_code}",
        }
        meta_data = {
            'project_name': "GSJ Japan 200k Regional Skeleton",
            'organization': "Macrostrat Japan Project / GSJ",
            'url': "https://macrostrat.org",
            'project_id': "GSJ_200K_SKELETON",
            'compile_date': "2026-08-14",
            'compiler_name': "MacroStrat Automated Pipeline",
            'compiler_orcid': "",
            'col_type': "column",
            'axis_type': "age",
            'b_int': 'Holocene',
            't_int': 'Holocene',
            'b_prop': 0.0,
            't_prop': 1.0,
            'rgeom': col_data['rgeom'],
            'position_unit': "meters",
            'time_unit': "Ma",
            'timescale': "ICS 2020",
            'srid': "4326",
            'comments': "Phase 1 Marine Skeleton generated from GSJ 200k Grid",
        }
        return [{
            'metadata': meta_data,
            'column': col_data,
            'units': units,
            'ref': ref_data,
            'sheet_meta': sheet_meta,
            'domain_name': 'Marine Geological Succession'
        }]

    # 陸域地質凡例を地質帯（Domain）ごとにクラスタリング
    domains = {}
    for leg in legends:
        dom = classify_legend_domain(leg)
        domains.setdefault(dom, []).append(leg)

    generated_cols = []
    curr_col_id = start_col_id

    for dom_name, dom_legends in domains.items():
        raw_units = []
        seen_keys = set()

        for leg in dom_legends:
            age_en = leg.get('formationAge_en', '')
            lith_en = leg.get('lithology_en', '')
            group_en = leg.get('group_en', '')
            age_ja = leg.get('formationAge_ja', '')
            lith_ja = leg.get('lithology_ja', '')

            key = (age_en, lith_en, group_en)
            if key in seen_keys:
                continue
            seen_keys.add(key)

            b_int, t_int, bp, tp, b_age, t_age = parse_seamless_age(age_en)
            main_lith, minor_lith = extract_macrostrat_lithologies(lith_en)
            env = infer_environment(dom_name, lith_en)

            unit_name = f"{t_int} {main_lith.capitalize()} ({name_en} {dom_name.split()[0]})"
            strat_name = f"{name_en} {dom_name}"

            raw_units.append({
                'symbol': leg.get('symbol', ''),
                'unit_name': unit_name,
                'strat_name': strat_name,
                'source_unit_name_ja': f"{age_ja} {lith_ja}",
                'b_int': b_int,
                'b_prop': bp,
                't_int': t_int,
                't_prop': tp,
                'b_age': b_age,
                't_age': t_age,
                'environment': env,
                'lithology': main_lith,
                'minor_lith': minor_lith or '',
                'unit_description': f"{age_en}; {lith_en} [{dom_name}]",
            })

        # 年代順ソート (Bottom: 最古 → Top: 最新)
        # 中心年代 (b_age + t_age)/2 および b_age, t_age の降順（古い順）
        raw_units.sort(key=lambda u: ((u['b_age'] + u['t_age']) / 2.0, u['b_age'], u['t_age']), reverse=True)

        units_data = []
        section_id = 1

        for pos, u in enumerate(raw_units, start=1):
            basal_surface = infer_basal_surface(dom_name)

            units_data.append({
                'unit_id': pos,
                'col_id': curr_col_id,
                'section_id': section_id,
                'position': pos,
                'b_int': u['b_int'],
                'b_prop': u['b_prop'],
                't_int': u['t_int'],
                't_prop': u['t_prop'],
                'unit_name': u['unit_name'],
                'strat_name': u['strat_name'],
                'environment': u['environment'],
                'unit_description': u['unit_description'],
                'lithology': u['lithology'],
                'minor_lith': u['minor_lith'],
                'min_thickness': '',
                'max_thickness': '',
                'basal_surface': basal_surface,
                'lateral_relationship': '',
                'comments': f"GSJ 200k Seamless V2 {u['symbol']}",
                't_pos': '',
                'source_unit_id': u['symbol'],
                'source_unit_name_ja': u['source_unit_name_ja'],
            })

        if raw_units:
            oldest_u = max(raw_units, key=lambda u: u['b_age'])
            youngest_u = min(raw_units, key=lambda u: u['t_age'])
            col_b_int = oldest_u['b_int']
            col_t_int = youngest_u['t_int']
        else:
            col_b_int = 'Phanerozoic'
            col_t_int = 'Holocene'

        dom_clean_tag = dom_name.replace(' ', '_').replace('&', 'and')
        col_name = f"{name_en} ({dom_name}) Column"

        col_data = {
            'col_id': curr_col_id,
            'col_name': col_name,
            'col_group': f"GSJ_200K_{sheet_code.replace('-', '_')}_{dom_clean_tag}",
            'ref_ids': 1,
            'date_collected': f"{pub_year}-01-01" if pub_year else "2010-01-01",
            'col_type': 'column',
            'axis_type': 'age',
            'b_int': col_b_int,
            't_int': col_t_int,
            'b_prop': 0.0,
            't_prop': 1.0,
            'geom': sheet_meta.get('wkt_geom', ''),
            'rgeom': f"POINT({sheet_meta['center'][1]} {sheet_meta['center'][0]})" if 'center' in sheet_meta else '',
            'comments': f"GSJ 200k Tectonic Skeleton: {dom_name} in {name_ja} ({sheet_code})",
        }

        ref_data = {
            'ref_id': 1,
            'title': f"1:200,000 Geological Map of Japan: {name_ja} ({name_en})",
            'authors': "Geological Survey of Japan, AIST",
            'publication': "Geological Map of Japan 1:200,000",
            'compilation': "GSJ Seamless Geological Map of Japan V2",
            'organization': "Geological Survey of Japan, AIST",
            'date': str(pub_year or 2010),
            'doi': "",
            'url': f"https://www.gsj.jp/Map/JP/geology2.html",
            'comments': f"IMW Sheet {sheet_code} - {dom_name}",
        }

        meta_data = {
            'project_name': "GSJ Japan 200k Regional Skeleton",
            'organization': "Macrostrat Japan Project / GSJ",
            'url': "https://macrostrat.org",
            'project_id': "GSJ_200K_SKELETON",
            'compile_date': "2026-08-14",
            'compiler_name': "MacroStrat Automated Pipeline",
            'compiler_orcid': "",
            'col_type': "column",
            'axis_type': "age",
            'b_int': col_b_int,
            't_int': col_t_int,
            'b_prop': 0.0,
            't_prop': 1.0,
            'rgeom': col_data['rgeom'],
            'position_unit': "meters",
            'time_unit': "Ma",
            'timescale': "ICS 2020",
            'srid': "4326",
            'comments': f"Phase 1 Multi-Column Tectonic Skeleton ({dom_name})",
        }

        generated_cols.append({
            'metadata': meta_data,
            'column': col_data,
            'units': units_data,
            'ref': ref_data,
            'sheet_meta': sheet_meta,
            'domain_name': dom_name,
        })
        curr_col_id += 1

    return generated_cols

def export_sheet_workbook(sheet_meta, columns_for_sheet, output_path):
    """
    1つの図幅に含まれるすべての地質帯Columnを1つのExcelファイル（v0.1.1準拠）にエクスポート
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # 1. metadata
    ws_meta = wb.create_sheet(title="metadata")
    ws_meta.append(["key", "value"])
    for k in common.METADATA_KEYS:
        v = columns_for_sheet[0]['metadata'].get(k, '')
        ws_meta.append([k, v])

    # 2. columns
    ws_cols = wb.create_sheet(title="columns")
    ws_cols.append(common.SUBMISSION_COLUMN_COLS)
    for c in columns_for_sheet:
        ws_cols.append([c['column'].get(col, '') for col in common.SUBMISSION_COLUMN_COLS])

    # 3. units
    ws_units = wb.create_sheet(title="units")
    ws_units.append(common.SUBMISSION_UNIT_COLS)
    global_unit_id = 1
    for c in columns_for_sheet:
        for u in c['units']:
            u_row = dict(u)
            u_row['unit_id'] = global_unit_id
            global_unit_id += 1
            ws_units.append([u_row.get(col, '') for col in common.SUBMISSION_UNIT_COLS])

    # 4. refs
    ws_refs = wb.create_sheet(title="refs")
    ws_refs.append(common.SUBMISSION_REF_COLS)
    ws_refs.append([columns_for_sheet[0]['ref'].get(col, '') for col in common.SUBMISSION_REF_COLS])

    # 5. images
    ws_imgs = wb.create_sheet(title="images")
    ws_imgs.append(common.SUBMISSION_IMAGE_COLS)

    # スタイル適用
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for ws in [ws_meta, ws_cols, ws_units, ws_refs, ws_imgs]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

def build_all_skeletons():
    """
    全200k図幅のスケルトンColumn（地質帯分割版）を一括生成
    """
    if not os.path.exists(CONFIG_PATH):
        print(f"[ERROR] {CONFIG_PATH} not found.")
        return

    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        sheets = json.load(f)

    all_columns = []
    global_col_id = 1
    total_sheets_processed = 0

    print(f"Building Multi-Column Tectonic Skeletons for {len(sheets)} sheets (Land + Sea)...")

    for sheet in sheets:
        sheet_code = sheet['sheet_code']
        cache_file = os.path.join(CACHE_DIR, f"{sheet_code}.json")

        legends = []
        if os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as cf:
                legends = json.load(cf)

        cols_for_sheet = build_columns_for_sheet(sheet, legends, start_col_id=global_col_id)
        global_col_id += len(cols_for_sheet)
        all_columns.extend(cols_for_sheet)

        name_en = sheet.get('name_en', sheet_code) or sheet_code
        region = sheet.get('region', '00_Other')

        # 個別図幅 Excel 出力（その図幅に含まれる全地質帯Columnを含む）
        out_excel = os.path.join(OUTPUT_BASE_DIR, region, f"{sheet_code}_{name_en}", f"{name_en}_200k_MultiColumn.xlsx")
        export_sheet_workbook(sheet, cols_for_sheet, out_excel)

        total_sheets_processed += 1
        dom_summary = [c['domain_name'] for c in cols_for_sheet]
        print(f"[{total_sheets_processed}/{len(sheets)}] {sheet_code} ({sheet.get('name_ja', '')}): {len(cols_for_sheet)} Columns -> {dom_summary}")

    # 全国統合マスター Excel の出力
    if all_columns:
        master_wb_path = os.path.join(OUTPUT_BASE_DIR, "Japan_200k_National_Skeleton_MultiColumn.xlsx")
        master_wb = openpyxl.Workbook()
        master_wb.remove(master_wb.active)

        # metadata
        ws_meta = master_wb.create_sheet(title="metadata")
        ws_meta.append(["key", "value"])
        for k in common.METADATA_KEYS:
            v = all_columns[0]['metadata'].get(k, '')
            ws_meta.append([k, v])

        # columns
        ws_cols = master_wb.create_sheet(title="columns")
        ws_cols.append(common.SUBMISSION_COLUMN_COLS)
        for c in all_columns:
            ws_cols.append([c['column'].get(col, '') for col in common.SUBMISSION_COLUMN_COLS])

        # units
        ws_units = master_wb.create_sheet(title="units")
        ws_units.append(common.SUBMISSION_UNIT_COLS)
        global_unit_id = 1
        for c in all_columns:
            for u in c['units']:
                u_copy = dict(u)
                u_copy['unit_id'] = global_unit_id
                global_unit_id += 1
                ws_units.append([u_copy.get(col, '') for col in common.SUBMISSION_UNIT_COLS])

        # refs
        ws_refs = master_wb.create_sheet(title="refs")
        ws_refs.append(common.SUBMISSION_REF_COLS)
        seen_refs = set()
        for c in all_columns:
            r = c['ref']
            if r['ref_id'] not in seen_refs:
                seen_refs.add(r['ref_id'])
                ws_refs.append([r.get(col, '') for col in common.SUBMISSION_REF_COLS])

        # images
        ws_imgs = master_wb.create_sheet(title="images")
        ws_imgs.append(common.SUBMISSION_IMAGE_COLS)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for ws in [ws_meta, ws_cols, ws_units, ws_refs, ws_imgs]:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

        master_wb.save(master_wb_path)
        total_units = sum(len(c['units']) for c in all_columns)
        print(f"\n=======================================================")
        print(f"SUCCESS: Generated {len(all_columns)} Tectonic Columns across {len(sheets)} Sheets (Total {total_units} Units)")
        print(f"Master Workbook: {master_wb_path}")
        print(f"=======================================================")

if __name__ == '__main__':
    build_all_skeletons()
