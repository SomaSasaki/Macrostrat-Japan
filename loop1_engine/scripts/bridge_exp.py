# -*- coding: utf-8 -*-
"""Bridge Module: Connecting Loop 1 (Execution) to Loop 2 (Verification) & Loop 3 (Review).

Sequentially executes:
1. make  (Fetch ZFK/PDF & initialize review workbook)
2. llm   (Run LLM extraction on text/abstract)
3. check (Execute deterministic invariant & monotonicity verification)
4. export (Generate Macrostrat v0.1.1 submission Excel)

Then compiles results into bilingual experimental notebooks:
- experiments/EXP-<id>_<name>/README.ja.md (Japanese: Developer debugging & evidence)
- experiments/EXP-<id>_<name>/README.md    (English: Domain peer review & Macrostrat export)
and updates root EXPERIMENTS catalogs with Obsidian bidirectional links.
"""

import os
import sys
import io
import re
import datetime
import subprocess
from pathlib import Path
from typing import Dict, Any, List, Optional

HERE = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(HERE / "scripts"))

try:
    import common
except ImportError:
    common = None


def run_command_capture(cmd_args: List[str]) -> tuple[int, str, str]:
    """Execute a python CLI command within the workspace and capture stdout/stderr."""
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    proc = subprocess.Popen(
        cmd_args,
        cwd=str(HERE),
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
        errors="replace",
        env=env,
    )
    stdout, stderr = proc.communicate()
    return proc.returncode, stdout, stderr


def parse_review_excel(excel_path: Path) -> List[Dict[str, Any]]:
    """Parse unit records from review Excel file if openpyxl or pandas is available."""
    records = []
    if not excel_path.exists():
        return records

    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        sheet = wb["Review"] if "Review" in wb.sheetnames else wb.active
        
        headers = [str(cell.value or "").strip() for cell in sheet[1]]
        header_map = {h: idx for idx, h in enumerate(headers) if h}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            
            def get_val(col_name: str, default: str = "") -> str:
                idx = header_map.get(col_name)
                if idx is not None and idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip()
                return default

            unit_name = get_val("unit_name") or get_val("REF_unit_name_ja") or get_val("unit_id")
            if not unit_name:
                continue

            records.append({
                "unit_id": get_val("unit_id"),
                "unit_name": unit_name,
                "b_age": get_val("b_age") or get_val("REF_b_age") or "-",
                "t_age": get_val("t_age") or get_val("REF_t_age") or "-",
                "lithology": get_val("lithology") or get_val("REF_lithology") or "-",
                "verbatim_quote": get_val("verbatim_quote") or get_val("REF_verbatim_quote") or "記載なし",
                "status": get_val("status") or "CHECK",
            })
    except Exception as e:
        # Fallback if openpyxl fails or is missing
        pass

    return records


def generate_bilingual_notes(
    exp_id: str,
    sheet_id: str,
    sheet_name: str,
    check_output: str,
    check_code: int,
    units: List[Dict[str, Any]],
    output_dir: Path,
) -> tuple[Path, Path]:
    """Generate both Japanese (README.ja.md) and English (README.md) experimental records."""
    output_dir.mkdir(parents=True, exist_ok=True)
    today_str = datetime.date.today().strftime("%Y-%m-%d")

    # Table rows in Japanese
    ja_table_rows = []
    en_table_rows = []

    for u in units:
        b_age = u.get("b_age", "-")
        t_age = u.get("t_age", "-")
        age_str = f"`{b_age} - {t_age} Ma`" if b_age != "-" and t_age != "-" else "要確認 (Stage 5)"
        quote = u.get("verbatim_quote", "記載なし").replace("\n", " ")
        status = "適合" if u.get("status") in ("OK", "PASS") else "要確認"
        status_en = "Validated" if status == "適合" else "Check"

        ja_table_rows.append(
            f"| **{u['unit_name']}** | {age_str} | `{u.get('lithology', '-')}` | {status} | {quote} |"
        )
        en_table_rows.append(
            f"| **{u['unit_name']}** | {age_str} | `{u.get('lithology', '-')}` | {status_en} | {quote} |"
        )

    ja_table = "\n".join(ja_table_rows) if ja_table_rows else "| - | - | - | - | - |"
    en_table = "\n".join(en_table_rows) if en_table_rows else "| - | - | - | - | - |"

    validation_summary_ja = "全不変条件をクリア (Pass)" if check_code == 0 else "警告または制約違反が検出されました (Review Required)"
    validation_summary_en = "All Invariants Satisfied (Pass)" if check_code == 0 else "Warnings / Constraint Violations Detected (Review Required)"

    # 1. README.ja.md (Japanese Note for Developer & Local Context)
    ja_content = f"""# {exp_id}: GSJ 5万分の1図幅「{sheet_name}」処理実験記録

- **関連全体仕様**: [[PROJECT_STRUCTURE.ja|Macrostrat Japan プロジェクト全体構造と不変条件]]
- **親目録**: [[EXPERIMENTS.ja|実験目録]]

---

## 1. 実験メタデータ
- **実験ID**: `{exp_id}`
- **対象図幅**: `{sheet_name}` (図幅コード: `{sheet_id}`)
- **実行日**: `{today_str}`
- **検証ステータス (`check`)**: **{validation_summary_ja}**

---

## 2. 抽出ファクト ⇄ 原文証拠 対照表 (Claim-Evidence Matrix)
> GSJ地質説明書の原文引用 (`verbatim_quote`) と、決定論的5段階年代解決ルールによる確定年代値の1対1対照です。

| 地層名 (Formation) | 確定年代 (b/t Ma) | 岩相 (Lithology) | ステータス | 原文引用 (GSJ説明書) |
| :--- | :--- | :--- | :--- | :--- |
{ja_table}

---

## 3. 整合性検証ログ (`python run.py check`)
```text
{check_output.strip() if check_output.strip() else "エラーなし (Clean)"}
```

---

## 4. エンジニア検証ログ（第2ループ作業記録）
- [ ] 警告・年代未解決地層の上下層関係を確認。
- [ ] 必要に応じて `scripts/common.py` の補間ルールまたはプロンプトを改修。
"""

    # 2. README.md (English Note for Macrostrat Community & Peers)
    en_content = f"""# {exp_id}: Geological Map Sheet "{sheet_name}" (1:50,000)

- **Architecture Reference**: [[PROJECT_STRUCTURE|Project Architecture & System Invariants]]
- **Catalog Index**: [[EXPERIMENTS|Experiments Catalog]]

---

## 1. Experiment Metadata
- **Experiment ID**: `{exp_id}`
- **Quadrangle**: `{sheet_name}` (Sheet ID: `{sheet_id}`)
- **Execution Date**: `{today_str}`
- **Invariant Validation (`check`)**: **{validation_summary_en}**

---

## 2. Stratigraphic Chronology & Claim-Evidence Table
> Verbatim Japanese citation quotes paired with derived chronostratigraphic ages.

| Formation Name | Age Interval (b/t Ma) | Lithology | Status | Verbatim Evidence Quote |
| :--- | :--- | :--- | :--- | :--- |
{en_table}

---

## 3. Validation Diagnostic Log (`check`)
```text
{check_output.strip() if check_output.strip() else "All invariants passed cleanly."}
```

---

## 4. Artifacts & Deliverables
- **Japanese Record**: [[README.ja|日本語実験ノート]]
- **Submission Excel**: `submission_v011.xlsx` (Macrostrat v0.1.1 schema)
"""

    ja_path = output_dir / "README.ja.md"
    en_path = output_dir / "README.md"

    ja_path.write_text(ja_content, encoding="utf-8")
    en_path.write_text(en_content, encoding="utf-8")

    return ja_path, en_path


def update_experiments_catalogs(
    exp_id: str,
    sheet_id: str,
    sheet_name: str,
    unit_count: int,
    check_code: int,
) -> None:
    """Update root EXPERIMENTS.ja.md and EXPERIMENTS.md files."""
    status_ja = "合格" if check_code == 0 else "要確認"
    status_en = "Pass" if check_code == 0 else "Review"

    ja_row = f"| **[[{exp_id}]]** | `{sheet_id}` | {sheet_name} | {unit_count} | {status_ja} | [[experiments/{exp_id}/README.ja\\|日本語]] | [[experiments/{exp_id}/README\\|English]] |"
    en_row = f"| **[[{exp_id}]]** | `{sheet_id}` | {sheet_name} | {unit_count} | {status_en} | [[experiments/{exp_id}/README.ja\\|日本語]] | [[experiments/{exp_id}/README\\|English]] |"

    def append_row(file_path: Path, new_row: str) -> None:
        if not file_path.exists():
            return
        content = file_path.read_text(encoding="utf-8")
        if exp_id in content:
            return  # Already present
        lines = content.splitlines()
        # Find where table ends
        inserted = False
        for i, line in enumerate(lines):
            if line.strip().startswith("|") and i + 1 < len(lines) and not lines[i + 1].strip().startswith("|"):
                lines.insert(i + 1, new_row)
                inserted = True
                break
        if not inserted:
            lines.append(new_row)
        file_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    append_row(HERE / "EXPERIMENTS.ja.md", ja_row)
    append_row(HERE / "EXPERIMENTS.md", en_row)


def execute_bridge_experiment(target_map: str) -> int:
    """Main bridge orchestration function."""
    print(f"===================================================")
    print(f"[Bridge] Macrostrat 3-Loop Execution: {target_map}")
    print(f"===================================================")

    py_exe = sys.executable

    # Step 1: run.py make
    print(f"\n[1/4] Initializing Review Workbook (run.py make)...")
    code1, out1, err1 = run_command_capture([py_exe, "run.py", "make", target_map])
    if code1 != 0:
        print(f"Warning: make step output: {err1 or out1}")

    # Step 2: run.py llm
    print(f"[2/4] Extracting Ages via LLM (run.py llm)...")
    code2, out2, err2 = run_command_capture([py_exe, "run.py", "llm", target_map])

    # Step 3: run.py check
    print(f"[3/4] Checking Invariants & Monotonicity (run.py check)...")
    code3, out3, err3 = run_command_capture([py_exe, "run.py", "check", target_map])
    check_combined = out3 + ("\n" + err3 if err3 else "")

    # Step 4: run.py export
    print(f"[4/4] Exporting Submission Format (run.py export)...")
    code4, out4, err4 = run_command_capture([py_exe, "run.py", "export", target_map])

    # Resolve sheet ID and files
    sheet_id = target_map
    sheet_name = target_map

    # Try resolving via common if available
    review_file = None
    for p in HERE.glob(f"m*_review.xlsx"):
        if target_map.lower() in p.name.lower() or target_map in p.name:
            review_file = p
            m = re.search(r"m(\d+)_review", p.name)
            if m:
                sheet_id = m.group(1)
            break

    if not review_file:
        for p in HERE.glob(f"data/processed/m*_review.xlsx"):
            review_file = p
            break

    units = parse_review_excel(review_file) if review_file else []

    exp_id = f"EXP-{sheet_id}_{sheet_name}"
    exp_dir = HERE / "experiments" / exp_id

    print(f"\n[Bridge] Compiling Bilingual Experimental Records...")
    ja_path, en_path = generate_bilingual_notes(
        exp_id=exp_id,
        sheet_id=sheet_id,
        sheet_name=sheet_name,
        check_output=check_combined,
        check_code=code3,
        units=units,
        output_dir=exp_dir,
    )

    update_experiments_catalogs(
        exp_id=exp_id,
        sheet_id=sheet_id,
        sheet_name=sheet_name,
        unit_count=len(units),
        check_code=code3,
    )

    print(f"\n[Bridge] Execution Completed:")
    print(f"  - Japanese Record (Loop 2): {ja_path.relative_to(HERE)}")
    print(f"  - English Record (Loop 3):  {en_path.relative_to(HERE)}")
    print(f"  - Catalogs updated: EXPERIMENTS.ja.md & EXPERIMENTS.md")
    print(f"===================================================\n")

    return 0


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("使用法: python scripts/bridge_exp.py <図幅名またはID>")
        sys.exit(1)
    target = sys.argv[1]
    sys.exit(execute_bridge_experiment(target))
