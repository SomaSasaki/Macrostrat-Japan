# -*- coding: utf-8 -*-
"""One-command, legacy-workbook-free GSJ 1:50,000 review pipeline.

The durable workflow is JSON-first:

ZFK + Shapefile + PDF -> canonical JSON -> cached PDF/LLM enrichment
-> Column PNG/KML -> compact Review Excel -> automatic QA.

Only the final Review workbook is intended for human editing.  Runtime code
never uses an existing workbook as a data source.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import subprocess
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence

from column_map_bundle import generate_column_map_from_bundle
from column_geography import (
    apply_vision_assignments,
    append_candidate_evidence,
    canonical_units,
    complete_missing_assignments_from_shape,
    expected_columns,
    select_representative_points,
)
from column_geometry_proposal import (
    build_geometry_proposal,
    build_source_only_box_result,
    discover_source_only_base_proposal,
)
from common import best_interval_for_age, fits_interval, is_blank, props_from_ages
from age_resolution import apply_age_interpolation
from compiled_layer import build_canonical_layer
from derived_previews import write_derived_previews
from llm_extract import BudgetExceeded, MODEL, QuotaExhaustedError, today_usage
from llm_column_vision import ColumnVisionError, run_column_vision
from local_age_extract import apply_local_age_notes
from local_abstract_science import apply_local_abstract_science
from local_japanese_science import apply_local_japanese_science
from map_thumbnail import (
    discover_stratigraphic_legend,
    render_thumbnail,
    write_map_metadata,
)
from map_workspace import WorkspaceError, WorkspaceInfo, prepare_map_workspace
from pdf_image_extract import extract_columnar_images, extract_environment_images
from pdf_body_fallback import completion_evidence_rows, find_missing_unit_body_evidence
from pdf_context_router import build_unit_aliases, route_pdf_contexts
from pdf_alias_mapping import PDFAliasError, run_alias_mapping
from pdf_field_extract import PDFFieldError, run_body_enrichment
from pdf_environment import PDFEnvironmentError, build_targets as build_environment_targets, run_environment_enrichment
from pdf_unit_bootstrap import bootstrap_pdf_units
from pilot_llm import (
    PilotLLMError,
    SourceDocument,
    _read_pdf_index,
    build_queue,
    load_cached_job,
    load_source,
    migrate_compatible_cached_job,
    run_stage,
)
from pilot_raw import build_raw_bundle
from review_v2 import (
    DEFAULT_HELPER,
    ReviewV2Error,
    WorkflowPaths,
    find_artifact_runtime,
    run_spreadsheet_builder,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
QA_SCRIPT = Path(__file__).with_name("qa_review_v2.mjs")


class PilotError(RuntimeError):
    """Actionable pilot failure."""


@dataclass(frozen=True)
class PilotSources:
    map_id: str
    workspace: Path
    zfk_root: Path
    references: Path
    publication: Path | None
    pdf: Path | None
    abstract: Path
    pdf_index: Path
    shape: Path | None
    column_config: Path | None
    llm_cache: Path
    source_manifest: Path


def _utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# GOLD fixtureがsha256で束縛している派生ファイル。上書きの前に必ず退避する。
BOUND_SOURCES = (
    Path("system") / "raw" / "raw_bundle.json",
    Path("system") / "compiled.json",
    Path("system") / "pdf_enrichment" / "routed_contexts.mapped.json",
    Path("system") / "environment_analysis" / "figures" / "environment_figure_candidates.json",
)


def _backup_bound_sources(output_dir: Path, stamp: str) -> None:
    """GOLDが束縛する派生JSONを、再生成で失う前に退避する。

    2026-08-12に、LLMが全滅した --force 実行が raw_bundle.json と
    compiled.json を48 unitから1 unitへ書き換え、GOLD fixtureのsha束縛が
    外れた。復元できる原本がどこにも無かった。
    """

    destination = output_dir / "system" / f"backup-{stamp}"
    saved = 0
    for relative in BOUND_SOURCES:
        source = output_dir / relative
        if not source.is_file():
            continue
        target = destination / relative.name
        target.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, target)
        saved += 1
    if saved:
        print(f"[BACKUP] GOLD束縛ファイル {saved}件 -> system/{destination.name}/")


def _workbook_unit_count(path: Path) -> int:
    """Review Excelのunit行数を数える（読めなければ0）。"""

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True)
        try:
            if "Review" not in workbook.sheetnames:
                return 0
            rows = workbook["Review"].iter_rows(values_only=True)
            header = next(rows, None)
            if not header or "unit_name" not in header:
                return 0
            index = header.index("unit_name")
            return sum(1 for row in rows if row and row[index])
        finally:
            workbook.close()
    except Exception:
        return 0


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_json(path: Path) -> dict[str, Any]:
    try:
        value = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise PilotError(f"JSONを読めません: {path} ({exc})") from exc
    if not isinstance(value, dict):
        raise PilotError(f"JSON objectではありません: {path}")
    return value


def _write_json(path: Path, value: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(
        json.dumps(value, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    os.replace(temporary, path)


def _write_text(path: Path, value: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_name(path.name + ".tmp")
    temporary.write_text(value, encoding="utf-8")
    os.replace(temporary, path)


def _prepare_pdf_derivatives(sources: PilotSources) -> None:
    """Create reusable local Abstract/page-index files when a PDF is present."""
    if sources.pdf is None:
        return
    if not sources.abstract.is_file():
        from extract_abstract import extract

        text, _page_range = extract(str(sources.pdf))
        if text.strip():
            _write_text(sources.abstract, text)
    if not sources.pdf_index.is_file():
        from pdf_locate import build_index

        build_index(str(sources.pdf), str(sources.pdf_index), quiet=True)


def _seed_legacy_towada_cache(cache_dir: Path, map_id: str) -> int:
    """One-time, non-overwriting migration of the already-paid m1050 cache."""
    if str(map_id) != "1050":
        return 0
    legacy = PROJECT_ROOT / "outputs" / "pilot" / "m1050" / "llm_cache"
    if not legacy.is_dir() or legacy.resolve() == cache_dir.resolve():
        return 0
    cache_dir.mkdir(parents=True, exist_ok=True)
    copied = 0
    for source in sorted(legacy.glob("pll_*.json")):
        target = cache_dir / source.name
        if target.exists():
            continue
        try:
            with target.open("xb") as handle:
                handle.write(source.read_bytes())
            copied += 1
        except FileExistsError:
            pass
    return copied


def _asset_path(workspace: WorkspaceInfo, name: str) -> Path | None:
    asset = workspace.assets.get(name)
    if not asset or not asset.available or asset.path is None:
        return None
    path = Path(asset.path).resolve()
    return path if path.exists() else None


def resolve_sources(map_id: str, *, fetch_missing: bool = True) -> PilotSources:
    """Prepare and resolve one map workspace; existing workbooks are ignored."""
    map_id = str(map_id).strip().lstrip("mM")
    if not map_id.isdigit():
        raise PilotError(f"map_idは数字で指定してください: {map_id!r}")
    workspace = prepare_map_workspace(
        map_id, root=PROJECT_ROOT, fetch_missing=fetch_missing
    )
    references = workspace.references_dir.resolve()
    publication = workspace.publication_json.resolve() if workspace.publication_json.is_file() else None
    pdf = _asset_path(workspace, "pdf")
    shape = _asset_path(workspace, "shape")
    zfk_root = workspace.zfk_dir.resolve()
    abstract = references / f"m{map_id}_abstract.txt"
    pdf_index = references / f"m{map_id}_pdfpages.json"
    column_config = PROJECT_ROOT / "config" / "pilots" / f"m{map_id}_columns.json"
    return PilotSources(
        map_id=map_id,
        workspace=workspace.workspace_dir.resolve(),
        zfk_root=zfk_root.resolve(),
        references=references.resolve(),
        publication=publication,
        pdf=pdf,
        abstract=abstract.resolve(),
        pdf_index=pdf_index.resolve(),
        shape=shape,
        column_config=column_config.resolve() if column_config.is_file() else None,
        llm_cache=workspace.llm_cache_dir.resolve(),
        source_manifest=workspace.manifest_path.resolve(),
    )


def apply_column_config(bundle: dict[str, Any], config: Mapping[str, Any],
                        *, config_path: Path) -> dict[str, Any]:
    """Apply reviewed PDF column assignments without consulting a workbook."""
    if str(config.get("map_id") or "") != str(bundle.get("map_id") or ""):
        raise PilotError("Column configのmap_idがraw bundleと一致しません。")
    policy = config.get("runtime_policy") or {}
    if policy.get("legacy_workbook_dependency") is not False or policy.get("legacy_workbook_read_allowed") is not False:
        raise PilotError("Column configがlegacy-workbook-freeを保証していません。")

    units = list(bundle.get("units") or bundle.get("review_v2_input", {}).get("unit_rows") or [])
    assignments = {str(row.get("unit_id")): row for row in config.get("units") or []}
    unit_ids = {str(row.get("unit_id")) for row in units}
    if unit_ids != set(assignments):
        missing = sorted(unit_ids - set(assignments))
        extra = sorted(set(assignments) - unit_ids)
        raise PilotError(
            "Column configがunit inventoryを過不足なく覆っていません"
            f" (missing={missing}, extra={extra})。"
        )

    evidence_rows = list(bundle.get("source_evidence") or bundle.get("review_v2_input", {}).get("evidence_rows") or [])
    evidence_specs = {str(row.get("evidence_id")): row for row in config.get("evidence") or []}
    uncertain = {str(row.get("unit_id")): row for row in config.get("uncertainties") or []}
    for row in units:
        assignment = assignments[str(row.get("unit_id"))]
        column_ids = [str(value) for value in assignment.get("column_ids") or []]
        if not column_ids:
            raise PilotError(f"Column assignmentが空です: {row.get('unit_id')}")
        row["column_id"] = ", ".join(column_ids)
        row["sort_order"] = assignment.get("sort_order")
        if assignment.get("review_status") == "CHECK":
            row["comments"] = "COLUMN CHECK: " + str(assignment.get("rationale") or "Review PDF Fig. 1.")
        else:
            row["comments"] = None
        for record in evidence_rows:
            if str(record.get("unit_id") or "") == str(row.get("unit_id")):
                record["column_id"] = row["column_id"]

        fallback_basis = next(iter(evidence_specs), f"m{bundle.get('map_id')}_pdf_column_split")
        basis_id = next(iter(assignment.get("basis") or []), fallback_basis)
        source = evidence_specs.get(basis_id) or evidence_specs.get(fallback_basis) or {}
        locator = "; ".join(
            bit for bit in (
                source.get("section_or_figure"),
                f"PDF page {source.get('pdf_page')}" if source.get("pdf_page") else None,
                f"printed page {source.get('printed_page')}" if source.get("printed_page") else None,
            ) if bit
        )
        evidence_rows.append({
            "evidence_id": f"{basis_id}_{row['unit_id']}_column",
            "unit_id": row["unit_id"],
            "column_id": row["column_id"],
            "field": "column_id",
            "candidate": row["column_id"],
            "source_type": "PDF",
            "source_file": source.get("file"),
            "source_locator": locator,
            "PDF_page": source.get("pdf_page"),
            "printed_page": source.get("printed_page"),
            "section_or_table": source.get("section_or_figure"),
            "matched_sentence": source.get("matched_text"),
            "full_context_quote": (
                str(source.get("matched_text") or "PDF stratigraphic figure")
                + " Interpretation: " + str(assignment.get("rationale") or "")
            ),
            "confidence_class": assignment.get("assignment_confidence") or "C",
            "explicit": assignment.get("assignment_confidence") == "A",
            "selection": "candidate",
            "extraction_method": "PDF stratigraphic Fig. 1 transcription; review-required subdivision",
        })

    columns = []
    default_ref_id = str(
        ((bundle.get("refs") or [{}])[0]).get("ref_id") or f"gsj{bundle.get('map_id')}"
    )
    for source_column in config.get("columns") or []:
        column = dict(source_column)
        column.setdefault("ref_ids", default_ref_id)
        column.setdefault("col_type", "column")
        column.setdefault("axis_type", "age")
        columns.append(column)
    if not columns:
        raise PilotError("Column configにColumnsがありません。")

    for image in bundle.get("images") or []:
        image["col_ids"] = ", ".join(str(row["col_id"]) for row in columns)
    metadata = dict(bundle.get("review_v2_input", {}).get("project") or {})
    metadata.update({
        "column_split_status": "candidate_review",
        "column_config_file": str(config_path),
        "column_config_sha256": _sha256(config_path),
        "column_assignment_basis": config.get("assignment_basis") or "GSJ PDF stratigraphic classification",
        "column_check_units": sorted(uncertain),
    })
    generated_at = str(bundle.get("compiled", {}).get("generated_at") or _utc_now())
    compiled, evidence = build_canonical_layer(
        units,
        column_rows=columns,
        evidence_rows=evidence_rows,
        metadata=metadata,
        map_id=bundle["map_id"],
        source_review=None,
        generated_at=generated_at,
    )
    bundle["units"] = units
    bundle["columns"] = columns
    bundle["source_evidence"] = evidence_rows
    bundle["compiled"] = compiled
    bundle["evidence"] = evidence
    bundle["review_v2_input"] = {
        "unit_rows": units,
        "column_rows": columns,
        "evidence_rows": evidence_rows,
        "project": metadata,
    }
    bundle["gaps"] = [
        gap for gap in bundle.get("gaps") or []
        if "column split" not in str(gap).casefold()
    ]
    bundle.setdefault("sources", []).append({
        "type": "COLUMN_CONFIG",
        "path": str(config_path),
        "sha256": _sha256(config_path),
    })
    return bundle


def apply_unsplit_column_fallback(
    bundle: dict[str, Any],
    *,
    reason: str = "No reviewed PDF column configuration is available.",
) -> dict[str, Any]:
    """Create a visible one-column review candidate for any unconfigured map."""
    units = [
        dict(row)
        for row in (bundle.get("units") or bundle.get("review_v2_input", {}).get("unit_rows") or [])
    ]
    existing = list(bundle.get("columns") or [])
    seed = dict(existing[0]) if existing else {}
    default_ref_id = str(
        ((bundle.get("refs") or [{}])[0]).get("ref_id") or f"gsj{bundle.get('map_id')}"
    )
    column = {
        **seed,
        "col_id": "unsplit",
        "col_name": seed.get("col_name") or "Unsplit candidate",
        "status": "CHECK",
        "ref_ids": seed.get("ref_ids") or default_ref_id,
        "col_type": seed.get("col_type") or "column",
        "axis_type": seed.get("axis_type") or "age",
        "comments": reason,
    }
    evidence_rows = list(
        bundle.get("source_evidence")
        or bundle.get("review_v2_input", {}).get("evidence_rows")
        or []
    )
    for index, row in enumerate(units, start=1):
        row["column_id"] = "unsplit"
        if row.get("sort_order") in (None, ""):
            row["sort_order"] = index
        prior = str(row.get("comments") or "").strip()
        marker = "COLUMN CHECK: " + reason
        row["comments"] = f"{prior} {marker}".strip() if marker not in prior else prior
        for record in evidence_rows:
            if str(record.get("unit_id") or "") == str(row.get("unit_id") or ""):
                record["column_id"] = "unsplit"
                record["column_ids"] = ["unsplit"]

    metadata = dict(bundle.get("review_v2_input", {}).get("project") or {})
    metadata.update({
        "column_split_status": "unresolved_review",
        "column_assignment_basis": reason,
        "column_check_units": [str(row.get("unit_id")) for row in units],
    })
    generated_at = str(bundle.get("compiled", {}).get("generated_at") or _utc_now())
    compiled, evidence = build_canonical_layer(
        units,
        column_rows=[column],
        evidence_rows=evidence_rows,
        metadata=metadata,
        map_id=bundle.get("map_id"),
        source_review=None,
        generated_at=generated_at,
    )
    bundle.update({
        "units": units,
        "columns": [column],
        "source_evidence": evidence_rows,
        "compiled": compiled,
        "evidence": evidence,
        "review_v2_input": {
            "unit_rows": units,
            "column_rows": [column],
            "evidence_rows": evidence_rows,
            "project": metadata,
        },
    })
    return bundle


def load_compatible_column_proposal(
    vision_dir: Path,
    *,
    map_id: str,
    source_pdf: Path,
    current_units: Sequence[Mapping[str, Any]],
) -> tuple[dict[str, Any], dict[str, Any], Path] | None:
    """Load a prior accepted proposal when a same-run retry cannot call an LLM.

    Cold-start resume may expand the unit inventory after Column Vision has
    already succeeded.  A provider budget failure must not erase those earlier
    source-derived assignments.  Reuse is deliberately strict: the proposal
    must belong to the same map, PDF and rendered figure, and every referenced
    unit ID/name must still be present in the current (possibly larger)
    inventory.  Newly discovered units remain explicitly unassigned.
    """

    proposal_path = vision_dir / "column_proposal.json"
    manifest_path = vision_dir / "vision_manifest.json"
    if not proposal_path.is_file() or not manifest_path.is_file():
        return None
    try:
        proposal = _read_json(proposal_path)
        manifest = _read_json(manifest_path)
        cache_value = str(manifest.get("cache_file") or "").strip()
        if cache_value:
            cache_path = Path(cache_value).resolve()
            workspace = vision_dir.resolve().parents[1]
            if cache_path.is_file() and cache_path.is_relative_to(workspace):
                cache_document = _read_json(cache_path)
                cached_proposal = cache_document.get("proposal")
                cache_identity_matches = (
                    cache_document.get("status") == "complete"
                    and cache_document.get("job_id") == manifest.get("job_id")
                    and cache_document.get("pdf_sha256") == manifest.get("pdf_sha256")
                    and cache_document.get("image_sha256") == manifest.get("image_sha256")
                    and cache_document.get("prompt_sha256") == manifest.get("prompt_sha256")
                )
                if cache_identity_matches and isinstance(cached_proposal, Mapping):
                    # The accepted cache is the immutable model output.  The
                    # visible proposal may contain later source-based
                    # completions, which are recomputed on every resume.
                    proposal = dict(cached_proposal)
        figure_path = Path(str(manifest.get("figure_image") or "")).resolve()
        resolved_vision = vision_dir.resolve()
        if not figure_path.is_file() or not figure_path.is_relative_to(resolved_vision):
            return None
        if str(manifest.get("map_id") or "") != str(map_id):
            return None
        if str(manifest.get("pdf_sha256") or "") != _sha256(source_pdf):
            return None
        if str(manifest.get("image_sha256") or "") != _sha256(figure_path):
            return None
        if not proposal.get("assignment_ready") or proposal.get("status") != "candidate_review":
            return None

        current_by_id = {
            str(row.get("unit_id") or "").strip(): str(row.get("unit_name") or "").strip()
            for row in current_units
            if str(row.get("unit_id") or "").strip()
        }
        columns = proposal.get("columns")
        proposal_units = proposal.get("units")
        if not isinstance(columns, list) or not isinstance(proposal_units, list):
            return None
        column_ids = [str(row.get("column_id") or "").strip() for row in columns]
        if not column_ids or any(not value for value in column_ids):
            return None
        if len(column_ids) != len(set(column_ids)):
            return None

        seen_units: set[str] = set()
        for row in proposal_units:
            unit_id = str(row.get("unit_id") or "").strip()
            unit_name = str(row.get("unit_name") or "").strip()
            if not unit_id or unit_id in seen_units or current_by_id.get(unit_id) != unit_name:
                return None
            seen_units.add(unit_id)
            memberships = row.get("memberships")
            if not isinstance(memberships, list) or not memberships:
                return None
            for membership in memberships:
                if str(membership.get("column_id") or "").strip() not in column_ids:
                    return None
                try:
                    if int(membership.get("sort_order")) <= 0:
                        return None
                except (TypeError, ValueError):
                    return None

        refreshed = json.loads(json.dumps(proposal))
        missing = sorted(set(current_by_id) - seen_units)
        refreshed["unassigned_units"] = missing
        validation = dict(refreshed.get("validation") or {})
        validation.update({
            "canonical_units": len(current_by_id),
            "matched_units": len(seen_units),
            "unassigned_units": len(missing),
        })
        refreshed["validation"] = validation
        return refreshed, dict(manifest), figure_path
    except (OSError, TypeError, ValueError, json.JSONDecodeError):
        return None


def _write_raw_bundle(bundle: Mapping[str, Any], raw_dir: Path) -> None:
    raw_dir.mkdir(parents=True, exist_ok=True)
    _write_json(raw_dir / "raw_bundle.json", bundle)
    _write_json(raw_dir / "compiled.json", bundle["compiled"])
    _write_json(raw_dir / "evidence.json", bundle["evidence"])


def _canonical_evidence_row(record: Mapping[str, Any]) -> dict[str, Any]:
    source = record.get("source") if isinstance(record.get("source"), Mapping) else {}
    confidence = record.get("confidence") if isinstance(record.get("confidence"), Mapping) else {}
    return {
        "evidence_id": record.get("evidence_id"),
        "unit_id": record.get("unit_id"),
        "row_key": record.get("row_key"),
        "column_ids": record.get("column_ids"),
        "scope": record.get("scope"),
        "field": record.get("field"),
        "candidate": record.get("candidate"),
        "source_type": source.get("type"),
        "source_file": source.get("file"),
        "source_locator": source.get("locator"),
        "PDF_page": source.get("pdf_page"),
        "printed_page": source.get("printed_page"),
        "section_or_table": source.get("section"),
        "matched_sentence": source.get("matched_sentence"),
        "full_context_quote": source.get("quote"),
        "confidence_class": confidence.get("class"),
        "assertion": record.get("assertion"),
        "selection": record.get("selection"),
        "conflict": record.get("conflict"),
        "conflict_detail": record.get("conflict_detail"),
        "extraction_method": record.get("extraction_method"),
    }


def normalize_numeric_intervals(output_dir: Path, *, generated_at: str) -> list[dict[str, Any]]:
    """Align coarse GSJ interval labels to verified numeric ages for Review display."""
    compiled = _read_json(output_dir / "compiled.json")
    evidence = _read_json(output_dir / "evidence.json")
    rows: list[dict[str, Any]] = []
    additions: list[dict[str, Any]] = []
    corrections: list[dict[str, Any]] = []
    for unit in compiled.get("units") or []:
        row = dict(unit.get("review_values") or {})
        values = unit.get("values") or {}
        for int_field, age_field in (("t_int", "t_age_ma"), ("b_int", "b_age_ma")):
            age = values.get(age_field)
            current = values.get(int_field)
            if age in (None, ""):
                continue
            interval_fit = fits_interval(age, current)
            if current not in (None, "") and interval_fit is not False:
                continue
            corrected = best_interval_for_age(age, current)
            if not corrected or corrected == current:
                continue
            row[int_field] = corrected
            correction = {
                "unit_id": unit.get("unit_id"), "field": int_field,
                "from": current, "to": corrected, "age_ma": age,
            }
            corrections.append(correction)
            identity = json.dumps(correction, ensure_ascii=False, sort_keys=True)
            additions.append({
                "evidence_id": "ev_" + hashlib.sha256(identity.encode("utf-8")).hexdigest()[:16],
                "unit_id": unit.get("unit_id"),
                "scope_type": "unit_global",
                "field": int_field,
                "candidate": corrected,
                "source_type": "Macrostrat",
                "source_file": "config/intervals.json",
                "source_locator": "international intervals; numeric-age containment",
                "full_context_quote": (
                    f"{age_field}={age} Ma "
                    f"{'has no interval' if current in (None, '') else f'is outside {current}'}; "
                    f"the matching Macrostrat interval is {corrected}."
                ),
                "confidence_class": "A",
                "explicit": True,
                "selection": "selected",
                "extraction_method": "deterministic Macrostrat interval normalization",
            })
        rows.append(row)
    if not corrections:
        return []
    evidence_rows = [
        _canonical_evidence_row(record)
        for record in evidence.get("evidence") or []
    ] + additions
    map_doc = compiled.get("map") or {}
    normalized_compiled, normalized_evidence = build_canonical_layer(
        rows,
        column_rows=map_doc.get("columns") or [],
        evidence_rows=evidence_rows,
        metadata=map_doc.get("metadata") or {},
        map_id=map_doc.get("map_id"),
        source_review=map_doc.get("source_review"),
        generated_at=generated_at,
    )
    _write_json(output_dir / "compiled.json", normalized_compiled)
    _write_json(output_dir / "evidence.json", normalized_evidence)
    return corrections


def _usage_day() -> dict[str, int]:
    _path, _all, _date, day = today_usage()
    return {"calls": int(day.get("calls") or 0), "tokens": int(day.get("tokens") or 0)}


def _finalize_workbook(building: Path, requested: Path) -> tuple[Path, str | None]:
    """Publish the verified workbook without overwriting an Excel-locked file."""
    if not building.exists():
        if requested.exists():
            return requested, None
        alternate = requested.with_name(f"{requested.stem}.updated{requested.suffix}")
        if alternate.exists():
            return alternate, f"{requested.name} is open or locked."
        return requested, None
    try:
        os.replace(building, requested)
        return requested, None
    except (PermissionError, OSError):
        alternate = requested.with_name(f"{requested.stem}.updated{requested.suffix}")
        if alternate.exists():
            stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
            alternate = requested.with_name(
                f"{requested.stem}.updated-{stamp}{requested.suffix}"
            )
        try:
            import shutil
            shutil.copyfile(building, alternate)
            building.unlink(missing_ok=True)
        except Exception:
            pass
        return alternate, (
            f"{requested.name} is open or locked in Excel; the verified update was saved as "
            f"{alternate.name}."
        )


def _builder_input(bundle: Mapping[str, Any], llm_manifest: Mapping[str, Any],
                   compiled: Mapping[str, Any]) -> dict[str, Any]:
    project_meta = list(bundle.get("project_meta") or [])
    project_meta.extend([
        {"key": "pipeline_scope", "value": f"GSJ 1:50,000 map m{bundle.get('map_id')}"},
        {"key": "source_priority", "value": "ZFK > Shapefile > PDF/LLM"},
        {"key": "llm_model", "value": llm_manifest.get("model")},
        {"key": "llm_external_calls", "value": llm_manifest.get("external_calls")},
        {"key": "llm_cache_hits", "value": llm_manifest.get("cache_hits")},
        {"key": "interval_normalizations", "value": len(llm_manifest.get("interval_normalizations") or [])},
        {"key": "review_instruction", "value": "Confirm yellow candidate cells and all CHECK rows before submission."},
    ])
    # Derived fields are visible reference values only.  Review-v2 export
    # deliberately clears and recalculates them from the reviewed ages.
    units = [dict(row) for row in bundle.get("units") or []]
    canonical = list(compiled.get("units") or [])
    for index, row in enumerate(units):
        if index >= len(canonical) or canonical[index].get("unit_id") != row.get("unit_id"):
            continue
        values = canonical[index].get("values") or {}
        derived = canonical[index].get("derived") or {}
        row["t_int"] = values.get("t_int")
        row["b_int"] = values.get("b_int")
        b_prop = derived.get("b_prop")
        t_prop = derived.get("t_prop")
        if b_prop is None and t_prop is None:
            b_prop, t_prop, _is_event = props_from_ages(
                values.get("unit_name"), values.get("t_int"), values.get("b_int"),
                values.get("t_age_ma"), values.get("b_age_ma"),
                values.get("strat_name"), values.get("unit_description"),
            )
        row["b_prop"] = b_prop
        row["t_prop"] = t_prop
    return {
        "schema_version": "pilot-review-input/1.0",
        "map_id": bundle.get("map_id"),
        "units": units,
        "columns": bundle.get("columns") or [],
        "refs": bundle.get("refs") or [],
        "images": bundle.get("images") or [],
        "project_meta": project_meta,
        "gsj_meta": bundle.get("gsj_meta") or [],
    }


def _run_qa(workbook: Path, qa_dir: Path, evidence_json: Path) -> str:
    if not QA_SCRIPT.is_file():
        raise PilotError(f"QA scriptがありません: {QA_SCRIPT}")
    runtime = find_artifact_runtime()
    if runtime.powershell is None:
        raise PilotError("Artifact Tool QAにPowerShellが必要です。")
    command = [
        str(runtime.powershell), "-NoProfile", "-NonInteractive",
        "-ExecutionPolicy", "Bypass", "-File", str(DEFAULT_HELPER),
        "-NodePath", str(runtime.node),
        "-NodeModulesPath", str(runtime.node_modules),
        "-BuilderPath", str(QA_SCRIPT),
        "-InputPath", str(workbook),
        "-OutputPath", str(qa_dir),
        "-MapPath", "", "-MapJsonPath", "", "-KmlPath", "",
        "-EvidenceJsonPath", str(evidence_json),
    ]
    completed = subprocess.run(
        command, capture_output=True, text=True, encoding="utf-8", errors="replace"
    )
    (qa_dir / "qa_command.log").write_text(
        (completed.stdout or "") + ("\nSTDERR\n" + completed.stderr if completed.stderr else ""),
        encoding="utf-8",
    )
    required = [
        qa_dir / "Review.png", qa_dir / "Columns.png",
        qa_dir / "Evidence_first.png", qa_dir / "Project.png",
        qa_dir / "inspection.ndjson",
    ]
    # On Windows the bundled renderer can terminate with 0xC0000409 during
    # native cleanup after every requested PNG/inspection has been flushed.
    # Treat that post-output crash as a warning, never as proof of success.
    if completed.returncode != 0 and not all(path.is_file() and path.stat().st_size for path in required):
        raise PilotError("Excelの描画QAに失敗しました: " + (completed.stderr or completed.stdout)[-1000:])
    return str(qa_dir / "inspection.ndjson")


def _submission_check(
    workbook: Path,
    output_dir: Path,
    *,
    strict: bool = False,
) -> dict[str, Any]:
    completed = subprocess.run(
        [sys.executable, "-X", "utf8", str(Path(__file__).with_name("export_submission.py")),
         str(workbook), "--check-only"],
        capture_output=True, text=True, encoding="utf-8", errors="replace",
    )
    log = (completed.stdout or "") + ("\nSTDERR\n" + completed.stderr if completed.stderr else "")
    (output_dir / "submission_check.txt").write_text(log, encoding="utf-8")
    errors = [line.strip() for line in log.splitlines() if "[ERROR]" in line]
    warnings = [line.strip() for line in log.splitlines() if "[warn]" in line.casefold()]
    if strict and (completed.returncode != 0 or errors):
        raise PilotError("Macrostrat提出前チェックで必須エラーが見つかりました。submission_check.txtを確認してください。")
    return {
        "return_code": completed.returncode,
        "errors": len(errors),
        "warnings": len(warnings),
        "ready_for_submission": completed.returncode == 0 and not errors,
    }


def _apply_semantic_submission_gate(
    check: dict[str, Any],
    compiled: Mapping[str, Any],
    *,
    column_mode: str,
    log_path: Path,
) -> dict[str, Any]:
    """Reject syntactically valid but scientifically unresolved output."""

    semantic_errors: list[str] = []
    semantic_warnings: list[str] = []
    unit_names = []
    missing_lithology: list[str] = []
    missing_age: list[str] = []
    missing_membership: list[str] = []
    for row in compiled.get("units") or []:
        values = row.get("values") if isinstance(row, Mapping) else {}
        if not isinstance(values, Mapping):
            values = row.get("review_values") if isinstance(row, Mapping) else {}
        if isinstance(values, Mapping):
            name = str(values.get("unit_name") or "").strip()
            unit_names.append(name)
            column_ids = [
                str(value).strip()
                for value in (row.get("column_ids") or [])
                if str(value).strip()
            ]
            if not column_ids or any(value.casefold() == "unassigned" for value in column_ids):
                missing_membership.append(name or str(row.get("unit_id") or "(unknown)"))
            if not str(values.get("lithology") or "").strip():
                missing_lithology.append(name or str(row.get("unit_id") or "(unknown)"))
            if not any(
                values.get(field) not in (None, "")
                for field in ("b_int", "t_int", "b_age_ma", "t_age_ma")
            ):
                missing_age.append(name or str(row.get("unit_id") or "(unknown)"))
    if not unit_names or any(name == "NO_DATA" for name in unit_names):
        semantic_errors.append(
            "[ERROR] Canonical unit inventory is a NO_DATA/empty placeholder."
        )
    if column_mode in {"pending", "unsplit_fallback"}:
        semantic_errors.append(
            "[ERROR] No validated Column topology was generated; unsplit fallback is review-only."
        )
    if missing_membership:
        semantic_errors.append(
            f"[ERROR] {len(missing_membership)} canonical unit(s) have no resolved Column membership: "
            + ", ".join(missing_membership[:8])
        )
    if missing_lithology:
        semantic_errors.append(
            f"[ERROR] {len(missing_lithology)} canonical unit(s) have no resolved primary lithology: "
            + ", ".join(missing_lithology[:8])
        )
    if missing_age:
        semantic_errors.append(
            f"[ERROR] {len(missing_age)} canonical unit(s) have no resolved age boundary: "
            + ", ".join(missing_age[:8])
        )

    map_doc = compiled.get("map") if isinstance(compiled.get("map"), Mapping) else {}
    unresolved_columns: list[str] = []
    for column in map_doc.get("columns") or []:
        if not isinstance(column, Mapping):
            continue
        column_id = str(column.get("col_id") or "(unknown)")
        has_location = bool(
            str(column.get("geom") or column.get("rgeom") or "").strip()
            or (column.get("lat") not in (None, "") and column.get("lng") not in (None, ""))
        )
        status = str(column.get("status") or "").strip().casefold()
        evidence = str(
            column.get("coordinate_evidence") or column.get("evidence") or ""
        ).casefold()
        verified = status in {"verified", "ok", "ready"}
        fallback = "fallback" in evidence or "requires review" in evidence
        if not has_location or not verified or fallback:
            unresolved_columns.append(column_id)
    if unresolved_columns:
        semantic_errors.append(
            "[ERROR] Column geometry is absent or unverified for: "
            + ", ".join(unresolved_columns)
        )

    orphaned = int((compiled.get("summary") or {}).get("orphaned_evidence_count") or 0)
    if orphaned:
        semantic_errors.append(
            f"[ERROR] Canonical evidence contains {orphaned} orphaned row(s)."
        )

    status_counts = (compiled.get("summary") or {}).get("status_counts") or {}
    review_rows = int(status_counts.get("CHECK") or 0) + int(status_counts.get("MISSING") or 0)
    if review_rows:
        semantic_errors.append(
            f"[ERROR] {review_rows} canonical unit row(s) still carry CHECK/MISSING status; "
            "scientific review or higher-confidence source resolution is required."
        )

    coverage = {
        "units": len(unit_names),
        "missing_column_membership": len(missing_membership),
        "missing_lithology": len(missing_lithology),
        "missing_age_boundary": len(missing_age),
        "unverified_columns": len(unresolved_columns),
        "orphaned_evidence": orphaned,
    }
    check = dict(check)
    check["scientific_coverage"] = coverage
    if semantic_warnings:
        check["warnings"] = int(check.get("warnings") or 0) + len(semantic_warnings)
        check["semantic_warnings"] = semantic_warnings
    if semantic_errors:
        check["errors"] = int(check.get("errors") or 0) + len(semantic_errors)
        check["ready_for_submission"] = False
        check["semantic_errors"] = semantic_errors
        with log_path.open("a", encoding="utf-8") as handle:
            handle.write("\n\n--- Semantic pipeline gate ---\n")
            handle.write("\n".join(semantic_errors) + "\n")
            if semantic_warnings:
                handle.write("\n".join(semantic_warnings) + "\n")
    return check


def run_pilot(map_id: str, *, output_dir: Path | None = None,
              model: str = MODEL, force: bool = False,
              use_llm: bool = True,
              sources: PilotSources | None = None,
              allow_legacy_cache_migration: bool = True) -> dict[str, Any]:
    """Run one map pipeline, optionally with an explicitly isolated source set.

    ``sources`` is used by the cold-start harness.  Normal CLI callers leave it
    unset and keep the established workspace discovery behavior.
    """
    requested_map_id = str(map_id).strip().lstrip("mM")
    sources = sources or resolve_sources(requested_map_id)
    if str(sources.map_id) != requested_map_id:
        raise PilotError(
            f"Injected sources are for m{sources.map_id}, not m{requested_map_id}."
        )
    output_dir = (
        Path(output_dir).expanduser().resolve()
        if output_dir is not None
        else sources.workspace
    )
    final_workbook = output_dir / f"m{sources.map_id}_review.xlsx"
    building_workbook = output_dir / f"m{sources.map_id}_review.building.xlsx"
    protected = sorted({
        path.resolve()
        for pattern in (f"m{sources.map_id}_review.xlsx", f"m{sources.map_id}_pilot_review.xlsx")
        for path in output_dir.glob(pattern)
        if path.is_file()
    })
    # 既存のReview Excelがある場合、以前は --force を要求して実行自体を止めていた。
    # その結果 --force が常用され、2026-08-12にはLLMが全滅した実行が
    # 48 unitのレビュー結果と派生JSONを1 unitのNO_DATAで上書きした。
    # 実行は止めず、成果物を候補ファイルへ書く。人が編集する本体は触らない。
    stamp = _utc_now().replace(":", "").replace("-", "").replace("+0000", "Z")
    guarded = bool(protected) and not force
    if protected:
        # system/ の派生JSONは実行のたびに作り直される。GOLDがsha256で束縛して
        # いるので、作り直す前に必ず退避する（--force の有無に関わらず）。
        _backup_bound_sources(output_dir, stamp)
    if guarded:
        final_workbook = output_dir / f"m{sources.map_id}_review.candidate-{stamp}.xlsx"
        print("[GUARD] 既存のReview Excelは上書きしません:")
        for path in protected:
            print(f"        {path.name}")
        print(f"        新しい結果はこちらへ書きます: {final_workbook.name}")
    elif protected and force:
        for path in protected:
            backup = path.with_name(f"{path.stem}.before-{stamp}{path.suffix}")
            shutil.copy2(path, backup)
            print(f"[BACKUP] {path.name} -> {backup.name}")
    output_dir.mkdir(parents=True, exist_ok=True)
    system_dir = output_dir / "system"
    system_dir.mkdir(parents=True, exist_ok=True)
    if building_workbook.exists():
        building_workbook.unlink()

    generated_at = _utc_now()
    _prepare_pdf_derivatives(sources)
    review_thumbnail = None
    try:
        review_thumbnail = render_thumbnail(
            sources.references, output_dir / "column_map.png"
        )
    except (OSError, RuntimeError, ValueError) as exc:
        print(f"  [CHECK] GSJ map thumbnail was not prepared: {exc}")
    print("[1/5] 利用可能なZFK・Shape・PDF evidenceを統合")
    bundle = build_raw_bundle(
        sources.map_id,
        sources.zfk_root,
        sources.references,
        sources.publication,
        generated_at=generated_at,
    )
    inventory_source = str(
        (bundle.get("review_v2_input", {}).get("project") or {}).get(
            "unit_inventory_source"
        ) or ""
    )
    usage_before = _usage_day()
    cache_dir = sources.llm_cache
    migrated_cache = (
        _seed_legacy_towada_cache(cache_dir, sources.map_id)
        if allow_legacy_cache_migration else 0
    )
    llm_available = bool(sources.pdf is not None and sources.abstract.is_file())
    source_doc = None
    quota_exhausted_error = ""
    # Loading primary-source text is a local operation.  Keep it available in
    # ``--no-llm`` runs so deterministic PDF vector parsing and cache-only
    # inventory reconstruction are not accidentally disabled with providers.
    if llm_available:
        source_doc = load_source(
            sources.pdf,
            sources.abstract,
            pdf_index=_read_pdf_index(sources.pdf_index),
        )
        # Column subdivision is intentionally not inferred here.  A reviewed
        # config is applied below; maps without one remain an explicit
        # ``Unsplit candidate``.  The previous experimental Vision hook sent
        # the first large PDF image (often a locator map), made an untracked
        # API call, and applied one global geocode centroid to every column.
        # Keep image/Vision/geocoding as a separately cached, reviewable stage
        # before it is allowed to write canonical candidates.

    # PDF-only maps need an inventory before ordinary per-unit enrichment is
    # possible.  The bootstrap prompt already asks for supported unit fields,
    # so the ordinary stage is never layered on top in the same run.
    bootstrap_attempted = bool(
        source_doc is not None and inventory_source == "PDF_PENDING"
    )
    bootstrap_error = ""
    bootstrap_manifest: dict[str, Any] | None = None
    if bootstrap_attempted:
        print("[2/5] PDF/LLM: unit inventoryを1回で作成")
        try:
            bootstrap = bootstrap_pdf_units(
                bundle,
                source_doc,
                cache_dir,
                generated_at=generated_at,
                allow_external_calls=use_llm,
            )
            bundle = bootstrap.bundle
            bootstrap_manifest = {**bootstrap.manifest, "status": "complete"}
        except (PilotLLMError, OSError, ValueError) as exc:
            if isinstance(exc, QuotaExhaustedError):
                quota_exhausted_error = str(exc)
            bootstrap_error = f"{type(exc).__name__}: {exc}"
            route_attempts = list(getattr(exc, "attempts", []) or [])
            bootstrap_manifest = {
                "schema_version": "pdf-unit-bootstrap/1.0",
                "stage": "pdf_unit_bootstrap",
                "status": "failed_review_placeholder",
                "map_id": sources.map_id,
                "model": model,
                "external_calls": sum(
                    1 for row in route_attempts if row.get("attempt_id")
                ),
                "cache_hits": 0,
                "added_evidence": 0,
                "error": bootstrap_error,
                "route_attempts": route_attempts,
            }
            print(
                "  [CHECK] PDF unit inventoryを確定できないため、"
                f"NO_DATA Reviewを生成します: {bootstrap_error}"
            )

    column_mode = "pending"
    column_warning = ""
    column_config_applied = False
    if sources.column_config is not None:
        try:
            bundle = apply_column_config(
                bundle,
                _read_json(sources.column_config),
                config_path=sources.column_config,
            )
            column_mode = "reviewed_config"
            column_config_applied = True
        except PilotError as exc:
            column_warning = str(exc)
            print(f"  [CHECK] Column configは適用しません: {exc}")

    vision_manifest: dict[str, Any] = {
        "stage": "column_geography_vision",
        "status": "not_run",
        "external_calls": 0,
        "cache_hits": 0,
    }
    geography_manifest: dict[str, Any] = {
        "schema_version": "column-geography/1.0",
        "status": "not_run",
        "columns": [],
    }
    proposal: Mapping[str, Any] | None = None
    vision_units = canonical_units(bundle)
    # Source-only geometry reads the current run's canonical inventory from
    # disk.  Persist it before Column extraction instead of accidentally
    # reading a stale prior-run raw bundle.
    raw_dir = system_dir / "raw"
    _write_raw_bundle(bundle, raw_dir)

    if (
        not column_config_applied
        and not bootstrap_error
        and source_doc is not None
        and sources.pdf is not None
        and vision_units
        and all(str(row.get("unit_name") or "").strip() != "NO_DATA" for row in vision_units)
    ):
        vision_dir = system_dir / "column_vision"
        try:
            base_proposal = discover_source_only_base_proposal(sources.workspace)
            box_result = build_source_only_box_result(
                sources.workspace, base_proposal,
            )
            proposal = build_geometry_proposal(
                workspace=sources.workspace,
                box_result=box_result,
                base_proposal=base_proposal,
            )
            _write_json(vision_dir / "column_box_assignments.json", box_result)
            _write_json(vision_dir / "column_proposal.json", proposal)
            vision_manifest = {
                "schema_version": "column-vision/1.0",
                "stage": "column_geography_local_vector",
                "status": "source_only_geometry",
                "map_id": sources.map_id,
                "source_pdf": str(sources.pdf),
                "pdf_sha256": _sha256(sources.pdf),
                "external_calls": 0,
                "cache_hits": 0,
                "assignment_ready": bool(proposal.get("assignment_ready")),
                "matched_units": int(
                    (proposal.get("validation") or {}).get("matched_units") or 0
                ),
                "memberships": int(
                    (proposal.get("validation") or {}).get("memberships") or 0
                ),
                "unassigned_source_inventory": list(
                    proposal.get("unassigned_units") or []
                ),
                "gold_inputs_used": False,
            }
            _write_json(vision_dir / "vision_manifest.json", vision_manifest)
            bundle = apply_vision_assignments(
                bundle,
                proposal,
                source_pdf=str(sources.pdf),
                source_figure=str(sources.pdf),
                # The source summary chart defines the Column deliverable.
                # Member-level bootstrap rows and explicitly omitted deposits
                # remain auditable in the proposal instead of becoming a fake
                # fourth UNASSIGNED Column.
                preserve_unassigned=False,
            )
            bundle, geography_manifest = select_representative_points(
                bundle,
                proposal,
                shape_path=sources.shape,
                output_dir=vision_dir,
                geocode_context=(
                    f"{sources.workspace.name}, {sources.workspace.parent.name}, Japan"
                ),
                map_bbox=review_thumbnail.bbox if review_thumbnail else None,
            )
            column_mode = "local_pdf_vector"
            print(
                "  Local PDF vector Column: "
                f"{len(proposal.get('columns') or [])} Columns / "
                f"{vision_manifest['matched_units']} units / "
                f"{vision_manifest['memberships']} memberships"
            )
        except (OSError, RuntimeError, ValueError) as exc:
            message = f"{type(exc).__name__}: {exc}"
            column_warning = "; ".join(
                value for value in (column_warning, message) if value
            )
            vision_manifest = {
                **vision_manifest,
                "stage": "column_geography_local_vector",
                "status": "failed_review_required",
                "error": message,
                "external_calls": 0,
                "cache_hits": 0,
            }
            print(f"  [CHECK] Local PDF vector Column was not applied: {message}")
    if (
        use_llm
        and column_mode == "pending"
        and not bootstrap_error
        and source_doc is not None
        and sources.pdf is not None
        and vision_units
    ):
        vision_dir = system_dir / "column_vision"
        candidate_dir = vision_dir / "candidates"
        try:
            official_legend = discover_stratigraphic_legend(sources.references)
            try:
                images = extract_columnar_images(
                    str(sources.pdf),
                    str(candidate_dir),
                    max_images=1,
                    pdf_index=source_doc.pdf_index,
                )
            except RuntimeError as exc:
                images = []
                print(f"  [CHECK] PDF層序図の自動選択に失敗: {exc}")
            # The report summary figure is the primary Column source.  An
            # official L1 map legend is useful only when the PDF has no
            # locally rankable/renderable stratigraphic figure.
            using_legend_fallback = not images and official_legend is not None
            if using_legend_fallback:
                images = [str(official_legend)]
            if not images:
                raise ColumnVisionError("No stratigraphic figure page passed local ranking.")
            figure_document = (
                {"candidates": [{"pdf_page": 0, "source": "official GSJ L1 legend fallback"}]}
                if using_legend_fallback
                else _read_json(candidate_dir / "figure_candidates.json")
            )
            figure = (
                figure_document.get("selected_candidate")
                or (figure_document.get("candidates") or [{}])[0]
            )
            pdf_page = int(figure.get("pdf_page") or 0)
            printed_page = None
            printed_pages = (
                source_doc.pdf_index.get("printed")
                if isinstance(source_doc.pdf_index, Mapping)
                else None
            )
            if (
                pdf_page > 0
                and isinstance(printed_pages, list)
                and pdf_page <= len(printed_pages)
            ):
                value = printed_pages[pdf_page - 1]
                if isinstance(value, int) and value > 0:
                    printed_page = value
            if printed_page is None and sources.column_config is not None:
                for record in (_read_json(sources.column_config).get("evidence") or []):
                    if int(record.get("pdf_page") or 0) == pdf_page:
                        printed_page = record.get("printed_page")
                        break
            vision_result = run_column_vision(
                map_id=sources.map_id,
                pdf_path=sources.pdf,
                image_path=images[0],
                pdf_page=pdf_page,
                printed_page=printed_page,
                report_text=source_doc.text,
                units=vision_units,
                expected_columns=expected_columns(bundle) if column_config_applied else (),
                pdf_index=source_doc.pdf_index,
                cache_dir=cache_dir,
                output_dir=vision_dir,
                generated_at=generated_at,
                # Reviewed configurations use the safer closed-world detector.
                # A true cold start has no reviewed Column IDs, so it must use
                # the validated open detector that discovers columns from the
                # selected primary-source figure itself.
                constrained=column_config_applied,
            )
            proposal = vision_result.proposal
            vision_manifest = dict(vision_result.manifest)
            geometry_recovered = False
            if not column_config_applied:
                try:
                    box_result = build_source_only_box_result(
                        sources.workspace, proposal,
                    )
                    recovered = build_geometry_proposal(
                        workspace=sources.workspace,
                        box_result=box_result,
                        base_proposal=proposal,
                    )
                    recovered_count = int(
                        (recovered.get("validation") or {}).get("memberships") or 0
                    )
                    current_count = sum(
                        len(row.get("memberships") or [])
                        for row in proposal.get("units") or []
                        if isinstance(row, Mapping)
                    )
                    if recovered_count >= current_count and recovered.get("assignment_ready"):
                        proposal = recovered
                        geometry_recovered = True
                        _write_json(
                            vision_dir / "column_box_assignments.json", box_result,
                        )
                        _write_json(vision_dir / "column_proposal.json", proposal)
                        vision_manifest["source_only_geometry_recovery"] = {
                            "status": "applied",
                            "memberships": recovered_count,
                            "figure_pdf_page": box_result.get("figure_pdf_page"),
                            "gold_inputs_used": False,
                        }
                except (OSError, ValueError) as exc:
                    vision_manifest["source_only_geometry_recovery"] = {
                        "status": "unavailable",
                        "error": f"{type(exc).__name__}: {exc}",
                        "gold_inputs_used": False,
                    }
            body_matches = find_missing_unit_body_evidence(
                bundle,
                proposal,
                pdf_index=source_doc.pdf_index,
                zfk_root=sources.zfk_root,
                source_pdf=sources.pdf,
            )
            if body_matches:
                _write_json(
                    vision_dir / "japanese_body_matches.json",
                    {
                        "schema_version": "japanese-body-fallback/1.0",
                        "map_id": sources.map_id,
                        "matches": body_matches,
                    },
            )
            shape_completions: list[dict[str, Any]] = []
            if proposal.get("unassigned_units") and not geometry_recovered:
                proposal, shape_completions = complete_missing_assignments_from_shape(
                    bundle,
                    proposal,
                    shape_path=sources.shape,
                    body_matches=body_matches,
                )
                if shape_completions or body_matches:
                    _write_json(
                        vision_dir / "column_proposal_completed.json",
                        {
                            "proposal": proposal,
                            "shape_completions": shape_completions,
                            "japanese_body_matches": body_matches,
                        },
                    )
            candidate_evidence = completion_evidence_rows(
                body_matches,
                shape_completions,
            )
            if candidate_evidence:
                bundle = append_candidate_evidence(
                    bundle,
                    candidate_evidence,
                    metadata_updates={
                        "japanese_body_fallback": (
                            f"{len(body_matches)} page-verified unit context(s); "
                            f"{len(shape_completions)} Shape-completed Column candidate(s)"
                        )
                    },
                )
            vision_manifest.update(
                {
                    "japanese_body_matches": len(body_matches),
                    "shape_fallback_units": len(shape_completions),
                    "assignment_ready_after_shape": bool(proposal.get("assignment_ready")),
                    "completed_proposal_file": (
                        str(vision_dir / "column_proposal_completed.json")
                        if shape_completions or body_matches
                        else None
                    )
                }
            )
            _write_json(vision_dir / "vision_manifest.json", vision_manifest)
            if not column_config_applied and proposal.get("assignment_ready"):
                bundle = apply_vision_assignments(
                    bundle,
                    proposal,
                    source_pdf=str(sources.pdf),
                    source_figure=str(images[0]),
                )
                column_mode = "vision_candidate"
            if column_config_applied or proposal.get("assignment_ready"):
                bundle, geography_manifest = select_representative_points(
                    bundle,
                    proposal,
                    shape_path=sources.shape,
                    output_dir=vision_dir,
                    geocode_context=f"{sources.workspace.name}, {sources.workspace.parent.name}, Japan",
                    map_bbox=review_thumbnail.bbox if review_thumbnail else None,
                )
        except (BudgetExceeded, ColumnVisionError, OSError, ValueError) as exc:
            if isinstance(exc, QuotaExhaustedError):
                quota_exhausted_error = str(exc)
            message = f"{type(exc).__name__}: {exc}"
            reusable = (
                load_compatible_column_proposal(
                    system_dir / "column_vision",
                    map_id=sources.map_id,
                    source_pdf=sources.pdf,
                    current_units=vision_units,
                )
                if proposal is None and not column_config_applied
                else None
            )
            if reusable is not None:
                proposal, prior_manifest, prior_figure = reusable
                _write_json(system_dir / "column_vision" / "column_proposal.json", proposal)
                vision_manifest = {
                    **prior_manifest,
                    "status": "reused_prior_compatible",
                    "external_calls": 0,
                    "cache_hits": 1,
                    "reused_from_generated_at": prior_manifest.get("generated_at"),
                    "retry_error": message,
                    "assignment_ready": True,
                    "unassigned_units": len(proposal.get("unassigned_units") or []),
                }
                for stale_key in (
                    "japanese_body_matches",
                    "member_parent_fallback_units",
                    "shape_fallback_units",
                    "assignment_ready_after_shape",
                    "completed_proposal_file",
                ):
                    vision_manifest.pop(stale_key, None)
                _write_json(
                    system_dir / "column_vision" / "vision_manifest.json",
                    vision_manifest,
                )
                bundle = apply_vision_assignments(
                    bundle,
                    proposal,
                    source_pdf=str(sources.pdf),
                    source_figure=str(prior_figure),
                )
                column_mode = "vision_candidate_reused"
                bundle, geography_manifest = select_representative_points(
                    bundle,
                    proposal,
                    shape_path=sources.shape,
                    output_dir=system_dir / "column_vision",
                    geocode_context=(
                        f"{sources.workspace.name}, {sources.workspace.parent.name}, Japan"
                    ),
                    map_bbox=review_thumbnail.bbox if review_thumbnail else None,
                )
                print(
                    "  [CHECK] Current Column call was unavailable; reused the "
                    f"validated same-source proposal: {message}"
                )
            else:
                column_warning = "; ".join(
                    value for value in (column_warning, message) if value
                )
                vision_manifest = {
                    **vision_manifest,
                    "status": "failed_review_required",
                    "error": message,
                }
                print(f"  [CHECK] Column/PDF geography stage was not applied: {message}")
    if column_mode == "pending":
        reason = column_warning or "No validated Column proposal is available."
        bundle = apply_unsplit_column_fallback(bundle, reason=reason)
        column_mode = "unsplit_fallback"
    compatible_cache_migrations = int(
        bool(vision_manifest.get("compatible_cache_migration"))
    )
    if bootstrap_attempted:
        llm_manifest = bootstrap_manifest or {
            "stage": "pdf_unit_bootstrap",
            "status": "failed_review_placeholder",
            "model": model,
            "external_calls": 0,
            "cache_hits": 0,
            "added_evidence": 0,
        }
        _write_json(system_dir / "compiled.json", bundle["compiled"])
        _write_json(system_dir / "evidence.json", bundle["evidence"])
        _write_json(system_dir / "pilot_llm_stage.json", llm_manifest)
    elif use_llm and source_doc is not None:
        queue = build_queue(
            bundle["compiled"], source_doc, model=model, map_id=sources.map_id
        )
        pending = [job for job in queue if load_cached_job(cache_dir, job) is None]
        for job in list(pending):
            if migrate_compatible_cached_job(cache_dir, job, source_doc) is not None:
                compatible_cache_migrations += 1
        pending = [job for job in queue if load_cached_job(cache_dir, job) is None]
        estimated_tokens = sum(job.estimated_total_tokens for job in pending)
        print(
            f"[2/5] PDF/LLM: {len(pending)} call予定・"
            f"推定{estimated_tokens:,} tokens（cache {len(queue)-len(pending)}）"
        )
        llm_result = run_stage(
            raw_dir, source_doc, system_dir,
            cache_dir=cache_dir, model=model,
            generated_at=generated_at,
            map_id=sources.map_id,
        )
        llm_manifest = _read_json(Path(llm_result.manifest_path))
    else:
        stage = "disabled" if not use_llm else "source_unavailable"
        detail = "テストモード" if not use_llm else "PDFまたは英文Abstractなし"
        print(f"[2/5] PDF/LLMを省略（{detail}）")
        _write_json(system_dir / "compiled.json", bundle["compiled"])
        _write_json(system_dir / "evidence.json", bundle["evidence"])
        llm_manifest = {
            "stage": stage, "model": model, "map_id": sources.map_id,
            "external_calls": 0, "cache_hits": 0, "added_evidence": 0,
            "reason": detail,
        }
        _write_json(system_dir / "pilot_llm_stage.json", llm_manifest)
    usage_after = _usage_day()
    llm_manifest["compatible_cache_jobs_migrated"] = compatible_cache_migrations
    if bootstrap_attempted and bootstrap_error:
        llm_manifest["usage_delta_at_failure"] = {
            "calls": usage_after["calls"] - usage_before["calls"],
            "tokens": usage_after["tokens"] - usage_before["tokens"],
        }
        if llm_manifest["usage_delta_at_failure"]["calls"] > 0:
            llm_manifest["external_calls"] = llm_manifest["usage_delta_at_failure"]["calls"]
        _write_json(system_dir / "pilot_llm_stage.json", llm_manifest)
    enrichment_dir = system_dir / "pdf_enrichment"
    enrichment_dir.mkdir(parents=True, exist_ok=True)
    routed_contexts: dict[str, Any] = {
        "schema_version": "pdf-context-router/1.0",
        "map_id": sources.map_id,
        "contexts": [],
        "unresolved": [],
        "context_characters": 0,
    }
    if source_doc is not None:
        pre_body_compiled = _read_json(system_dir / "compiled.json")
        unit_aliases = build_unit_aliases(pre_body_compiled, sources.zfk_root)
        alias_manifest: dict[str, Any] = {
            "schema_version": "pdf-unit-aliases/1.0",
            "stage": "pdf_unit_alias_mapping",
            "status": "not_needed",
            "external_calls": 0,
            "cache_hits": 0,
            "mapped_units": 0,
        }
        if any(
            row.get("status") == "alias_mapping_required"
            for row in unit_aliases.get("units") or []
        ):
            try:
                unit_aliases, alias_manifest = run_alias_mapping(
                    unit_aliases,
                    source_doc.pdf_index or {},
                    source_sha256=_sha256(sources.pdf),
                    cache_dir=cache_dir,
                    allow_external_calls=use_llm,
                )
            except (PDFAliasError, OSError, ValueError) as exc:
                alias_manifest = {
                    **alias_manifest,
                    "status": "failed_review_required",
                    "error": f"{type(exc).__name__}: {exc}",
                }
                print(f"  [CHECK] PDF unit alias mapping was not applied: {exc}")
        routed_contexts = route_pdf_contexts(
            pre_body_compiled,
            unit_aliases,
            pdf_index=source_doc.pdf_index,
        )
        _write_json(enrichment_dir / "unit_aliases.json", unit_aliases)
        _write_json(enrichment_dir / "routed_contexts.json", routed_contexts)
        _write_json(enrichment_dir / "alias_mapping_manifest.json", alias_manifest)
    else:
        alias_manifest = {
            "schema_version": "pdf-unit-aliases/1.0",
            "stage": "pdf_unit_alias_mapping",
            "status": "source_unavailable",
            "external_calls": 0,
            "cache_hits": 0,
            "mapped_units": 0,
        }
    llm_manifest["pdf_body_routing"] = {
        "contexts": len(routed_contexts.get("contexts") or []),
        "unresolved_units": len(routed_contexts.get("unresolved") or []),
        "context_characters": int(routed_contexts.get("context_characters") or 0),
    }
    local_age_manifest: dict[str, Any] = {
        "schema_version": "local-age-notes/1.0",
        "stage": "local_age_notes",
        "status": "source_unavailable" if source_doc is None else "no_contexts",
        "external_calls": 0,
        "added_evidence": 0,
    }
    if source_doc is not None and routed_contexts.get("contexts"):
        try:
            local_age_manifest = apply_local_age_notes(
                system_dir,
                routed_contexts,
                source_file=sources.pdf,
                generated_at=generated_at,
            )
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            local_age_manifest = {
                **local_age_manifest,
                "status": "failed_review_required",
                "error": f"{type(exc).__name__}: {exc}",
            }
            print(f"  [CHECK] Local analytical-age notes were not applied: {exc}")
    llm_manifest["local_age_notes"] = local_age_manifest
    abstract_science_manifest: dict[str, Any] = {
        "schema_version": "local-abstract-science/1.0",
        "stage": "local_abstract_science",
        "status": "source_unavailable",
        "external_calls": 0,
        "added_evidence": 0,
    }
    if sources.abstract.is_file():
        try:
            abstract_science_manifest = apply_local_abstract_science(
                system_dir,
                sources.abstract,
                generated_at=generated_at,
            )
        except (OSError, ValueError, json.JSONDecodeError) as exc:
            abstract_science_manifest = {
                **abstract_science_manifest,
                "status": "failed_review_required",
                "error": f"{type(exc).__name__}: {exc}",
            }
            print(f"  [CHECK] Local abstract science extraction was not applied: {exc}")
    llm_manifest["local_abstract_science"] = abstract_science_manifest
    japanese_science_manifest: dict[str, Any] = {
        "schema_version": "local-japanese-science/1.0",
        "stage": "local_japanese_science",
        "status": "source_unavailable" if source_doc is None else "no_contexts",
        "external_calls": 0,
        "added_evidence": 0,
    }
    if source_doc is not None and routed_contexts.get("contexts"):
        try:
            japanese_science_manifest = apply_local_japanese_science(
                system_dir,
                routed_contexts,
                source_file=sources.pdf,
                generated_at=generated_at,
            )
        except (OSError, RuntimeError, ValueError, json.JSONDecodeError) as exc:
            japanese_science_manifest = {
                **japanese_science_manifest,
                "status": "failed_review_required",
                "error": f"{type(exc).__name__}: {exc}",
            }
            print(f"  [CHECK] Local Japanese PDF science extraction was not applied: {exc}")
    llm_manifest["local_japanese_science"] = japanese_science_manifest
    body_manifest: dict[str, Any] = {
        "schema_version": "pdf-field-enrichment/1.0",
        "stage": "pdf_body_field_enrichment",
        "status": "not_run",
        "external_calls": 0,
        "cache_hits": 0,
        "added_evidence": 0,
    }
    if source_doc is not None and routed_contexts.get("contexts"):
        try:
            body_manifest = run_body_enrichment(
                system_dir,
                routed_contexts,
                source_file=sources.pdf,
                source_sha256=_sha256(sources.pdf),
                cache_dir=cache_dir,
                generated_at=generated_at,
                allow_external_calls=use_llm,
            )
        except (PDFFieldError, OSError, ValueError) as exc:
            if isinstance(exc, QuotaExhaustedError):
                quota_exhausted_error = str(exc)
            body_manifest = {
                **body_manifest,
                "status": "failed_review_required",
                "error": f"{type(exc).__name__}: {exc}",
            }
            print(f"  [CHECK] Targeted PDF body enrichment was not applied: {exc}")
    llm_manifest["pdf_body"] = body_manifest
    llm_manifest["pdf_alias_mapping"] = alias_manifest
    environment_manifest: dict[str, Any] = {
        "schema_version": "pdf-environment-analysis/1.0",
        "stage": "pdf_environment_multimodal",
        "status": "not_run",
        "external_calls": 0,
        "cache_hits": 0,
        "added_evidence": 0,
    }
    if use_llm and source_doc is not None and routed_contexts.get("contexts"):
        environment_compiled = _read_json(system_dir / "compiled.json")
        environment_targets, environment_unresolved = build_environment_targets(
            environment_compiled, routed_contexts
        )
        if environment_targets:
            environment_dir = system_dir / "environment_analysis"
            figure_dir = environment_dir / "figures"
            try:
                environment_images = extract_environment_images(
                    str(sources.pdf),
                    str(figure_dir),
                    max_images=3,
                    pdf_index=source_doc.pdf_index,
                )
                figure_manifest_path = figure_dir / "environment_figure_candidates.json"
                figure_manifest = (
                    _read_json(figure_manifest_path)
                    if figure_manifest_path.is_file()
                    else {"candidates": []}
                )
                environment_manifest = run_environment_enrichment(
                    system_dir,
                    routed_contexts,
                    source_file=sources.pdf,
                    source_sha256=_sha256(sources.pdf),
                    image_paths=environment_images,
                    figure_manifest=figure_manifest,
                    cache_dir=cache_dir,
                    generated_at=generated_at,
                    constrained=True,
                )
                print(
                    "  PDF environment: "
                    f"{environment_manifest.get('accepted', 0)} accepted / "
                    f"{environment_manifest.get('target_units', 0)} targets / "
                    f"{environment_manifest.get('external_calls', 0)} call"
                )
            except (PDFEnvironmentError, OSError, RuntimeError, ValueError) as exc:
                if isinstance(exc, QuotaExhaustedError):
                    quota_exhausted_error = str(exc)
                environment_manifest = {
                    **environment_manifest,
                    "status": "failed_review_required",
                    "target_units": len(environment_targets),
                    "unresolved": environment_unresolved,
                    "error": f"{type(exc).__name__}: {exc}",
                }
                print(f"  [CHECK] Multimodal environment analysis was not applied: {exc}")
        else:
            environment_manifest = {
                **environment_manifest,
                "status": "no_targets",
                "unresolved": environment_unresolved,
            }
    elif source_doc is None:
        environment_manifest["status"] = "source_unavailable"
    elif not use_llm:
        environment_manifest["status"] = "disabled"
    llm_manifest["pdf_environment"] = environment_manifest
    llm_manifest["quota_exhausted"] = bool(quota_exhausted_error)
    llm_manifest["quota_error"] = quota_exhausted_error or None
    llm_manifest["external_calls"] = (
        int(llm_manifest.get("external_calls") or 0)
        + int(alias_manifest.get("external_calls") or 0)
        + int(body_manifest.get("external_calls") or 0)
        + int(environment_manifest.get("external_calls") or 0)
    )
    llm_manifest["cache_hits"] = (
        int(llm_manifest.get("cache_hits") or 0)
        + int(alias_manifest.get("cache_hits") or 0)
        + int(body_manifest.get("cache_hits") or 0)
        + int(environment_manifest.get("cache_hits") or 0)
    )
    llm_manifest["added_evidence"] = (
        int(llm_manifest.get("added_evidence") or 0)
        + int(local_age_manifest.get("added_evidence") or 0)
        + int(abstract_science_manifest.get("added_evidence") or 0)
        + int(japanese_science_manifest.get("added_evidence") or 0)
        + int(body_manifest.get("added_evidence") or 0)
        + int(environment_manifest.get("added_evidence") or 0)
    )
    usage_after = _usage_day()
    age_interpolation_manifest = apply_age_interpolation(
        system_dir, generated_at=generated_at
    )
    llm_manifest["age_interpolation"] = age_interpolation_manifest
    llm_manifest["added_evidence"] = (
        int(llm_manifest.get("added_evidence") or 0)
        + int(age_interpolation_manifest.get("added_evidence") or 0)
    )
    interval_corrections = normalize_numeric_intervals(
        system_dir, generated_at=generated_at
    )
    derived_previews = write_derived_previews(system_dir)
    scoped_compiled = _read_json(system_dir / "compiled.json")
    orphaned_evidence = int(
        (scoped_compiled.get("summary") or {}).get("orphaned_evidence_count") or 0
    )
    if orphaned_evidence:
        raise PilotError(
            "Canonical evidence binding failed: "
            f"{orphaned_evidence} evidence record(s) are not attached to a unit/Column row."
        )
    llm_manifest["interval_normalizations"] = interval_corrections
    llm_manifest["derived_preview_rows"] = len(derived_previews.get("rows") or [])
    _write_json(system_dir / "pilot_llm_stage.json", llm_manifest)

    map_result = None
    map_error = ""
    if sources.shape is not None:
        print("[3/5] Column色分け地図・Google Earth KMLを生成")
        try:
            map_result = generate_column_map_from_bundle(
                system_dir, sources.shape, output_dir,
                stem="column_map",
                title=f"m{sources.map_id} {sources.workspace.name} Column candidates",
            )
            root_map_json = output_dir / "column_map.json"
            system_map_json = system_dir / "column_map.json"
            if root_map_json.is_file():
                os.replace(root_map_json, system_map_json)
                map_document = _read_json(system_map_json)
                map_document["json_path"] = str(system_map_json)
                _write_json(system_map_json, map_document)
        except Exception as exc:
            map_error = f"{type(exc).__name__}: {exc}"
            print(f"  [CHECK] 地図生成を省略してExcel作成を続行します: {map_error}")
    elif review_thumbnail is not None:
        print("[3/5] Shapeなし：GSJ GeoTIFF thumbnail と候補点を生成")
        try:
            review_thumbnail = render_thumbnail(
                sources.references,
                output_dir / "column_map.png",
                columns=bundle.get("columns") or [],
            )
            if review_thumbnail is not None:
                write_map_metadata(
                    system_dir / "column_map.json",
                    review_thumbnail,
                    bundle.get("columns") or [],
                )
                map_result = {"mode": "geotiff_thumbnail"}
        except (OSError, RuntimeError, ValueError) as exc:
            map_error = f"{type(exc).__name__}: {exc}"
            print(f"  [CHECK] GSJ map thumbnail was not finalized: {map_error}")
    else:
        print("[3/5] Shape/GeoTIFFなし：地図/KMLを省略")

    final_compiled = _read_json(system_dir / "compiled.json")
    builder_input = _builder_input(bundle, llm_manifest, final_compiled)
    builder_input_path = system_dir / "review_input.json"
    _write_json(builder_input_path, builder_input)
    workflow = WorkflowPaths(
        source_review=builder_input_path,
        output_dir=output_dir,
        review_v2=building_workbook,
        compiled_json=system_dir / "compiled.json",
        evidence_json=system_dir / "evidence.json",
        map_png=output_dir / "column_map.png",
        map_json=system_dir / "column_map.json",
        map_kml=output_dir / "column_map.kml",
    )
    print("[4/5] Review Excelを生成")
    run_spreadsheet_builder(workflow, map_generated=map_result is not None)

    print("[5/5] Excel描画・数式・提出前チェック")
    qa_dir = system_dir / "qa"
    qa_dir.mkdir(parents=True, exist_ok=True)
    inspection = _run_qa(building_workbook, qa_dir, workflow.evidence_json)
    check = _submission_check(building_workbook, system_dir, strict=False)
    check = _apply_semantic_submission_gate(
        check,
        final_compiled,
        column_mode=column_mode,
        log_path=system_dir / "submission_check.txt",
    )
    # 中身が痩せた結果で人の作業を置き換えない。LLMが全滅した実行は
    # unitが1件のNO_DATAになるが、これは既存の48 unitより価値が低い。
    if not guarded and protected:
        built_units = len(final_compiled.get("units") or [])
        existing_units = max(
            (_workbook_unit_count(path) for path in protected), default=0
        )
        if built_units < existing_units:
            final_workbook = output_dir / f"m{sources.map_id}_review.candidate-{stamp}.xlsx"
            print(
                f"  [GUARD] 生成結果は {built_units} unit で、既存の {existing_units} unit より"
                " 少ないため本体を置き換えません。"
            )
            print(f"          結果はこちらに残します: {final_workbook.name}")
    final_workbook, workbook_warning = _finalize_workbook(
        building_workbook, final_workbook
    )
    if workbook_warning:
        print(f"  [CHECK] {workbook_warning}")
    builder_sidecar = Path(str(building_workbook) + ".inspect.ndjson")
    build_inspection = qa_dir / "build_inspection.ndjson"
    if builder_sidecar.is_file():
        # The builder owns this exact sidecar name.  Keep diagnostics with the
        # other machine QA artifacts instead of beside the human workbook.
        os.replace(builder_sidecar, build_inspection)

    compiled = _read_json(system_dir / "compiled.json")
    evidence = _read_json(system_dir / "evidence.json")
    source_counts = Counter(
        str(record.get("source", {}).get("type") or "Unknown")
        for record in evidence.get("evidence") or []
    )
    manifest = {
        "schema_version": "gsj-50k-pilot/2.0",
        "status": "complete",
        "generated_at": generated_at,
        "map_id": sources.map_id,
        "legacy_workbook_read": False,
        "source_priority": ["ZFK", "Shapefile", "PDF/LLM"],
        "sources": {
            "workspace_manifest": str(sources.source_manifest),
            "pdf": ({"available": True, "path": str(sources.pdf), "sha256": _sha256(sources.pdf)}
                    if sources.pdf else {"available": False}),
            "shape": ({"available": True, "path": str(sources.shape), "sha256": _sha256(sources.shape)}
                      if sources.shape else {"available": False}),
            "zfk": {
                "available": any((sources.zfk_root / "units").glob("*.json")),
                "path": str(sources.zfk_root),
                "unit_json": len(list((sources.zfk_root / "units").glob("*.json"))),
                "geojson": len(list((sources.zfk_root / "geojson").glob("*.geojson"))),
            },
            "column_config": (
                {"available": True, "path": str(sources.column_config), "sha256": _sha256(sources.column_config)}
                if sources.column_config else {"available": False}
            ),
            "column_mode": column_mode,
            "column_warning": column_warning,
        },
        "llm": {
            "model": model,
            "external_calls": (
                int(llm_manifest.get("external_calls", 0) or 0)
                + int(vision_manifest.get("external_calls", 0) or 0)
            ),
            "cache_hits": (
                int(llm_manifest.get("cache_hits", 0) or 0)
                + int(vision_manifest.get("cache_hits", 0) or 0)
            ),
            "added_evidence": llm_manifest.get("added_evidence", 0),
            "usage_before": usage_before,
            "usage_after": usage_after,
            "usage_delta": {
                "calls": usage_after["calls"] - usage_before["calls"],
                "tokens": usage_after["tokens"] - usage_before["tokens"],
            },
            "legacy_cache_files_migrated": migrated_cache,
            "compatible_cache_jobs_migrated": compatible_cache_migrations,
            "stage": llm_manifest.get("stage"),
            "status": llm_manifest.get("status", "complete"),
            "error": llm_manifest.get("error"),
            "quota_exhausted": llm_manifest.get("quota_exhausted", False),
            "quota_error": llm_manifest.get("quota_error"),
            "pdf_body_routing": llm_manifest.get("pdf_body_routing"),
            "pdf_alias_mapping": llm_manifest.get("pdf_alias_mapping"),
            "local_age_notes": llm_manifest.get("local_age_notes"),
            "local_abstract_science": llm_manifest.get("local_abstract_science"),
            "local_japanese_science": llm_manifest.get("local_japanese_science"),
            "pdf_body": llm_manifest.get("pdf_body"),
            "pdf_environment": llm_manifest.get("pdf_environment"),
            "age_interpolation": llm_manifest.get("age_interpolation"),
            "column_vision": vision_manifest,
        },
        "column_geography": geography_manifest,
        "canonical": {
            "units": compiled.get("summary", {}).get("unit_count"),
            "columns": compiled.get("summary", {}).get("column_count"),
            "evidence": compiled.get("summary", {}).get("evidence_count"),
            "orphaned_evidence_count": compiled.get("summary", {}).get(
                "orphaned_evidence_count"
            ),
            "evidence_by_source": dict(sorted(source_counts.items())),
            "status_counts": compiled.get("summary", {}).get("status_counts"),
            "interval_normalizations": interval_corrections,
        },
        "map": ({
            "generated": True,
            "matched_units": map_result.get("matched_units", 0) if isinstance(map_result, dict) else getattr(map_result, "matched_units", 0),
            "unmatched_units": map_result.get("unmatched_units", 0) if isinstance(map_result, dict) else getattr(map_result, "unmatched_units", 0),
            "warnings": list(map_result.get("warnings", [])) if isinstance(map_result, dict) else list(getattr(map_result, "warnings", [])),
        } if map_result is not None else {
            "generated": False,
            "matched_units": 0,
            "unmatched_units": 0,
            "warnings": [map_error] if map_error else ["Shapefile unavailable"],
        }),
        "qa": {
            "inspection": inspection,
            "build_inspection": str(build_inspection) if build_inspection.is_file() else None,
            "submission_check": check,
            "workbook_warning": workbook_warning,
        },
        "artifacts": {
            "review_excel": str(final_workbook),
            "compiled_json": str(system_dir / "compiled.json"),
            "evidence_json": str(system_dir / "evidence.json"),
            "derived_previews_json": str(system_dir / "derived_previews.json"),
            "environment_manifest": str(
                system_dir / "environment_analysis" / "environment_manifest.json"
            ),
            "pilot_manifest": str(system_dir / "pilot_manifest.json"),
        },
    }
    vision_dir = system_dir / "column_vision"
    for key, path in {
        "column_proposal": vision_dir / "column_proposal.json",
        "coordinate_candidates": vision_dir / "coordinate_candidates.json",
        "vision_manifest": vision_dir / "vision_manifest.json",
    }.items():
        if path.is_file():
            manifest["artifacts"][key] = str(path)
    completed_proposal = vision_manifest.get("completed_proposal_file")
    if completed_proposal and Path(str(completed_proposal)).is_file():
        manifest["artifacts"]["column_proposal_completed"] = str(completed_proposal)
    if map_result is not None:
        for key, path in {
            "column_map_png": output_dir / "column_map.png",
            "column_map_kml": output_dir / "column_map.kml",
        }.items():
            if path.is_file():
                manifest["artifacts"][key] = str(path)
    _write_json(system_dir / "pilot_manifest.json", manifest)
    return manifest


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("map_id")
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--model", default=MODEL)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--no-llm", action="store_true", help=argparse.SUPPRESS)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        result = run_pilot(
            args.map_id, output_dir=args.output_dir, model=args.model,
            force=args.force, use_llm=not args.no_llm,
        )
    except (PilotError, PilotLLMError, ReviewV2Error, WorkspaceError,
            FileNotFoundError, ValueError, OSError) as exc:
        print(f"[STOP] {exc}", file=sys.stderr)
        return 2
    submission = (result.get("qa") or {}).get("submission_check") or {}
    summary_status = (
        result["status"]
        if submission.get("ready_for_submission", True)
        else "complete_with_review_required"
    )
    print(json.dumps({
        "status": summary_status,
        "pipeline_status": result["status"],
        "review_excel": result["artifacts"]["review_excel"],
        "column_map_kml": result["artifacts"].get("column_map_kml"),
        "submission": submission,
        "llm": result["llm"],
        "canonical": result["canonical"],
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
