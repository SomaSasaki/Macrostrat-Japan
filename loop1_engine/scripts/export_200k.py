from pathlib import Path
# -*- coding: utf-8 -*-
"""
export_200k.py — GSJ 20万分の1レビュー用 Excel からの厳格検証 (check) および提出用 Excel 出力 (export)

Macrostrat column-ingestion format v0.1.1 準拠。
- 厳格なバリデータ: 層序単調性（年代逆転）、prop範囲（0.0〜1.0）、Macrostrat公式語彙（214岩相/83環境/接触関係）、Column年代包含関係、ID一意性の全数検証
- Python直接prop計算: Excel数式キャッシュに依存せず、Python側で確実に数値を計算・エクスポート
"""

import io
import json
import os
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, BASE_DIR)

import scripts.common as common

CONFIG_PATH = os.path.join(BASE_DIR, 'config', 'map_index_200k.json')
REVIEW_BASE_DIR = os.path.join(BASE_DIR, 'data', '200k', '02_review')
SUBMISSION_BASE_DIR = os.path.join(BASE_DIR, 'data', '200k', '03_submission', '00_SKELETON_200K')
VOCAB_PATH = os.path.join(Path(__file__).resolve().parents[2], 'loop2_governance', 'config', 'vocab.json')

intervals_dict = common.load_intervals()
int_names = set(intervals_dict.keys())

with open(VOCAB_PATH, 'r', encoding='utf-8') as f:
    vocab_data = json.load(f)
lith_vocab = {v.lower(): v for v in vocab_data.get('lithology', [])}
env_vocab = {v.lower(): v for v in vocab_data.get('environment', [])}

VALID_BASAL_SURFACES = {
    "conformable", "unconformity", "angular unconformity", "disconformity",
    "paraconformity", "nonconformity", "intrusive", "faulted", "unknown", ""
}

def find_review_file(sheet_code):
    """
    sheet_code または name_en に対応するレビューExcelファイルを探索
    """
    for root, dirs, files in os.walk(REVIEW_BASE_DIR):
        for f in files:
            if (f.startswith(f"m200k_{sheet_code}_review") or f"_{sheet_code}_" in f) and f.endswith(".xlsx"):
                return os.path.join(root, f)
    return None

def compute_unit_prop(int_name, age_ma, is_bottom=True):
    """
    指定された interval と年代(Ma)から prop (0.000〜1.000) を計算
    """
    if not int_name or int_name not in intervals_dict:
        return 0.0 if is_bottom else 1.0
    
    b_bounds = common.interval_bounds(int_name)
    if not b_bounds:
        return 0.0 if is_bottom else 1.0
    
    b_age, t_age = b_bounds
    if age_ma is None or str(age_ma).strip() == '':
        return 0.0 if is_bottom else 1.0
    
    try:
        age_val = float(age_ma)
    except (ValueError, TypeError):
        return 0.0 if is_bottom else 1.0

    denom = b_age - t_age
    if denom <= 0:
        return 0.5
    
    prop = (b_age - age_val) / denom
    return round(max(0.0, min(1.0, prop)), 3)

def check_200k_workbook(file_path):
    """
    レビュー用Excelの厳格な整合性を全数検証する（真のバリデータ）
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    errors = []
    warnings = []

    if 'units_review' not in wb.sheetnames or 'columns_review' not in wb.sheetnames:
        errors.append("必須シート (units_review, columns_review) が見つかりません。")
        return errors, warnings

    ws_cols = wb['columns_review']
    ws_units = wb['units_review']

    col_headers = [c.value for c in ws_cols[1]]
    unit_headers = [c.value for c in ws_units[1]]

    # 1. columns_review 検証 & Column年代範囲収集
    col_ids = set()
    col_ranges = {}
    for row_idx, row in enumerate(ws_cols.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        c_dict = dict(zip(col_headers, row))
        c_id = str(c_dict.get('col_id') or '')
        c_name = c_dict.get('col_name')
        b_int = c_dict.get('b_int')
        t_int = c_dict.get('t_int')

        if not c_id:
            errors.append(f"[columns_review L{row_idx}] col_id が空です。")
        elif c_id in col_ids:
            errors.append(f"[columns_review L{row_idx}] 重複した col_id '{c_id}' が存在します。")
        else:
            col_ids.add(c_id)

        if not c_name:
            warnings.append(f"[columns_review L{row_idx}] col_name が空です。")

        # Column の年代範囲
        b_b = common.interval_bounds(b_int) if b_int else None
        t_b = common.interval_bounds(t_int) if t_int else None
        col_max_age = b_b[0] if b_b else 541.0
        col_min_age = t_b[1] if t_b else 0.0
        col_ranges[c_id] = (col_max_age, col_min_age, b_int, t_int)

    # 2. units_review 検証 (単調性・語彙・prop・ID・Column包含・sort_order)
    seen_unit_ids = set()
    column_units = {}

    for row_idx, row in enumerate(ws_units.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        u_dict = dict(zip(unit_headers, row))
        u_id = str(u_dict.get('unit_id') or '').strip()
        u_col_id = str(u_dict.get('col_id') or '').strip()
        u_name = str(u_dict.get('unit_name') or '').strip()
        sort_order_raw = u_dict.get('sort_order')
        b_int = str(u_dict.get('b_int') or '').strip()
        t_int = str(u_dict.get('t_int') or '').strip()
        b_age_ma = u_dict.get('b_age_ma')
        t_age_ma = u_dict.get('t_age_ma')
        b_prop_raw = u_dict.get('b_prop')
        t_prop_raw = u_dict.get('t_prop')
        lith = str(u_dict.get('lithology') or '').strip()
        minor_lith = str(u_dict.get('minor_lith') or '').strip()
        env = str(u_dict.get('environment') or '').strip()
        basal = str(u_dict.get('basal_surface') or '').strip()

        # ID一意性と空値検査
        if not u_id:
            errors.append(f"[units_review L{row_idx}] unit_id が空です。")
        elif u_id in seen_unit_ids:
            errors.append(f"[units_review L{row_idx}] 重複した unit_id '{u_id}'")
        else:
            seen_unit_ids.add(u_id)

        if not u_col_id:
            errors.append(f"[units_review L{row_idx}] col_id が未指定です。")
        elif u_col_id not in col_ids:
            errors.append(f"[units_review L{row_idx}] 未定義の col_id '{u_col_id}' が参照されています。")

        if not u_name:
            errors.append(f"[units_review L{row_idx}] unit_name が空です。")

        # sort_order 整数検査
        try:
            sort_order = int(sort_order_raw)
            if sort_order < 1:
                errors.append(f"[units_review L{row_idx}] sort_order '{sort_order_raw}' は 1 以上の整数でなければなりません。")
        except (ValueError, TypeError):
            errors.append(f"[units_review L{row_idx}] sort_order '{sort_order_raw}' は有効な整数ではありません。")
            sort_order = 1

        # Intervals 語彙照合
        if not b_int or b_int not in intervals_dict:
            errors.append(f"[units_review L{row_idx}] 不明な b_int '{b_int}' (Macrostrat Intervals外)")
        if not t_int or t_int not in intervals_dict:
            errors.append(f"[units_review L{row_idx}] 不明な t_int '{t_int}' (Macrostrat Intervals外)")

        # 年代境界の実測値取得
        b_bounds = common.interval_bounds(b_int) if b_int in intervals_dict else None
        t_bounds = common.interval_bounds(t_int) if t_int in intervals_dict else None

        # b_age_ma / t_age_ma 数値検証
        u_b_age = None
        u_t_age = None
        if b_age_ma is not None and str(b_age_ma).strip() != '':
            try:
                u_b_age = float(b_age_ma)
                if b_bounds and (u_b_age > b_bounds[0] + 0.5 or u_b_age < b_bounds[1] - 0.5):
                    warnings.append(f"[units_review L{row_idx}] b_age_ma({u_b_age}) が b_int '{b_int}' の年代範囲 ({b_bounds[0]}-{b_bounds[1]} Ma) 外です。")
            except ValueError:
                errors.append(f"[units_review L{row_idx}] b_age_ma '{b_age_ma}' は有効な数値ではありません。")
        else:
            u_b_age = b_bounds[0] if b_bounds else 541.0

        if t_age_ma is not None and str(t_age_ma).strip() != '':
            try:
                u_t_age = float(t_age_ma)
                if t_bounds and (u_t_age > t_bounds[0] + 0.5 or u_t_age < t_bounds[1] - 0.5):
                    warnings.append(f"[units_review L{row_idx}] t_age_ma({u_t_age}) が t_int '{t_int}' の年代範囲 ({t_bounds[0]}-{t_bounds[1]} Ma) 外です。")
            except ValueError:
                errors.append(f"[units_review L{row_idx}] t_age_ma '{t_age_ma}' は有効な数値ではありません。")
        else:
            u_t_age = t_bounds[1] if t_bounds else 0.0

        # 単元内での年代整合性 (b_age >= t_age)
        if u_b_age is not None and u_t_age is not None:
            if u_b_age < u_t_age:
                errors.append(f"[units_review L{row_idx}] 年代逆転: b_age({u_b_age} Ma) < t_age({u_t_age} Ma)")

        # Column 年代包含検査
        if u_col_id in col_ranges and u_b_age is not None and u_t_age is not None:
            col_max_age, col_min_age, col_b_int, col_t_int = col_ranges[u_col_id]
            if col_max_age is not None and u_b_age is not None and u_b_age > col_max_age + 0.5:
                errors.append(f"[units_review L{row_idx}] Unit '{u_name}' の底面年代 ({u_b_age} Ma) が Column '{u_col_id}' の最古年代 ({col_max_age} Ma, {col_b_int}) を超えています。")
            if col_min_age is not None and u_t_age is not None and u_t_age < col_min_age - 0.5:
                errors.append(f"[units_review L{row_idx}] Unit '{u_name}' の上面年代 ({u_t_age} Ma) が Column '{u_col_id}' の最新年代 ({col_min_age} Ma, {col_t_int}) より若くなっています。")

        # prop 厳格範囲検査 (フェイルオープンの排除: 負値や1超、文字列は即エラー)
        if b_prop_raw is not None and str(b_prop_raw).strip() != '':
            try:
                bp_val = float(b_prop_raw)
                if not (0.0 <= bp_val <= 1.0):
                    errors.append(f"[units_review L{row_idx}] b_prop ({bp_val}) が 0.0〜1.0 の範囲外です。")
            except ValueError:
                errors.append(f"[units_review L{row_idx}] b_prop '{b_prop_raw}' は有効な数値ではありません。")

        if t_prop_raw is not None and str(t_prop_raw).strip() != '':
            try:
                tp_val = float(t_prop_raw)
                if not (0.0 <= tp_val <= 1.0):
                    errors.append(f"[units_review L{row_idx}] t_prop ({tp_val}) が 0.0〜1.0 の範囲外です。")
            except ValueError:
                errors.append(f"[units_review L{row_idx}] t_prop '{t_prop_raw}' は有効な数値ではありません。")

        # 岩相語彙照合 (214件)
        if not lith or lith.lower() not in lith_vocab:
            errors.append(f"[units_review L{row_idx}] Macrostrat語彙外の岩相名 '{lith}'")
        if minor_lith and minor_lith.lower() not in lith_vocab:
            warnings.append(f"[units_review L{row_idx}] Macrostrat語彙外の副岩相名 '{minor_lith}'")

        # 環境語彙照合 (83件)
        if env and env.lower() not in env_vocab and env.lower() != 'unknown':
            warnings.append(f"[units_review L{row_idx}] Macrostrat語彙外の environment '{env}'")

        # 接触関係語彙照合
        if basal and basal.lower() not in VALID_BASAL_SURFACES:
            warnings.append(f"[units_review L{row_idx}] 非標準の basal_surface '{basal}'")

        # Column単位の単調性チェック用データ収集
        column_units.setdefault(u_col_id, []).append({
            'row_idx': row_idx,
            'sort_order': sort_order,
            'b_age': u_b_age,
            't_age': u_t_age,
            'unit_name': u_name,
        })

    # 3. Column 内の層序単調性 ＆ sort_order 一意性全数検査
    for c_id, u_list in column_units.items():
        # sort_order 重複検査
        s_orders = [u['sort_order'] for u in u_list]
        if len(s_orders) != len(set(s_orders)):
            errors.append(f"[Column {c_id}] 重複した sort_order が存在します: {s_orders}")

        u_list_sorted = sorted(u_list, key=lambda x: x['sort_order'])
        for i in range(len(u_list_sorted) - 1):
            curr_u = u_list_sorted[i]
            next_u = u_list_sorted[i + 1]
            
            # 真の年代逆転エラー: 上位の層(next_u)の最新上面年代(t_age)すら、下位の層(curr_u)の最古底面年代(b_age)より古い場合
            if curr_u['b_age'] is not None and next_u['t_age'] is not None:
                if (next_u['t_age'] - curr_u['b_age']) > 0.1:
                    errors.append(
                        f"[Column {c_id} 層序逆転] 下位 position {curr_u['sort_order']} ({curr_u['b_age']}-{curr_u['t_age']} Ma, {curr_u['unit_name']}) "
                        f"の上に、完全に古い時代の上位 position {next_u['sort_order']} ({next_u['b_age']}-{next_u['t_age']} Ma, {next_u['unit_name']}) が積み重なっています。"
                    )

    return errors, warnings

def export_200k_workbook(review_file_path, out_excel_path):
    """
    レビューExcelから提出用Excel（v0.1.1 5シート）を出力（Python直接prop計算）
    """
    wb_rev = openpyxl.load_workbook(review_file_path, data_only=True)
    ws_cols_rev = wb_rev['columns_review']
    ws_units_rev = wb_rev['units_review']
    ws_refs_rev = wb_rev['refs_review'] if 'refs_review' in wb_rev.sheetnames else None
    ws_meta_rev = wb_rev['project_meta'] if 'project_meta' in wb_rev.sheetnames else None

    col_headers = [c.value for c in ws_cols_rev[1]]
    unit_headers = [c.value for c in ws_units_rev[1]]

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    # 1. metadata
    ws_meta = wb_out.create_sheet(title="metadata")
    ws_meta.append(["key", "value"])
    if ws_meta_rev:
        for row in ws_meta_rev.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                ws_meta.append([row[0], row[1] if len(row) > 1 else ''])
    else:
        for k in common.METADATA_KEYS:
            ws_meta.append([k, common.METADATA_DEFAULTS.get(k, '')])

    # 2. columns
    ws_cols = wb_out.create_sheet(title="columns")
    ws_cols.append(common.SUBMISSION_COLUMN_COLS)
    col_id_map = {}
    numeric_col_id = 1

    for row in ws_cols_rev.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        c_dict = dict(zip(col_headers, row))
        orig_id = str(c_dict.get('col_id', ''))
        col_id_map[orig_id] = numeric_col_id

        out_row = [
            numeric_col_id,
            c_dict.get('col_name', ''),
            c_dict.get('col_group', ''),
            1,  # ref_ids
            c_dict.get('date_collected', '2010-01-01'),
            'column',
            'age',
            c_dict.get('b_int', 'Phanerozoic'),
            c_dict.get('t_int', 'Holocene'),
            0.0,  # b_prop
            1.0,  # t_prop
            c_dict.get('geom', ''),
            c_dict.get('rgeom', ''),
            c_dict.get('comments', ''),
        ]
        ws_cols.append(out_row)
        numeric_col_id += 1

    # 3. units (Python側で厳密にprop直接計算)
    ws_units = wb_out.create_sheet(title="units")
    ws_units.append(common.SUBMISSION_UNIT_COLS)
    global_u_id = 1

    for row in ws_units_rev.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        u_dict = dict(zip(unit_headers, row))
        orig_col_id = str(u_dict.get('col_id', ''))
        mapped_col_id = col_id_map.get(orig_col_id, 1)

        b_int = u_dict.get('b_int', 'Phanerozoic')
        t_int = u_dict.get('t_int', 'Holocene')
        b_age_ma = u_dict.get('b_age_ma')
        t_age_ma = u_dict.get('t_age_ma')

        # Python側で直接propを厳密計算
        bp_val = compute_unit_prop(b_int, b_age_ma, is_bottom=True)
        tp_val = compute_unit_prop(t_int, t_age_ma, is_bottom=False)

        # position の整数取得
        s_ord = u_dict.get('sort_order')
        pos_val = int(s_ord) if s_ord and str(s_ord).isdigit() else global_u_id

        out_u_row = [
            global_u_id,
            mapped_col_id,
            u_dict.get('section_id', 1),
            pos_val,
            b_int,
            bp_val,
            t_int,
            tp_val,
            u_dict.get('unit_name', ''),
            u_dict.get('strat_name', ''),
            u_dict.get('environment', ''),
            u_dict.get('unit_description', ''),
            u_dict.get('lithology', ''),
            u_dict.get('minor_lith', ''),
            u_dict.get('min_thickness', ''),
            u_dict.get('max_thickness', ''),
            u_dict.get('basal_surface', 'unknown'),
            u_dict.get('lateral_relationship', ''),
            u_dict.get('comments', ''),
            u_dict.get('t_pos', ''),
            u_dict.get('REF_symbol', ''),
            f"{u_dict.get('REF_age_ja', '')} {u_dict.get('REF_lith_ja', '')}",
        ]
        ws_units.append(out_u_row)
        global_u_id += 1

    # 4. refs
    ws_refs = wb_out.create_sheet(title="refs")
    ws_refs.append(common.SUBMISSION_REF_COLS)
    if ws_refs_rev:
        for row in ws_refs_rev.iter_rows(min_row=2, values_only=True):
            if any(row):
                ws_refs.append(list(row))

    # 5. images
    ws_imgs = wb_out.create_sheet(title="images")
    ws_imgs.append(common.SUBMISSION_IMAGE_COLS)

    header_font = Font(name="Meiryo UI", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    for ws in [ws_meta, ws_cols, ws_units, ws_refs, ws_imgs]:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    os.makedirs(os.path.dirname(out_excel_path), exist_ok=True)
    wb_out.save(out_excel_path)
    return len(col_id_map), global_u_id - 1

def run_check_cli(target_code):
    if target_code.lower() == 'all':
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            sheets = json.load(f)
        total_errs = 0
        total_warns = 0
        failed_sheets = []
        for s in sheets:
            code = s['sheet_code']
            rf = find_review_file(code)
            if not rf:
                print(f"[SKIP] {code}: レビューファイルが見つかりません。")
                continue
            errs, warns = check_200k_workbook(rf)
            total_errs += len(errs)
            total_warns += len(warns)
            status = "PASS" if not errs else "FAIL"
            if errs:
                failed_sheets.append((code, errs))
            print(f"[{status}] {code} ({s.get('name_ja', '')}): {len(errs)} errors, {len(warns)} warnings")
        print(f"\n=======================================================")
        print(f"Total Errors: {total_errs}, Total Warnings: {total_warns}")
        if failed_sheets:
            print(f"Failed Sheets ({len(failed_sheets)}):")
            for c, es in failed_sheets[:10]:
                print(f"  - {c}: {es[0]}")
        print(f"=======================================================")
    else:
        rf = find_review_file(target_code)
        if not rf:
            print(f"[ERROR] {target_code} のレビューファイルが見つかりません。")
            return
        errs, warns = check_200k_workbook(rf)
        print(f"=== Check Results for {target_code} ({os.path.basename(rf)}) ===")
        if not errs and not warns:
            print("OK: すべての検証項目（単調性・prop・語彙・ID一意性）をパスしました。")
        else:
            for e in errs:
                print(f"[ERROR] {e}")
            for w in warns:
                print(f"[WARN] {w}")

def run_export_cli(target_code):
    if target_code.lower() == 'all':
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            sheets = json.load(f)
        total_cols = 0
        total_units = 0
        for s in sheets:
            code = s['sheet_code']
            name_en = s.get('name_en', code) or code
            region = s.get('region', '00_Other')
            rf = find_review_file(code)
            if not rf:
                continue
            out_p = os.path.join(SUBMISSION_BASE_DIR, region, f"m200k_{code}_{name_en}", f"{name_en}_200k_MultiColumn.xlsx")
            c_cnt, u_cnt = export_200k_workbook(rf, out_p)
            total_cols += c_cnt
            total_units += u_cnt
            print(f"[EXPORT] {code} ({s.get('name_ja', '')}): {c_cnt} Cols, {u_cnt} Units -> {os.path.basename(out_p)}")
        print(f"\n=======================================================")
        print(f"SUCCESS: Exported {total_cols} Columns ({total_units} Units) to {SUBMISSION_BASE_DIR}")
        print(f"=======================================================")
    else:
        rf = find_review_file(target_code)
        if not rf:
            print(f"[ERROR] {target_code} のレビューファイルが見つかりません。")
            return
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            sheets = json.load(f)
        s_meta = next((s for s in sheets if s['sheet_code'] == target_code or s.get('name_en') == target_code), None)
        name_en = s_meta.get('name_en', target_code) if s_meta else target_code
        region = s_meta.get('region', '00_Other') if s_meta else '00_Other'
        out_p = os.path.join(SUBMISSION_BASE_DIR, region, f"m200k_{target_code}_{name_en}", f"{name_en}_200k_MultiColumn.xlsx")
        c_cnt, u_cnt = export_200k_workbook(rf, out_p)
        print(f"[SUCCESS] Exported {c_cnt} Columns ({u_cnt} Units) to {out_p}")

if __name__ == '__main__':
    if len(sys.argv) > 2 and sys.argv[1] == 'check':
        run_check_cli(sys.argv[2])
    elif len(sys.argv) > 2 and sys.argv[1] == 'export':
        run_export_cli(sys.argv[2])
    else:
        print("Usage: python scripts/export_200k.py [check|export] [sheet_code|all]")
