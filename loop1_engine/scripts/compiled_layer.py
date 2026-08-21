# -*- coding: utf-8 -*-
"""Canonical, auditable data layer for a single GSJ map sheet.

The review workbook is a user interface, not the durable data store.  This
module converts either that workbook or already-normalized row dictionaries
into two JSON documents:

``compiled.json``
    Human-entered values, safely resolved fallback values, unit-level review
    status, and three compact evidence summaries.

``evidence.json``
    Long-form evidence with stable identifiers and complete provenance.

No workbook is changed.  A non-blank value in a review/input column is always
authoritative over an automatically collected candidate.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import re
import tempfile
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

try:
    from common import (
        check_vocab,
        normalize_vocab,
        parse_lithology_relations,
        resolve_lithology_value,
    )
except ImportError:  # pragma: no cover - package-style import
    from .common import (
        check_vocab,
        normalize_vocab,
        parse_lithology_relations,
        resolve_lithology_value,
    )




SCHEMA_VERSION = "1.0.0"
VALID_STATUSES = ("OK", "CHECK", "MISSING")

# Fields shown/edited in the review interface.  Unknown non-REF columns are
# retained as well, so a workbook can evolve without losing human input.
CANONICAL_VALUE_FIELDS = (
    "unit_id",
    "column_id",
    "sort_order",
    "position",
    "section_id",
    "t_pos",
    "unit_name",
    "t_int",
    "b_int",
    "t_age_ma",
    "b_age_ma",
    "t_prop",
    "b_prop",
    "strat_name",
    "environment",
    "unit_description",
    "lithology",
    "minor_lith",
    "min_thickness",
    "max_thickness",
    "basal_surface",
    "lateral_relationship",
    "comments",
)

EVIDENCE_GROUP_FIELDS = {
    "age_evidence": {
        "age_text", "local_age_notes", "t_int", "b_int", "t_age_ma", "b_age_ma", "t_prop", "b_prop",
    },
    "context_evidence": {
        "unit_name", "unit_name_ja", "strat_name", "environment",
        "environment_applicability",
        "unit_description", "description_context", "lithology", "minor_lith",
        "lithology_context", "column_id", "sort_order", "place_names",
    },
    "physical_evidence": {
        "min_thickness", "max_thickness", "thickness", "thickness_context",
        "basal_surface", "lateral_relationship",
    },
}

FIELD_TO_GROUP = {
    field: group
    for group, fields in EVIDENCE_GROUP_FIELDS.items()
    for field in fields
}

# Only candidates already expressed as a final-schema field may fill a blank
# resolved value.  Context blobs such as REF_thickness are evidence, not values.
AUTO_RESOLVE_FIELDS = {
    "unit_name", "strat_name", "environment", "unit_description",
    "lithology", "minor_lith", "min_thickness", "max_thickness",
    "basal_surface", "lateral_relationship", "t_int", "b_int",
    "t_age_ma", "b_age_ma",
}

VOCAB_AUTO_RESOLVE_FIELDS = {"lithology", "minor_lith", "environment"}
NUMERIC_AUTO_RESOLVE_FIELDS = {
    "min_thickness", "max_thickness", "t_age_ma", "b_age_ma",
}

EVIDENCE_SCOPE_TYPES = {"unit_global", "column_specific", "map_global"}
UNIT_GLOBAL_FIELDS = {
    "unit_name", "unit_name_ja", "strat_name", "unit_description",
    "description_context", "lithology", "minor_lith", "lithology_context",
    "age_text", "t_int", "b_int", "t_age_ma", "b_age_ma", "t_prop", "b_prop",
    "local_age_notes",
}

# Evidence extractors intentionally keep their source note beside the raw
# candidate.  That is useful in evidence.json, but the note must never leak
# into a resolved Macrostrat value.  Only a trailing, explicitly labelled note
# is removed; ordinary parenthetical geology text remains untouched.
_TRAILING_SOURCE_NOTE_PATTERNS = (
    re.compile(
        r"\s*[\(（]\s*(?:source|evidence|citation|reference|basis|quote|excerpt|"
        r"出典|根拠|引用|資料)\s*[:：][\s\S]*[\)）]\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"\s*[\[【]\s*(?:source|evidence|citation|reference|basis|quote|excerpt|"
        r"出典|根拠|引用|資料)\s*[:：][\s\S]*[\]】]\s*$",
        re.IGNORECASE,
    ),
    re.compile(
        r"\s*(?:[—–-]\s*)?(?:source|evidence|citation|reference|basis|"
        r"出典|根拠|引用|資料)\s*[:：][\s\S]*$",
        re.IGNORECASE,
    ),
)

_NUMERIC_LITERAL = re.compile(
    r"^[+-]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][+-]?\d+)?$"
)

CONFIDENCE_SCORES = {"A": 0.98, "B": 0.85, "C": 0.65, "D": 0.35}
SOURCE_PRIORITY = {
    "ZFK": 40,
    "SHAPEFILE": 30,
    "SHAPE": 30,
    "PDF": 20,
    "LLM": 10,
    "GSJ": 20,
    "REVIEW": 50,
}
FIELD_SOURCE_PRIORITY = {
    "environment": {"PDF": 45, "ZFK": 35, "SHAPEFILE": 5, "LLM": 20},
    "unit_description": {"PDF": 45, "ZFK": 35, "SHAPEFILE": 5, "LLM": 20},
    "min_thickness": {"ZFK": 45, "PDF": 40, "SHAPEFILE": 5, "LLM": 20},
    "max_thickness": {"ZFK": 45, "PDF": 40, "SHAPEFILE": 5, "LLM": 20},
    "basal_surface": {"ZFK": 45, "PDF": 40, "SHAPEFILE": 5, "LLM": 20},
}

_REF_EVIDENCE_SPECS = (
    # ref column, canonical field, source, confidence, assertion, locator column
    ("REF_unit_name_en", "unit_name", "ZFK", "A", "explicit", "REF_source"),
    ("REF_unit_name_ja", "unit_name_ja", "ZFK", "A", "explicit", "REF_source"),
    ("REF_age_text", "age_text", "ZFK", "A", "explicit", "REF_source"),
    ("REF_age_from_abstract", "age_text", "PDF", "C", "explicit", "REF_source"),
    ("REF_desc", "description_context", "ZFK", "B", "explicit", "REF_source"),
    ("REF_thickness", "thickness_context", "ZFK", "B", "explicit", "REF_source"),
    ("REF_lith_text", "lithology_context", "ZFK", "A", "explicit", "REF_source"),
    ("REF_lithology_gsj", "lithology", "ZFK", "B", "explicit", "REF_source"),
    ("REF_minor_lith_gsj", "minor_lith", "ZFK", "B", "explicit", "REF_source"),
    ("REF_strat_name", "strat_name", "ZFK", "B", "explicit", "REF_source"),
    ("REF_basal_surface", "basal_surface", "ZFK", "B", "explicit", "REF_source"),
    ("REF_lithology", "lithology", "PDF", "C", "explicit", "REF_source"),
    ("REF_minor_lith", "minor_lith", "PDF", "C", "explicit", "REF_source"),
    ("REF_environment", "environment", "PDF", "C", "explicit", "REF_source"),
    ("REF_unit_description", "unit_description", "PDF", "C", "explicit", "REF_source"),
    ("REF_shape_unit_name", "unit_name", "Shapefile", "A", "explicit", "REF_shape_source"),
    ("REF_shape_age_text", "age_text", "Shapefile", "A", "explicit", "REF_shape_source"),
    ("REF_shape_lith_text", "lithology_context", "Shapefile", "A", "explicit", "REF_shape_source"),
    ("REF_column_id", "column_id", "Vision", "C", "inferred", "REF_source"),
    ("REF_sort_order", "sort_order", "Vision", "C", "inferred", "REF_source"),
    ("REF_place_names", "place_names", "Vision", "C", "inferred", "REF_source"),
)

_EVIDENCE_SUMMARY_COLUMNS = set(EVIDENCE_GROUP_FIELDS) | {
    "status", "review_status", "age_evidence", "context_evidence", "physical_evidence",
}


def is_blank(value: Any) -> bool:
    """Return True for empty spreadsheet/JSON values without treating 0 as empty."""
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str):
        return value.strip().lower() in {"", "nan", "none", "nat"}
    return False


def json_value(value: Any) -> Any:
    """Convert pandas/numpy/Excel-friendly scalars to stable JSON values."""
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Path):
        return str(value)
    # numpy scalar support without importing numpy.
    item = getattr(value, "item", None)
    if callable(item) and not isinstance(value, (str, bytes, bytearray)):
        try:
            return json_value(item())
        except (TypeError, ValueError):
            pass
    if isinstance(value, Mapping):
        return {str(k): json_value(v) for k, v in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [json_value(v) for v in value]
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def normalize_row(row: Mapping[str, Any] | Any) -> dict[str, Any]:
    """Normalize a dict, pandas Series, or mapping-like row."""
    if isinstance(row, Mapping):
        items = row.items()
    elif hasattr(row, "to_dict"):
        items = row.to_dict().items()
    else:
        items = dict(row).items()
    return {str(k).strip(): json_value(v) for k, v in items if str(k).strip()}


def split_column_ids(value: Any) -> list[str]:
    if is_blank(value):
        return []
    if isinstance(value, (list, tuple, set)):
        parts: Iterable[Any] = value
    else:
        parts = re.split(r"[,;]", str(value))
    out: list[str] = []
    for part in parts:
        text = str(part).strip()
        if text and text not in out:
            out.append(text)
    return out


def _truthy(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return value != 0
    return str(value or "").strip().lower() in {
        "1", "true", "yes", "y", "selected", "best", "conflict", "check",
    }


def _selection(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"yes", "true", "1", "selected", "best"}:
        return "selected"
    if text in {"validation", "validate", "supporting"}:
        return "validation"
    if text in {"no", "false", "0", "rejected", "unselected"}:
        return "unselected"
    return "candidate"


def _confidence_class(value: Any, default: str = "D") -> str:
    text = str(value or "").strip().upper()
    if text in CONFIDENCE_SCORES:
        return text
    try:
        score = float(value)
    except (TypeError, ValueError):
        return default
    if score >= 0.95:
        return "A"
    if score >= 0.80:
        return "B"
    if score >= 0.55:
        return "C"
    return "D"


def _assertion_type(row: Mapping[str, Any], confidence: str, source_type: str) -> str:
    explicit_value = row.get("explicit")
    if explicit_value is not None:
        return "explicit" if _truthy(explicit_value) else "inferred"
    raw = str(row.get("assertion") or row.get("evidence_type") or row.get("basis_type") or "")
    if "infer" in raw.lower() or "推定" in raw:
        return "inferred"
    if "explicit" in raw.lower() or "明記" in raw:
        return "explicit"
    if confidence == "D" or "INFER" in source_type.upper():
        return "inferred"
    return "explicit"


def _first(row: Mapping[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if not is_blank(value):
            return value
    return None


def _int_or_text(value: Any) -> int | str | None:
    if is_blank(value):
        return None
    try:
        number = float(value)
        if number.is_integer():
            return int(number)
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def parse_locator(locator: Any) -> dict[str, Any]:
    """Extract common PDF/printed-page labels while retaining the raw locator."""
    text = "" if is_blank(locator) else str(locator).strip()
    result: dict[str, Any] = {"locator": text or None}
    if not text:
        result.update({"pdf_page": None, "printed_page": None, "section": None})
        return result

    pdf_page = None
    printed_page = None
    for pattern in (
        r"(?i)\bPDF\s*(?:page|p\.?|頁)?\s*[:#]?\s*(\d+)",
        r"PDF\s*[頁ページ]+\s*[:#]?\s*(\d+)",
    ):
        match = re.search(pattern, text)
        if match:
            pdf_page = int(match.group(1))
            break
    for pattern in (
        r"(?i)\bprinted\s*(?:page|p\.?)?\s*[:#]?\s*(\d+)",
        r"(?:印刷|冊子|本文)\s*(?:page|p\.?|頁|ページ)?\s*[:#]?\s*(\d+)",
    ):
        match = re.search(pattern, text)
        if match:
            printed_page = int(match.group(1))
            break
    section = None
    match = re.search(r"(?i)(?:section|table|fig(?:ure)?|節|表|図)\s*[:#]?\s*([^;|\n]+)", text)
    if match:
        section = match.group(0).strip()
    result.update({"pdf_page": pdf_page, "printed_page": printed_page, "section": section})
    return result


def _normalize_source_type(value: Any) -> str:
    text = str(value or "Unknown").strip()
    upper = text.upper()
    if "SHAPE" in upper:
        return "Shapefile"
    if "ZFK" in upper:
        return "ZFK"
    if "PDF" in upper:
        return "PDF"
    if "LLM" in upper:
        return "LLM"
    if "REVIEW" in upper or "HUMAN" in upper:
        return "Review"
    return text or "Unknown"


def _stable_evidence_id(record: Mapping[str, Any]) -> str:
    identity = {
        "unit_id": record.get("unit_id"),
        "row_key": record.get("row_key"),
        "column_ids": record.get("column_ids"),
        "scope": record.get("scope"),
        "field": record.get("field"),
        "candidate": record.get("candidate"),
        "source_type": record.get("source", {}).get("type"),
        "source_locator": record.get("source", {}).get("locator"),
        "quote": record.get("source", {}).get("quote"),
        "assertion": record.get("assertion"),
    }
    payload = json.dumps(identity, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return "ev_" + hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def _evidence_scope(
    row: Mapping[str, Any],
    *,
    unit_id: str | None,
    row_key: Any,
    column_ids: Sequence[str],
    field: str,
    source_type: str,
) -> dict[str, Any]:
    """Normalize evidence binding independently from temporary Column names.

    Older PDF extraction records used the placeholder Column ``unsplit``.
    Those records describe a unit, not a future Column row, so this migration
    treats them as unit-global evidence.  Explicit scope always wins.
    """
    raw_scope = row.get("scope")
    scope_type = None
    scope_columns: list[str] = []
    if isinstance(raw_scope, Mapping):
        scope_type = str(raw_scope.get("type") or "").strip().lower()
        scope_columns = split_column_ids(raw_scope.get("column_ids"))
    if not scope_type:
        scope_type = str(row.get("scope_type") or "").strip().lower()
    if not scope_columns:
        scope_columns = split_column_ids(row.get("scope_column_ids"))

    if scope_type not in EVIDENCE_SCOPE_TYPES:
        if not unit_id:
            scope_type = "map_global"
        elif (
            [value.casefold() for value in column_ids] == ["unsplit"]
            or "::unsplit" in str(row_key or "").casefold()
        ):
            scope_type = "unit_global"
        elif field in UNIT_GLOBAL_FIELDS and not row_key:
            scope_type = "unit_global"
        else:
            scope_type = "column_specific"

    if scope_type == "unit_global":
        scope_columns = []
    elif scope_type == "column_specific":
        if not scope_columns:
            scope_columns = [value for value in column_ids if value.casefold() != "unsplit"]
    else:  # map_global
        scope_columns = []
    return {"type": scope_type, "column_ids": scope_columns}


def _normalized_candidate(field: str, value: Any) -> str:
    if is_blank(value):
        return ""
    if field in {"lithology", "minor_lith"}:
        stripped = _strip_trailing_source_note(value)
        resolved = resolve_lithology_value(stripped)
        if resolved.get("value"):
            return str(resolved["value"]).casefold().replace("; ", ";")
    text = re.sub(r"\s+", " ", str(value)).strip().casefold()
    if field == "unit_name":
        # ZFK display names commonly append lithology in parentheses whereas
        # the editable unit_name intentionally omits it.
        text = re.sub(r"\s*[\(（][^\(\)（）]*[\)）]\s*$", "", text).strip()
    if field in {"unit_name", "lithology", "minor_lith", "age_text"}:
        text = re.sub(r"[;,/・、]+", ";", text)
        text = re.sub(r"\s*;\s*", ";", text)
    return text


def _labeled_context_value(context: Any, label: str) -> str | None:
    """Read a single ``label: value`` line from Review-v2 compact evidence."""
    if is_blank(context):
        return None
    match = re.search(
        rf"(?im)^\s*{re.escape(label)}\s*:\s*(.*?)\s*$",
        str(context),
    )
    return match.group(1).strip() if match and match.group(1).strip() else None


def _normalize_evidence_row(row: Mapping[str, Any]) -> dict[str, Any] | None:
    unit_id = _first(row, "unit_id", "source_unit_id")
    field = _first(row, "field", "field_name")
    candidate = _first(row, "candidate", "candidate_value", "value")
    compact_context = _first(row, "source_and_full_context")
    quote = _first(
        row, "full_context_quote", "quote", "context", "matched_sentence",
        "source_and_full_context",
    )
    if is_blank(field) or (is_blank(candidate) and is_blank(quote)):
        return None

    source_raw = _first(row, "source_type", "source", "provider")
    if is_blank(source_raw):
        source_raw = _labeled_context_value(compact_context, "source_type")
    source_type = _normalize_source_type(source_raw)
    confidence_raw = _first(row, "confidence_class", "confidence")
    if is_blank(confidence_raw):
        confidence_raw = _labeled_context_value(compact_context, "confidence")
        if confidence_raw:
            confidence_raw = confidence_raw.split(" ", 1)[0]
    confidence = _confidence_class(confidence_raw, "D")
    flag = str(_first(row, "flag") or "").strip().upper()
    assertion = (
        "inferred" if flag == "INFERRED"
        else _assertion_type(row, confidence, source_type)
    )
    locator_raw = _first(row, "source_locator", "locator")
    if is_blank(locator_raw):
        locator_raw = _labeled_context_value(compact_context, "locator")
    parsed = parse_locator(locator_raw)
    pdf_page = _int_or_text(_first(row, "PDF_page", "pdf_page", "page", "page_no"))
    if pdf_page is None:
        pdf_page = _int_or_text(_labeled_context_value(compact_context, "PDF page"))
    printed_page = _int_or_text(_first(row, "printed_page", "printed_page_no"))
    if printed_page is None:
        printed_page = _int_or_text(_labeled_context_value(compact_context, "printed page"))
    section = _first(row, "section_or_table", "section", "table", "figure")
    if is_blank(section):
        section = _labeled_context_value(compact_context, "section/table")

    conflict_flag_value = row.get("conflict")
    conflict_detail_value = row.get("conflict_detail")
    conflict = flag == "CONFLICT" or _truthy(conflict_flag_value) or (
        not is_blank(conflict_detail_value)
        and str(conflict_detail_value).strip().lower() not in {"no", "none", "false", "0"}
    )
    normalized_unit_id = None if is_blank(unit_id) else str(unit_id).strip()
    normalized_row_key = _first(row, "row_key")
    normalized_column_ids = split_column_ids(
        _first(row, "column_ids", "column_id", "col_id")
    )
    normalized_field = str(field).strip()
    scope = _evidence_scope(
        row,
        unit_id=normalized_unit_id,
        row_key=normalized_row_key,
        column_ids=normalized_column_ids,
        field=normalized_field,
        source_type=source_type,
    )
    result = {
        "unit_id": normalized_unit_id,
        "source_unit_id": json_value(_first(row, "source_unit_id")),
        "row_key": normalized_row_key if scope["type"] == "column_specific" else None,
        "column_ids": scope["column_ids"],
        "scope": scope,
        "field": normalized_field,
        "group": FIELD_TO_GROUP.get(str(field).strip(), "other_evidence"),
        "candidate": json_value(candidate),
        "source": {
            "type": source_type,
            "file": json_value(
                _first(row, "source_file", "file", "filename")
                or _labeled_context_value(compact_context, "source_file")
            ),
            "locator": parsed["locator"],
            "pdf_page": pdf_page if pdf_page is not None else parsed["pdf_page"],
            "printed_page": printed_page if printed_page is not None else parsed["printed_page"],
            "section": json_value(section) if section is not None else parsed["section"],
            "matched_sentence": json_value(_first(row, "matched_sentence")),
            "quote": json_value(quote),
        },
        "confidence": {
            "class": confidence,
            "score": CONFIDENCE_SCORES[confidence],
        },
        "assertion": assertion,
        "selection": _selection(_first(row, "selected", "selection", "flag")),
        "conflict": conflict,
        "conflict_detail": (
            None if is_blank(conflict_detail_value) else str(conflict_detail_value).strip()
        ),
        "extraction_method": json_value(_first(row, "extraction_method", "method", "model")),
        "resolution_state": json_value(_first(row, "resolution_state", "field_state")),
        "parse": {
            "raw_phrase": json_value(_first(row, "raw_phrase")),
            "normalized_terms": json_value(_first(row, "normalized_terms")),
            "role": json_value(_first(row, "role")),
            "role_cue": json_value(_first(row, "role_cue")),
            "dropped_modifiers": json_value(_first(row, "dropped_modifiers")),
            "parser": json_value(_first(row, "parser", "parser_version")),
            "source_span": json_value(_first(row, "source_span")),
        },
    }
    if not any(value not in (None, "", []) for value in result["parse"].values()):
        result["parse"] = None
    supplied_id = _first(row, "evidence_id") or _labeled_context_value(
        compact_context, "evidence_id"
    )
    result["evidence_id"] = (
        str(supplied_id).strip() if not is_blank(supplied_id) else _stable_evidence_id(result)
    )
    return result


def _ref_source_type(row: Mapping[str, Any], default: str) -> str:
    if default in {"PDF", "Shapefile"}:
        return default
    if str(row.get("REF_shape_match") or "").strip().lower() == "shape_only":
        return "Shapefile"
    return default


def _ref_evidence(row: Mapping[str, Any], row_key: str, unit_id: str,
                  column_ids: Sequence[str]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    row_confidence = _confidence_class(row.get("REF_confidence_class"), "")
    conflict_detail = row.get("REF_conflict")
    final_values = {
        key: value for key, value in row.items()
        if not key.startswith("REF_") and key not in _EVIDENCE_SUMMARY_COLUMNS
    }

    for ref_col, field, source_default, confidence_default, assertion, locator_col in _REF_EVIDENCE_SPECS:
        candidate = row.get(ref_col)
        if is_blank(candidate):
            continue
        source_type = _ref_source_type(row, source_default)
        confidence = row_confidence if row_confidence in CONFIDENCE_SCORES else confidence_default
        locator = row.get(locator_col)
        selection = "candidate"
        if not is_blank(final_values.get(field)):
            if _normalized_candidate(field, final_values[field]) == _normalized_candidate(field, candidate):
                selection = "selected"
            else:
                selection = "validation"
        parsed = parse_locator(locator)
        # Long REF context is retained as the quote.  For short structured
        # labels the same text is still the auditable source statement.
        record = {
            "unit_id": unit_id,
            "row_key": row_key,
            "column_ids": list(column_ids),
            "scope": {"type": "column_specific", "column_ids": list(column_ids)},
            "field": field,
            "group": FIELD_TO_GROUP.get(field, "other_evidence"),
            "candidate": candidate,
            "source": {
                "type": source_type,
                "file": None,
                "locator": parsed["locator"],
                "pdf_page": parsed["pdf_page"],
                "printed_page": parsed["printed_page"],
                "section": parsed["section"],
                "matched_sentence": None,
                "quote": candidate,
            },
            "confidence": {"class": confidence, "score": CONFIDENCE_SCORES[confidence]},
            "assertion": assertion,
            "selection": selection,
            "conflict": not is_blank(conflict_detail),
            "conflict_detail": None if is_blank(conflict_detail) else str(conflict_detail),
            "extraction_method": "review_ref_column",
        }
        record["evidence_id"] = _stable_evidence_id(record)
        out.append(record)
    return out


def _source_rank(field: str, source_type: str) -> int:
    source = source_type.upper()
    overrides = FIELD_SOURCE_PRIORITY.get(field, {})
    if source in overrides:
        return overrides[source]
    if source == "SHAPE":
        source = "SHAPEFILE"
    return SOURCE_PRIORITY.get(source, 0)


def _evidence_rank(record: Mapping[str, Any]) -> tuple[int, float, int, int]:
    selection_rank = {
        "selected": 4,
        "validation": 3,
        "candidate": 2,
        "unselected": 0,
    }.get(str(record.get("selection")), 1)
    confidence_score = float(record.get("confidence", {}).get("score") or 0)
    assertion_rank = 1 if record.get("assertion") == "explicit" else 0
    source_rank = _source_rank(
        str(record.get("field") or ""), str(record.get("source", {}).get("type") or "")
    )
    return selection_rank, confidence_score, assertion_rank, source_rank


def _dedupe_evidence(records: Iterable[dict[str, Any]]) -> list[dict[str, Any]]:
    best: dict[str, dict[str, Any]] = {}
    for record in records:
        evidence_id = str(record["evidence_id"])
        previous = best.get(evidence_id)
        if previous is None or _evidence_rank(record) > _evidence_rank(previous):
            best[evidence_id] = record
    return list(best.values())


def _mark_automatic_conflicts(records: list[dict[str, Any]]) -> None:
    """Flag divergent A/B structured candidates from different sources."""
    by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for record in records:
        if record["field"] not in {"unit_name", "lithology", "minor_lith", "age_text"}:
            continue
        if record["assertion"] != "explicit":
            continue
        if record["confidence"]["class"] not in {"A", "B"}:
            continue
        key = (str(record.get("row_key") or ""), str(record.get("unit_id") or ""), record["field"])
        by_key.setdefault(key, []).append(record)

    for items in by_key.values():
        values: dict[str, set[str]] = {}
        for item in items:
            normalized = _normalized_candidate(item["field"], item["candidate"])
            source = item["source"]["type"]
            if normalized:
                values.setdefault(normalized, set()).add(source)
        source_types = {item["source"]["type"] for item in items}
        if len(values) <= 1 or len(source_types) <= 1:
            continue
        field = str(items[0].get("field") or "")
        if field in {"lithology", "minor_lith"}:
            term_sets = [
                {term for term in normalized.split(";") if term}
                for normalized in values
            ]
            # A terse legend and a detailed body description often differ only
            # by extra compatible terms.  Treat a pure subset as corroboration,
            # not disagreement; incompatible cross-terms still conflict.
            if all(
                left.issubset(right) or right.issubset(left)
                for index, left in enumerate(term_sets)
                for right in term_sets[index + 1:]
            ):
                continue
        detail = "Structured sources disagree: " + " | ".join(sorted(values))
        for item in items:
            item["conflict"] = True
            if not item.get("conflict_detail"):
                item["conflict_detail"] = detail


def _display_candidate(value: Any, limit: int = 120) -> str:
    if is_blank(value):
        return "(context only)"
    text = re.sub(r"\s+", " ", str(value)).strip()
    return text if len(text) <= limit else text[: limit - 1] + "…"


def _summarize_group(group: str, records: Sequence[dict[str, Any]],
                     values: Mapping[str, Any], origins: Mapping[str, str]) -> dict[str, Any]:
    grouped = [record for record in records if record.get("group") == group]
    by_field: dict[str, list[dict[str, Any]]] = {}
    for record in grouped:
        by_field.setdefault(record["field"], []).append(record)

    best_by_field: dict[str, dict[str, Any]] = {}
    for field, items in by_field.items():
        best = max(items, key=_evidence_rank)
        best_by_field[field] = {
            "evidence_id": best["evidence_id"],
            "candidate": best["candidate"],
            "source_type": best["source"]["type"],
            "confidence_class": best["confidence"]["class"],
            "assertion": best["assertion"],
            "conflict": best["conflict"],
        }

    has_conflict = any(record.get("conflict") for record in grouped)
    inferred_best = any(item["assertion"] == "inferred" for item in best_by_field.values())
    candidate_origin = any(
        origins.get(field) == "evidence_candidate"
        for field in EVIDENCE_GROUP_FIELDS[group]
    )

    if group == "age_evidence":
        has_value = any(not is_blank(values.get(field)) for field in ("t_int", "b_int", "t_age_ma", "b_age_ma"))
    elif group == "context_evidence":
        has_value = not is_blank(values.get("unit_name"))
    else:
        has_value = any(
            not is_blank(values.get(field))
            for field in ("min_thickness", "max_thickness", "basal_surface", "lateral_relationship")
        )

    if not has_value and not grouped:
        status = "MISSING"
    elif not has_value or has_conflict or inferred_best or candidate_origin:
        status = "CHECK"
    else:
        status = "OK"

    display_lines = []
    for field, item in sorted(best_by_field.items(), key=lambda pair: pair[0])[:8]:
        tags = [item["confidence_class"], item["source_type"]]
        if item["assertion"] == "inferred":
            tags.append("INFERRED")
        prefix = "⚠ " if item["conflict"] else ""
        display_lines.append(
            f"{prefix}[{'|'.join(tags)}] {field}: {_display_candidate(item['candidate'])}"
        )
    if not display_lines:
        display_lines.append("No supporting evidence")

    return {
        "status": status,
        "summary": "\n".join(display_lines),
        "best_by_field": best_by_field,
        "evidence_ids": [record["evidence_id"] for record in grouped],
        "conflict_count": sum(bool(record.get("conflict")) for record in grouped),
        "inferred_count": sum(record.get("assertion") == "inferred" for record in grouped),
    }


def _unit_status(values: Mapping[str, Any], origins: Mapping[str, str],
                 records: Sequence[dict[str, Any]], summaries: Mapping[str, Mapping[str, Any]],
                 field_resolution: Mapping[str, Mapping[str, Any]] | None = None) -> tuple[str, list[str]]:
    reasons: list[str] = []
    missing_required = []
    if is_blank(values.get("unit_name")):
        missing_required.append("unit_name")
    if not split_column_ids(values.get("column_id")):
        missing_required.append("column_id")
    if missing_required:
        reasons.extend(f"missing_required:{field}" for field in missing_required)
        return "MISSING", reasons

    if any(record.get("conflict") for record in records):
        reasons.append("source_conflict")
    if any(record.get("assertion") == "inferred" and record.get("selection") != "unselected"
           for record in records):
        reasons.append("inferred_evidence")
    if any(origin == "evidence_candidate" for origin in origins.values()):
        reasons.append("unconfirmed_candidate")
    environment_not_applicable = any(
        record.get("field") == "environment_applicability"
        and str(record.get("candidate") or "").strip().casefold() == "not_applicable"
        and record.get("selection") != "unselected"
        for record in records
    )
    if is_blank(values.get("environment")) and not environment_not_applicable:
        reasons.append("missing_optional:environment")
    if summaries["physical_evidence"]["status"] == "MISSING":
        reasons.append("missing_optional:physical_evidence")
    resolution = field_resolution or {}
    if (resolution.get("lithology") or {}).get("state") == "unresolved":
        reasons.append("unresolved:lithology")
    if (resolution.get("minor_lith") or {}).get("state") == "unresolved":
        reasons.append("unresolved:minor_lith")
    if (resolution.get("age") or {}).get("state") == "unresolved":
        reasons.append("unresolved:chronology")
    elif (resolution.get("age") or {}).get("requires_review"):
        reasons.append("age_requires_review")
    if reasons:
        return "CHECK", reasons
    return "OK", reasons


def _row_key(unit_id: str, column_ids: Sequence[str], index: int,
             seen: dict[str, int]) -> str:
    base = f"{unit_id or 'unit'}::{'|'.join(column_ids) or 'no-column'}"
    seen[base] = seen.get(base, 0) + 1
    return base if seen[base] == 1 else f"{base}::{seen[base]}"


def _related_evidence(record: Mapping[str, Any], row_key: str, unit_id: str,
                      column_ids: Sequence[str]) -> bool:
    scope = record.get("scope") if isinstance(record.get("scope"), Mapping) else {}
    scope_type = str(scope.get("type") or "column_specific")
    if scope_type == "map_global":
        return False
    if str(record.get("unit_id") or "") != unit_id:
        return False
    if scope_type == "unit_global":
        return True
    if record.get("row_key"):
        return record["row_key"] == row_key
    evidence_cols = set(scope.get("column_ids") or record.get("column_ids") or [])
    return not evidence_cols or bool(evidence_cols.intersection(column_ids))


def _review_values(row: Mapping[str, Any]) -> tuple[dict[str, Any], dict[str, str]]:
    keys = list(CANONICAL_VALUE_FIELDS)
    keys.extend(
        key for key in row
        if not key.startswith("REF_")
        and not key.startswith("_")
        and key not in _EVIDENCE_SUMMARY_COLUMNS
        and key not in keys
    )
    values = {key: json_value(row.get(key)) for key in keys}
    # Excel joins over entirely empty numeric source cells can materialize as
    # punctuation-only strings such as `", "` or `", , "`.  They are not
    # scientific measurements and must not block verified numeric evidence.
    for field in NUMERIC_AUTO_RESOLVE_FIELDS:
        value = values.get(field)
        if isinstance(value, str) and not re.sub(r"[\s,;|/]+", "", value):
            values[field] = None
    formulas = row.get("_formulas") or {}
    return values, {str(k): str(v) for k, v in formulas.items() if not is_blank(v)}


def _normalize_lithology_roles(
    values: dict[str, Any], origins: Mapping[str, str]
) -> dict[str, Any] | None:
    """Prevent evidence-resolved major terms from also remaining minor.

    This is a role normalization, not a new scientific assertion.  Human
    review values remain authoritative; the rule is applied only when at least
    one of the two role fields was resolved from evidence candidates.
    """
    if not (
        origins.get("lithology") == "evidence_candidate"
        or origins.get("minor_lith") == "evidence_candidate"
    ):
        return None
    major = [term.strip() for term in str(values.get("lithology") or "").split(";") if term.strip()]
    minor = [term.strip() for term in str(values.get("minor_lith") or "").split(";") if term.strip()]
    if not major or not minor:
        return None
    major_keys = {term.casefold() for term in major}
    removed = [term for term in minor if term.casefold() in major_keys]
    if not removed:
        return None
    retained = [term for term in minor if term.casefold() not in major_keys]
    values["minor_lith"] = "; ".join(retained) if retained else None
    return {
        "rule": "minor_lith_minus_lithology",
        "removed_from_minor_lith": removed,
    }


def _best_field_evidence(field: str, records: Sequence[dict[str, Any]]) -> dict[str, Any] | None:
    candidates = [
        record for record in records
        if record.get("field") == field
        and not is_blank(record.get("candidate"))
        and record.get("selection") != "unselected"
        and record.get("resolution_state") != "explicitly_absent"
    ]
    # A conflict remains a CHECK condition, but it must not allow a lower-grade
    # non-conflicting PDF/LLM candidate to bypass the official structured
    # source.  Select by the normal deterministic rank and expose the conflict
    # through field_resolution.requires_review.
    for candidate in sorted(candidates, key=_evidence_rank, reverse=True):
        if not is_blank(_resolved_candidate_value(
            field,
            candidate.get("candidate"),
            allow_environment_free_text=(
                candidate.get("resolution_state") == "source_verified_free_text"
            ),
        )):
            return candidate
    return None


def _field_resolution(
    field: str,
    values: Mapping[str, Any],
    origins: Mapping[str, str],
    records: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """Describe whether and how one canonical lithology field was resolved."""
    related = [
        record for record in records
        if record.get("field") == field and record.get("selection") != "unselected"
    ]
    conflicts = [record for record in related if record.get("conflict")]
    value = values.get(field)
    if not is_blank(value):
        origin = origins.get(field, "review")
        best = _best_field_evidence(field, related) if origin == "evidence_candidate" else None
        method = "human_review" if origin == "review" else (
            (best or {}).get("extraction_method")
            or str(((best or {}).get("source") or {}).get("type") or "evidence_candidate")
        )
        evidence_ids = []
        if best:
            evidence_ids.append(best.get("evidence_id"))
        requires_review = bool(conflicts)
        if best and (
            best.get("assertion") != "explicit"
            or str((best.get("confidence") or {}).get("class") or "") not in {"A", "B"}
        ):
            requires_review = True
        return {
            "state": "present",
            "method": method,
            "origin": origin,
            "evidence_ids": [value for value in evidence_ids if value],
            "requires_review": requires_review,
            "conflict_count": len(conflicts),
        }

    nonblank_candidates = [record for record in related if not is_blank(record.get("candidate"))]
    absences = [
        record for record in related
        if str(record.get("resolution_state") or "").casefold() == "explicitly_absent"
        and not record.get("conflict")
    ]
    if field == "minor_lith" and absences and not nonblank_candidates and not conflicts:
        best_absence = max(absences, key=_evidence_rank)
        return {
            "state": "explicitly_absent",
            "method": best_absence.get("extraction_method") or "source_review",
            "origin": "evidence_state",
            "evidence_ids": [best_absence.get("evidence_id")],
            "requires_review": best_absence.get("assertion") != "explicit",
            "conflict_count": 0,
        }

    reasons = []
    if conflicts:
        reasons.append("source_conflict")
    if nonblank_candidates:
        reasons.append("candidate_not_safely_resolved")
    else:
        reasons.append("no_candidate")
    return {
        "state": "unresolved",
        "method": None,
        "origin": None,
        "evidence_ids": [record.get("evidence_id") for record in related if record.get("evidence_id")],
        "requires_review": True,
        "conflict_count": len(conflicts),
        "reasons": reasons,
    }


def _age_resolution(
    values: Mapping[str, Any],
    origins: Mapping[str, str],
    records: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    """Describe the disposition of the Formation-level chronology.

    Local analytical ages and biozones are deliberately excluded: those are
    retained as ``local_age_notes`` evidence and cannot fill a boundary.
    """

    fields = ("t_int", "b_int", "t_age_ma", "b_age_ma")
    present_fields = [field for field in fields if not is_blank(values.get(field))]
    related = [
        record for record in records
        if record.get("field") in fields and record.get("selection") != "unselected"
    ]
    conflicts = [record for record in related if record.get("conflict")]
    if not present_fields:
        return {
            "state": "unresolved",
            "method": None,
            "origin": None,
            "evidence_ids": [
                record.get("evidence_id") for record in related if record.get("evidence_id")
            ],
            "requires_review": True,
            "conflict_count": len(conflicts),
            "reasons": ["source_conflict" if conflicts else "no_candidate"],
        }

    evidence_ids: list[str] = []
    methods: list[str] = []
    requires_review = bool(conflicts)
    for field in present_fields:
        origin = origins.get(field, "review")
        if origin == "review":
            methods.append("human_review")
            continue
        best = _best_field_evidence(field, related)
        if best:
            if best.get("evidence_id"):
                evidence_ids.append(str(best["evidence_id"]))
            methods.append(
                str(best.get("extraction_method") or
                    ((best.get("source") or {}).get("type") or "evidence_candidate"))
            )
            if (
                best.get("assertion") != "explicit"
                or str((best.get("confidence") or {}).get("class") or "") not in {"A", "B"}
            ):
                requires_review = True
        else:
            methods.append(str(origin))
            requires_review = True

    if is_blank(values.get("t_int")) or is_blank(values.get("b_int")):
        requires_review = True
    unique_methods = list(dict.fromkeys(methods))
    return {
        "state": "present",
        "method": unique_methods[0] if len(unique_methods) == 1 else "mixed",
        "origin": "review" if all(origins.get(field, "review") == "review" for field in present_fields)
        else "evidence_candidate",
        "evidence_ids": list(dict.fromkeys(evidence_ids)),
        "requires_review": requires_review,
        "conflict_count": len(conflicts),
        "present_fields": present_fields,
    }


def _strip_trailing_source_note(value: Any) -> Any:
    """Return a candidate value without a labelled trailing source note.

    The evidence record itself is never changed.  This helper only prepares a
    possible resolved value, keeping source prose in ``evidence.json`` while
    preventing strings such as ``"sandstone（出典: ...）"`` from entering a
    Macrostrat field.
    """
    if not isinstance(value, str):
        return value
    text = value.strip()
    while text:
        previous = text
        for pattern in _TRAILING_SOURCE_NOTE_PATTERNS:
            text = pattern.sub("", text).strip()
        # A separator immediately before the removed note is not part of the
        # value (for example ``"sandstone; （出典: ...）"``).
        text = re.sub(r"[\s;|,/—–-]+$", "", text).strip()
        if text == previous:
            break
    return text or None


def _safe_numeric_candidate(value: Any) -> int | float | None:
    """Convert only an unambiguous, finite numeric scalar.

    Units, ranges, inequalities, and prose are deliberately rejected.  They
    remain available as evidence but require a human decision before becoming
    a resolved value.
    """
    if isinstance(value, bool) or is_blank(value):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    text = str(value).strip()
    if not _NUMERIC_LITERAL.fullmatch(text):
        return None
    try:
        number = float(text)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(number):
        return None
    if not any(marker in text.lower() for marker in (".", "e")) and number.is_integer():
        return int(number)
    return number


def _resolved_candidate_value(
    field: str,
    raw_candidate: Any,
    *,
    allow_environment_free_text: bool = False,
) -> Any:
    """Prepare one evidence candidate for safe automatic promotion."""
    candidate = _strip_trailing_source_note(raw_candidate)
    if is_blank(candidate):
        return None

    if field in NUMERIC_AUTO_RESOLVE_FIELDS:
        return _safe_numeric_candidate(candidate)

    if field in VOCAB_AUTO_RESOLVE_FIELDS:
        # Macrostrat accepts comma separators for lithology, but the review
        # convention and final submission use semicolons.  Environment keeps
        # commas intact because they may be part of free prose; unknown prose
        # is rejected below in any case.
        vocab_input = str(candidate)
        if field != "environment":
            resolved = resolve_lithology_value(vocab_input)
            return resolved.get("value")
        normalized, _ = normalize_vocab(vocab_input, field, sep=";")
        known, unknown = check_vocab(normalized, field, sep=";")
        if allow_environment_free_text and unknown:
            free_text = re.sub(r"\s+", " ", str(candidate)).strip()
            if (
                free_text
                and len(free_text) <= 160
                and not any(ord(character) < 32 for character in free_text)
            ):
                return free_text
        if unknown or not known:
            return None
        return "; ".join(known)

    return json_value(candidate)


def _shape_lithology_candidates(value: Any) -> dict[str, Any]:
    """Return role-aware, vocabulary-checked GSJ Shape candidates."""
    parsed = parse_lithology_relations(value)
    parsed["secondary_term_fallback"] = False
    if (
        not parsed.get("minor")
        and len(parsed.get("major") or []) > 1
        and (parsed.get("role_cues") or {}).get("major") == "legend_list"
    ):
        # A flat GSJ legend has no abundance verbs.  Under the proposal's
        # 100%-fill rule its leading term is the primary lithology and later
        # terms are secondary candidates.  Mark the inference explicitly.
        leading, *secondary = parsed["major"]
        parsed["major"] = [leading]
        parsed["minor"] = secondary
        parsed["major_value"] = leading
        parsed["minor_value"] = "; ".join(secondary) or None
        parsed["role_cues"]["major"] = "leading legend term"
        parsed["role_cues"]["minor"] = "secondary legend term"
        parsed["secondary_term_fallback"] = True
        for detail in parsed.get("details") or []:
            if detail.get("term") in secondary:
                detail["role"] = "minor"
    parsed["safe"] = bool(parsed.get("major")) and not parsed.get("unknown")
    return parsed


def _shape_lithology_candidate(value: Any) -> str | None:
    """Backward-compatible flattened candidate for older callers.

    New code should use :func:`_shape_lithology_candidates` so subordinate
    terms are emitted as ``minor_lith`` evidence instead of being flattened.
    """
    parsed = _shape_lithology_candidates(value)
    if not parsed.get("safe"):
        return None
    terms = [*parsed.get("major", []), *parsed.get("minor", [])]
    return "; ".join(dict.fromkeys(terms)) or None


def _shape_evidence_rows(
    unit_rows: Sequence[Mapping[str, Any]],
    shape_root: str | os.PathLike[str],
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Create auditable Shapefile candidates without mutating review values.

    The official ``geo_A.dbf`` English unit name is used for an exact normalized
    match.  Duplicate names (for example two mapped facies of one volcanic unit)
    are consumed in DBF order.  Raw lithology stays ``lithology_context`` so that
    it cannot bypass the later Macrostrat controlled-vocabulary mapping step.
    """
    try:
        from shape_source import load_shape_units
    except ImportError:  # pragma: no cover - package-style import
        from .shape_source import load_shape_units

    shape = load_shape_units(shape_root)
    if not shape.get("available"):
        return [], {"available": False}

    by_name: dict[str, list[dict[str, Any]]] = {}
    for item in shape.get("units", []):
        key = _normalized_candidate("unit_name", item.get("unit_name_en"))
        if key:
            by_name.setdefault(key, []).append(item)

    rows: list[dict[str, Any]] = []
    matched = 0
    for review in unit_rows:
        unit_id = review.get("unit_id")
        display_name = review.get("unit_name") or review.get("REF_unit_name_en")
        key = _normalized_candidate("unit_name", display_name)
        candidates = by_name.get(key) or []
        if not candidates:
            continue
        item = candidates.pop(0)
        matched += 1
        locator = (
            f"geo_A.dbf record {item.get('record_index')}; "
            f"MAJOR_CODE={item.get('major_code')}; SYMBOL={item.get('symbol')}"
        )
        context = " | ".join(
            str(value).strip() for value in (
                item.get("display_name_en"), item.get("age_text_en"), item.get("lithology_en")
            ) if not is_blank(value)
        )
        common = {
            "unit_id": unit_id,
            "column_id": review.get("column_id") or review.get("col_id"),
            "source_type": "Shapefile",
            "source_file": shape.get("dbf_path"),
            "source_locator": locator,
            "full_context_quote": context,
            "confidence_class": "A",
            "assertion": "explicit",
            "selection": "validation",
            "extraction_method": "geo_A.dbf exact normalized English-name match",
        }
        for field, value in (
            ("unit_name", item.get("unit_name_en")),
            ("unit_name_ja", item.get("unit_name_ja")),
            ("age_text", item.get("age_text_en")),
            ("lithology_context", item.get("lithology_en")),
        ):
            if not is_blank(value):
                rows.append({**common, "field": field, "candidate": value})
        parsed_lithology = _shape_lithology_candidates(item.get("lithology_en"))
        unknown = [str(value) for value in parsed_lithology.get("unknown") or []]
        parse_conflict = bool(unknown)
        conflict_detail = (
            "Unresolved Shape lithology terms: " + " | ".join(unknown)
            if unknown else None
        )
        for role, field in (("major", "lithology"), ("minor", "minor_lith")):
            terms = list(parsed_lithology.get(role) or [])
            if not terms:
                continue
            role_details = [
                detail for detail in parsed_lithology.get("details") or []
                if detail.get("role") == role
            ]
            dropped = [
                modifier
                for detail in role_details
                for modifier in detail.get("dropped_modifiers") or []
            ]
            rows.append({
                **common,
                "field": field,
                "candidate": "; ".join(terms),
                "explicit": not (
                    role == "minor" and parsed_lithology.get("secondary_term_fallback")
                ),
                "raw_phrase": item.get("lithology_en"),
                "normalized_terms": terms,
                "role": role,
                "role_cue": (parsed_lithology.get("role_cues") or {}).get(role),
                "dropped_modifiers": dropped,
                "parser": "lithology_relation_parser/v2",
                "source_span": item.get("lithology_en"),
                "conflict": parse_conflict,
                "conflict_detail": conflict_detail,
            })

    metadata = {
        "available": True,
        "dbf_path": shape.get("dbf_path"),
        "shp_path": shape.get("shp_path"),
        "bbox": json_value(shape.get("bbox")),
        "centroid": json_value(shape.get("centroid")),
        "unit_count": len(shape.get("units", [])),
        "matched_review_units": matched,
        "unmatched_review_units": max(0, len(unit_rows) - matched),
        "excluded_records": shape.get("excluded_records", 0),
    }
    return rows, metadata


def _infer_map_id(map_id: Any, metadata: Mapping[str, Any], source_review: str | None,
                  units: Sequence[Mapping[str, Any]]) -> str:
    if not is_blank(map_id):
        return str(map_id).strip().lstrip("mM")
    for key in ("map_id", "zfk_map_id", "source_map_id", "project_id"):
        value = metadata.get(key)
        if not is_blank(value):
            match = re.search(r"(?:^|\D)m?(\d{3,5})(?:\D|$)", str(value))
            if match:
                return match.group(1)
    if source_review:
        match = re.search(r"m(\d{3,5})", Path(source_review).name, re.IGNORECASE)
        if match:
            return match.group(1)
    for row in units:
        match = re.search(r"m(\d{3,5})", str(row.get("unit_id") or ""), re.IGNORECASE)
        if match:
            return match.group(1)
    return ""


def build_canonical_layer(
    unit_rows: Iterable[Mapping[str, Any] | Any],
    *,
    column_rows: Iterable[Mapping[str, Any] | Any] | None = None,
    evidence_rows: Iterable[Mapping[str, Any] | Any] | None = None,
    metadata: Mapping[str, Any] | None = None,
    map_id: Any = None,
    source_review: str | os.PathLike[str] | None = None,
    generated_at: str | None = None,
) -> tuple[dict[str, Any], dict[str, Any]]:
    """Build canonical compiled/evidence documents from normalized rows."""
    units = [normalize_row(row) for row in unit_rows]
    columns = [normalize_row(row) for row in (column_rows or [])]
    metadata_clean = normalize_row(metadata or {})
    source_review_text = str(source_review) if source_review is not None else None
    map_id_text = _infer_map_id(map_id, metadata_clean, source_review_text, units)
    timestamp = generated_at or datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    seen: dict[str, int] = {}

    row_descriptors: list[tuple[dict[str, Any], str, str, list[str]]] = []
    all_evidence: list[dict[str, Any]] = []
    for index, row in enumerate(units, start=1):
        unit_id = "" if is_blank(row.get("unit_id")) else str(row["unit_id"]).strip()
        column_ids = split_column_ids(row.get("column_id") or row.get("col_id"))
        row_key = _row_key(unit_id, column_ids, index, seen)
        row_descriptors.append((row, row_key, unit_id, column_ids))
        all_evidence.extend(_ref_evidence(row, row_key, unit_id, column_ids))

    for raw in evidence_rows or []:
        record = _normalize_evidence_row(normalize_row(raw))
        if record is not None:
            all_evidence.append(record)

    all_evidence = _dedupe_evidence(all_evidence)
    _mark_automatic_conflicts(all_evidence)
    all_evidence.sort(key=lambda item: (
        str(item.get("unit_id") or ""), str(item.get("row_key") or ""),
        str(item.get("field") or ""), str(item.get("evidence_id") or ""),
    ))

    compiled_units: list[dict[str, Any]] = []
    status_counts = {status: 0 for status in VALID_STATUSES}
    for row, row_key, unit_id, column_ids in row_descriptors:
        review_values, formulas = _review_values(row)
        values = dict(review_values)
        origins: dict[str, str] = {
            key: "review" for key, value in review_values.items() if not is_blank(value)
        }
        related = [
            record for record in all_evidence
            if _related_evidence(record, row_key, unit_id, column_ids)
        ]
        for field in AUTO_RESOLVE_FIELDS:
            if not is_blank(values.get(field)):
                continue
            best = _best_field_evidence(field, related)
            if best is not None:
                resolved = _resolved_candidate_value(
                    field,
                    best["candidate"],
                    allow_environment_free_text=(
                        best.get("resolution_state") == "source_verified_free_text"
                    ),
                )
                if not is_blank(resolved):
                    values[field] = resolved
                    origins[field] = "evidence_candidate"

        lithology_role_normalization = _normalize_lithology_roles(values, origins)

        summaries = {
            group: _summarize_group(group, related, values, origins)
            for group in EVIDENCE_GROUP_FIELDS
        }
        field_resolution = {
            field: _field_resolution(field, values, origins, related)
            for field in ("lithology", "minor_lith")
        }
        field_resolution["age"] = _age_resolution(values, origins, related)
        status, reasons = _unit_status(
            values, origins, related, summaries, field_resolution
        )
        status_counts[status] += 1
        human_status = _first(row, "status", "review_status")
        compiled_units.append({
            "row_key": row_key,
            "unit_id": unit_id or None,
            "column_ids": column_ids,
            "review_values": review_values,
            "values": values,
            "value_origins": origins,
            "field_resolution": field_resolution,
            "lithology_role_normalization": lithology_role_normalization,
            "formulas": formulas,
            "age_evidence": summaries["age_evidence"],
            "context_evidence": summaries["context_evidence"],
            "physical_evidence": summaries["physical_evidence"],
            "status": status,
            "status_reasons": reasons,
            "human_status": json_value(human_status),
        })

    bound_evidence_ids = {
        str(record["evidence_id"])
        for record in all_evidence
        if str((record.get("scope") or {}).get("type") or "") == "map_global"
        or any(
            _related_evidence(record, row_key, unit_id, column_ids)
            for _, row_key, unit_id, column_ids in row_descriptors
        )
    }
    orphaned_evidence_ids = [
        str(record["evidence_id"])
        for record in all_evidence
        if str(record["evidence_id"]) not in bound_evidence_ids
    ]

    resolution_counts = {
        field: {"present": 0, "explicitly_absent": 0, "unresolved": 0}
        for field in ("lithology", "minor_lith", "age")
    }
    for unit in compiled_units:
        for field in resolution_counts:
            state = str((unit.get("field_resolution", {}).get(field) or {}).get("state") or "unresolved")
            resolution_counts[field][state] = resolution_counts[field].get(state, 0) + 1
    unit_count = len(compiled_units)
    resolution_metrics = {
        "lithology_present_rate": (
            resolution_counts["lithology"]["present"] / unit_count if unit_count else 0.0
        ),
        "minor_lith_disposition_rate": (
            (unit_count - resolution_counts["minor_lith"]["unresolved"]) / unit_count
            if unit_count else 0.0
        ),
        "age_present_rate": (
            resolution_counts["age"]["present"] / unit_count if unit_count else 0.0
        ),
        "age_interval_pair_rate": (
            sum(
                not is_blank((unit.get("values") or {}).get("t_int"))
                and not is_blank((unit.get("values") or {}).get("b_int"))
                for unit in compiled_units
            ) / unit_count if unit_count else 0.0
        ),
    }

    compiled = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": timestamp,
        "map": {
            "map_id": map_id_text,
            "source_review": source_review_text,
            "metadata": metadata_clean,
            "columns": columns,
        },
        "units": compiled_units,
        "summary": {
            "unit_count": len(compiled_units),
            "column_count": len(columns),
            "evidence_count": len(all_evidence),
            "orphaned_evidence_count": len(orphaned_evidence_ids),
            "orphaned_evidence_ids": orphaned_evidence_ids,
            "status_counts": status_counts,
            "field_resolution_counts": resolution_counts,
            "field_resolution_metrics": resolution_metrics,
            "unresolved_age_unit_ids": [
                unit.get("unit_id") for unit in compiled_units
                if (unit.get("field_resolution", {}).get("age") or {}).get("state") == "unresolved"
            ],
        },
    }
    evidence_doc = {
        "schema_version": SCHEMA_VERSION,
        "generated_at": timestamp,
        "map_id": map_id_text,
        "source_review": source_review_text,
        "evidence": all_evidence,
    }
    return compiled, evidence_doc


def _sheet_name(workbook: Any, candidates: Sequence[str]) -> str | None:
    lookup = {str(name).casefold(): name for name in workbook.sheetnames}
    for candidate in candidates:
        found = lookup.get(candidate.casefold())
        if found:
            return found
    return None


def _read_rows(formula_sheet: Any, value_sheet: Any | None = None) -> list[dict[str, Any]]:
    rows = formula_sheet.iter_rows()
    try:
        header_cells = next(rows)
    except StopIteration:
        return []
    headers = ["" if cell.value is None else str(cell.value).strip() for cell in header_cells]
    if not any(headers):
        return []

    value_rows = value_sheet.iter_rows() if value_sheet is not None else None
    if value_rows is not None:
        try:
            next(value_rows)
        except StopIteration:
            value_rows = None

    out: list[dict[str, Any]] = []
    for formula_cells in rows:
        cached_cells = next(value_rows, ()) if value_rows is not None else ()
        row: dict[str, Any] = {}
        formulas: dict[str, str] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            formula_value = formula_cells[index].value if index < len(formula_cells) else None
            cached_value = cached_cells[index].value if index < len(cached_cells) else None
            if isinstance(formula_value, str) and formula_value.startswith("="):
                formulas[header] = formula_value
                row[header] = json_value(cached_value)
            else:
                row[header] = json_value(formula_value)
        if formulas:
            row["_formulas"] = formulas
        if any(not is_blank(v) for key, v in row.items() if not key.startswith("_")):
            out.append(row)
    return out


def _read_key_values(rows: Sequence[Mapping[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for row in rows:
        key = _first(row, "key", "項目", "name")
        if is_blank(key):
            continue
        value = _first(row, "value", "値")
        result[str(key).strip()] = json_value(value)
    return result


def _read_project_metadata_sheet(worksheet: Any) -> dict[str, Any]:
    """Read the PROJECT_METADATA block from a Review-v2 Project sheet."""
    matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
    marker_rows: list[tuple[int, str]] = []
    for index, row in enumerate(matrix):
        first = next((str(value).strip() for value in row if not is_blank(value)), "")
        marker = first.upper()
        if marker in {"PROJECT_METADATA", "REFERENCES", "IMAGES"}:
            marker_rows.append((index, marker))
    metadata_marker = next(
        ((position, row_index) for position, (row_index, marker) in enumerate(marker_rows)
         if marker == "PROJECT_METADATA"),
        None,
    )
    if metadata_marker is None:
        return {}
    marker_position, marker_row = metadata_marker
    end = (
        marker_rows[marker_position + 1][0]
        if marker_position + 1 < len(marker_rows)
        else len(matrix)
    )
    header_row = next(
        (index for index in range(marker_row + 1, end)
         if any(not is_blank(value) for value in matrix[index])),
        None,
    )
    if header_row is None:
        return {}
    headers = [str(value).strip().lower() if not is_blank(value) else "" for value in matrix[header_row]]
    key_index = headers.index("key") if "key" in headers else 0
    value_index = headers.index("value") if "value" in headers else min(1, len(headers) - 1)
    metadata: dict[str, Any] = {}
    gsj: dict[str, Any] = {}
    for row in matrix[header_row + 1:end]:
        key = row[key_index] if key_index < len(row) else None
        if is_blank(key):
            continue
        name = str(key).strip()
        value = row[value_index] if value_index < len(row) else None
        if name.lower().startswith("gsj_"):
            gsj[name[4:]] = json_value(value)
        else:
            metadata[name] = json_value(value)
    if gsj:
        metadata["gsj"] = gsj
    return metadata


def load_review_workbook(review_path: str | os.PathLike[str]) -> dict[str, Any]:
    """Read old or Review-v2 workbook layouts without modifying the file."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - project already depends on it
        raise RuntimeError("openpyxl is required to read an existing review workbook") from exc

    path = Path(review_path)
    if not path.exists():
        raise FileNotFoundError(path)
    formula_book = openpyxl.load_workbook(path, read_only=True, data_only=False)
    value_book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        def read(names: Sequence[str]) -> list[dict[str, Any]]:
            name = _sheet_name(formula_book, names)
            if name is None:
                return []
            value_name = _sheet_name(value_book, (name,))
            return _read_rows(
                formula_book[name], value_book[value_name] if value_name is not None else None
            )

        units = read(("Review", "units_review", "units"))
        columns = read(("Columns", "columns_review", "columns"))
        evidence = read(("Evidence", "source_evidence"))
        project_rows = read(("project_meta", "metadata"))
        gsj_rows = read(("gsj_meta",))
        project_sheet_name = _sheet_name(value_book, ("Project",))
        metadata = (
            _read_project_metadata_sheet(value_book[project_sheet_name])
            if project_sheet_name is not None
            else _read_key_values(project_rows)
        )
        gsj_metadata = _read_key_values(gsj_rows)
        if gsj_metadata:
            metadata["gsj"] = gsj_metadata
        return {
            "unit_rows": units,
            "column_rows": columns,
            "evidence_rows": evidence,
            "metadata": metadata,
            "source_review": str(path),
        }
    finally:
        formula_book.close()
        value_book.close()


def _atomic_json_write(path: Path, document: Mapping[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(document, ensure_ascii=False, indent=2, sort_keys=False) + "\n"
    with tempfile.NamedTemporaryFile(
        mode="w", encoding="utf-8", dir=path.parent, delete=False, suffix=".tmp"
    ) as handle:
        handle.write(payload)
        temporary = Path(handle.name)
    os.replace(temporary, path)


def write_canonical_layer(
    compiled: Mapping[str, Any], evidence: Mapping[str, Any],
    output_dir: str | os.PathLike[str],
) -> dict[str, str]:
    """Atomically write ``compiled.json`` and ``evidence.json``."""
    destination = Path(output_dir)
    compiled_path = destination / "compiled.json"
    evidence_path = destination / "evidence.json"
    _atomic_json_write(compiled_path, compiled)
    _atomic_json_write(evidence_path, evidence)
    return {"compiled": str(compiled_path), "evidence": str(evidence_path)}


def compile_review_workbook(
    review_path: str | os.PathLike[str],
    *,
    output_dir: str | os.PathLike[str] | None = None,
    map_id: Any = None,
    generated_at: str | None = None,
    shape_root: str | os.PathLike[str] | None = None,
) -> tuple[dict[str, Any], dict[str, Any], dict[str, str] | None]:
    """Read a review workbook, build documents, and optionally write them."""
    bundle = load_review_workbook(review_path)
    resolved_shape_root = (
        Path(shape_root)
        if shape_root is not None
        else Path(review_path).resolve().parent / "references"
    )
    shape_rows, shape_metadata = _shape_evidence_rows(
        bundle["unit_rows"], resolved_shape_root
    )
    metadata = dict(bundle["metadata"])
    if shape_metadata.get("available"):
        metadata["shape"] = shape_metadata
    compiled, evidence = build_canonical_layer(
        bundle["unit_rows"],
        column_rows=bundle["column_rows"],
        evidence_rows=[*bundle["evidence_rows"], *shape_rows],
        metadata=metadata,
        map_id=map_id,
        source_review=bundle["source_review"],
        generated_at=generated_at,
    )
    paths = write_canonical_layer(compiled, evidence, output_dir) if output_dir else None
    return compiled, evidence, paths


def main(argv: Sequence[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Build per-map compiled.json and evidence.json from a review workbook."
    )
    parser.add_argument("review", help="Path to m####_review.xlsx")
    parser.add_argument(
        "--output-dir",
        help="Destination directory (default: <review folder>/compiled)",
    )
    parser.add_argument("--map-id", default=None, help="Override inferred map id")
    parser.add_argument(
        "--shape-root",
        default=None,
        help="Directory containing geo_A.dbf (default: sibling references directory)",
    )
    args = parser.parse_args(argv)
    output_dir = args.output_dir or str(Path(args.review).resolve().parent / "compiled")
    compiled, _evidence, paths = compile_review_workbook(
        args.review, output_dir=output_dir, map_id=args.map_id, shape_root=args.shape_root
    )
    print(
        f"map {compiled['map']['map_id'] or '(unknown)'}: "
        f"{compiled['summary']['unit_count']} units, "
        f"{compiled['summary']['evidence_count']} evidence records"
    )
    print(paths["compiled"])
    print(paths["evidence"])
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
