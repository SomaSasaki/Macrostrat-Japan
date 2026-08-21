# -*- coding: utf-8 -*-
"""
apply_llm_candidates.py — LLMが拾った候補をレビューExcelに書き込む

2種類の列に書く。

  REF_ 列（グレー・参照専用）
      候補と **原文引用** をそのまま残す。あとから根拠を辿るための記録。

  編集列（黄色・緑）
      実際に提出物へ行く値。ここも自動入力する。
      人の作業を減らすのが目的なので、空欄待ちにはしない。

★ 上書きについて
  自動入力は既存の値を上書きする。ただしLLMの出力は実行ごとに変わるので、
  黙って消えると気づけない。そこで必ず

      1. 書き込む前にバックアップを作る（.bak_日時.xlsx）
      2. 変わった値を画面に一覧で出す（旧 → 新）

  上書きしたくないときは apply(..., keep=True)（CLI では --keep）。

★ 情報源の優先順位
  同じ項目に2つの情報源があるときは、確かなほうを優先する。

      [本文] ZFKの derived  … GSJが日本語本文から抽出済み。LLMを通さない。
                              make の時点で strat_name / basal_surface /
                              min_thickness / max_thickness に入っている。
      [要約] 英文Abstract    … LLMが読んだもの。要約なので粗い。

  本文由来の値が既にあるなら、要約由来では上書きしない（格下げになるため）。
  年代・environment・unit_description は英文Abstractにしか無いので要約由来。

★ lithology / minor_lith は自動入力しない
  本文由来（REF_lithology_gsj）と要約由来（REF_lithology）を並べてあるので、
  どちらを採るかは本文を読んで人が決める。主／副の振り分けは機械には決められない。
"""

import os
import re
import shutil
import sys
import unicodedata
from datetime import datetime

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import (  # noqa: E402
    PROP_NUMBER_FORMAT, best_interval_for_age, check_vocab, fits_interval,
    is_blank, normalize_vocab, props_from_ages, truncate_for_cell, vocab_quality)
from llm_extract import format_candidate, format_field  # noqa: E402


def norm_name(s):
    """突き合わせ用に地層名を正規化する。'the Shitazaki Formation' -> 'shitazaki'"""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = s.encode("ascii", "ignore").decode().lower()
    s = re.sub(r"^\s*the\s+", "", s)
    # 種別語を落として固有名だけにする
    s = re.sub(r"\b(formation|fm\.?|member|group|pluton|lava|volcanics?|complex|"
               r"deposits?|tuff|terrace|pyroclastic|flow|fan|volcanic)\b", " ", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " ".join(s.split())


def match_rows(name, row_names):
    """
    LLMの地層名に対応する行を **全部** 返す。完全一致 → 部分一致 の順。

    ★ 同じ地層名の行が複数あることがある。十和田図幅には
      「Tsukihiyama Volcanics」の行が2つある（ZFK の u018 と u019）。
      1件しか返さないと2つ目の行が永遠に空のままになる（実際になった）。

    ★ 部分一致は「単語の集合が片方に含まれること」＋「短いほうが2語以上」を
      条件にする。ただの部分文字列で判定すると、正規化後に1語だけ残る名前
      （十和田段丘堆積物 → "towada"）が、"towada caldera forming stage tephra"
      のような別物にも当たってしまう。実際に十和田テフラ群3件が
      十和田段丘堆積物の行に流れ込み、年代を3回上書きした。
    """
    key = norm_name(name)
    if not key:
        return []
    exact = [i for i, rn in row_names.items() if norm_name(rn) == key]
    if exact:
        return exact

    kt = set(key.split())
    hits = []
    for i, rn in row_names.items():
        rt = set(norm_name(rn).split())
        if not rt or min(len(kt), len(rt)) < 2:
            continue                      # 1語しか残らない名前で部分一致はしない
        if kt <= rt or rt <= kt:
            hits.append(i)
    # 部分一致は1件に絞れたときだけ採用する（取り違えを避ける）
    return hits if len(hits) == 1 else []


def match_row(name, row_names):
    """後方互換。最初の1件だけ返す。"""
    hits = match_rows(name, row_names)
    return hits[0] if hits else None


# --- LLM の抽出結果 -> REF_ 列（引用つきの記録） ---
REF_TARGETS = {
    "REF_age_from_abstract": lambda u: format_candidate(u),
    "REF_lithology": lambda u: format_field(u, "lithology"),
    "REF_minor_lith": lambda u: format_field(u, "minor_lith"),
    "REF_environment": lambda u: format_field(u, "environment"),
    "REF_unit_description": lambda u: format_field(u, "unit_description"),
}

# --- LLM の抽出結果 -> 編集列（提出物に行く値） ---
# (列名, LLMのキー, 本文由来の値があっても上書きするか)
AUTOFILL = [
    ("t_age_ma",         "t_age_ma",         True),
    ("b_age_ma",         "b_age_ma",         True),
    ("environment",      "environment",      True),   # Abstract にしか無い
    ("unit_description", "unit_description", True),   # Abstract にしか無い
    ("strat_name",       "strat_name",       False),  # 本文（凡例）優先
    ("basal_surface",    "basal_surface",    False),  # 本文（contacts）優先
    ("min_thickness",    "min_thickness",    False),  # 本文（derived）優先
    ("max_thickness",    "max_thickness",    False),  # 本文（derived）優先
]

# lithology / minor_lith は意図的に入れていない（本文と要約を見比べて人が決める）


def _desc_score(text, unit_name=""):
    """
    unit_description の内容の濃さを点数にする。上書きの可否を決めるためだけの目安。

    重く見るもの:
      ・地層名が入っているか（'They are mainly composed of gravel' では
        どの地層の話か分からない）
      ・年代だけの文でないか（'The age is ca. 15 ka.' は記載ではない）
      ・長さ（情報量の代わり）
    """
    s = " ".join(str(text or "").split())
    if not s:
        return -1.0
    score = min(len(s), 400) / 400.0            # 0〜1
    key = norm_name(unit_name)
    if key and any(w in norm_name(s).split() for w in key.split()):
        score += 1.0                            # 地層名に触れている
    if re.match(r"^(the age|k-ar age|ft age|the radiometric|they are|it is|"
                r"this is|part of)\b", s.lower()):
        score -= 0.6                            # 主語が地層でない・年代だけ
    if re.search(r"\b(composed of|consists? of|distributed|overlie|underlie|"
                 r"sequence|deposits? (are|is))\b", s.lower()):
        score += 0.3                            # 記載らしい語
    return score


def _norm_cell(v):
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return " ".join(str(v).split())


def apply(review_path, candidates, dry=False, keep=False):
    """
    (書き込んだ数, 新規行数, 対応づかなかった名前) を返す。

    keep=True なら編集列の既存値には触れない（REF_ 列だけ更新）。
    """
    wb = openpyxl.load_workbook(review_path)
    if "units_review" not in wb.sheetnames:
        print("[ERROR] units_review シートがありません。")
        return 0, 0, []

    ws = wb["units_review"]
    header = [c.value for c in ws[1]]

    # 足りない REF_ 列は「末尾に」足す。
    # ★ 途中に insert_cols してはいけない。openpyxl は数式の参照を書き換えないので、
    #   t_prop/b_prop の VLOOKUP が別の列を指すようになってしまう。
    #   正しい位置に並べ直すのは `python run.py repair --migrate` の役目。
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    from common import column_width
    added_cols = []
    for name in REF_TARGETS:
        if name in header:
            continue
        pos = len(header) + 1
        ws.cell(row=1, column=pos).value = name
        ws.cell(row=1, column=pos).fill = PatternFill("solid", start_color="F2F2F2")
        ws.cell(row=1, column=pos).font = Font(bold=True)
        ws.column_dimensions[get_column_letter(pos)].width = column_width(name)
        header = [c.value for c in ws[1]]
        added_cols.append(f"{name}({get_column_letter(pos)}列)")
    if added_cols:
        print(f"  [notice] 列を末尾に追加: {', '.join(added_cols)}")
        print("           正しい位置に並べ直すには "
              "`python run.py repair --migrate` を実行してください。")

    ci = {h: i for i, h in enumerate(header, 1)}

    # 既存行の地層名（unit_name 優先、無ければ REF_unit_name_en）
    row_names = {}
    for r in range(2, ws.max_row + 1):
        nm = ws.cell(row=r, column=ci.get("unit_name", 1)).value
        if not nm and "REF_unit_name_en" in ci:
            nm = ws.cell(row=r, column=ci["REF_unit_name_en"]).value
        if nm and str(nm).strip() and str(nm).strip() != "NO_DATA":
            row_names[r] = nm

    changes = []          # (行, 地層名, 列, 旧, 新)
    written, unmatched = 0, []

    kept_better = []

    def put(r, col, new, name, overwrite):
        """1セル書き込む。変更があれば記録する。"""
        if col not in ci or new in (None, ""):
            return
        cell = ws.cell(row=r, column=ci[col])
        old = cell.value
        if not overwrite and not is_blank(old):
            return                                   # 本文由来の値を守る

        # ★ 公式語彙に当たっている値を、当たっていない値で上書きしない。
        #   LLMは実行ごとに揺れる。前回 'fluvial indet.'（公式語）だったものが
        #   今回 'fluvial'（自由記述）になる、という格下げが実際に起きた。
        if col in ("environment", "lithology", "minor_lith"):
            kind = "environment" if col == "environment" else "lithology"
            new, fixes = normalize_vocab(new, kind)
            if fixes:
                kept_better.append(f"行{r} {col}: " + " / ".join(fixes))
            if not is_blank(old) and vocab_quality(new, kind) < vocab_quality(old, kind):
                kept_better.append(
                    f"行{r} {col}: 公式語の '{old}' を残しました（候補 '{new}'）")
                return

        # ★ 記載文を、より内容の薄いもので上書きしない。
        #   LLMが 'The age is ca. 15 ka.' や 'They are mainly composed of
        #   gravel and sand.'（どの地層の話か分からない）を返すことがある。
        if col == "unit_description" and not is_blank(old):
            if _desc_score(new, name) < _desc_score(old, name):
                kept_better.append(
                    f"行{r} unit_description: 内容の濃い元の文を残しました"
                    f"（候補 '{str(new)[:44]}…'）")
                return

        if _norm_cell(old) == _norm_cell(new):
            return
        if not dry:
            cell.value = truncate_for_cell(new, col)
        changes.append((r, name, col, old, new))

    def fill_row(u, r, name):
        """候補1件を1行に書き込む。"""
        # 1) REF_ 列（引用つきの記録）
        for col, fmt in REF_TARGETS.items():
            v = fmt(u)
            if v and col in ci and not dry:
                ws.cell(row=r, column=ci[col]).value = truncate_for_cell(v, col)

        # 2) 編集列（提出物に行く値）
        if not keep:
            for col, key, overwrite in AUTOFILL:
                put(r, col, u.get(key), name, overwrite)

            # 2b) 年代に合わせて t_int / b_int を直す。
            #     ★ ZFKの時代区分は粗い（「更新世」）ので age_mapping で
            #       Early Pleistocene などに落ちるが、Abstractの数値年代
            #       （引用照合済み）と食い違うことがある。実際、十和田では
            #       0.4 Ma の火砕流が Early Pleistocene(2.58–1.8Ma) に当たり、
            #       b_prop が 2.79 という無効な値になっていた。
            #     数値年代のほうが確かな証拠なので、そちらに合わせる。
            for int_col, age_key in (("t_int", "t_age_ma"), ("b_int", "b_age_ma")):
                if int_col not in ci:
                    continue
                age = u.get(age_key)
                if age is None or age == "":
                    continue
                cur = ws.cell(row=r, column=ci[int_col]).value
                if fits_interval(age, cur):
                    continue
                new_iv = best_interval_for_age(age, cur)
                if new_iv and _norm_cell(new_iv) != _norm_cell(cur):
                    if not dry:
                        ws.cell(row=r, column=ci[int_col]).value = new_iv
                    changes.append((r, name, int_col, cur, new_iv))

            # 3) t_prop / b_prop
            #    噴火イベント（年代1点＋火砕流などの名前）は表示桁で同じ値に
            #    丸まる範囲を入れる。ふつうの地層は数式のまま触らない。
            unit_nm = ws.cell(row=r, column=ci.get("unit_name", 1)).value
            ja = (ws.cell(row=r, column=ci["REF_unit_name_ja"]).value
                  if "REF_unit_name_ja" in ci else "")
            bp, tp, is_event = props_from_ages(
                unit_nm,
                ws.cell(row=r, column=ci["t_int"]).value if "t_int" in ci else None,
                ws.cell(row=r, column=ci["b_int"]).value if "b_int" in ci else None,
                u.get("t_age_ma"), u.get("b_age_ma"), ja)
            if is_event and bp is not None:
                for col, val in (("b_prop", bp), ("t_prop", tp)):
                    if col not in ci:
                        continue
                    cell = ws.cell(row=r, column=ci[col])
                    old = cell.value
                    if not dry:
                        cell.value = val
                        cell.number_format = PROP_NUMBER_FORMAT
                    if not (isinstance(old, str) and old.startswith("=")):
                        if _norm_cell(old) != _norm_cell(val):
                            changes.append((r, name, col, old, val))
                    else:
                        changes.append((r, name, col, "（数式）", val))

    filled_rows = set()
    for u in candidates:
        rows = match_rows(u["unit_name"], row_names)
        if not rows:
            unmatched.append(u)
            continue
        # ★ 同名の行が複数あれば **全部** に入れる。
        #   十和田図幅には「Tsukihiyama Volcanics」の行が2つあり（ZFK u018/u019）、
        #   1件しか埋めていなかったので2つ目が空のままだった。
        for r in rows:
            fill_row(u, r, row_names[r])
            filled_rows.add(r)
        written += 1

    # ★ Abstract に出てこない地層を知らせる。
    #   十和田の「崖錐堆積物（Talus deposits）」は英文Abstractに一言も無い。
    #   捏造しないので空欄のままになる。それが正しい挙動だと分かるように、
    #   どの地層が該当するかを明示する（本文＝REF_desc から手で書く目印）。
    no_abstract = [(r, n) for r, n in row_names.items() if r not in filled_rows]
    added = 0
    if not row_names and unmatched:
        start = 2
        if ws.max_row >= 2 and str(ws.cell(row=2, column=ci.get(
                "REF_unit_name_en", 1)).value or "").strip() == "NO_DATA":
            if not dry:
                ws.delete_rows(2)
        for k, u in enumerate(unmatched):
            r = start + k
            if not dry:
                ws.cell(row=r, column=ci["unit_name"]).value = u["unit_name"]
                for col, fmt in REF_TARGETS.items():
                    v = fmt(u)
                    if v and col in ci:
                        ws.cell(row=r, column=ci[col]).value = truncate_for_cell(v, col)
                if not keep:
                    for col, key, _ in AUTOFILL:
                        if col in ci and u.get(key) not in (None, ""):
                            ws.cell(row=r, column=ci[col]).value = \
                                truncate_for_cell(u[key], col)
                if "sort_order" in ci:
                    ws.cell(row=r, column=ci["sort_order"]).value = k + 1
                if "column_id" in ci:
                    ws.cell(row=r, column=ci["column_id"]).value = \
                        ws.cell(row=2, column=ci["column_id"]).value or 1
            added += 1
        unmatched = []

    report_changes(changes)
    if no_abstract:
        print(f"\n  --- 英文Abstractに記載が無い地層 {len(no_abstract)} 件 ---")
        print("      Abstractに一言も出てこないので、候補を作れません（捏造しません）。")
        print("      REF_desc（本文全文）と REF_source（PDFページ）を見て、"
              "手で書いてください。")
        for _, n in no_abstract[:12]:
            print(f"        ・{n}")
        if len(no_abstract) > 12:
            print(f"        ... 他 {len(no_abstract) - 12} 件")
    if kept_better:
        print(f"\n  --- 語彙の調整 {len(kept_better)} 件 ---")
        for m in kept_better[:15]:
            print(f"    {m}")
        if len(kept_better) > 15:
            print(f"    ... 他 {len(kept_better) - 15} 件")
    warn_offvocab(ws, ci, changes)

    if not dry:
        # ★ 上書きするので、保存の前に必ず退避しておく。
        #   LLMの出力は実行ごとに変わる。前回のほうが良かったときに戻せるように。
        if changes:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            bak = review_path.replace(".xlsx", f".bak_{stamp}.xlsx")
            try:
                shutil.copy2(review_path, bak)
                print(f"  [backup] 上書き前の状態を退避: {os.path.basename(bak)}")
            except Exception as e:
                print(f"  [warn] バックアップを作れませんでした（{type(e).__name__}）。"
                      "上書きは中止します。")
                return -1, 0, [u["unit_name"] for u in unmatched]
        try:
            wb.save(review_path)
        except PermissionError:
            print()
            print("[ERROR] Excelファイルが開かれているため保存できませんでした。")
            print(f"        {review_path}")
            print("        Excelで閉じてから、もう一度同じコマンドを実行してください。")
            print("        （Geminiの結果は上に出ているので、無料枠の消費は1回分だけです）")
            return -1, 0, [u["unit_name"] for u in unmatched]

    return written, added, [u["unit_name"] for u in unmatched]


def report_changes(changes, limit=40):
    """自動入力で変わった値を一覧にする。★ 黙って上書きしないための仕組み。"""
    if not changes:
        return
    from common import pad
    filled = [c for c in changes if is_blank(c[3])]
    over = [c for c in changes if not is_blank(c[3])]
    print(f"\n  --- 自動入力 {len(changes)} 件"
          f"（新規 {len(filled)} / 上書き {len(over)}）---")
    if over:
        print("  ★ 上書きしたもの（元の値が消えています。必要なら戻してください）")
        for r, nm, col, old, new in over[:limit]:
            print(f"    行{r:<4}{pad(str(nm)[:26], 28)}{pad(col, 18)}"
                  f"{str(old)[:24]!r} → {str(new)[:34]!r}")
        if len(over) > limit:
            print(f"    ... 他 {len(over) - limit} 件")
    if filled:
        by_col = {}
        for _, _, col, _, _ in filled:
            by_col[col] = by_col.get(col, 0) + 1
        print("  新しく埋まったもの: "
              + " / ".join(f"{k} {v}件" for k, v in sorted(by_col.items())))


def warn_offvocab(ws, ci, changes):
    """自動入力した environment が Macrostrat 公式表に無いときに知らせる。"""
    terms = {}
    for r, _, col, _, new in changes:
        if col != "environment":
            continue
        for t in check_vocab(new, "environment")[1]:
            terms[t] = terms.get(t, 0) + 1
    if terms:
        print("  [notice] 公式表に無い environment: "
              + ", ".join(f"'{k}'×{v}" for k, v in sorted(terms.items()))
              + "（自由記述は仕様上許容。そのままで構いません）")
