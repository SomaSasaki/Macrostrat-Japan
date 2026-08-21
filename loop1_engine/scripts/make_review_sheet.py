# -*- coding: utf-8 -*-
"""
make_review_sheet.py — GSJ 5万分の1図幅からレビュー用 Excel を生成する

Macrostrat column-ingestion format v0.1.1 準拠のレビューシートを作る。
生成されるシート:
  Instructions   使い方
  units_review   地層データ本体（REF_* = 参照専用 / それ以外 = 編集対象）
  columns_review Column 定義
  refs_review    文献（ZFK メタデータから自動下書き）
  images_review  図版（references/ 内の画像から自動下書き）
  project_meta   プロジェクトメタデータ（key-value）
  gsj_meta       GSJ 由来の出典情報（自動・参照専用）
"""

import argparse
import glob
import json
import os
import re
import shutil
import sys
import urllib.request
import zipfile
from datetime import datetime

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import (  # noqa: E402
    FORMAT_VERSION,
    METADATA_DEFAULTS,
    METADATA_KEYS,
    NO_WRAP_COLS,
    PROP_NUMBER_FORMAT,
    REVIEW_COLUMN_COLS,
    REVIEW_UNIT_COLS,
    SUBMISSION_IMAGE_COLS,
    SUBMISSION_REF_COLS,
    canonical_map_title,
    extract_thickness_notes,
    is_blank,
    format_gsj_author,
    get_region_folder,
    column_width,
    truncate_for_cell,
    intervals_for_excel,
    gsj_doc_url,
    join_authors,
    load_json,
    make_ref_id,
    normalize_sheet_code,
    safe_folder_name,
    slugify_col_id,
    strip_trailing_paren,
)

UA = {"User-Agent": "Mozilla/5.0"}


# ---------------------------------------------------------------------------
# API 取得
# ---------------------------------------------------------------------------

def fetch_json(url, label=""):
    req = urllib.request.Request(url, headers=UA)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        if label:
            print(f"  [notice] {label} を取得できませんでした: {e}")
        return None


def fetch_unit_details(unit_id, local_dir):
    """ZFK の地層詳細。ローカルキャッシュ優先。エンドポイントは2系統を試す。"""
    local_path = os.path.join(local_dir, "units", f"{unit_id}.json")
    if os.path.exists(local_path):
        return load_json(local_path)

    for url in (
        f"https://gbank.gsj.jp/ld/resource/zfk/units/{unit_id}.json",
        f"https://gbank.gsj.jp/ld/resource/zfk/unit/{unit_id}.json",
    ):
        data = fetch_json(url)
        if data:
            # 次回以降のためにキャッシュしておく（API 負荷軽減）
            try:
                os.makedirs(os.path.dirname(local_path), exist_ok=True)
                with open(local_path, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False)
            except Exception:
                pass
            return data
    print(f"  [warn] unit {unit_id} の詳細を取得できませんでした")
    return {}


def extract_candidates(text, mapping_dict):
    if not text:
        return ""
    candidates = []
    for k in sorted(mapping_dict.keys(), key=len, reverse=True):
        if k in text:
            v = mapping_dict[k]
            if v not in candidates:
                candidates.append(v)
    return ", ".join(candidates)


def download_assets(map_id, output_dir):
    print("--- 参照ファイルをダウンロード中 ---")
    assets_dir = os.path.join(output_dir, "references")
    os.makedirs(assets_dir, exist_ok=True)

    pub_data = fetch_json(
        f"https://gbank.gsj.jp/ld/resource/publication/map/g050/map{map_id}.json",
        "出版メタデータ",
    )
    if not pub_data:
        return

    for d in pub_data.get("downloadData", []):
        file_url = d.get("@id")
        if not file_url:
            continue
        data_type = (d.get("data_type") or "").lower()
        title = d.get("title", "Unnamed")
        filename = file_url.split("/")[-1]
        filepath = os.path.join(assets_dir, filename)

        is_pdf = "pdf" in data_type or file_url.endswith(".pdf")
        is_zip = "shapefile" in data_type or "geotiff" in data_type or file_url.endswith(".zip")
        if not (is_pdf or is_zip):
            continue

        extract_dir = os.path.join(assets_dir, filename.replace(".zip", ""))
        # 既に取得済みならスキップ（再実行を高速化）
        if (is_pdf and os.path.exists(filepath)) or (is_zip and os.path.isdir(extract_dir)):
            print(f"  [skip] 取得済み: {title}")
            continue

        print(f"  ダウンロード: {title} ({data_type})")
        try:
            urllib.request.urlretrieve(file_url, filepath)
            if is_zip:
                with zipfile.ZipFile(filepath, "r") as zf:
                    zf.extractall(extract_dir)
                os.remove(filepath)
        except Exception as e:
            print(f"  [warn] {file_url} の取得に失敗: {e}")


# ---------------------------------------------------------------------------
# refs / images の自動下書き
# ---------------------------------------------------------------------------

IMAGE_ROLE_HINTS = [
    (re.compile(r"_L(\d+)\.(jpg|jpeg|png)$", re.I), "Map legend (凡例)"),
    (re.compile(r"_S(\d+)\.(jpg|jpeg|png)$", re.I), "Cross-section (断面図)"),
    (re.compile(r"_F(\d+)\.(jpg|jpeg|png)$", re.I), "Map face (図面)"),
]


def build_refs_draft(map_meta, sheet_code, map_id):
    """ZFK map.json の書誌情報から refs シートの下書きを作る。不明値は空欄のまま。"""
    authors_en = [a.get("name_en") for a in map_meta.get("authors", []) if a.get("name_en")]
    authors_ja = [a.get("name_ja") for a in map_meta.get("authors", []) if a.get("name_ja")]
    if not authors_en and map_meta.get("author"):
        authors_ja = authors_ja or list(map_meta.get("author") or [])

    ref_id = make_ref_id(authors_en, map_meta.get("pub_year"), map_id)
    title_en = map_meta.get("title_en") or ""
    series = map_meta.get("series") or ""
    publication = "Quadrangle Series, 1:50,000" if "5万分の1" in series else series

    return pd.DataFrame([{
        "ref_id": ref_id,
        "title": title_en,  # title_en が null の GSJ 図幅が多い。空欄なら手動入力する。
        "authors": join_authors([format_gsj_author(a) for a in authors_en]),
        "publication": publication,
        "compilation": "",
        "organization": "Geological Survey of Japan, AIST" if map_meta.get("publisher") else "",
        "date": map_meta.get("pub_year") or "",
        "doi": "",
        "url": gsj_doc_url(sheet_code),
        "comments": (
            f"ZFK map.json より自動生成。原題(和): {map_meta.get('title_ja', '')} / "
            f"著者(和): {'、'.join(authors_ja)}"
        ) if (map_meta.get("title_ja") or authors_ja) else "",
    }], columns=SUBMISSION_REF_COLS)


def build_images_draft(output_dir, ref_id, col_ids):
    """references/ 以下の凡例・断面画像から images シートの下書きを作る。"""
    rows = []
    assets_dir = os.path.join(output_dir, "references")
    if not os.path.isdir(assets_dir):
        return pd.DataFrame(columns=SUBMISSION_IMAGE_COLS)

    found = []
    for pattern in ("*.jpg", "*.jpeg", "*.png"):
        found.extend(glob.glob(os.path.join(assets_dir, "**", pattern), recursive=True))

    for path in sorted(found):
        name = os.path.basename(path)
        role = ""
        for rx, label in IMAGE_ROLE_HINTS:
            if rx.search(name):
                role = label
                break
        if not role:
            continue
        rows.append({
            "col_ids": ",".join(col_ids),
            "image_name": name,
            "ref_id": ref_id,
            "page_no": "",
            "fig_no": "",
            "description": role,  # ★ 役割のみ自動判定。中身の説明は手動で追記すること
            "comments": f"references/ から自動検出: {os.path.relpath(path, output_dir)}",
        })

    if not rows:
        rows.append({c: "" for c in SUBMISSION_IMAGE_COLS})
        rows[0]["col_ids"] = ",".join(col_ids)
        rows[0]["ref_id"] = ref_id
        rows[0]["comments"] = "図版が自動検出されませんでした。手動で追加してください。"
    return pd.DataFrame(rows, columns=SUBMISSION_IMAGE_COLS)


# ---------------------------------------------------------------------------
# 英文Abstract
# ---------------------------------------------------------------------------

def extract_abstract_for(map_id, output_dir):
    """
    references/ の説明書PDFから英文Abstractを取り出す。

    GSJ説明書の巻末Abstractには、地層の英語名・年代（Ma）・上下関係・岩相が
    まとまっている。一戸図幅では完成形の b_prop/t_prop の元になった年代が
    すべてここにあった（本文200ページを読む必要はない）。

    戻り値: (本文, (開始ページ, 終了ページ) or None, 使ったPDFのパス or "")
    """
    print("--- 英文Abstractを抽出中 ---")
    assets = os.path.join(output_dir, "references")
    pdfs = sorted(glob.glob(os.path.join(assets, "*_D.pdf"))) or \
        sorted(glob.glob(os.path.join(assets, "*.pdf")))
    if not pdfs:
        print("  [notice] 説明書PDFが見つかりません。Abstractはスキップします。")
        return "", None, ""

    try:
        from extract_abstract import extract, summarize
    except Exception as e:
        print(f"  [notice] Abstract抽出モジュールを読めません: {e}")
        return "", None, ""

    try:
        text, rng = extract(pdfs[0])
    except FileNotFoundError:
        print("  [notice] pdftotext が見つかりません（poppler-utils が必要）。スキップします。")
        return "", None, pdfs[0]
    except Exception as e:
        print(f"  [notice] Abstract抽出に失敗: {e}")
        return "", None, pdfs[0]

    if not text.strip():
        print("  [notice] 英文Abstractを検出できませんでした。")
        return "", None, pdfs[0]

    out_txt = os.path.join(assets, f"m{map_id}_abstract.txt")
    try:
        with open(out_txt, "w", encoding="utf-8") as f:
            f.write(text)
    except Exception:
        pass

    s = summarize(text)
    print(f"  p.{rng[0]}-{rng[1]} / {s['chars']}文字 / "
          f"年代レンジ {s['年代レンジ (A–B Ma)']} / 地層名 {s['地層名 (Formation/Group等)']}")
    return text, rng, pdfs[0]


def abstract_to_rows(text):
    """Abstractを段落ごとに分けて DataFrame にする（Excelの1セル上限を避けるため）。"""
    if not text.strip():
        return pd.DataFrame([{"no": 1, "text":
                              "英文Abstractを取得できませんでした。"}])
    paras, buf = [], []
    for line in text.splitlines():
        if line.strip():
            buf.append(line.strip())
        elif buf:
            paras.append(" ".join(buf))
            buf = []
    if buf:
        paras.append(" ".join(buf))
    paras = [p for p in paras if len(p) > 2]
    return pd.DataFrame([{"no": i, "text": p[:32000]} for i, p in enumerate(paras, 1)])


# ---------------------------------------------------------------------------
# メイン
# ---------------------------------------------------------------------------

def build_review_excel(map_id, force=False, n_columns=1, compiler_name=""):
    print(f"=== レビュー用 Excel を生成: Map {map_id} ===")

    # --- 1. 出版APIからタイトルと図幅コード ---
    title_ja = f"Map_{map_id}"
    title_en_guess = ""
    sheet_code = ""
    pub_data = fetch_json(
        f"https://gbank.gsj.jp/ld/resource/publication/map/g050/map{map_id}.json",
        "出版メタデータ",
    )
    if pub_data:
        title_ja = pub_data.get("title_j", pub_data.get("label", title_ja))
        if "GeoMap" in title_ja:
            title_ja = title_ja.replace("1:50K GeoMap: ", "").replace("1:50K GeoMap '", "").replace("'", "")
        if "5万分の1地質図幅" in title_ja:
            title_ja = title_ja.replace("5万分の1地質図幅", "").replace("「", "").replace("」", "").strip()
        label_en = pub_data.get("label", "")
        m = re.search(r"GeoMap[:']\s*([A-Za-z\- ]+)", label_en)
        if m:
            title_en_guess = m.group(1).strip()
        for ld in pub_data.get("downloadData", []):
            m2 = re.search(r"G050_(\d{5})", ld.get("@id", ""))
            if m2:
                sheet_code = m2.group(1)
                break

    # --- 2. ZFK API ---
    local_dir = os.path.join("data", "raw", "zfk", f"m{map_id}")
    local_map_path = os.path.join(local_dir, "map.json")
    map_data = load_json(local_map_path) if os.path.exists(local_map_path) else None
    if map_data is None:
        map_data = fetch_json(
            f"https://gbank.gsj.jp/ld/resource/zfk/maps/m{map_id}.json", "ZFK メタデータ"
        ) or {}
        if map_data:
            try:
                os.makedirs(local_dir, exist_ok=True)
                with open(local_map_path, "w", encoding="utf-8") as f:
                    json.dump(map_data, f, ensure_ascii=False)
            except Exception:
                pass

    map_meta = map_data.get("map", {})

    # ★ 表示名を取得元に依存しない形へ正規化する。
    #   出版API「十和田 2005」/ ZFK「十和田地域の地質」→ どちらでも「十和田 2005」
    canonical = canonical_map_title(
        pub_title="" if title_ja == f"Map_{map_id}" else title_ja,
        zfk_title=map_meta.get("title_ja", ""),
        pub_year=map_meta.get("pub_year", ""),
    )
    if canonical:
        title_ja = canonical
    elif title_ja == f"Map_{map_id}":
        title_ja = map_meta.get("title_ja", title_ja)

    if not sheet_code:
        sheet_code = normalize_sheet_code(map_meta.get("sheet_code", ""))
    sheet_code = normalize_sheet_code(sheet_code)

    lat = lng = ""
    if "geom" in map_data and "centroid" in map_data.get("geom", {}):
        lat = map_data["geom"]["centroid"].get("lat", "")
        lng = map_data["geom"]["centroid"].get("lon", "")

    # --- 3. 地層リスト ---
    units_list = []
    u_resp = fetch_json(
        f"https://gbank.gsj.jp/ld/resource/zfk/query/unitsInMap?map_id={map_id}", "地層リスト"
    )
    if u_resp:
        units_list = u_resp.get("result", {}).get("units", [])
    if not units_list and os.path.isdir(os.path.join(local_dir, "units")):
        files = sorted(glob.glob(os.path.join(local_dir, "units", "*.json")))
        units_list = [{"id": os.path.basename(f)[:-5]} for f in files]

    # --- 4. Column の定義 ---
    base_slug = slugify_col_id(title_en_guess or "") or f"m{map_id}"
    if n_columns > 1:
        suffixes = ["west", "central", "east", "north", "south"][:n_columns]
        col_ids = [slugify_col_id(base_slug, s) for s in suffixes]
        col_names = [f"{title_en_guess or title_ja} District, {s}ern area" for s in suffixes]
    else:
        col_ids = [base_slug]
        col_names = [f"{title_en_guess or title_ja} District"]

    # --- 5. 出力先を決めてから参照ファイルを取る ---
    region_folder = get_region_folder(sheet_code)
    map_folder = f"m{map_id}_{safe_folder_name(title_ja)}".strip().rstrip("_")

    # ★ 先に既存フォルダを探す。作ってから探すと、使わない空フォルダが残ってしまう。
    #   図幅タイトルは出版APIとZFKで表記が違う（例: 「十和田 2005」と「十和田地域の地質」）。
    #   どちらが取れるかは通信状況で変わるため、フォルダ名だけで同一性を判断してはいけない。
    #   m{id}_ の一致だけで既存フォルダとみなし、必ず再利用する。
    review_root = os.path.join("data", "50k", "02_review")
    submission_root = os.path.join("data", "50k", "03_submission")
    existing = sorted(
        p for p in glob.glob(os.path.join(review_root, "**", f"m{map_id}_*"), recursive=True)
        if os.path.isdir(p)
    )
    if existing:
        output_dir = existing[0]
        rel = os.path.relpath(output_dir, review_root)
        submission_dir = os.path.join(submission_root, rel)
        print(f"  [notice] 既存フォルダを再利用します: {output_dir}")
        if len(existing) > 1:
            print(f"  [warn] m{map_id}_* のフォルダが {len(existing)} 個あります。整理を検討してください:")
            for p in existing:
                print(f"         - {p}")
    else:
        output_dir = os.path.join(review_root, region_folder, map_folder)
        submission_dir = os.path.join(submission_root, region_folder, map_folder)

    os.makedirs(output_dir, exist_ok=True)
    os.makedirs(submission_dir, exist_ok=True)

    # ★ ダウンロード前に上書き判定する（再実行時に無駄な通信をしないため）
    excel_path = os.path.join(output_dir, f"m{map_id}_review.xlsx")
    if os.path.exists(excel_path) and not force:
        print("\n[STOP] 既にレビューファイルが存在します:")
        print(f"       {excel_path}")
        print("       あなたの編集内容を失わないため、上書きせずに終了します。")
        print("")
        print("       ・内容を確認したい          -> python run.py check " + str(map_id))
        print("       ・提出ファイルを作りたい     -> python run.py export " + str(map_id))
        print("       ・新しい書式で作り直したい   -> python run.py make " + str(map_id) + " --force")
        print("         （--force でも .bak_日時.xlsx として自動バックアップします）")
        return excel_path

    download_assets(map_id, output_dir)

    # --- 5b. GSJ Shapefile のネイティブ属性 ---
    # geo_A.dbf は LLM を通さず、時代・地層名・岩相（和英）を読める。
    from shape_source import load_shape_units, match_status
    ref_dir = os.path.join(output_dir, "references")
    shape_data = load_shape_units(ref_dir)
    shape_units = shape_data.get("units", [])
    shape_by_code = {str(u["major_code"]): u for u in shape_units}
    shape_matched_codes = set()
    shape_conflicts = 0
    if shape_data.get("available"):
        print(f"  Shapefile geo_A.dbf: {len(shape_units)} 地質ユニット"
              f"（除外 {shape_data.get('excluded_records', 0)} レコード）")
        if (lat == "" or lng == "") and shape_data.get("centroid"):
            lat = shape_data["centroid"]["lat"]
            lng = shape_data["centroid"]["lng"]
            print("    Column 座標は geo_A.shp の図郭中心を使用します。")

    # ZFK が無い図幅では、空1行ではなく shape の凡例属性から全行を作る。
    if not units_list and shape_units:
        units_list = [{"id": "", "_shape_unit": unit} for unit in shape_units]
        print("    ZFK なし: shape 属性を units_review の一次ソースにします。")

    # --- 6. units 行の構築 ---
    age_map = load_json(os.path.join("config", "age_mapping.json"))
    lith_map = load_json(os.path.join("config", "lithology_mapping.json"))

    # ★ 図幅PDFのページ索引。層厚や記載が「PDFの何ページ」なのかを示すために使う。
    #   ZFK にはページ番号が入っていないので、手元のPDFと文字列照合して求める。
    #   1回7秒ほど。references/ にキャッシュされるので2回目以降は一瞬。
    import gsj_derived as G
    from pdf_locate import index_for
    print("  PDFのページ索引を作成中（出典ページの表示に使います）...")
    pdf_index = index_for(map_id, ref_dir)
    if pdf_index:
        print(f"    {len(pdf_index['pages'])} ページ")

    rows, full_docs, thick_notes, evidence_rows = [], [], [], []
    for i, u_item in enumerate(units_list, 1):
        unit_id = u_item.get("id") or ""
        u = fetch_unit_details(unit_id, local_dir) if unit_id else {}

        legend = u.get("legend", {})
        focus = legend.get("focus", {})
        parent_facies = legend.get("parent_facies", {})
        parent_age = legend.get("parent_age", {})
        target = u.get("target", {})

        zfk_lith_ja = focus.get("label_ja", "") or ""     # focus = 岩相
        zfk_lith_en = focus.get("label_en", "") or ""
        zfk_name_ja = parent_facies.get("label_ja", "") or ""   # parent_facies = 地層名
        zfk_name_en = parent_facies.get("label_en", "") or ""
        zfk_age_ja = parent_age.get("label_ja", "") or ""

        major_code = focus.get("major_code", "") or u.get("self", {}).get("major_code", "")
        try:
            major_code = str(int(float(major_code))) if major_code != "" else ""
        except (TypeError, ValueError):
            major_code = str(major_code).strip()
        shape_unit = u_item.get("_shape_unit") or shape_by_code.get(major_code)
        if shape_unit:
            shape_matched_codes.add(str(shape_unit["major_code"]))

        # 情報源の優先順位は ZFK > shape。shape は ZFK の空欄補完と競合検出にも使う。
        lith_ja = zfk_lith_ja or (shape_unit or {}).get("lithology_ja", "")
        lith_en = zfk_lith_en or (shape_unit or {}).get("lithology_en", "")
        name_ja = zfk_name_ja or (shape_unit or {}).get("unit_name_ja", "")
        name_en = zfk_name_en or (shape_unit or {}).get("unit_name_en", "")

        ref_name_ja = f"{name_ja} ({lith_ja})" if name_ja and lith_ja else (lith_ja or name_ja)
        ref_name_en = f"{name_en} ({lith_en})" if name_en and lith_en else (lith_en or name_en)

        # ★ unit_name = (lithology) を除いた地層名。
        #   parent_facies があればそれを直接使う（文字列パース不要で確実）。
        #   なければ REF 文字列から末尾括弧を除去してフォールバック。
        clean_unit_name = (name_en or strip_trailing_paren(ref_name_en)
                           or name_ja or strip_trailing_paren(ref_name_ja))

        desc = target.get("text", "") or ""
        age_text = zfk_age_ja or (shape_unit or {}).get("age_ja", "")
        mapped = (age_map.get(age_text, {})
                  or age_map.get((shape_unit or {}).get("age_text_ja", ""), {}))

        combined = f"{ref_name_ja} {lith_ja} {desc}"

        # --- GSJ が本文から抽出済みの構造化データ（derived）を読む ---
        if u:
            g_mn, g_mx = G.best_thickness(u)
            g_li = G.lithologies(u, pdf_index)
            g_bs = G.basal_surface(u, pdf_index)
            g_sn = G.strat_name(u)
        else:
            g_mn = g_mx = None
            g_li = g_bs = None
            g_sn = ""

        if shape_unit:
            shape_rel = os.path.relpath(shape_data["dbf_path"], output_dir)
            shape_locator = (f"{shape_rel} record {shape_unit['record_index']}"
                             f"; MAJOR_CODE={shape_unit['major_code']}")
            if u:
                shape_match, shape_conflict = match_status(shape_unit, {
                    "unit_name_ja": zfk_name_ja,
                    "lithology_ja": zfk_lith_ja,
                    "age_ja": zfk_age_ja,
                })
            else:
                shape_match, shape_conflict = "shape_only", ""
            if shape_conflict:
                shape_conflicts += 1
        else:
            shape_locator = ""
            shape_match, shape_conflict = "not_available", ""

        if not unit_id:
            unit_id = f"m{map_id}_s{str(shape_unit['major_code']).zfill(3)}"

        # 層厚は数値候補＋GSJ抽出値＋本文の全記述を、1文ごとの出典ページつきで入れる。
        # 場所ごとに値が違うことが多く（西部10m・東部25m）、判断には文脈とページが要る。

        row = {c: "" for c in REVIEW_UNIT_COLS}
        row.update({
            "REF_unit_name_en": ref_name_en,
            "REF_unit_name_ja": ref_name_ja,
            "REF_source": G.describe_source(u, pdf_index) if u else shape_locator,
            "REF_age_text": age_text,
            "REF_desc": desc,                       # ★ 全文（切らない）
            "REF_thickness": G.thickness_block(u, pdf_index),   # ★ 全文＋出典ページ
            "REF_lith_text": lith_ja,
            "REF_lith_candidates": extract_candidates(combined, lith_map),
            "REF_lithology_gsj": (g_li or {}).get("major", ""),
            "REF_minor_lith_gsj": (g_li or {}).get("minor", ""),
            "REF_strat_name": g_sn or "",
            "REF_basal_surface": (g_bs or {}).get("text", ""),
            "REF_shape_source": shape_locator,
            "REF_shape_match": shape_match,
            "REF_shape_unit_name": ((shape_unit or {}).get("display_name_en", "")
                                    or (shape_unit or {}).get("display_name_ja", "")),
            "REF_shape_age_text": ((shape_unit or {}).get("age_text_en", "")
                                   or (shape_unit or {}).get("age_text_ja", "")),
            "REF_shape_lith_text": ((shape_unit or {}).get("lithology_en", "")
                                    or (shape_unit or {}).get("lithology_ja", "")),
            "REF_confidence_class": "A" if (u or shape_unit) else "D",
            "REF_conflict": shape_conflict,
            "unit_id": unit_id,
            "column_id": col_ids[0],
            "sort_order": i,   # ZFK の並び順を初期値に。最上位=1 の想定で要確認。
            "unit_name": clean_unit_name,
            "t_int": mapped.get("t_int", ""),
            "b_int": mapped.get("b_int", ""),
            # --- 本文由来の値をそのまま編集列にも入れておく（自動入力） ---
            #     LLM を通していないので幻覚が入らない。出典は REF_ 列に残る。
            "strat_name": g_sn or "",
            "basal_surface": (g_bs or {}).get("value", ""),
            # ★ 層厚は「分布及び層厚」節から読む。GSJの derived は下位の1枚を
            #   指していることが多く、そのまま入れると誤った値が入る。
            "min_thickness": "" if g_mn is None else g_mn,
            "max_thickness": "" if g_mx is None else g_mx,
            "comments": ("shape の DBF レコード順を sort_order の初期候補に使用。"
                         "凡例の上下関係で要確認。" if shape_unit and not u else ""),
        })
        rows.append(row)
        full_docs.append({"unit_id": unit_id, "unit_name": clean_unit_name,
                          "unit_name_ja": name_ja or lith_ja, "text": desc})
        for note in extract_thickness_notes(desc, max_items=12):
            thick_notes.append({"unit_id": unit_id, "unit_name": clean_unit_name,
                                "thickness_note": note})

        # 候補値を長形式で残す。自動値の根拠と競合をセル単位で追跡できる。
        if u:
            zfk_locator = f"https://gbank.gsj.jp/ld/resource/zfk/units/{unit_id}.json"
            for field_name, value in (
                ("unit_name", ref_name_en or ref_name_ja),
                ("age_text", age_text),
                ("lithology", zfk_lith_en or zfk_lith_ja),
            ):
                if value:
                    evidence_rows.append({
                        "unit_id": unit_id, "field_name": field_name,
                        "candidate_value": value, "source_type": "ZFK",
                        "source_locator": zfk_locator, "confidence_class": "A",
                        "selected": "yes", "conflict": shape_conflict,
                    })
        if shape_unit:
            for field_name, value in (
                ("unit_name", shape_unit.get("display_name_en") or shape_unit.get("display_name_ja")),
                ("age_text", shape_unit.get("age_text_en") or shape_unit.get("age_text_ja")),
                ("lithology", shape_unit.get("lithology_en") or shape_unit.get("lithology_ja")),
                ("symbol", shape_unit.get("symbol")),
            ):
                if value:
                    evidence_rows.append({
                        "unit_id": unit_id, "field_name": field_name,
                        "candidate_value": value, "source_type": "Shapefile",
                        "source_locator": shape_locator, "confidence_class": "A",
                        "selected": "yes" if not u else "validation",
                        "conflict": shape_conflict,
                    })

    if not rows:
        print("  [notice] ZFK データなし。空テンプレート行を作成します。")
        blank = {c: "" for c in REVIEW_UNIT_COLS}
        blank.update({
            "REF_unit_name_en": "NO_DATA",
            "REF_unit_name_ja": "ZFKデータなし",
            "REF_desc": "自動入力データがありません。図幅PDFの地質総括図から手入力してください。",
            "unit_id": f"m{map_id}_u001",
            "column_id": col_ids[0],
            "sort_order": 1,
            "REF_confidence_class": "D",
            "REF_conflict": "ZFK/shape ともに利用不可",
        })
        rows.append(blank)

    # ★ t_pos の自動入力。
    #   公式仕様「Units that are unbounded at the top ... are dropped during
    #   ingestion」により、最上位の層に t_pos が無いと取り込み時に落ちる。
    #   sort_order → position（反転）→ t_pos の順で求める。
    #   column_id は「1, 2」と複数指せるので、必ず展開してから Column ごとに計算する。
    from common import auto_t_pos
    tps = auto_t_pos([r.get("column_id") for r in rows],
                     [r.get("sort_order") for r in rows])
    for k, tp in enumerate(tps):
        if tp != "" and is_blank(rows[k].get("t_pos")):
            rows[k]["t_pos"] = tp


    df_units = pd.DataFrame(rows, columns=REVIEW_UNIT_COLS)

    # Excel のセル上限（32,767字）だけ守る。長文列は切らずに全文入れる。
    for c in df_units.columns:
        df_units[c] = [truncate_for_cell(v, c) for v in df_units[c]]

    # --- 7. columns_review ---
    col_rows = []
    for cid, cname in zip(col_ids, col_names):
        r = {c: "" for c in REVIEW_COLUMN_COLS}
        r.update({
            "col_id": cid,
            "col_name": cname,
            "col_group": f"{title_en_guess or title_ja} District" if n_columns > 1 else "",
            "lat": lat,
            "lng": lng,
            "col_type": "column",
            "axis_type": "age",
            "comments": "図幅中心の座標を初期値としています。分割する場合は代表地点に修正してください。",
        })
        col_rows.append(r)
    df_cols = pd.DataFrame(col_rows, columns=REVIEW_COLUMN_COLS)


    # --- 7b. 英文Abstract（PDF巻末）を同時に取り込む ---
    abstract_text, abstract_rng, abstract_pdf = extract_abstract_for(map_id, output_dir)

    # --- 8. refs / images ---
    df_refs = build_refs_draft(map_meta, sheet_code, map_id)
    ref_id = df_refs.iloc[0]["ref_id"]
    df_cols["ref_ids"] = ref_id
    df_images = build_images_draft(output_dir, ref_id, col_ids)

    # --- 9. project_meta / gsj_meta ---
    meta_values = dict(METADATA_DEFAULTS)
    meta_values["project_name"] = title_en_guess or title_ja
    meta_values["compile_date"] = datetime.now().strftime("%Y-%m-%d")
    meta_values["compiler_name"] = compiler_name
    df_project_meta = pd.DataFrame(
        [{"key": k, "value": meta_values.get(k, "")} for k in METADATA_KEYS]
    )

    # --- intervals 参照表（t_prop/b_prop の自動計算に使う） ---
    # 全1715件ではなく、実際に使う193件に絞る（Excelの肥大化とノイズを避ける）
    iv = intervals_for_excel()
    df_intervals = pd.DataFrame(
        sorted(
            ({"interval": k, "b_age_ma": v["b_age"], "t_age_ma": v["t_age"],
              "int_type": v.get("int_type", "")} for k, v in iv.items()),
            key=lambda r: (-float(r["b_age_ma"]), r["interval"]),
        ),
        columns=["interval", "b_age_ma", "t_age_ma", "int_type"],
    )

    df_gsj_meta = pd.DataFrame([
        {"key": "map_id", "value": map_id},
        {"key": "title_ja", "value": title_ja},
        {"key": "title_en", "value": map_meta.get("title_en") or title_en_guess},
        {"key": "sheet_code", "value": sheet_code},
        {"key": "pub_year", "value": map_meta.get("pub_year", "")},
        {"key": "series", "value": map_meta.get("series", "")},
        {"key": "publisher", "value": map_meta.get("publisher", "")},
        {"key": "centroid_lat", "value": lat},
        {"key": "centroid_lng", "value": lng},
        {"key": "source_zfk_map", "value": f"https://gbank.gsj.jp/ld/resource/zfk/maps/m{map_id}.json"},
        {"key": "source_publication", "value": f"https://gbank.gsj.jp/ld/resource/publication/map/g050/map{map_id}.json"},
        {"key": "source_doc_page", "value": gsj_doc_url(sheet_code)},
        {"key": "abstract_pdf", "value": os.path.basename(abstract_pdf) if abstract_pdf else ""},
        {"key": "abstract_pages",
         "value": f"p.{abstract_rng[0]}-{abstract_rng[1]}" if abstract_rng else ""},
        {"key": "abstract_chars", "value": len(abstract_text)},
        {"key": "shape_available", "value": bool(shape_data.get("available"))},
        {"key": "shape_geo_a_dbf", "value": os.path.relpath(shape_data["dbf_path"], output_dir)
         if shape_data.get("dbf_path") else ""},
        {"key": "shape_unit_count", "value": len(shape_units)},
        {"key": "shape_matched_count", "value": len(shape_matched_codes)},
        {"key": "shape_unmatched_codes", "value": ", ".join(sorted(
            set(shape_by_code) - shape_matched_codes,
            key=lambda x: (0, int(x)) if str(x).isdigit() else (1, str(x))))},
        {"key": "shape_conflict_count", "value": shape_conflicts},
        {"key": "shape_bbox", "value": ", ".join(str(v) for v in (shape_data.get("bbox") or []))},
        {"key": "generated_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"key": "format_version", "value": FORMAT_VERSION},
    ])

    df_abstract = abstract_to_rows(abstract_text)

    # REF_desc はセルでは抜粋しか読めないので、全文はここに置く
    df_desc = pd.DataFrame(full_docs or [{"unit_id": "", "unit_name": "",
                                          "unit_name_ja": "", "text": ""}],
                           columns=["unit_id", "unit_name", "unit_name_ja", "text"])
    # 層厚は場所によって変わる。判断できるように文ごと一覧にする
    df_thick = pd.DataFrame(thick_notes or [{"unit_id": "", "unit_name": "",
                                             "thickness_note": "本文に層厚の記述が見つかりませんでした。"}],
                            columns=["unit_id", "unit_name", "thickness_note"])
    evidence_cols = ["unit_id", "field_name", "candidate_value", "source_type",
                     "source_locator", "confidence_class", "selected", "conflict"]
    df_evidence = pd.DataFrame(
        evidence_rows or [{c: "" for c in evidence_cols}], columns=evidence_cols)

    # --- 10. 書き出し（既存ファイルは絶対に消さない） ---
    if os.path.exists(excel_path) and force:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup = excel_path.replace(".xlsx", f".bak_{stamp}.xlsx")
        shutil.copy2(excel_path, backup)
        print(f"  [backup] 既存ファイルを退避: {os.path.basename(backup)}")

    instructions = pd.DataFrame([
        {"項目": "0. このファイルの役割",
         "説明": "GSJの構造化データ（ZFKを優先、次にShapefile）から自動取得した下書きです。ここを編集して `python run.py export <図幅名>` を実行すると、Macrostrat公式フォーマット(v0.1.1)の提出ファイルが 03_submission に作られます。"},
        {"項目": "1. 色の意味",
         "説明": "グレー(REF_*)=自動取得の参照専用。編集しても提出物には影響しません。 / 黄色=必ず確認・入力する列。 / 緑=そのまま提出物に転記される列。"},
        {"項目": "2. unit_name（重要）",
         "説明": "REF_unit_name_en は「地層名 (岩相)」の形ですが、unit_name 列には岩相を除いた地層名だけを自動で入れてあります。この列の値がそのまま提出ファイルの unit_name になります。誤りがあれば直接書き換えてください。"},
        {"項目": "3. sort_order（重要）",
         "説明": "最上位(いちばん新しい地層)を 1 とし、下に行くほど大きい数にします。提出時に Macrostrat の position（最下位=1、古い→新しい）へ自動で反転変換されます。同じ数字を複数行に付けると、重なり合うユニットとして扱われます。"},
        {"項目": "4. t_prop / b_prop（自動計算）",
         "説明": "水色の t_prop・b_prop は数式です。直接入力せず、隣の t_age_ma・b_age_ma に地層の年代を Ma 単位で入れてください（15.3ka なら 0.0153）。t_int・b_int で指定した時代の中での相対位置が自動で計算されます。0＝その時代の下端（古い側）、1＝上端（若い側）。数式を消して直接数値を入れても構いません。"},
        {"項目": "4b. abstract シート",
         "説明": "図幅PDF巻末の英文Abstractを自動で取り込んであります。地層の英語名・年代（例: the Shitazaki: 10.5–8.5 Ma）・上下関係・岩相がまとまっているので、t_age_ma / b_age_ma を埋めるときの一次資料として使ってください。references/m<ID>_abstract.txt にも同じものを保存しています。"},
        {"項目": "4c. area（Column）ごとに値が違うとき",
         "説明": "column_id が「1, 2」の行では、値もカンマで分けると Column ごとに割り当てられます。例: min_thickness に「10, 20」と書くと Column1 が10m、Column2 が20m。値が1つなら全Columnに同じ値が入ります。★ 個数が一致しないときは分解しません（lithology の「gravel, sand」を壊さないため）。1つのColumnの中で複数の岩相を書くときは「gravel; sand」とセミコロンを使ってください。"},
        {"項目": "4d. REF_desc / REF_thickness は全文が入っています",
         "説明": "この2列は途中で切らず、全文をセルに入れてあります。折り返しはオフにしてあるので、セルを選んで数式バー（画面上部の入力欄）を見るか、セルを広げて読んでください。REF_thickness には【GSJ抽出 ○m】に続けて本文中の層厚の記述をすべて並べてあります。層厚は場所で変わるので（「南半部では最大20m、北半部では最大10m」）、どの地点の話かを本文で確かめてから min/max を決めてください。同じ内容は descriptions / thickness_notes シートにも1行ずつ並べてあります。"},
        {"項目": "4f. REF_source（出典ページ）",
         "説明": "その地層が図幅説明書のどこに書いてあるかです。「§4.12 十和田段丘堆積物（Tw）｜ PDF p.59（印刷 p.51）｜ 小見出し: 地層名 / 模式地 / 分布及び層厚 / 層序関係 / 岩相 / 時代」の形。PDFページは手元のPDFと本文を文字列照合して求めています（ZFKにページ番号は入っていません）。REF_thickness や REF_lithology_gsj の各値にも、それぞれの該当ページを付けてあります。"},
        {"項目": "4h. Shapefile と確信度",
         "説明": "REF_shape_* は geo_A.dbf のネイティブ属性です。ZFKがある場合は MAJOR_CODE で照合し、ZFKを採用、shapeを検証用に残します。ZFKが無い場合はshapeを初期値にします。REF_confidence_class=A はGSJ構造化原値、Bは規則抽出、CはLLM+原文照合、Dは推定です。REF_conflict が空でない行だけ人が確認してください。source_evidence シートにはセル単位の候補・出典・採否を長形式で保存しています。"},
        {"項目": "4e. section_id / t_pos は自動計算",
         "説明": "section_id は export のときに自動で入ります（年代に大きな断絶があるときだけ切り、微小なすき間では切りません）。t_pos は各Columnの最上位の層に必ず入れます。公式仕様に「上端が決まらない層は取り込み時に落とされる」とあり、最上位の層は上に隣接する層が無いので明示が要るためです（一戸完成形も同じ入れ方をしています）。手入力があればそちらを優先します。"},
        {"項目": "4g. 噴火イベントの t_prop / b_prop",
         "説明": "火砕流・テフラ・溶岩など、年代が1点で決まる瞬間的な堆積は、本来 b_prop と t_prop が同じ値になります。しかし公式仕様は b_prop < t_prop を要求するため、表示桁（小数第3位）に四捨五入すると同じ値になる範囲を上下端にしています。例: 割合が 0.132 なら b_prop=0.1315 / t_prop=0.13249。どちらも 0.132 と表示されます。判定は「年代が1点」かつ「地層名に火砕流・テフラ・溶岩などが入っている」の両方を満たすときだけです。"},
        {"項目": "5. 岩相は2つの情報源を並べてあります",
         "説明": "REF_lithology_gsj / REF_minor_lith_gsj は【本文】由来。GSJ自身が日本語本文から抽出した構造化データで、日本語の語・該当文・出典ページが付いています。LLMを通していないので幻覚がありません。REF_lithology / REF_minor_lith は【要約】由来で、英文Abstractから拾ったものです。最初からMacrostrat向けの英語ですが、要約なので粗くなります。この2系統を見比べて、lithology / minor_lith にどちらを採るか決めてください（ここは自動入力しません）。GSJ側の主／副の振り分けは確信度による機械的な仮置きにすぎません。"},
        {"項目": "6. column_id",
         "説明": "columns_review の col_id を書きます（例: ichinohe-west）。「ichinohe-west, ichinohe-east」のようにカンマ区切りで書くと、提出時に自動で両方のColumnに複製されます。"},
        {"項目": "7. Columnの分割",
         "説明": "地域を分けたいときは columns_review に行を追加し、col_id をスラッグ形式（例: ichinohe-central）で付けます。geom または lat/lng のどちらかは必須です。"},
        {"項目": "7b. `python run.py llm <図幅名>` で埋まるもの",
         "説明": "英文Abstractを読ませて REF_age_from_abstract / REF_lithology / REF_minor_lith / REF_environment / REF_basal_surface / REF_unit_description に候補を入れ、さらに t_age_ma / b_age_ma / environment / basal_surface / unit_description / min_thickness / max_thickness にも自動入力します（t_prop / b_prop は年代から自動で計算されます）。すべて原文の引用つきで、引用が原文に無い候補・引用に数値が無い候補は自動で捨てられます。environment と lithology は config/vocab.json（Macrostrat公式APIの語彙表）から選ばせています。★ 上書きも行いますが、書き込み前に必ずバックアップを作り、変わった値を画面に一覧表示します。LLMの出力は実行ごとに変わるので、消えた値がないか確認してください。上書きしたくないときは --keep を付けてください。"},
        {"項目": "8. refs_review / images_review",
         "説明": "ZFKの書誌情報と references/ 内の凡例・断面画像から下書きを作ってあります。title が空の場合は英語タイトルを、description は図版の中身を手動で追記してください。"},
        {"項目": "9. project_meta",
         "説明": "compiler_name（あなたの氏名）と b_int / t_int（図幅全体の最古・最新の時代）を入力してください。"},
        {"項目": "10. 事前チェック",
         "説明": "`python run.py check <図幅名>` で、未入力や矛盾を出力前に一覧できます。"},
        {"項目": "11. 出典の原則",
         "説明": "推測で値を埋めないこと。不明なものは空欄のままにし、根拠は comments 列に出典・ページ・図表番号を書いてください。"},
    ])

    try:
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            instructions.to_excel(writer, sheet_name="Instructions", index=False)
            df_units.to_excel(writer, sheet_name="units_review", index=False)
            df_cols.to_excel(writer, sheet_name="columns_review", index=False)
            df_refs.to_excel(writer, sheet_name="refs_review", index=False)
            df_images.to_excel(writer, sheet_name="images_review", index=False)
            df_project_meta.to_excel(writer, sheet_name="project_meta", index=False)
            df_gsj_meta.to_excel(writer, sheet_name="gsj_meta", index=False)
            df_abstract.to_excel(writer, sheet_name="abstract", index=False)
            df_desc.to_excel(writer, sheet_name="descriptions", index=False)
            df_thick.to_excel(writer, sheet_name="thickness_notes", index=False)
            df_evidence.to_excel(writer, sheet_name="source_evidence", index=False)
            df_intervals.to_excel(writer, sheet_name="intervals", index=False)
            _write_prop_formulas(writer, df_units)
            _format_workbook(writer, df_units, df_cols, df_refs, df_images,
                             df_project_meta, df_gsj_meta, instructions, df_intervals,
                              df_abstract, df_desc, df_thick, df_evidence)
        print(f"\n完了: {excel_path}")
        print(f"      units {len(df_units)} 行 / columns {len(df_cols)} 件 / "
              f"images {len(df_images)} 件 / abstract {len(df_abstract)} 段落 / "
              f"層厚の記述 {len(df_thick)} 件")
    except PermissionError:
        print(f"\n[ERROR] 保存できません: {excel_path}")
        print("        Excel でこのファイルを開いていませんか？ 閉じてから再実行してください。")
        sys.exit(1)

    return excel_path


def write_prop_formulas(ws, columns, n_rows):
    """
    t_prop / b_prop に「年代(Ma)を入れると割合が自動で出る」数式を書き込む。

        prop = (interval の b_age − 地層の年代) / (interval の b_age − interval の t_age)

    interval の Ma は intervals シートから VLOOKUP で引くので、
    t_int / b_int を書き換えれば分母も自動で追従する。
    年代欄が空なら空欄のままにする（勝手に値を作らない）。

    ★ 列を挿入すると openpyxl は数式の参照を書き換えないため、
      列構成を変えたら必ずこの関数で数式を作り直すこと（repair_layout が呼ぶ）。
    """
    from openpyxl.utils import get_column_letter as L

    cols = {name: i for i, name in enumerate(columns, 1)}
    written = 0
    for prop_col, int_col, age_col in (("t_prop", "t_int", "t_age_ma"),
                                       ("b_prop", "b_int", "b_age_ma")):
        if not all(c in cols for c in (prop_col, int_col, age_col)):
            continue
        ci, ca, cp = L(cols[int_col]), L(cols[age_col]), cols[prop_col]
        for r in range(2, n_rows + 2):
            iv, age = f"${ci}{r}", f"${ca}{r}"
            b = f"VLOOKUP({iv},intervals!$A:$C,2,FALSE)"
            t = f"VLOOKUP({iv},intervals!$A:$C,3,FALSE)"
            ws.cell(row=r, column=cp).value = (
                f'=IF(OR({iv}="",{age}=""),"",IFERROR(({b}-{age})/({b}-{t}),""))'
            )
            ws.cell(row=r, column=cp).number_format = PROP_NUMBER_FORMAT
            written += 1
    return written


def _write_prop_formulas(writer, df_units):
    write_prop_formulas(writer.sheets["units_review"], list(df_units.columns), len(df_units))


def _format_workbook(writer, df_units, df_cols, df_refs, df_images, df_meta, df_gsj,
                     df_inst, df_intervals=None, df_abstract=None,
                     df_desc=None, df_thick=None, df_evidence=None):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    fill_ref = PatternFill("solid", start_color="F2F2F2")        # グレー: 参照専用
    fill_input = PatternFill("solid", start_color="FFF2CC")      # 黄色: 要入力
    fill_out = PatternFill("solid", start_color="D9EAD3")        # 緑: 提出物へ転記
    bold = Font(bold=True)
    wrap = Alignment(wrapText=True, vertical="top")
    nowrap = Alignment(wrapText=False, vertical="top")

    ws = writer.sheets["Instructions"]
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 105
    ws.freeze_panes = "A2"
    for r in range(1, len(df_inst) + 2):
        ws.cell(row=r, column=1).alignment = wrap
        ws.cell(row=r, column=2).alignment = wrap
    ws.cell(row=1, column=1).font = bold
    ws.cell(row=1, column=2).font = bold

    # --- units_review ---
    ws = writer.sheets["units_review"]
    # ★ 固定するのは見出し行だけ（列は固定しない）。
    #   列を固定すると、幅の広い REF_* 列が画面を占有して右へスクロールできなくなる。
    #   実際 G2（A〜F固定）で横スクロール不能、B2（A列固定）でも
    #   「なぜかA列だけ動かない」と混乱の元になった。列固定は入れないこと。
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions   # 絞り込み・並べ替えを使えるように
    required = {"unit_name", "sort_order", "column_id", "lithology", "t_int", "b_int"}
    fill_calc = PatternFill("solid", start_color="DEEAF6")   # 水色: 自動計算（数式）
    for i, name in enumerate(df_units.columns, 1):
        letter = get_column_letter(i)
        cell = ws.cell(row=1, column=i)
        cell.font = bold
        if name.startswith("REF_"):
            cell.fill = fill_ref
        elif name in ("t_prop", "b_prop"):
            cell.fill = fill_calc
        elif name in required or name in ("t_age_ma", "b_age_ma"):
            cell.fill = fill_input
        else:
            cell.fill = fill_out
        ws.column_dimensions[letter].width = column_width(name)
        # ★ 長文列は折り返さない。全文を入れているので折り返すと1行が数百行になり、
        #   かえって読めなくなる。1行表示にしておけば、セルを選べば数式バーで全文が読める。
        align = nowrap if name in NO_WRAP_COLS else wrap
        for r in range(2, len(df_units) + 2):
            ws.cell(row=r, column=i).alignment = align

    # --- 残りのシート ---
    for sheet, df in [
        ("columns_review", df_cols),
        ("refs_review", df_refs),
        ("images_review", df_images),
        ("project_meta", df_meta),
        ("gsj_meta", df_gsj),
        ("intervals", df_intervals),
        ("abstract", df_abstract),
        ("descriptions", df_desc),
        ("thickness_notes", df_thick),
        ("source_evidence", df_evidence),
    ]:
        if df is None:
            continue
        w = writer.sheets[sheet]
        w.freeze_panes = "A2"
        fill = (fill_ref if sheet in ("gsj_meta", "intervals", "abstract",
                               "descriptions", "thickness_notes", "source_evidence")
            else fill_input)
        for i, name in enumerate(df.columns, 1):
            c = w.cell(row=1, column=i)
            c.fill = fill
            c.font = bold
            w.column_dimensions[get_column_letter(i)].width = column_width(name)
            for r in range(2, len(df) + 2):
                w.cell(row=r, column=i).alignment = wrap


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="GSJ図幅からレビュー用Excelを生成")
    p.add_argument("map_id", type=str, help="ZFK map ID (例: 1050)")
    p.add_argument("--force", action="store_true", help="既存レビューファイルを上書き（自動バックアップあり）")
    p.add_argument("--columns", type=int, default=1, help="初期Column数（例: 3 で west/central/east を用意）")
    p.add_argument("--compiler", type=str, default="", help="compiler_name に入れる氏名")
    a = p.parse_args()
    build_review_excel(a.map_id, force=a.force, n_columns=a.columns, compiler_name=a.compiler)
