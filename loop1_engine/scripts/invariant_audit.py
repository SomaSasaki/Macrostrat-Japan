# -*- coding: utf-8 -*-
"""レビュー簿に 5 大不変条件を機械的に当て、未解決ユニットを根拠つきで並べる。

    python run.py audit ichinohe          一戸を監査して要約を表示
    python run.py audit ichinohe --json   監査結果を system/audit/ に保存
    python run.py audit --all             02_review 配下すべてを監査

検査する不変条件（specs/MEMORY.md）:

    1. 年代の単調性        b_age >= t_age >= 0.0 Ma。区間名から引いた年代とも突き合わせる。
    2. 証拠の保持          年代・環境・岩相を埋めたなら、Evidence シートに原文が要る。
    3. 公式統制語彙        lithology / environment は config/vocab.json と照合する。
    4. 識別子の不変性      unit_id を system/unit_id_registry.json に記録し、変化を検出する。
    5. 1 Formation = 1 Row 同一 formation_key が複数行に割れていないか見る。

LLM も外部ネットワークも使わない。判定はすべてローカルの表と設定ファイルで完結する。
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
REVIEW_ROOT = (ROOT / "loop2_governance" / "data" / "50k" / "02_review") if (ROOT / "loop2_governance" / "data" / "50k" / "02_review").is_dir() else (ROOT / "data" / "50k" / "02_review")
INTERVALS = (ROOT / "loop2_governance" / "config" / "intervals.json") if (ROOT / "loop2_governance" / "config" / "intervals.json").is_file() else (ROOT / "config" / "intervals.json")
VOCAB = (ROOT / "loop2_governance" / "config" / "vocab.json") if (ROOT / "loop2_governance" / "config" / "vocab.json").is_file() else (ROOT / "config" / "vocab.json")

REVIEW_SHEETS = ("Review", "units_review", "units")
# 空でも違反にはしないが、埋まっていないと図幅を締められないフィールド。
REQUIRED_FOR_CLOSE = ("unit_name", "lithology", "t_int", "b_int", "environment", "b_prop")
SEVERITY_ORDER = {"error": 0, "warning": 1, "info": 2}

# 貫入岩体（深成岩体・岩脈など）は堆積してできた地層ではないため、堆積環境
# （environment）という属性がそもそも存在しない。necessary な欄として扱うと、
# 埋めようのない欄を埋めるために原典に無い語を入れる圧力が生じる。
# そこで貫入岩体に限り environment を必須項目から外す。
# 判定は「岩体名に貫入岩を示す語がある」か「主岩相がすべて深成岩である」場合のみ。
# 本文の記述（例：「一戸深成岩体に貫入されている」）で誤判定しないよう、
# unit_description は判定に使わない。
_INTRUSIVE_NAME = re.compile(
    r"pluton|intrusion|intrusive|batholith|laccolith|stock\b|dike|dyke|sill\b|深成岩体|貫入岩体",
    re.IGNORECASE)
_INTRUSIVE_LITHOLOGIES = frozenset({
    "granite", "granodiorite", "gabbro", "diorite", "quartz diorite",
    "quartz monzonite", "monzonite", "monzodiorite", "monzogabbro", "tonalite",
    "syenite", "norite", "anorthosite", "peridotite", "pyroxenite",
    "dolerite", "diabase", "aplite", "pegmatite", "granophyre",
})


def is_intrusive(unit: dict) -> bool:
    """貫入岩体（深成岩体・岩脈など）かどうか。"""
    name = f"{_clean(unit.get('unit_name'))} {_clean(unit.get('strat_name'))}"
    if _INTRUSIVE_NAME.search(name):
        return True
    liths = {t.casefold() for t in split_terms(unit.get("lithology"))}
    return bool(liths) and liths <= _INTRUSIVE_LITHOLOGIES


def required_for_close(unit: dict) -> tuple[str, ...]:
    """そのユニットを締めるために埋まっている必要のある項目。"""
    if is_intrusive(unit):
        return tuple(f for f in REQUIRED_FOR_CLOSE if f != "environment")
    return REQUIRED_FOR_CLOSE


def _load_json(path: Path, default: Any = None) -> Any:
    try:
        with path.open(encoding="utf-8") as handle:
            return json.load(handle)
    except (OSError, json.JSONDecodeError):
        return default


def _clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _number(value: Any) -> float | None:
    text = _clean(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", text)
        return float(match.group(0)) if match else None


def split_terms(value: Any) -> list[str]:
    """';' 区切りの語彙値を項目に割る（config/vocab.json の _区切り に従う）。"""
    return [part.strip() for part in re.split(r"[;；]", _clean(value)) if part.strip()]


# ---------------------------------------------------------------------------
# 読み込み
# ---------------------------------------------------------------------------

def load_review(path: Path) -> dict[str, Any]:
    """レビュー簿を、行の辞書と Evidence の索引に開く。"""
    from openpyxl import load_workbook

    book = load_workbook(path, read_only=True, data_only=True)
    try:
        name = next((n for n in REVIEW_SHEETS if n in book.sheetnames), "")
        if not name:
            raise ValueError(f"層シートがありません: {book.sheetnames}")
        rows = list(book[name].iter_rows(values_only=True))
        header = [_clean(cell) for cell in rows[0]] if rows else []
        units = []
        for raw in rows[1:]:
            if not any(v not in (None, "") for v in raw):
                continue
            units.append({header[i]: raw[i] for i in range(min(len(header), len(raw)))})

        evidence: dict[tuple[str, str], list[dict[str, str]]] = {}
        if "Evidence" in book.sheetnames:
            erows = list(book["Evidence"].iter_rows(values_only=True))
            eheader = [_clean(c) for c in erows[0]] if erows else []
            for raw in erows[1:]:
                if not any(v not in (None, "") for v in raw):
                    continue
                record = {eheader[i]: _clean(raw[i]) for i in range(min(len(eheader), len(raw)))}
                key = (record.get("unit_id", ""), record.get("field", ""))
                evidence.setdefault(key, []).append(record)

        columns = []
        if "Columns" in book.sheetnames:
            crows = list(book["Columns"].iter_rows(values_only=True))
            cheader = [_clean(c) for c in crows[0]] if crows else []
            for raw in crows[1:]:
                if not any(v not in (None, "") for v in raw):
                    continue
                columns.append({cheader[i]: _clean(raw[i]) for i in range(min(len(cheader), len(raw)))})
    finally:
        book.close()
    return {"units": units, "evidence": evidence, "columns": columns, "header": header}


# ---------------------------------------------------------------------------
# 不変条件
# ---------------------------------------------------------------------------

def check_monotonicity(units: list[dict], intervals: dict) -> list[dict]:
    findings = []
    for unit in units:
        uid = _clean(unit.get("unit_id"))
        b_age, t_age = _number(unit.get("b_age_ma")), _number(unit.get("t_age_ma"))
        if b_age is not None and t_age is not None and b_age < t_age:
            findings.append(_finding("error", "monotonicity", uid, "b_age_ma",
                                     f"b_age {b_age} < t_age {t_age}"))
        for field, value in (("b_age_ma", b_age), ("t_age_ma", t_age)):
            if value is not None and value < 0:
                findings.append(_finding("error", "monotonicity", uid, field,
                                         f"{field} が負の値 {value}"))

        # 区間名から引いた年代とも突き合わせる。表記ゆれは警告に留める。
        b_int, t_int = _clean(unit.get("b_int")), _clean(unit.get("t_int"))
        b_def, t_def = intervals.get(b_int), intervals.get(t_int)
        if b_int and b_def is None:
            findings.append(_finding("warning", "monotonicity", uid, "b_int",
                                     f"config/intervals.json に無い区間名: {b_int}"))
        if t_int and t_def is None:
            findings.append(_finding("warning", "monotonicity", uid, "t_int",
                                     f"config/intervals.json に無い区間名: {t_int}"))
        if b_def and t_def and _number(b_def.get("b_age")) is not None:
            if _number(b_def["b_age"]) < _number(t_def.get("t_age") or 0):
                findings.append(_finding("error", "monotonicity", uid, "b_int",
                                         f"区間 {b_int}({b_def['b_age']}Ma) が "
                                         f"{t_int}({t_def.get('t_age')}Ma) より新しい"))
        if b_def and b_age is not None and not (
                _number(b_def.get("t_age")) - 0.51 <= b_age <= _number(b_def.get("b_age")) + 0.51):
            findings.append(_finding("warning", "monotonicity", uid, "b_age_ma",
                                     f"b_age {b_age} が区間 {b_int} "
                                     f"[{b_def.get('t_age')}, {b_def.get('b_age')}] の外"))
    return findings


def check_evidence(units: list[dict], evidence: dict) -> list[dict]:
    """数値や語彙を埋めたなら、Evidence 側に原文が残っているはず。"""
    watched = ("b_age_ma", "t_age_ma", "b_int", "t_int", "environment", "lithology")
    findings = []
    for unit in units:
        uid = _clean(unit.get("unit_id"))
        inline = _clean(unit.get("age_evidence")) + _clean(unit.get("context_evidence")) \
            + _clean(unit.get("physical_evidence"))
        for field in watched:
            if not _clean(unit.get(field)):
                continue
            rows = evidence.get((uid, field)) or []
            has_quote = any(len(_clean(r.get("source_and_full_context"))) >= 12 for r in rows)
            if not has_quote and not inline:
                findings.append(_finding("error", "evidence", uid, field,
                                         "値が入っているのに Evidence にも本文にも根拠が無い"))
            elif not has_quote:
                findings.append(_finding("warning", "evidence", uid, field,
                                         "Evidence シートに専用の原文行が無い（本文欄のみ）"))
    return findings


def check_vocabulary(units: list[dict], vocab: dict) -> list[dict]:
    lithologies = {t.strip().casefold() for t in (vocab.get("lithology") or [])}
    environments = {t.strip().casefold() for t in (vocab.get("environment") or [])}
    findings = []
    for unit in units:
        uid = _clean(unit.get("unit_id"))
        for field, allowed in (("lithology", lithologies), ("minor_lith", lithologies),
                               ("environment", environments)):
            if not allowed:
                continue
            for term in split_terms(unit.get(field)):
                if term.casefold() not in allowed:
                    findings.append(_finding("warning", "vocabulary", uid, field,
                                             f"公式語彙に無い語: {term}"))
    return findings


def check_unit_ids(units: list[dict], registry_path: Path, write: bool) -> list[dict]:
    """unit_id は恒久。過去の台帳と突き合わせ、消えた ID・増えた ID を報告する。"""
    current = [_clean(u.get("unit_id")) for u in units if _clean(u.get("unit_id"))]
    findings = []
    duplicates = {uid for uid in current if current.count(uid) > 1}
    for uid in sorted(duplicates):
        findings.append(_finding("error", "unit_id", uid, "unit_id", "同じ unit_id が複数行にある"))

    previous = _load_json(registry_path, {}) or {}
    known = list(previous.get("unit_ids") or [])
    if known:
        for uid in sorted(set(known) - set(current)):
            findings.append(_finding("error", "unit_id", uid, "unit_id",
                                     f"台帳にあった unit_id が今回の簿から消えている"
                                     f"（台帳 {previous.get('updated_at', '')}）"))
        for uid in sorted(set(current) - set(known)):
            findings.append(_finding("info", "unit_id", uid, "unit_id", "新しく増えた unit_id"))
    else:
        findings.append(_finding("info", "unit_id", "", "unit_id",
                                 "unit_id 台帳が無いため、今回の一覧を初期台帳として扱う"))
    if write:
        registry_path.parent.mkdir(parents=True, exist_ok=True)
        merged = sorted(set(known) | set(current))
        with registry_path.open("w", encoding="utf-8") as handle:
            json.dump({"updated_at": _now(), "unit_ids": merged,
                       "current_count": len(current)}, handle, ensure_ascii=False, indent=1)
    return findings


def check_one_row_per_formation(units: list[dict]) -> list[dict]:
    """1 Formation = 1 Row。側方変化は column_id 側で表し、行を割らない。"""
    seen: dict[str, list[str]] = {}
    for unit in units:
        key = _clean(unit.get("formation_key")) or _clean(unit.get("strat_name")) \
            or _clean(unit.get("unit_name"))
        if not key:
            continue
        seen.setdefault(key.casefold(), []).append(_clean(unit.get("unit_id")))
    findings = []
    for key, ids in sorted(seen.items()):
        if len(ids) > 1:
            findings.append(_finding("warning", "one_row_per_formation", ", ".join(ids),
                                     "strat_name", f"同じ層名が {len(ids)} 行に割れている: {key}"))
    return findings


def _finding(severity: str, rule: str, unit_id: str, field: str, message: str) -> dict:
    return {"severity": severity, "rule": rule, "unit_id": unit_id,
            "field": field, "message": message}


def _now() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


# ---------------------------------------------------------------------------
# 未解決ユニットの棚卸し
# ---------------------------------------------------------------------------

def collect_unresolved(units: list[dict], evidence: dict) -> list[dict]:
    """埋まっていない必須フィールドを、Evidence の候補と一緒に並べる。

    次に人（または LLM 段）が読むべき原文がその場に出るので、
    PDF を頭から読み直さずに解決作業へ入れる。
    """
    out = []
    for unit in units:
        uid = _clean(unit.get("unit_id"))
        missing = [f for f in required_for_close(unit) if not _clean(unit.get(f))]
        if not missing:
            continue
        hints: dict[str, list[dict[str, str]]] = {}
        for field in missing:
            rows = evidence.get((uid, field)) or []
            if rows:
                hints[field] = [{
                    "candidate": r.get("candidate", ""),
                    "flag": r.get("flag", ""),
                    "quote": r.get("source_and_full_context", "")[:400],
                } for r in rows[:3]]
        out.append({
            "unit_id": uid,
            "unit_name": _clean(unit.get("unit_name")),
            "strat_name": _clean(unit.get("strat_name")),
            "column_id": _clean(unit.get("column_id")),
            "missing": missing,
            "evidence_hints": hints,
            "description": _clean(unit.get("unit_description"))[:240],
        })
    return out


# ---------------------------------------------------------------------------
# 図幅単位の実行
# ---------------------------------------------------------------------------

def find_workbook(folder: Path) -> Path | None:
    def usable(path: Path) -> bool:
        return path.is_file() and not path.name.startswith("~$") and ".bak_" not in path.name

    primary = [p for p in folder.glob("m*_review.xlsx") if usable(p)]
    candidates = [p for p in folder.glob("m*_review.candidate-*.xlsx") if usable(p)]
    newest = max(candidates, key=lambda p: p.stat().st_mtime) if candidates else None
    if primary:
        main = max(primary, key=lambda p: p.stat().st_mtime)
        if newest and newest.stat().st_size > main.stat().st_size * 2:
            return newest
        return main
    return newest


def _map_ids_by_name() -> dict[str, str]:
    """図幅名（和英）から map_id を引く表。config/gsj_50k_grid.json から作る。"""
    grid = _load_json(ROOT / "config" / "gsj_50k_grid.json", {}) or {}
    table: dict[str, str] = {}
    for sheet in grid.get("sheets") or []:
        map_id = str(sheet.get("latest_map_id") or "")
        if not map_id:
            continue
        for raw in (sheet.get("title_ja"), sheet.get("title_en")):
            match = re.search(r"[「']([^」']+)[」']", str(raw or ""))
            name = (match.group(1) if match else "").casefold().strip()
            if name:
                table.setdefault(name, map_id)
        table.setdefault(str(sheet.get("sheet_code") or ""), map_id)
    return table


def iter_workspaces(selector: str = "") -> list[Path]:
    """図幅名（和英）・map_id・図幅コード・フォルダ名の部分一致で選ぶ。"""
    if not REVIEW_ROOT.is_dir():
        return []
    folders = [p for region in sorted(REVIEW_ROOT.iterdir()) if region.is_dir()
               for p in sorted(region.iterdir()) if p.is_dir() and re.match(r"m\d+", p.name)]
    if not selector:
        return folders
    needle = selector.casefold().strip()
    hits = [p for p in folders
            if needle in p.name.casefold()
            or re.match(rf"m{re.escape(needle.lstrip('m'))}(?:[_\s]|$)", p.name)]
    if hits:
        return hits
    map_id = _map_ids_by_name().get(needle)
    if not map_id:
        return []
    return [p for p in folders if re.match(rf"m{map_id}(?:[_\s]|$)", p.name)]


def audit_workspace(folder: Path, write_registry: bool = True) -> dict[str, Any]:
    workbook = find_workbook(folder)
    if workbook is None:
        return {"folder": str(folder.relative_to(ROOT)).replace(os.sep, "/"),
                "error": "レビュー簿が見つかりません", "findings": [], "unresolved": []}
    try:
        loaded = load_review(workbook)
    except Exception as exc:
        return {"folder": str(folder.relative_to(ROOT)).replace(os.sep, "/"),
                "workbook": workbook.name, "error": f"読込エラー: {exc}",
                "findings": [], "unresolved": []}

    intervals = _load_json(INTERVALS, {}) or {}
    vocab = _load_json(VOCAB, {}) or {}
    units, evidence = loaded["units"], loaded["evidence"]
    registry = folder / "system" / "unit_id_registry.json"

    findings = (check_monotonicity(units, intervals)
                + check_evidence(units, evidence)
                + check_vocabulary(units, vocab)
                + check_unit_ids(units, registry, write_registry)
                + check_one_row_per_formation(units))
    findings.sort(key=lambda f: (SEVERITY_ORDER.get(f["severity"], 9), f["rule"], f["unit_id"]))

    counts = {"error": 0, "warning": 0, "info": 0}
    for finding in findings:
        counts[finding["severity"]] = counts.get(finding["severity"], 0) + 1

    filled = {field: sum(1 for u in units if _clean(u.get(field))) for field in REQUIRED_FOR_CLOSE
              if field in loaded["header"]}
    return {
        "folder": str(folder.relative_to(ROOT)).replace(os.sep, "/"),
        "workbook": workbook.name,
        "audited_at": _now(),
        "units": len(units),
        "columns": [c.get("col_name") or c.get("col_id") for c in loaded["columns"]],
        "filled": filled,
        "counts": counts,
        "findings": findings,
        "unresolved": collect_unresolved(units, evidence),
        "invariants_pass": counts["error"] == 0,
    }


def write_report(report: dict[str, Any]) -> Path:
    folder = ROOT / report["folder"]
    out_dir = folder / "system" / "audit"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    path = out_dir / f"audit-{stamp}.json"
    with path.open("w", encoding="utf-8") as handle:
        json.dump(report, handle, ensure_ascii=False, indent=1)
    latest = out_dir / "audit-latest.json"
    with latest.open("w", encoding="utf-8") as handle:
        json.dump(report, handle, ensure_ascii=False, indent=1)
    return path


def print_report(report: dict[str, Any], limit: int = 25) -> None:
    print(f"\n--- {report['folder']} ---")
    if report.get("error"):
        print(f"  {report['error']}")
        return
    counts = report["counts"]
    print(f"  簿 {report['workbook']} / {report['units']} 層 / "
          f"カラム {', '.join(c for c in report['columns'] if c) or '—'}")
    print(f"  不変条件: {'PASS' if report['invariants_pass'] else 'FAIL'} "
          f"(error {counts['error']} / warning {counts['warning']} / info {counts['info']})")
    for field, count in report["filled"].items():
        print(f"    {field:<12} {count}/{report['units']}")
    shown = 0
    for finding in report["findings"]:
        if finding["severity"] == "info":
            continue
        print(f"    [{finding['severity']:<7}] {finding['rule']:<22} "
              f"{finding['unit_id']:<14} {finding['field']:<12} {finding['message']}")
        shown += 1
        if shown >= limit:
            print(f"    … ほか {len([f for f in report['findings'] if f['severity'] != 'info']) - shown} 件")
            break
    print(f"  未解決 {len(report['unresolved'])} ユニット")
    for item in report["unresolved"][:limit]:
        hint = "; ".join(f"{k}: {v[0]['candidate']}" for k, v in item["evidence_hints"].items() if v)
        print(f"    {item['unit_id']:<14} {item['unit_name'][:34]:<36} "
              f"欠落 {','.join(item['missing'])}{'  | 候補 ' + hint if hint else ''}")
    if len(report["unresolved"]) > limit:
        print(f"    … ほか {len(report['unresolved']) - limit} ユニット")


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="レビュー簿の不変条件を監査する")
    parser.add_argument("target", nargs="*", help="図幅名・map_id・フォルダ名の一部")
    parser.add_argument("--all", action="store_true", help="02_review 配下すべて")
    parser.add_argument("--json", action="store_true", help="system/audit/ に保存")
    parser.add_argument("--limit", type=int, default=25)
    args = parser.parse_args(argv)

    selectors = args.target or ([""] if args.all else [])
    if not selectors:
        parser.error("図幅を指定するか --all を付けてください（例: run.py audit ichinohe）")

    folders: list[Path] = []
    for selector in selectors:
        hits = iter_workspaces(selector)
        if not hits:
            print(f"[skip] 該当するワークスペースがありません: {selector or '(all)'}")
        folders.extend(hits)
    folders = sorted(set(folders))
    if not folders:
        return 1

    failed = 0
    for folder in folders:
        report = audit_workspace(folder)
        print_report(report, args.limit)
        if args.json and not report.get("error"):
            print(f"  保存: {write_report(report).relative_to(ROOT)}")
        if report.get("error") or not report.get("invariants_pass"):
            failed += 1
    print(f"\n監査 {len(folders)} 図幅 / 不合格 {failed}")
    return 1 if failed else 0


if __name__ == "__main__":
    for _stream in (sys.stdout, sys.stderr):
        if hasattr(_stream, "reconfigure"):
            _stream.reconfigure(encoding="utf-8", errors="replace")
    raise SystemExit(main())
