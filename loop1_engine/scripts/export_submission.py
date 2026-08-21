# -*- coding: utf-8 -*-
"""
export_submission.py — レビュー用 Excel から Macrostrat 提出ファイルを生成する

出力形式は Macrostrat column-ingestion format v0.1.1。
Ichinohe_Composite_column.xlsx（完成形）と同じ 5 シート構成・同じ列順で出力する。

  metadata  key-value レイアウト（Documentation 行より下は取り込み時に無視される）
  units     unit_id, col_id, section_id, position, b_int, b_prop, t_int, t_prop,
            unit_name, strat_name, environment, unit_description, lithology,
            minor_lith, min_thickness, max_thickness, basal_surface,
            lateral_relationship, comments, t_pos
  columns   col_id, col_name, col_group, ref_ids, ...
  refs      ref_id, title, authors, ...
  images    col_ids, image_name, ref_id, ...

使い方:
  python scripts/export_submission.py <path/to/m1050_review.xlsx>
  python scripts/export_submission.py <path> --check-only
"""

import argparse
import os
import sys
from datetime import datetime

import pandas as pd

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common import (  # noqa: E402
    FORMAT_VERSION,
    METADATA_DEFAULTS,
    METADATA_KEYS,
    REQUIRED_CHRONO_FIELDS,
    OPTIONAL_REPORTED_FIELDS,
    REQUIRED_UNIT_FIELDS,
    SUBMISSION_COLUMN_COLS,
    SUBMISSION_IMAGE_COLS,
    SUBMISSION_REF_COLS,
    SUBMISSION_UNIT_COLS,
    age_from_prop,
    compute_prop,
    derive_sections,
    derive_t_pos,
    derive_positions,
    best_interval_for_age,
    fits_interval,
    interval_bounds,
    intervals_containing,
    is_blank,
    parse_age_ma,
    props_from_ages,
    COMMA_AMBIGUOUS_FIELDS,
    NEVER_SPLIT_FIELDS,
    PER_COLUMN_SPLIT_FIELDS,
    split_per_column,
    strip_trailing_paren,
)

DOC_NOTE = (
    "Note: everything below this row is dropped from the sheet on ingestion. "
    "This block records provenance for the compilation."
)

PROJECT_BLOCK_MARKERS = {
    "PROJECT_METADATA": "metadata",
    "REFERENCES": "refs",
    "IMAGES": "images",
}


# ---------------------------------------------------------------------------
# 読み込み
# ---------------------------------------------------------------------------

def read_sheet(path, name, fallback_names=()):
    for n in (name,) + tuple(fallback_names):
        try:
            df = pd.read_excel(path, sheet_name=n, dtype=object)
            df.columns = [str(c).strip() for c in df.columns]
            return df
        except Exception:
            continue
    return None


def has_sheet(path, name):
    """Return whether an exact worksheet name exists without changing the file."""
    try:
        with pd.ExcelFile(path) as workbook:
            return name in workbook.sheet_names
    except Exception:
        return False


def read_kv_sheet(path, name):
    """key-value レイアウトのシートを dict にする。
    'Documentation' 行より下は説明文なので読み飛ばす（公式テンプレートと同じ規約）。"""
    df = read_sheet(path, name)
    if df is None or df.empty:
        return {}
    cols = list(df.columns)
    kcol = "key" if "key" in cols else cols[0]
    vcol = "value" if "value" in cols else (cols[1] if len(cols) > 1 else cols[0])
    out = {}
    for _, r in df.iterrows():
        k = r.get(kcol)
        if is_blank(k):
            continue
        key = str(k).strip()
        if key.lower() == "documentation":
            break
        v = r.get(vcol)
        out[key] = "" if is_blank(v) else v
    return out


def read_project_blocks(path):
    """Review v2 の ``Project`` シートを読む。

    Project は人間が見るシート数を減らすため、次の3ブロックを
    縦に並べる。各 marker の次の非空行が header で、次の marker
    までが data である。

      PROJECT_METADATA -> key, value
      REFERENCES       -> SUBMISSION_REF_COLS
      IMAGES           -> SUBMISSION_IMAGE_COLS

    ``gsj_`` で始まる metadata key は GSJ provenance として分離する。
    ブロックがない、または壊れている場合は空値を返し、旧形式の
    project_meta / refs_review / images_review への fallback を妨げない。
    """
    try:
        raw = pd.read_excel(path, sheet_name="Project", header=None, dtype=object)
    except Exception:
        return {}, {}, None, None

    def text(v):
        return "" if is_blank(v) else str(v).strip()

    markers = []
    for row_idx, row in raw.iterrows():
        values = [text(v) for v in row.tolist()]
        first = next((v for v in values if v), "")
        marker = first.upper()
        if marker in PROJECT_BLOCK_MARKERS:
            markers.append((row_idx, marker))

    blocks = {}
    for pos, (marker_row, marker) in enumerate(markers):
        end = markers[pos + 1][0] if pos + 1 < len(markers) else len(raw)
        header_row = None
        for row_idx in range(marker_row + 1, end):
            if any(text(v) for v in raw.iloc[row_idx].tolist()):
                header_row = row_idx
                break
        if header_row is None:
            continue

        headers = [text(v) for v in raw.iloc[header_row].tolist()]
        used = [(i, h) for i, h in enumerate(headers) if h]
        if not used:
            continue
        records = []
        for row_idx in range(header_row + 1, end):
            record = {h: raw.iat[row_idx, i] for i, h in used}
            if any(not is_blank(v) for v in record.values()):
                records.append(record)
        blocks[PROJECT_BLOCK_MARKERS[marker]] = pd.DataFrame(
            records, columns=[h for _, h in used])

    meta, gsj = {}, {}
    metadata = blocks.get("metadata")
    if metadata is not None and not metadata.empty:
        cols = list(metadata.columns)
        kcol = "key" if "key" in cols else cols[0]
        vcol = "value" if "value" in cols else (cols[1] if len(cols) > 1 else cols[0])
        for _, row in metadata.iterrows():
            key = text(row.get(kcol))
            if not key or key.lower() == "documentation":
                continue
            value = "" if is_blank(row.get(vcol)) else row.get(vcol)
            if key.lower().startswith("gsj_"):
                gsj[key[4:]] = value
            else:
                meta[key] = value

    refs = blocks.get("refs")
    images = blocks.get("images")
    if refs is not None and "ref_id" not in refs.columns:
        refs = None
    if images is not None and "image_name" not in images.columns:
        images = None
    return meta, gsj, refs, images


def _formula_rows(path, col_name, n_rows, sheet_names=("Review", "units_review", "units")):
    """
    units_review の指定列で「数式が入っている行」を True にしたリストを返す。

    pandas はキャッシュされた計算結果しか見えないので、openpyxl で生の値を読む。
    読めなければ全部 False（=手入力扱い）にしておく。安全側に倒す。
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
        ws = next((wb[n] for n in sheet_names if n in wb.sheetnames), None)
        if ws is None:
            return [False] * n_rows
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if col_name not in header:
            return [False] * n_rows
        i = header.index(col_name)
        out = []
        for row in ws.iter_rows(min_row=2, max_row=n_rows + 1):
            v = row[i].value if i < len(row) else None
            out.append(isinstance(v, str) and v.startswith("="))
        wb.close()
        return (out + [False] * n_rows)[:n_rows]
    except Exception:
        return [False] * n_rows


def clean(v):
    return "" if is_blank(v) else (str(v).strip() if isinstance(v, str) else v)


def as_number(v):
    """数値なら float。数式文字列・空欄・文字列なら None。"""
    if is_blank(v):
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return None if v != v else float(v)   # NaN を弾く
    s = str(v).strip()
    if s.startswith("="):
        return None                            # 未評価の数式
    try:
        return float(s)
    except ValueError:
        return None


def resolve_props(df):
    """
    t_prop / b_prop を確定する。

    レビューExcel では t_prop/b_prop は VLOOKUP を使った数式になっている。
    pandas は数式を評価しないので（Excelで開いて保存していなければ
    キャッシュ値も無い）、ここで Python 側でも同じ計算をやり直す。

    優先順位:
      1. prop 欄に数値が直接入っている  -> それを尊重（手動上書き）
      2. t_age_ma / b_age_ma がある      -> interval から計算
      3. どちらも無い                    -> 空欄のまま（勝手に埋めない）

    戻り値: (件数の内訳 dict)
    """
    stats = {"manual": 0, "computed": 0, "blank": 0, "event": 0,
             "fixed_int": [], "event_names": [], "one_point_names": [],
             "failed": []}

    # ★ まず、年代に合わない t_int / b_int を直す。
    #   ZFKの時代区分は粗く（「更新世」）、age_mapping を通すと
    #   Early Pleistocene(2.58–1.8Ma) のような広い区分に落ちる。そこへ
    #   0.4 Ma という数値年代を入れると b_prop が 2.79 になり、公式仕様の
    #   「0〜1」に反した無効なデータが出てしまう（実際に出た）。
    #   数値年代のほうが確かな証拠なので、時代名をそちらに合わせる。
    for c in ("t_int", "b_int"):
        if c not in df.columns:
            continue
        icol, acol = df.columns.get_loc(c), ("t_age_ma" if c == "t_int" else "b_age_ma")
        if acol not in df.columns:
            continue
        for pos in range(len(df)):
            age = df.iloc[pos][acol]
            cur = df.iloc[pos][c]
            if is_blank(age) or fits_interval(age, cur) is not False:
                continue
            new_iv = best_interval_for_age(age, cur)
            if new_iv and str(new_iv) != str(cur):
                df.iloc[pos, icol] = new_iv
                stats["fixed_int"].append(
                    f"{df.iloc[pos].get('unit_name', '')}: {c} {cur} → {new_iv}"
                    f"（{parse_age_ma(age):g}Ma）")

    # ★ 次に「年代が1点しかない層」を処理する。
    #   火砕流・テフラ・溶岩のように年代が1点で決まる堆積は、本来 b_prop と
    #   t_prop が同じ値になる。しかし公式仕様は「b_prop must be less than
    #   t_prop」を求めるので、表示桁（小数第3位）に四捨五入すると同じ値になる
    #   最小の幅を上下端にする。0.132 → b_prop=0.1315 / t_prop=0.13249
    for c in ("t_prop", "b_prop"):
        if c not in df.columns:
            df[c] = ""
    # ★ 位置（iloc）で回す。explode したあとの index は重複しうるので
    #   df.loc[i] だと DataFrame が返ってきて壊れる（実際に ValueError が出た）。
    def cell(pos, name):
        return df.iloc[pos][name] if name in df.columns else None

    def typed(pos, name):
        """
        その prop が「尊重すべき手入力の数値」か。

        違うもの:
          - 数式のキャッシュ結果（Excelが計算した値。古い t_int のままのことがある）
          - 0〜1 の範囲外の値（公式仕様に反する無効値。過去の実行が書いた残骸）
        """
        if bool(cell(pos, f"_{name}_formula")):
            return False
        v = as_number(cell(pos, name))
        return v is not None and 0 <= v <= 1

    bcol, tcol = df.columns.get_loc("b_prop"), df.columns.get_loc("t_prop")
    for pos in range(len(df)):
        if typed(pos, "b_prop") or typed(pos, "t_prop"):
            continue                              # prop の手入力を尊重
        b_age = parse_age_ma(cell(pos, "b_age_ma"))
        t_age = parse_age_ma(cell(pos, "t_age_ma"))
        if b_age is None and t_age is None:
            continue
        if b_age is not None and t_age is not None and b_age != t_age:
            continue                              # 期間がある層はふつうに計算する

        bp, tp, is_event = props_from_ages(
            cell(pos, "unit_name"), cell(pos, "t_int"), cell(pos, "b_int"),
            cell(pos, "t_age_ma"), cell(pos, "b_age_ma"),
            cell(pos, "REF_unit_name_ja"), cell(pos, "REF_unit_name_en"))
        if bp is None or tp is None:
            continue
        df.iloc[pos, bcol] = bp
        df.iloc[pos, tcol] = tp
        # ★ このあとの「年代から計算し直す」ループに上書きされないよう、
        #   確定済みの印を付ける。付け忘れると b_prop == t_prop に戻り、
        #   公式仕様の b_prop < t_prop に反する（実際に戻された）。
        for name in ("b_prop", "t_prop"):
            c = f"_{name}_formula"
            if c in df.columns:
                df.iloc[pos, df.columns.get_loc(c)] = False
        (stats["event_names"] if is_event else stats["one_point_names"]).append(
            str(cell(pos, "unit_name") or ""))
        stats["event"] += 1

    for prop_col, int_col, age_col in (("t_prop", "t_int", "t_age_ma"),
                                       ("b_prop", "b_int", "b_age_ma")):
        if prop_col not in df.columns:
            df[prop_col] = ""
        out = []
        for i, (_, row) in enumerate(df.iterrows(), 2):
            explicit = (None if row.get(f"_{prop_col}_formula")
                        else as_number(row.get(prop_col)))
            if explicit is not None and not (0 <= explicit <= 1):
                explicit = None            # 無効値は尊重せず計算し直す
            if explicit is not None:
                stats["manual"] += 1
                out.append(explicit)
                continue

            age_raw = row.get(age_col) if age_col in df.columns else None
            if is_blank(age_raw):
                stats["blank"] += 1
                out.append("")
                continue

            interval = row.get(int_col)
            p = compute_prop(age_raw, interval)
            if p is None:
                age_ma = parse_age_ma(age_raw)
                if age_ma is None:
                    why = f"{age_col}={age_raw!r} を年代として解釈できない"
                elif interval_bounds(interval) == (None, None):
                    why = f"{int_col}={interval!r} が intervals に無い"
                else:
                    why = f"{int_col}={interval!r} の境界が不正"
                stats["failed"].append(f"{row.get('unit_name', '')or f'行{i}'}: {why}")
                out.append("")
            else:
                stats["computed"] += 1
                out.append(round(p, 6))
        df[prop_col] = out
    return stats


def restore_review_prop_previews(df, stats):
    """Use v2 reference values only when age-based submit-time calculation is impossible.

    Review v2 deliberately presents ``t_prop`` and ``b_prop`` as read-only reference
    values.  ``resolve_props`` recalculates them from numeric ages first.  Older review
    files can contain a valid derived value even when the original numeric age was not
    retained, so this final fallback preserves that information instead of silently
    turning it into a blank submission cell.
    """
    restored = 0
    for name in ("b_prop", "t_prop"):
        preview_name = f"_review_preview_{name}"
        if preview_name not in df.columns or name not in df.columns:
            continue
        # resolve_props can leave an all-blank column with pandas' strict string
        # dtype; convert it back to object before restoring numeric previews.
        df[name] = pd.Series(df[name].tolist(), index=df.index, dtype=object)
        target_col = df.columns.get_loc(name)
        for pos in range(len(df)):
            if as_number(df.iloc[pos][name]) is not None:
                continue
            preview = as_number(df.iloc[pos][preview_name])
            if preview is None or not 0 <= preview <= 1:
                continue
            df.iloc[pos, target_col] = preview
            restored += 1
    if restored:
        stats["preview_fallback"] = restored
        stats["blank"] = max(0, stats.get("blank", 0) - restored)
    return restored


def clean_frame(df):
    """pandas 2.1+ は DataFrame.map、それ以前は applymap。両方に対応する。"""
    if hasattr(pd.DataFrame, "map"):
        return df.map(clean)
    return df.applymap(clean)


# ---------------------------------------------------------------------------
# 検証
# ---------------------------------------------------------------------------

def validate(df_units, df_cols, df_refs, df_images, meta):
    """(errors, warnings) を返す。errors があれば出力しない。"""
    errors, warnings = [], []

    if df_units is None or df_units.empty:
        errors.append("units_review シートが空です。")
        return errors, warnings

    if "column_id" not in df_units.columns:
        errors.append("units_review に column_id 列がありません。")
        return errors, warnings

    valid_ids = set()
    if df_cols is not None and "col_id" in df_cols.columns:
        valid_ids = {str(c).strip() for c in df_cols["col_id"] if not is_blank(c)}
    if not valid_ids:
        errors.append("columns_review に有効な col_id がありません。")

    # column_id 参照チェック
    bad = sorted({cid for cid in df_units["_col_id"] if cid not in valid_ids})
    if bad:
        errors.append(
            f"units の column_id が columns_review の col_id と一致しません: {bad}\n"
            f"          有効な col_id: {sorted(valid_ids)}"
        )

    # unit_id の整合性。
    #
    # 同じ unit_id が複数行に現れること自体は正常。1つの地層が複数の Column に
    # またがる場合、Column ごとに行が展開されるため。
    # 異常なのは次の2つ。
    #   (1) 同じ unit_id に別の地層名が付いている
    #       → 別々の地層が1つのIDを共有している。記載文や年代が入れ替わる
    #   (2) 同じ (unit_id, column_id) の組が2行以上ある
    #       → 同一 Column 内の重複行
    #
    # どちらも空欄チェックには掛からない。見た目は埋まっているので気づけない。
    if "unit_id" in df_units.columns and "unit_name" in df_units.columns:
        names_by_id: dict[str, list[str]] = {}
        pairs: dict[tuple[str, str], int] = {}
        for _i, row in df_units.iterrows():
            uid = str(row.get("unit_id") or "").strip()
            if not uid:
                continue
            name = str(row.get("unit_name") or "").strip()
            if name and name not in names_by_id.setdefault(uid, []):
                names_by_id[uid].append(name)
            col = str(row.get("_col_id") if "_col_id" in df_units.columns
                      else row.get("column_id") or "").strip()
            pairs[(uid, col)] = pairs.get((uid, col), 0) + 1

        collided = {uid: names for uid, names in names_by_id.items() if len(names) > 1}
        if collided:
            lines = [f"別々の地層が同じ unit_id を共有しています（{len(collided)} 件）。"
                     f"記載文や年代が入れ替わります。"]
            for uid in sorted(collided)[:10]:
                lines.append(f"          {uid}: " + " / ".join(collided[uid]))
            if len(collided) > 10:
                lines.append(f"          ... 他 {len(collided) - 10} 件")
            errors.append("\n".join(lines))

        repeated = {key: n for key, n in pairs.items() if n > 1}
        if repeated:
            lines = [f"同じ unit_id が同じ Column に複数回現れます（{len(repeated)} 件）。"]
            for (uid, col) in sorted(repeated)[:10]:
                lines.append(f"          {uid} / column {col}: {repeated[(uid, col)]} 行")
            errors.append("\n".join(lines))

    # Column 未割当のまま提出させない。
    #
    # unassigned は「層序図から所属 Column を判定できなかった」ユニットの置き場で、
    # 地理的な実体ではない。Excel 上では status と comments で警告しているが、
    # 見落とされたまま提出されると意味のない Column が Macrostrat に入る。
    unassigned = [
        str(row.get("unit_name") or "(名称なし)").strip()
        for _i, row in df_units.iterrows()
        if str(row.get("_col_id") or "").strip().casefold() == "unassigned"
    ]
    if unassigned:
        lines = [f"Column が未割当のユニットが {len(unassigned)} 件あります。"
                 f"提出前に正しい Column へ割り当ててください。"]
        for name in unassigned[:10]:
            lines.append(f"          {name}")
        if len(unassigned) > 10:
            lines.append(f"          ... 他 {len(unassigned) - 10} 件")
        errors.append("\n".join(lines))

    # 実質的に必要なのは unit_name だけ
    for f in REQUIRED_UNIT_FIELDS:
        if f not in df_units.columns:
            warnings.append(f"units_review に {f} 列がありません。")
            continue
        n = int(df_units[f].apply(is_blank).sum())
        if n:
            warnings.append(f"{f} が未入力の行が {n} 件あります。")

    # 任意項目は「お知らせ」。空欄でも提出できるので警告には数えない。
    notes = []
    for f in OPTIONAL_REPORTED_FIELDS:
        if f in df_units.columns:
            n = int(df_units[f].apply(is_blank).sum())
            if n:
                notes.append(f"{f} 未入力 {n}")
    if notes:
        print("  未入力（任意項目・空欄のままで提出できます）: " + " / ".join(notes))

    # t_int と b_int は片方あればよい（公式仕様）。両方空のときだけ知らせる。
    if all(f in df_units.columns for f in REQUIRED_CHRONO_FIELDS):
        n = int((df_units["t_int"].apply(is_blank) & df_units["b_int"].apply(is_blank)).sum())
        if n:
            warnings.append(f"t_int と b_int が両方とも未入力の行が {n} 件あります"
                            f"（片方だけでも入っていれば年代モデルを組めます）。")

    # sort_order
    if "sort_order" in df_units.columns:
        n_blank = int(df_units["sort_order"].apply(is_blank).sum())
        if n_blank:
            warnings.append(f"sort_order が未入力の行が {n_blank} 件あります（行の並び順で代用します）。")
        for cid, grp in df_units.groupby("_col_id"):
            # None / NaN / 非整数 を除いてから欠番を調べる
            vals = []
            for v in grp["_sort"]:
                if v is None or (isinstance(v, float) and v != v):   # NaN 判定
                    continue
                if float(v).is_integer():
                    vals.append(int(v))
            if not vals:
                continue
            missing = sorted(set(range(1, max(vals) + 1)) - set(vals))
            if missing:
                warnings.append(f"col '{cid}': sort_order に欠番があります -> {missing}")
            dupes = sorted({v for v in vals if vals.count(v) > 1})
            if dupes:
                warnings.append(
                    f"col '{cid}': sort_order が重複しています -> {dupes} "
                    f"（意図的な「重なり合うユニット」ならこの警告は無視して構いません）"
                )

    # prop 範囲
    for f in ("b_prop", "t_prop"):
        if f not in df_units.columns:
            continue
        for i, v in enumerate(df_units[f], 2):
            if is_blank(v):
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                warnings.append(f"行{i}: {f} が数値ではありません ({v!r})")
                continue
            if not (0 <= fv <= 1):
                int_col = "b_int" if f == "b_prop" else "t_int"
                iv = df_units.iloc[i - 2].get(int_col)
                raw_age = parse_age_ma(df_units.iloc[i - 2].get(
                    "b_age_ma" if f == "b_prop" else "t_age_ma"))
                age = raw_age if raw_age is not None else age_from_prop(fv, iv)
                extra = ""
                if age is not None:
                    extra = f" - 年代 {age:g}Ma が {iv} の範囲外です"
                    # ★ どの時代に入れ直せばいいかまで出す。
                    #   ZFKの「更新世」が Late Pleistocene に落ちるなど、
                    #   age_mapping の粗さが原因のことが多い。
                    fit = intervals_containing(age)
                    if fit:
                        extra += f"。{' / '.join(fit[:3])} が該当します"
                warnings.append(f"行{i}: {f} が 0〜1 の範囲外です ({fv:.4f}){extra}")

    # b_prop < t_prop（同じ interval の場合のみ。公式仕様の制約）
    if all(c in df_units.columns for c in ("b_prop", "t_prop", "b_int", "t_int")):
        for i, (_, r) in enumerate(df_units.iterrows(), 2):
            bp, tp = as_number(r.get("b_prop")), as_number(r.get("t_prop"))
            if bp is None or tp is None:
                continue
            if str(r.get("b_int")).strip() == str(r.get("t_int")).strip() and bp >= tp:
                warnings.append(
                    f"行{i} '{clean(r.get('unit_name'))}': 同じ時代の中で b_prop({bp:.4f}) "
                    f">= t_prop({tp:.4f})。下限は上限より古い必要があります。"
                )

    # 年代の上下関係（b_age_ma は t_age_ma より古い＝大きい はず）
    if all(c in df_units.columns for c in ("b_age_ma", "t_age_ma")):
        for i, (_, r) in enumerate(df_units.iterrows(), 2):
            b, t = parse_age_ma(r.get("b_age_ma")), parse_age_ma(r.get("t_age_ma"))
            if b is None or t is None:
                continue
            if b < t:
                warnings.append(
                    f"行{i} '{clean(r.get('unit_name'))}': b_age_ma({b:g}Ma) が "
                    f"t_age_ma({t:g}Ma) より新しくなっています。入れ違いではありませんか。"
                )

    # Macrostrat 公式語彙との照合
    # ★ 一致しなくてもエラーにはしない。公式仕様の environment は
    #   "free text ... or Macrostrat environment" と自由記述を認めており、
    #   仕様の例に出てくる "shallow marine" 自体が公式表に無い。
    #   行ごとに出すと騒がしいので、語ごとに集約して1件の警告にまとめる。
    from common import check_vocab, load_vocab
    if load_vocab().get("environment"):
        for field, kind in (("environment", "environment"),
                            ("lithology", "lithology"),
                            ("minor_lith", "lithology")):
            if field not in df_units.columns:
                continue
            unmatched = {}
            for _, r in df_units.iterrows():
                for t in check_vocab(r.get(field), kind)[1]:
                    unmatched[t] = unmatched.get(t, 0) + 1
            if unmatched:
                shown = sorted(unmatched, key=lambda t: (-unmatched[t], t))[:8]
                more = f" 他{len(unmatched) - len(shown)}語" if len(unmatched) > len(shown) else ""
                warnings.append(
                    f"{field}: Macrostrat公式表に無い語が {len(unmatched)} 種 - "
                    + ", ".join(f"'{t}'×{unmatched[t]}" for t in shown) + more
                    + "（自由記述は仕様上許容。意図的ならそのままで構いません）")

    # columns の位置情報
    if df_cols is not None:
        for _, r in df_cols.iterrows():
            cid = clean(r.get("col_id"))
            if not cid:
                continue
            has_geom = not is_blank(r.get("geom"))
            has_ll = not is_blank(r.get("lat")) and not is_blank(r.get("lng"))
            if not has_geom and not has_ll:
                errors.append(f"col '{cid}': geom または lat/lng のどちらかが必須です。")

    # refs
    if df_refs is not None and not df_refs.empty:
        for _, r in df_refs.iterrows():
            if is_blank(r.get("ref_id")):
                continue
            if is_blank(r.get("title")):
                warnings.append(f"ref '{clean(r.get('ref_id'))}': title が空です（英語タイトルを手入力してください）。")

    # metadata
    if is_blank(meta.get("project_name")):
        warnings.append("project_meta: project_name が空です。")
    if is_blank(meta.get("compiler_name")) and is_blank(meta.get("compiler_orcid")):
        warnings.append("project_meta: compiler_name または compiler_orcid のどちらかが必要です。")

    return errors, warnings


# ---------------------------------------------------------------------------
# 変換
# ---------------------------------------------------------------------------

def export_to_macrostrat(input_excel, check_only=False, out_path=None):
    print(f"読み込み: {input_excel}")
    if not os.path.exists(input_excel):
        print("[ERROR] ファイルが見つかりません。")
        return None

    is_review_v2 = has_sheet(input_excel, "Review")

    # Review v2 を優先し、旧 review / submission 名にも後方互換する。
    df_units = read_sheet(input_excel, "Review", ("units_review", "units"))

    # ★ prop が「手入力の数値」か「Excelが計算した数式の結果」かを見分ける。
    #   pandas は数式ではなくキャッシュされた計算結果を返すので、区別できない。
    #   区別しないと、古い t_int で計算された値を「手入力だから尊重」して
    #   そのまま出してしまう（実際に b_prop=2.79 という無効値が残った）。
    if df_units is not None and not df_units.empty:
        for c in ("t_prop", "b_prop"):
            df_units[f"_{c}_formula"] = _formula_rows(input_excel, c, len(df_units))
    df_cols = read_sheet(input_excel, "Columns", ("columns_review", "columns", "cols"))
    df_refs = read_sheet(input_excel, "refs_review", ("refs",))
    df_images = read_sheet(input_excel, "images_review", ("images",))
    meta = read_kv_sheet(input_excel, "project_meta")
    gsj = read_kv_sheet(input_excel, "gsj_meta")

    # Review v2 では metadata / refs / images を Project 1枚にまとめる。
    # 旧個別シートがある場合はそちらを優先し、欠けた要素だけ
    # Project から補う。移行期のワークブックでも値を変えないため。
    project_meta, project_gsj, project_refs, project_images = read_project_blocks(input_excel)
    meta = {**project_meta, **meta}
    gsj = {**project_gsj, **gsj}
    if df_refs is None:
        df_refs = project_refs
    if df_images is None:
        df_images = project_images

    if df_units is None:
        print("[ERROR] Review / units_review シートが読めません。")
        return None
    df_units = df_units.dropna(how="all").reset_index(drop=True)

    # Review v2 の自動列は確認用であり、提出値の手入力欄ではない。
    # prop は年代から再計算し、年代が残っていない旧データだけ後段で
    # valid な確認値へフォールバックする。section_id / t_pos は Column を
    # 展開した後に必ず再導出する（position は従来から常に再導出）。
    if is_review_v2:
        for name in ("b_prop", "t_prop"):
            if name in df_units.columns:
                df_units[f"_review_preview_{name}"] = df_units[name].copy()
                df_units[name] = ""
        for name in ("section_id", "t_pos"):
            if name in df_units.columns:
                df_units[name] = ""

    # --- 旧フォーマットのレビューファイルへの後方互換 ---
    if "unit_name" not in df_units.columns:
        alt = next((c for c in df_units.columns if c.replace(" ", "_").lower() == "unit_name"), None)
        if alt:
            print(f"  [compat] 列 '{alt}' を unit_name として扱います。")
            df_units["unit_name"] = df_units[alt]
        elif "REF_unit_name_en" in df_units.columns:
            print("  [compat] unit_name 列がないため REF_unit_name_en から (岩相) を除いて生成します。")
            df_units["unit_name"] = df_units["REF_unit_name_en"].apply(strip_trailing_paren)
    # 個別セルの空欄も REF から補完
    if "REF_unit_name_en" in df_units.columns:
        df_units["unit_name"] = [
            strip_trailing_paren(ref) if is_blank(un) else un
            for un, ref in zip(df_units["unit_name"], df_units["REF_unit_name_en"])
        ]
    if "environment" not in df_units.columns and "env" in df_units.columns:
        df_units["environment"] = df_units["env"]
    if "min_thickness" not in df_units.columns and "min_thick" in df_units.columns:
        df_units["min_thickness"] = df_units["min_thick"]
    if "max_thickness" not in df_units.columns and "max_thick" in df_units.columns:
        df_units["max_thickness"] = df_units["max_thick"]
    if "comments" not in df_units.columns and "notes" in df_units.columns:
        df_units["comments"] = df_units["notes"]

    # --- column_id の展開（"a, b" -> 2行） ---
    df_units["_col_raw"] = df_units["column_id"].apply(
        lambda v: [s.strip() for s in str(v).split(",") if s.strip()] if not is_blank(v) else []
    )
    n_dropped = int((df_units["_col_raw"].apply(len) == 0).sum())
    if n_dropped:
        print(f"  [notice] column_id が空の行を {n_dropped} 件スキップします。")
    df_units = df_units[df_units["_col_raw"].apply(len) > 0].copy()

    # ★ area（Column）ごとに値が違う列を、行を複製する前に切り分ける。
    #   column_id が "1, 2" の行で min_thickness が "10, 20" なら
    #   Column1 に 10、Column2 に 20 を割り当てる。
    #   個数が一致しない場合は分解せず、同じ値を全Columnに使う（lithology の
    #   "gravel, sand" のようにカンマが本来の区切りである列を壊さないため）。
    # explode の対象は全列（行を複製するため）。ただし **カンマで分解してよいのは
    # ホワイトリストの列だけ**。散文（unit_description）まで分解すると、
    # 「, 」の数がたまたま Column 数と一致したときに文が真っ二つに割れる。
    per_col_fields = [c for c in df_units.columns
                      if not c.startswith(("_", "REF_")) and c != "column_id"]
    splittable = set(PER_COLUMN_SPLIT_FIELDS) - set(NEVER_SPLIT_FIELDS)
    split_log = {}
    counts = [len(v) for v in df_units["_col_raw"]]

    # ★ 列ごとにリストを組み立ててから、列まるごと差し替える。
    #   セル単位の df.at[idx, col] = [...] は、その列が文字列dtype
    #   （pandas が str 型と推論した列）だと
    #   "setting an array element with a sequence" で落ちる。
    #   pandas のバージョンや中身によって dtype が変わるので、
    #   セルにリストを入れる書き方はしないこと。
    for f in per_col_fields:
        col, hit = [], 0
        for val, n in zip(df_units[f].tolist(), counts):
            if n < 2 or f not in splittable:
                col.append([val] * max(n, 1))     # 分解せず、同じ値を全Columnへ
                continue
            parts, did = split_per_column(val, n)
            if did:
                col.append(parts)
                hit += 1
            else:
                col.append([val] * n)
        if hit:
            split_log[f] = hit
        df_units[f] = pd.Series(col, index=df_units.index, dtype=object)

    if split_log:
        detail = " / ".join(f"{k} {v}行" for k, v in sorted(split_log.items()))
        print(f"  area別の値に分解: {detail}")
        risky = sorted(set(split_log) & set(COMMA_AMBIGUOUS_FIELDS))
        if risky:
            print(f"  [warn] {', '.join(risky)} は本来カンマで複数値を並べる列です。")
            print("         area別に分けたのでなければ、区切りを ';' に変えてください。")
            print("         例: lithology は 'gravel; sand' と書くと1つのColumnの中の"
                  "2岩相として扱われます。")

    # ★ explode すると index が重複する。そのままだと df.at[i, col] が
    #   1セルではなく複数行を指してしまうので、必ず振り直す。
    # ★ あわせて object 型に揃える。pandas が「この列は文字列」と推論した列に
    #   数値（section_id / t_pos）を入れようとすると例外になるため。
    df_units = df_units.explode(["_col_raw"] + per_col_fields)
    df_units = df_units.reset_index(drop=True).astype(object)
    df_units["_col_id"] = df_units["_col_raw"].astype(str).str.strip()

    def to_num(v):
        if is_blank(v):
            return None
        try:
            f = float(v)
        except (TypeError, ValueError):
            return None
        return None if f != f else f   # NaN を弾く

    # .apply だと全 None の列が float64 NaN 化してしまうため、object 型のまま組み立てる
    if "sort_order" in df_units.columns:
        sorts = [to_num(v) for v in df_units["sort_order"]]
    else:
        sorts = [None] * len(df_units)
    df_units["_sort"] = pd.Series(sorts, index=df_units.index, dtype=object)

    # ★ t_prop / b_prop を確定（Excelの数式は pandas では評価されないので再計算）
    prop_stats = resolve_props(df_units)
    if is_review_v2:
        restore_review_prop_previews(df_units, prop_stats)
    if prop_stats["computed"] or prop_stats["manual"] or prop_stats.get("preview_fallback"):
        print(f"  prop: 年代から計算 {prop_stats['computed']} 件 / "
              f"手入力を尊重 {prop_stats['manual']} 件 / 未入力 {prop_stats['blank']} 件"
              + (f" / 参照値fallback {prop_stats['preview_fallback']} 件"
                 if prop_stats.get("preview_fallback") else "")
              + (f" / 年代が1点 {prop_stats['event']} 件（表示桁で同値になる幅を使用）"
                 if prop_stats.get("event") else ""))
    if prop_stats.get("event_names"):
        print(f"    噴火など瞬間的な堆積 {len(prop_stats['event_names'])} 件: "
              + ", ".join(prop_stats["event_names"][:5])
              + (" ..." if len(prop_stats["event_names"]) > 5 else ""))
    if prop_stats.get("one_point_names"):
        print(f"    年代が片側しか分からない {len(prop_stats['one_point_names'])} 件: "
              + ", ".join(prop_stats["one_point_names"][:5])
              + (" ..." if len(prop_stats["one_point_names"]) > 5 else ""))
        print("      → 上下の年代が別々に分かるなら t_age_ma / b_age_ma に入れてください。")
    if prop_stats["fixed_int"]:
        print(f"  時代名を年代に合わせて修正 {len(prop_stats['fixed_int'])} 件"
              "（ZFKの区分が粗く、数値年代と食い違っていたため）:")
        for m in prop_stats["fixed_int"][:12]:
            print(f"    - {m}")
        if len(prop_stats["fixed_int"]) > 12:
            print(f"    ... 他 {len(prop_stats['fixed_int']) - 12} 件")
    for f in prop_stats["failed"][:10]:
        print(f"  [warn] prop を計算できません - {f}")

    errors, warnings = validate(df_units, df_cols, df_refs, df_images, meta)

    if warnings:
        print(f"\n--- 警告 {len(warnings)} 件（出力は続行します） ---")
        for w in warnings:
            print(f"  [warn] {w}")
    if errors:
        print(f"\n--- エラー {len(errors)} 件（出力を中止します） ---")
        for e in errors:
            print(f"  [ERROR] {e}")
        return False
    if not warnings:
        print("  検証OK: 問題は見つかりませんでした。")
    if check_only:
        print("\n(--check-only のため出力はしていません)")
        return True

    # --- 並び替え: columns_review の順 -> 各Column内は sort_order 昇順（上=新しい） ---
    col_order = [str(c).strip() for c in df_cols["col_id"] if not is_blank(c)]
    df_units["_col_rank"] = df_units["_col_id"].apply(
        lambda c: col_order.index(c) if c in col_order else 9999
    )
    df_units["_row"] = range(len(df_units))
    df_units["_sort_key"] = df_units["_sort"].apply(lambda v: 1e9 if v is None else v)
    df_units = df_units.sort_values(["_col_rank", "_sort_key", "_row"]).reset_index(drop=True)

    # --- position の導出（sort_order を反転: 最下位=1、古い→新しい） ---
    positions = []
    for _, grp in df_units.groupby("_col_rank", sort=True):
        sorts = list(grp["_sort"])
        if all(s is None for s in sorts):
            # sort_order が全く無い Column は行順から生成（上が新しい前提）
            n = len(sorts)
            positions.extend([(idx, n - i) for i, idx in enumerate(grp.index)])
        else:
            filled = [s if s is not None else (i + 1) for i, s in enumerate(sorts)]
            for idx, p in zip(grp.index, derive_positions(filled)):
                positions.append((idx, p))
    pos_map = dict(positions)
    df_units["_position"] = [pos_map.get(i) for i in df_units.index]

    # --- section_id / t_pos を補う（手入力があればそちらを優先） ---
    auto_sec = auto_tpos = 0
    for _, grp in df_units.groupby("_col_rank", sort=True):
        idxs = list(grp.index)                       # 上（新しい）から下（古い）の順

        # section: 年代のすき間で区切る（公式仕様の "inferred from gaps"）
        bounds = [(parse_age_ma(df_units.at[i, "b_age_ma"])
                   if "b_age_ma" in df_units.columns else None,
                   parse_age_ma(df_units.at[i, "t_age_ma"])
                   if "t_age_ma" in df_units.columns else None) for i in idxs]
        for i, sec in zip(idxs, derive_sections(bounds)):
            if sec is None:
                continue
            if "section_id" not in df_units.columns or is_blank(df_units.at[i, "section_id"]):
                df_units.at[i, "section_id"] = sec
                auto_sec += 1

        # t_pos: 各Columnの最上位の層には必ず入れる（無いと取り込み時に落ちる）。
        #        position が重なっている層にも上端を明示する。
        for i, tp in zip(idxs, derive_t_pos([df_units.at[i, "_position"] for i in idxs])):
            if tp is None:
                continue
            if "t_pos" not in df_units.columns or is_blank(df_units.at[i, "t_pos"]):
                df_units.at[i, "t_pos"] = tp
                auto_tpos += 1

    # section が1つしかない Column では section_id は不要（仕様上も省略可）
    for _, grp in df_units.groupby("_col_rank", sort=True):
        if "section_id" in df_units.columns and grp["section_id"].nunique() <= 1:
            for i in grp.index:
                df_units.at[i, "section_id"] = ""
    if auto_sec or auto_tpos:
        n_sec = int(df_units["section_id"].apply(lambda v: not is_blank(v)).sum()) \
            if "section_id" in df_units.columns else 0
        print(f"  自動計算: section_id {n_sec} 行 / t_pos {auto_tpos} 行"
              "（section_id は年代のすき間から / t_pos は各Columnの最上位と重なり）")

    # --- units シートの構築 ---
    out = pd.DataFrame(columns=SUBMISSION_UNIT_COLS)
    n = len(df_units)
    out["unit_id"] = list(range(1, n + 1))          # 全体で一意な連番（Ichinohe と同じ流儀）
    out["col_id"] = list(df_units["_col_id"])
    out["position"] = list(df_units["_position"])

    direct = ["section_id", "b_int", "b_prop", "t_int", "t_prop", "unit_name",
              "strat_name", "environment", "unit_description", "lithology",
              "minor_lith", "min_thickness", "max_thickness", "basal_surface",
              "lateral_relationship", "comments", "t_pos"]
    for f in direct:
        out[f] = list(df_units[f]) if f in df_units.columns else ""

    out["source_unit_id"] = list(df_units["unit_id"]) if "unit_id" in df_units.columns else ""
    out["source_unit_name_ja"] = (
        list(df_units["REF_unit_name_ja"]) if "REF_unit_name_ja" in df_units.columns else ""
    )
    out = clean_frame(out)

    # --- columns シート ---
    cols_out = pd.DataFrame(columns=SUBMISSION_COLUMN_COLS)
    if df_cols is not None and not df_cols.empty:
        src = df_cols[df_cols["col_id"].apply(lambda v: not is_blank(v))].copy()
        for f in SUBMISSION_COLUMN_COLS:
            cols_out[f] = list(src[f]) if f in src.columns else [""] * len(src)
        # geom が空なら lat/lng から "lng, lat" を組み立てる（Ichinohe と同じ表記）
        geoms = []
        for i, (_, r) in enumerate(src.iterrows()):
            g = r.get("geom")
            if is_blank(g) and not is_blank(r.get("lat")) and not is_blank(r.get("lng")):
                g = f"{clean(r.get('lng'))}, {clean(r.get('lat'))}"
            geoms.append(clean(g))
        cols_out["geom"] = geoms
        cols_out = clean_frame(cols_out)

    # --- refs / images ---
    refs_out = pd.DataFrame(columns=SUBMISSION_REF_COLS)
    if df_refs is not None and not df_refs.empty:
        src = df_refs[df_refs["ref_id"].apply(lambda v: not is_blank(v))].copy()
        for f in SUBMISSION_REF_COLS:
            refs_out[f] = list(src[f]) if f in src.columns else [""] * len(src)
        refs_out = clean_frame(refs_out)

    images_out = pd.DataFrame(columns=SUBMISSION_IMAGE_COLS)
    if df_images is not None and not df_images.empty:
        src = df_images[df_images["image_name"].apply(lambda v: not is_blank(v))].copy()
        for f in SUBMISSION_IMAGE_COLS:
            images_out[f] = list(src[f]) if f in src.columns else [""] * len(src)
        images_out = clean_frame(images_out)

    # --- metadata（key-value + Documentation ブロック） ---
    meta_rows = []
    for k in METADATA_KEYS:
        v = meta.get(k, "")
        if is_blank(v):
            v = METADATA_DEFAULTS.get(k, "")
        meta_rows.append({"field": k, "value": clean(v)})
    meta_rows.append({"field": "", "value": ""})
    meta_rows.append({"field": "Documentation", "value": DOC_NOTE})
    meta_rows.append({"field": "format_version", "value": FORMAT_VERSION})
    meta_rows.append({"field": "exported_at", "value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
    meta_rows.append({"field": "source_review_file", "value": os.path.basename(input_excel)})
    for k in ("map_id", "title_ja", "title_en", "sheet_code", "pub_year",
              "source_zfk_map", "source_publication", "source_doc_page"):
        if k in gsj:
            meta_rows.append({"field": f"gsj_{k}", "value": clean(gsj[k])})
    meta_out = pd.DataFrame(meta_rows)

    # --- 出力先: 02_review と同じ地域フォルダ構成を 03_submission に鏡写しする ---
    if out_path is None:
        out_path = _mirror_output_path(input_excel, meta.get("project_name"), gsj)
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            meta_out.to_excel(writer, sheet_name="metadata", index=False, header=False)
            out.to_excel(writer, sheet_name="units", index=False)
            cols_out.to_excel(writer, sheet_name="columns", index=False)
            refs_out.to_excel(writer, sheet_name="refs", index=False)
            images_out.to_excel(writer, sheet_name="images", index=False)
            _format(writer, meta_out, out, cols_out, refs_out, images_out)
    except PermissionError:
        print(f"\n[ERROR] 保存できません: {out_path}")
        print("        Excel でこのファイルを開いていませんか？ 閉じてから再実行してください。")
        return None

    print("\n=== 出力成功 ===")
    print(f"  {out_path}")
    print(f"  units {len(out)} 行 / columns {len(cols_out)} 件 / refs {len(refs_out)} 件 / images {len(images_out)} 件")
    for cid in cols_out["col_id"]:
        sub = out[out["col_id"] == cid]
        if len(sub):
            print(f"    - {cid}: {len(sub)} 層 (position {sub['position'].min()}-{sub['position'].max()})")
    return out_path


def _mirror_output_path(input_excel, project_name, gsj):
    """data/02_review/<地域>/<図幅>/x_review.xlsx -> data/03_submission/<地域>/<図幅>/<name>.xlsx"""
    abs_in = os.path.abspath(input_excel)
    parts = abs_in.replace("\\", "/").split("/")
    if "02_review" in parts:
        i = parts.index("02_review")
        rel = parts[i + 1:-1]
        root = "/".join(parts[:i])
        out_dir = os.path.join(root, "03_submission", *rel)
    else:
        out_dir = os.path.join("data", "50k", "03_submission")

    # project_name > GSJ英名 > レビューファイル名(m1050) の順でフォールバック
    stem = os.path.basename(input_excel).replace("_review.xlsx", "").replace(".xlsx", "")
    base = project_name or gsj.get("title_en") or stem or "column"
    safe = "".join(c for c in str(base) if c.isalnum() or c in " -_").strip().replace(" ", "_")
    return os.path.join(out_dir, f"{safe}_Composite_column.xlsx")


def _format(writer, meta_out, units, cols, refs, images):
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    head = PatternFill("solid", start_color="D9EAD3")
    bold = Font(bold=True)
    wrap = Alignment(wrapText=True, vertical="top")

    ws = writer.sheets["metadata"]
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 70
    for r in range(1, len(meta_out) + 1):
        ws.cell(row=r, column=1).font = bold
        ws.cell(row=r, column=2).alignment = wrap

    wide = {"unit_name": 34, "strat_name": 34, "unit_description": 60, "lithology": 28,
            "minor_lith": 24, "comments": 30, "environment": 20, "col_name": 36,
            "title": 40, "authors": 45, "url": 45, "description": 55, "image_name": 34,
            "col_ids": 34, "source_unit_name_ja": 30, "geom": 26, "organization": 32}
    for sheet, df in [("units", units), ("columns", cols), ("refs", refs), ("images", images)]:
        w = writer.sheets[sheet]
        w.freeze_panes = "A2"
        for i, name in enumerate(df.columns, 1):
            c = w.cell(row=1, column=i)
            c.fill = head
            c.font = bold
            w.column_dimensions[get_column_letter(i)].width = wide.get(name, 15)
            for r in range(2, len(df) + 2):
                w.cell(row=r, column=i).alignment = wrap


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="レビューExcelをMacrostrat提出形式へ変換")
    p.add_argument("input_excel", type=str, help="*_review.xlsx のパス")
    p.add_argument("--check-only", action="store_true", help="検証のみ実行（ファイルを書き出さない）")
    p.add_argument("--out", type=str, default=None, help="出力パスを明示指定")
    a = p.parse_args()
    r = export_to_macrostrat(a.input_excel, check_only=a.check_only, out_path=a.out)
    sys.exit(0 if r else 1)
