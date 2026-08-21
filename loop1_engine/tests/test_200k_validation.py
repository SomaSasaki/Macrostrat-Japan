# -*- coding: utf-8 -*-
"""
test_200k_validation.py — 200k バリデータの自動回帰テストスイート
- prop 計算正常値・境界値・不正値拒否
- 欠落 unit ID、重複 sort order の検出
- Column 年代包含違反の検出
- 年代重複同位層の正常判定と真の年代逆転エラー検出
"""

import os
import sys
import unittest
import tempfile
import openpyxl
import pandas as pd

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
sys.path.insert(0, BASE_DIR)

from scripts.export_200k import compute_unit_prop, check_200k_workbook
import scripts.common as common

class Test200kValidation(unittest.TestCase):
    def test_compute_unit_prop_normal(self):
        # Pleistocene (2.58 - 0.0117 Ma): 1.0 Ma の prop 計算
        # (2.58 - 1.0) / (2.58 - 0.0117) = 1.58 / 2.5683 = 0.615
        p = compute_unit_prop('Pleistocene', 1.0, is_bottom=True)
        self.assertAlmostEqual(p, 0.615, places=2)

        # 境界値: 底面 (2.58 Ma) -> 0.0
        p_b = compute_unit_prop('Pleistocene', 2.58, is_bottom=True)
        self.assertEqual(p_b, 0.0)

        # 境界値: 上面 (0.0117 Ma) -> 1.0
        p_t = compute_unit_prop('Pleistocene', 0.0117, is_bottom=False)
        self.assertEqual(p_t, 1.0)

    def _create_mock_workbook(self, cols_rows, units_rows):
        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # columns_review
        ws_cols = wb.create_sheet("columns_review")
        col_headers = ["col_id", "col_name", "col_group", "b_int", "t_int", "b_prop", "t_prop", "geom", "rgeom", "comments"]
        ws_cols.append(col_headers)
        for r in cols_rows:
            ws_cols.append(r)

        # units_review
        ws_units = wb.create_sheet("units_review")
        unit_headers = [
            "unit_id", "col_id", "sort_order", "unit_name", "strat_name",
            "t_int", "t_age_ma", "t_prop", "b_int", "b_age_ma", "b_prop",
            "lithology", "minor_lith", "environment", "unit_description",
            "min_thickness", "max_thickness", "basal_surface", "lateral_relationship",
            "comments", "t_pos"
        ]
        ws_units.append(unit_headers)
        for r in units_rows:
            ws_units.append(r)

        tf = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tf.close()
        wb.save(tf.name)
        return tf.name

    def test_missing_unit_id_rejection(self):
        cols = [["col_1", "Test Column", "GRP1", "Cretaceous", "Neogene", 0.0, 1.0, "", "", ""]]
        units = [
            ["", "col_1", 1, "Unit A", "Strat A", "Neogene", 20.0, 0.5, "Cretaceous", 100.0, 0.0, "sandstone", "", "marine", "", "", "", "unknown", "", "", ""]
        ]
        wb_path = self._create_mock_workbook(cols, units)
        try:
            errors, warns = check_200k_workbook(wb_path)
            self.assertTrue(any("unit_id が空です" in e for e in errors))
        finally:
            if os.path.exists(wb_path): os.remove(wb_path)

    def test_duplicate_sort_order_rejection(self):
        cols = [["col_1", "Test Column", "GRP1", "Cretaceous", "Neogene", 0.0, 1.0, "", "", ""]]
        units = [
            ["u_1", "col_1", 1, "Unit A", "Strat A", "Neogene", 20.0, 0.5, "Cretaceous", 100.0, 0.0, "sandstone", "", "marine", "", "", "", "unknown", "", "", ""],
            ["u_2", "col_1", 1, "Unit B", "Strat B", "Neogene", 10.0, 0.8, "Neogene", 20.0, 0.2, "mudstone", "", "marine", "", "", "", "unknown", "", "", ""]
        ]
        wb_path = self._create_mock_workbook(cols, units)
        try:
            errors, warns = check_200k_workbook(wb_path)
            self.assertTrue(any("重複した sort_order" in e for e in errors))
        finally:
            if os.path.exists(wb_path): os.remove(wb_path)

    def test_invalid_prop_range_rejection(self):
        cols = [["col_1", "Test Column", "GRP1", "Cretaceous", "Neogene", 0.0, 1.0, "", "", ""]]
        units = [
            ["u_1", "col_1", 1, "Unit A", "Strat A", "Neogene", 20.0, 1.5, "Cretaceous", 100.0, -0.2, "sandstone", "", "marine", "", "", "", "unknown", "", "", ""]
        ]
        wb_path = self._create_mock_workbook(cols, units)
        try:
            errors, warns = check_200k_workbook(wb_path)
            self.assertTrue(any("b_prop (-0.2) が 0.0〜1.0 の範囲外" in e for e in errors))
            self.assertTrue(any("t_prop (1.5) が 0.0〜1.0 の範囲外" in e for e in errors))
        finally:
            if os.path.exists(wb_path): os.remove(wb_path)

    def test_column_age_containment_violation(self):
        # Column は Neogene (23.03 - 2.58 Ma)
        cols = [["col_1", "Neogene Column", "GRP1", "Neogene", "Neogene", 0.0, 1.0, "", "", ""]]
        # Unit が Cretaceous (100 Ma) -> Column 年代範囲を逸脱
        units = [
            ["u_1", "col_1", 1, "Cretaceous Unit", "Strat A", "Cretaceous", 80.0, 0.5, "Cretaceous", 100.0, 0.0, "sandstone", "", "marine", "", "", "", "unknown", "", "", ""]
        ]
        wb_path = self._create_mock_workbook(cols, units)
        try:
            errors, warns = check_200k_workbook(wb_path)
            self.assertTrue(any("Column 'col_1' の最古年代" in e for e in errors))
        finally:
            if os.path.exists(wb_path): os.remove(wb_path)

    def test_stratigraphic_true_inversion_detection(self):
        cols = [["col_1", "Test Column", "GRP1", "Paleozoic", "Holocene", 0.0, 1.0, "", "", ""]]
        # position 1: Neogene (20 - 10 Ma)
        # position 2: Cretaceous (100 - 80 Ma) -> 下より上が完全に古い（明らかな逆転）
        units = [
            ["u_1", "col_1", 1, "Young Unit", "Strat A", "Neogene", 10.0, 0.5, "Neogene", 20.0, 0.0, "sandstone", "", "marine", "", "", "", "unknown", "", "", ""],
            ["u_2", "col_1", 2, "Old Unit", "Strat B", "Cretaceous", 80.0, 0.5, "Cretaceous", 100.0, 0.0, "mudstone", "", "marine", "", "", "", "unknown", "", "", ""]
        ]
        wb_path = self._create_mock_workbook(cols, units)
        try:
            errors, warns = check_200k_workbook(wb_path)
            self.assertTrue(any("層序逆転" in e for e in errors))
        finally:
            if os.path.exists(wb_path): os.remove(wb_path)

if __name__ == '__main__':
    unittest.main()
